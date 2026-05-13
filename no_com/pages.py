"""
pages.py
========
세 개의 메인 페이지와 옵션 다이얼로그.
비즈니스 로직은 core / excel_io 로 위임한다.
"""

from __future__ import annotations

from copy import copy as _copy_style
import logging
import os
import re
import shutil
import traceback
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from PySide6.QtCore import Qt, QPointF, QBuffer, QByteArray, QIODevice, QThread, QTimer, Signal
from PySide6.QtGui import QBrush, QColor, QImage, QKeySequence, QPixmap, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView, QApplication, QCheckBox, QComboBox, QCompleter, QDialog,
    QDialogButtonBox, QDoubleSpinBox, QFileDialog,
    QFormLayout, QFrame, QGraphicsPixmapItem, QGraphicsScene, QGridLayout, QHBoxLayout,
    QHeaderView, QInputDialog, QLabel, QLineEdit,
    QListWidget, QListWidgetItem, QMessageBox,
    QProgressDialog, QPushButton, QSizePolicy, QSpinBox, QSplitter, QTabWidget,
    QTableWidget, QTableWidgetItem, QTextEdit,
    QVBoxLayout, QWidget,
)

import excel_io
from core import (
    QuoteState, SheetName,
    ensure_dir, exe_dir, fmt_krw, fmt_qty,
    normalize_token, parse_invest_info, s, safe_filename, to_float, unique_path,
)
from widgets import (
    DraggableItemsTable, DroppableQuoteTable, PdfView,
    SignatureItem, bold_label, centered_checkbox,
    get_checkbox_from_cell, info_label, labeled_frame,
    PasswordDialog, tint_button,
)

logger = logging.getLogger("QuoteApp")


def _make_plain_table(cols: int, headers: List[str]) -> QTableWidget:
    t = QTableWidget(0, cols)
    t.setHorizontalHeaderLabels(headers)
    t.verticalHeader().setVisible(False)
    t.setEditTriggers(QTableWidget.NoEditTriggers)
    t.setSelectionBehavior(QAbstractItemView.SelectRows)
    t.setSelectionMode(QAbstractItemView.SingleSelection)
    t.setWordWrap(False)
    return t


PROCESS_CHOICES = ["", "CVD", "DIFF", "IMP", "ETCH", "METAL", "CLEAN", "GCS"]
VENDOR_CHOICES  = [
    "", "AMAT", "ASM", "AXCELIS", "EUGENETECH", "WONIK_IPS",
    "KOKUSAI", "LAM", "SEMES", "TEL", "TES", "ULVAC", "GCS",
]


RACK_REQUEST_LEFT_LABELS = [
    "접수 일자", "납품 요청", "PR NO.", "담당자", "공 정", "세부 공정", "라인", "설 비",
    "설비MODEL", "설비호기", "PUMP MODEL", "수량(CH)", "5D", "FSC", "RACK CH",
    "GATE V/V TYPE", "METAL BELLOWS", "HOT-N2 TYPE", "IMS 중계기", "DA / LEP", "INVERTER",
    "ANGLE Valve", "Option", "비 고", "간섭여부", "Interface Cable", "Dual Cable",
    "통신모듈 Cable", "3단 모니터링 Cable",
]
RACK_REQUEST_REF_NOTES = [
    "작성일", "Main설비반입일 D-28", "견적의뢰DATA", "수기입력", "견적의뢰DATA", "견적의뢰DATA",
    "견적의뢰DATA", "견적의뢰DATA", "견적의뢰DATA", "견적의뢰DATA", "견적의뢰DATA", "견적의뢰DATA",
    "견적의뢰DATA", "5D 매칭",
] + ["공정+설비+5D 매칭"] * 15
RACK_REQUEST_MANAGERS = ["김종균", "박현준", "이기웅", "정상철", "황명선"]




# ──────────────────────────────────────────────
# 백그라운드 Excel → PDF 변환 스레드
# ──────────────────────────────────────────────
class _TemplateLoaderThread(QThread):
    """STEP1용: 통합양식 Excel을 백그라운드에서 파싱."""
    done    = Signal(object)   # (rows, by_class, price_by_spec, order, cmap) 또는 Exception
    def __init__(self, path: str, parent=None) -> None:
        super().__init__(parent)
        self.path = path
    def run(self) -> None:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.path, data_only=False)
            rows, by_class, price_by_spec, order = excel_io.parse_items_sheet(wb[SheetName.ITEMS])
            cmap = excel_io.parse_code_map_sheet(wb[SheetName.CODE_MAP])
            self.done.emit((rows, by_class, price_by_spec, order, cmap))
        except Exception as e:
            self.done.emit(e)


class _GenQuoteThread(QThread):
    """견적서 생성을 백그라운드에서 실행."""
    progress = Signal(str)          # 로그 메시지
    done     = Signal(object)       # List[(rd, path)] 또는 Exception

    def __init__(self, state, qtype: str, items, rds: list, parent=None) -> None:
        super().__init__(parent)
        self.state = state; self.qtype = qtype
        self.items = items; self.rds   = rds

    def run(self) -> None:
        try:
            if self.qtype == "국내" and len(self.rds) >= 2:
                def _cb(done, total, name):
                    self.progress.emit(f"[{done}/{total}] 저장: {name}")
                results = excel_io.generate_quote_multi(
                    self.state, self.items, self.rds, progress_cb=_cb)
            else:
                rd   = self.rds[0] if self.rds else {}
                path = excel_io.generate_quote(self.state, self.qtype, self.items, rd)
                self.progress.emit(f"저장: {os.path.basename(path)}")
                results = [(rd, path)]
            self.done.emit(results)
        except Exception as e:
            self.done.emit(e)


class _GenCoverThread(QThread):
    """갑지 생성을 백그라운드에서 실행."""
    progress = Signal(str)   # 로그 메시지
    done     = Signal(object)  # 저장 경로(str) 또는 Exception

    def __init__(self, template_path: str, folder: str, paths: list,
                 investor_name: str, parent=None) -> None:
        super().__init__(parent)
        self.template_path = template_path; self.folder = folder
        self.paths = paths; self.investor_name = investor_name

    def run(self) -> None:
        try:
            def _cb(done, total, name):
                self.progress.emit(f"[{done}/{total}] 읽는 중: {name}")
            out = excel_io.generate_cover(
                self.template_path, self.folder, self.paths,
                investor_name=self.investor_name, progress_cb=_cb)
            self.done.emit(out)
        except Exception as e:
            self.done.emit(e)


class _BgWorker(QThread):
    """범용 백그라운드 작업 스레드. UI 블로킹 없이 Excel 작업 실행."""
    result = Signal(object)
    error  = Signal(str)

    def __init__(self, fn, *args, parent=None, **kwargs):
        super().__init__(parent)
        self._fn, self._args, self._kwargs = fn, args, kwargs

    def run(self) -> None:
        try:
            self.result.emit(self._fn(*self._args, **self._kwargs))
        except Exception as e:
            import traceback
            self.error.emit(f"{e}\n\n{traceback.format_exc()}")

    @staticmethod
    def run_with_progress(parent, msg: str, fn, *args,
                          on_result=None, on_error=None, **kwargs) -> "_BgWorker":
        """진행 다이얼로그를 띄우고 fn을 백그라운드 실행. 완료 시 콜백 호출."""
        from PySide6.QtWidgets import QProgressDialog
        dlg = QProgressDialog(msg, None, 0, 0, parent)
        dlg.setWindowModality(Qt.WindowModal)
        dlg.setMinimumDuration(0)
        dlg.show()

        worker = _BgWorker(fn, *args, parent=parent, **kwargs)

        def _on_result(r):
            dlg.close()
            if on_result:
                on_result(r)

        def _on_error(e):
            dlg.close()
            if on_error:
                on_error(e)
            else:
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.critical(parent, "오류", e)

        worker.result.connect(_on_result)
        worker.error.connect(_on_error)
        worker.finished.connect(worker.deleteLater)
        worker.start()
        return worker


class _ExcelLoaderThread(QThread):
    """전자서명 페이지용: Excel 시트를 CopyPicture로 캡처 (ExportAsFixedFormat 미사용 → RenameFile 없음)."""
    progress = Signal(int, int, str)   # (완료수, 전체수, 현재파일명)

    def __init__(self, paths: List[str], tmp_dir: str, parent=None) -> None:
        super().__init__(parent)
        self.paths      = paths
        self.tmp_dir    = tmp_dir
        self.sheet_pngs : List[List[str]] = []   # 파일별 PNG 경로 리스트
        self._cancel    = False

    def cancel(self) -> None:
        self._cancel = True

    def run(self) -> None:
        total = len(self.paths)
        try:
            com_ctx = excel_io.ExcelCOM()
            com_ctx.__enter__()
            xl_app = com_ctx.app
        except Exception as e:
            logger.error("Excel COM 초기화 실패: %s", e, exc_info=True)
            com_ctx = None
            xl_app = None

        try:
            for i, xlsx in enumerate(self.paths):
                if self._cancel:
                    break
                self.progress.emit(i, total, os.path.basename(xlsx))
                try:
                    pngs = excel_io.excel_capture_sheets_to_pngs(
                        xlsx, self.tmp_dir, i + 1, xl_app)
                    self.sheet_pngs.append(pngs)
                except Exception as e:
                    logger.error("시트 캡처 실패 (%s): %s", xlsx, e, exc_info=True)
                    self.sheet_pngs.append([])
        finally:
            if com_ctx is not None:
                try:
                    com_ctx.__exit__(None, None, None)
                except Exception:
                    pass

        self.progress.emit(len(self.sheet_pngs), total, "완료")


# ══════════════════════════════════════════════
# ══════════════════════════════════════════════
# 중국 옵션 다이얼로그
# ══════════════════════════════════════════════

class ChinaOptionsDialog(QDialog):
    def __init__(self, parent, state: QuoteState) -> None:
        super().__init__(parent)
        self.setWindowTitle("중국 옵션 입력")
        self.state = state
        self.setModal(True)
        self.setFixedWidth(380)

        cn = state.cn_info or {}
        root = QVBoxLayout(self)
        form = QFormLayout()
        self._maker   = QLineEdit(cn.get("maker", ""))
        self._process = QLineEdit(cn.get("process", ""))
        self._tool    = QLineEdit(cn.get("tool", ""))
        self._line    = QLineEdit(cn.get("line", ""))
        form.addRow("Maker",    self._maker)
        form.addRow("Process",  self._process)
        form.addRow("설비호기", self._tool)
        form.addRow("라인",     self._line)
        root.addLayout(form)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self.accept); bb.rejected.connect(self.reject)
        root.addWidget(bb)

    def apply(self) -> None:
        self.state.cn_info = {
            "maker": self._maker.text().strip(), "process": self._process.text().strip(),
            "tool":  self._tool.text().strip(),  "line":    self._line.text().strip(),
        }


# ══════════════════════════════════════════════
# 미국 옵션 다이얼로그
# ══════════════════════════════════════════════

class USOptionsDialog(QDialog):
    def __init__(self, parent, state: QuoteState) -> None:
        super().__init__(parent)
        self.setWindowTitle("미국 옵션 입력")
        self.state = state
        self.setModal(True)
        self.setFixedWidth(420)

        us = state.us_info or {}
        root = QVBoxLayout(self)
        form = QFormLayout()
        self._maker   = QLineEdit(us.get("maker", ""))
        self._process = QLineEdit(us.get("process", ""))
        self._tool    = QLineEdit(us.get("tool", ""))
        self._exchange = QDoubleSpinBox()
        self._exchange.setRange(0, 1e9); self._exchange.setDecimals(2)
        self._exchange.setValue(float(us.get("exchange", 0) or 0))
        now = datetime.now()
        y, m = us.get("base_ym", (now.year, now.month))
        self._year  = QSpinBox(); self._year.setRange(2000, 2100); self._year.setValue(int(y))
        self._month = QSpinBox(); self._month.setRange(1, 12);      self._month.setValue(int(m))
        self._site  = QComboBox(); self._site.addItems(["Taylor", "Austin"])
        site = us.get("site", "Taylor")
        self._site.setCurrentText(site if site in ("Taylor", "Austin") else "Taylor")
        ym_row = QWidget()
        ym_h = QHBoxLayout(ym_row); ym_h.setContentsMargins(0,0,0,0)
        ym_h.addWidget(self._year); ym_h.addWidget(QLabel("년"))
        ym_h.addWidget(self._month); ym_h.addWidget(QLabel("월"))
        ym_h.addStretch(1)
        form.addRow("Maker",         self._maker)
        form.addRow("Process",       self._process)
        form.addRow("설비호기",      self._tool)
        form.addRow("환율",          self._exchange)
        form.addRow("기준일(년/월)", ym_row)
        form.addRow("Site",          self._site)
        root.addLayout(form)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self.accept); bb.rejected.connect(self.reject)
        root.addWidget(bb)

    def apply(self) -> None:
        self.state.us_info = {
            "maker": self._maker.text().strip(), "process": self._process.text().strip(),
            "tool":  self._tool.text().strip(),  "exchange": float(self._exchange.value()),
            "base_ym": (int(self._year.value()), int(self._month.value())),
            "site": self._site.currentText(),
        }


# ══════════════════════════════════════════════
# STEP5 테이블 관리 믹스인
# ══════════════════════════════════════════════

class Step5Manager:
    """
    STEP5 테이블의 행 CRUD + TOTAL/CREDIT 행 관리.
    QuoteBuilderPage 가 다중 상속하여 사용한다.

    하위 클래스가 반드시 제공해야 하는 속성:
        self.step5_table      : DroppableQuoteTable
        self.chk_step5_all    : QCheckBox
        self.state            : QuoteState
        self._step4_highlight : set
    하위 클래스가 구현해야 하는 메서드:
        self._render_items_table(scroll_top: bool)
    """

    # ── 행 종류 판별 ────────────────────────────
    def _row_role(self, row: int) -> str:
        it = self.step5_table.item(row, 1)
        return s(it.data(Qt.UserRole)) if it else ""

    def _is_total(self, row: int)  -> bool: return self._row_role(row) == "TOTAL"
    def _is_credit(self, row: int) -> bool: return self._row_role(row) == "CREDIT"
    def _is_item(self, row: int)   -> bool: return self._row_role(row) == "ITEM"

    # ── 셀 setter ───────────────────────────────
    def _set_qty(self, row: int, qty: float, blank: bool = False) -> None:
        it = QTableWidgetItem("" if blank else fmt_qty(qty))
        it.setData(Qt.UserRole, qty); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(row, 3, it)

    def _set_price(self, row: int, price: float, blank: bool = False) -> None:
        it = QTableWidgetItem("" if blank else fmt_krw(price))
        it.setData(Qt.UserRole, price); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(row, 4, it)

    def _set_amt(self, row: int, amt: float) -> None:
        it = QTableWidgetItem(fmt_krw(amt))
        it.setData(Qt.UserRole, amt); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(row, 5, it)

    def _get_float(self, row: int, col: int) -> float:
        it = self.step5_table.item(row, col)
        return float(it.data(Qt.UserRole) or 0.0) if it else 0.0

    # ── TOTAL 행 보장 ────────────────────────────
    def _ensure_total(self) -> None:
        n = self.step5_table.rowCount()
        if n == 0 or not self._is_total(n - 1):
            self._append_total()

    def _append_total(self) -> None:
        r = self.step5_table.rowCount()
        self.step5_table.insertRow(r)
        it_cat = QTableWidgetItem("TOTAL")
        it_cat.setData(Qt.UserRole, "TOTAL"); it_cat.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(r, 1, it_cat)
        it_spec = QTableWidgetItem("총 합계")
        it_spec.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(r, 2, it_spec)
        self._set_qty(r, 0, blank=True); self._set_price(r, 0, blank=True); self._set_amt(r, 0)

    # ── 합계 재계산 ──────────────────────────────
    def _recalc_totals(self) -> None:
        self._ensure_total()
        total_qty = total_amt = 0.0
        last = self.step5_table.rowCount() - 1
        for r in range(last):
            if self._is_total(r): continue
            total_amt += self._get_float(r, 5)
            if not self._is_credit(r):
                total_qty += self._get_float(r, 3)
        self._set_qty(last, total_qty, blank=True)
        self._set_amt(last, total_amt)

    def _recalc_row(self, row: int) -> None:
        self._set_amt(row, self._get_float(row, 3) * self._get_float(row, 4))

    # ── 행 추가 ─────────────────────────────────
    def _add_row(self, category: str, spec: str, qty: float, unit_price: float) -> None:
        """같은 spec 이면 수량 누적, 없으면 TOTAL 행 위에 새 행 삽입."""
        self._ensure_total()
        last = self.step5_table.rowCount() - 1
        for r in range(last):
            if self._is_total(r) or self._is_credit(r): continue
            it = self.step5_table.item(r, 2)
            if it and it.text().strip() == spec.strip():
                self._set_qty(r, self._get_float(r, 3) + qty)
                self._recalc_row(r); self._recalc_totals(); return

        ins = last
        self.step5_table.insertRow(ins)
        cb = QCheckBox()
        cb.stateChanged.connect(self._on_step5_check_changed)
        self.step5_table.setCellWidget(ins, 0, cb)
        it_cat = QTableWidgetItem(category)
        it_cat.setData(Qt.UserRole, "ITEM"); it_cat.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(ins, 1, it_cat)
        it_spec = QTableWidgetItem(spec)
        it_spec.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(ins, 2, it_spec)
        self._set_qty(ins, qty); self._set_price(ins, unit_price)
        self._set_amt(ins, qty * unit_price); self._recalc_totals()
        self._sort_step5_items()

    # ── Credit 행 ────────────────────────────────
    def _clear_credit_rows(self) -> None:
        for r in range(self.step5_table.rowCount() - 1, -1, -1):
            if self._is_credit(r): self.step5_table.removeRow(r)
        self._ensure_total()

    def _add_credit_row(self, label: str, amount: float) -> None:
        self._ensure_total()
        ins = self.step5_table.rowCount() - 1
        self.step5_table.insertRow(ins)
        red = QBrush(QColor(200, 0, 0))
        it_cat = QTableWidgetItem("CREDIT")
        it_cat.setData(Qt.UserRole, "CREDIT"); it_cat.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(ins, 1, it_cat)
        it_spec = QTableWidgetItem(label)
        it_spec.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable); it_spec.setForeground(red)
        self.step5_table.setItem(ins, 2, it_spec)
        self.step5_table.setItem(ins, 3, QTableWidgetItem(""))
        self.step5_table.setItem(ins, 4, QTableWidgetItem(""))
        it_amt = QTableWidgetItem(fmt_krw(amount))
        it_amt.setData(Qt.UserRole, float(amount)); it_amt.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        it_amt.setForeground(red); self.step5_table.setItem(ins, 5, it_amt)

    def _refresh_credits(self) -> None:
        self._clear_credit_rows()
        if self.state.pump_credit: self._add_credit_row("Pump Credit", -abs(self.state.pump_credit))
        if self.state.rack_credit: self._add_credit_row("Rack Credit", -abs(self.state.rack_credit))

    # ── STEP5 정렬 (STEP4 품목 순서 기준) ────────
    def _sort_step5_items(self) -> None:
        """ITEM 행을 self._spec_order 순서에 따라 제자리 정렬."""
        spec_order: Dict[str, int] = getattr(self, "_spec_order", {})
        if not spec_order:
            return
        item_rows = [r for r in range(self.step5_table.rowCount()) if self._is_item(r)]
        if len(item_rows) <= 1:
            return
        # 각 행 데이터 스냅샷
        snaps = []
        for r in item_rows:
            cb = self.step5_table.cellWidget(r, 0)
            cat_it  = self.step5_table.item(r, 1)
            spec_it = self.step5_table.item(r, 2)
            snaps.append({
                "checked": cb.isChecked() if isinstance(cb, QCheckBox) else False,
                "cat":     s(cat_it.text())  if cat_it  else "",
                "spec":    s(spec_it.text()) if spec_it else "",
                "qty":     self._get_float(r, 3),
                "price":   self._get_float(r, 4),
                "amt":     self._get_float(r, 5),
            })
        snaps.sort(key=lambda x: (0 if x["cat"] == "PUMP" else 1, spec_order.get(x["spec"], 10**9)))
        # 정렬된 데이터를 같은 위치에 재기입 (행 삽입/삭제 없이)
        self.step5_table.blockSignals(True)
        for r, d in zip(item_rows, snaps):
            cb = self.step5_table.cellWidget(r, 0)
            if isinstance(cb, QCheckBox):
                cb.blockSignals(True); cb.setChecked(d["checked"]); cb.blockSignals(False)
            cat_it  = self.step5_table.item(r, 1)
            spec_it = self.step5_table.item(r, 2)
            if cat_it:  cat_it.setText(d["cat"])
            if spec_it: spec_it.setText(d["spec"])
            self._set_qty(r, d["qty"])
            self._set_price(r, d["price"])
            self._set_amt(r, d["amt"])
        self.step5_table.blockSignals(False)

    # ── STEP5 스냅샷 ─────────────────────────────
    def _snapshot_items(self) -> List[Dict[str, Any]]:
        rows = []
        for r in range(self.step5_table.rowCount()):
            if self._is_total(r): continue
            cat_it  = self.step5_table.item(r, 1)
            spec_it = self.step5_table.item(r, 2)
            rows.append({
                "role":  s(cat_it.data(Qt.UserRole)) if cat_it else "",
                "cat":   s(cat_it.text())             if cat_it else "",
                "spec":  s(spec_it.text())            if spec_it else "",
                "qty":   self._get_float(r, 3),
                "price": self._get_float(r, 4),
                "amt":   self._get_float(r, 5),
            })
        return rows

    # ── 전체선택 동기화 ──────────────────────────
    def _sync_step5_all_chk(self) -> None:
        selectable = checked = 0
        last = self.step5_table.rowCount() - 1
        for r in range(last + 1):
            if self._is_total(r) or self._is_credit(r): continue
            cb = self.step5_table.cellWidget(r, 0)
            if isinstance(cb, QCheckBox):
                selectable += 1
                if cb.isChecked(): checked += 1
        self.chk_step5_all.blockSignals(True)
        if   selectable == 0 or checked == 0: self.chk_step5_all.setCheckState(Qt.Unchecked)
        elif checked == selectable:            self.chk_step5_all.setCheckState(Qt.Checked)
        else:                                  self.chk_step5_all.setCheckState(Qt.PartiallyChecked)
        self.chk_step5_all.blockSignals(False)

    def _on_step5_check_changed(self, _state: int) -> None:
        """개별 체크박스 변경 → 전체선택 상태 동기화."""
        self._sync_step5_all_chk()

    def _on_step5_select_all(self, checked: bool) -> None:
        """전체선택 체크박스 핸들러."""
        last = self.step5_table.rowCount() - 1
        for r in range(last + 1):
            if self._is_total(r) or self._is_credit(r): continue
            cb = self.step5_table.cellWidget(r, 0)
            if isinstance(cb, QCheckBox):
                cb.blockSignals(True); cb.setChecked(checked); cb.blockSignals(False)
        self._sync_step5_all_chk()

    # ── STEP4 하이라이트 동기화 ──────────────────
    def _sync_step4_highlight(self, scroll_top: bool = True) -> None:
        transferred: set = set()
        for r in range(self.step5_table.rowCount()):
            if self._is_total(r) or self._is_credit(r): continue
            it = self.step5_table.item(r, 2)
            if it and it.text().strip():
                transferred.add(it.text().strip())
        self._step4_highlight = transferred
        self._render_items_table(scroll_top=scroll_top)

    def _render_items_table(self, scroll_top: bool = True) -> None:
        """하위 클래스(QuoteBuilderPage)에서 구현."""
        raise NotImplementedError("_render_items_table must be implemented by subclass")


# ══════════════════════════════════════════════
# 견적서 작성 페이지
# ══════════════════════════════════════════════

class QuoteBuilderPage(Step5Manager, QWidget):

    def __init__(self) -> None:
        QWidget.__init__(self)
        self.state = QuoteState()

        self._step4_highlight     : set                     = set()
        self._filtered_items      : List[Tuple[str, float]] = []
        self._spec_order          : Dict[str, int]          = {}
        self._done_req_indices    : set                     = set()   # 견적 완료된 의뢰행 인덱스
        self._pending_req_indices : List[int]               = []      # 현재 생성 중인 의뢰행 인덱스

        self._filter_timer  : QTimer                          = QTimer(self)
        self._filter_timer.setSingleShot(True)
        self._filter_timer.timeout.connect(self._do_filter)
        self._tpl_thread    : Optional[_TemplateLoaderThread] = None
        self._gen_qt_thread : Optional[_GenQuoteThread]       = None
        self._gen_cv_thread : Optional[_GenCoverThread]       = None

        self._build_ui()
        self._ensure_total()
        self._log("준비: STEP1 → STEP2 → STEP4/드래그 → STEP5 → 생성")

        sc = QShortcut(QKeySequence.Delete, self)
        sc.activated.connect(self._delete_checked_rows)

    # ══════════════════════════════════════════
    # UI 구성
    # ══════════════════════════════════════════

    def _build_ui(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(14, 14, 14, 14)
        outer.setSpacing(12)

        grid = QGridLayout()
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(12)

        # Row 0: 버튼 4개
        self.btn_step1 = self._action_btn("STEP 1\n통합양식 LOAD", self._load_template)
        self.btn_step2 = self._action_btn("STEP 2\n의뢰파일 LOAD", self._load_request)
        tint_button(self.btn_step1, "#C8E6C9")   # 연초록
        tint_button(self.btn_step2, "#BBDEFB")   # 연파랑
        self.btn_step2.setEnabled(False)
        self._step3_frame = self._build_step3()
        self._step3_frame.setEnabled(False)
        guide = labeled_frame("견적서작성 사용법", min_h=110)
        guide.layout().addWidget(QLabel(
            "① STEP1: 통합양식 LOAD\n② STEP2: 의뢰파일 LOAD\n"
            "③ 의뢰행 더블클릭 → 조건 자동입력\n"
            "④ STEP4 더블클릭/드래그 → STEP5\n⑤ 견적서 생성 버튼"))
        grid.addWidget(self.btn_step1,    0, 0)
        grid.addWidget(self.btn_step2,    0, 1)
        grid.addWidget(self._step3_frame, 0, 2)
        grid.addWidget(guide,             0, 3)

        # Row 1: 의뢰 / STEP4 / 우측
        grid.addWidget(self._build_req_panel(),   1, 0)
        grid.addWidget(self._build_step4_panel(), 1, 1)
        grid.addWidget(self._build_right_panel(), 1, 2, 1, 2)

        # Row 2: 로그 / 옵션+견적정보
        grid.addWidget(self._build_worklog(),            2, 0, 1, 2)
        grid.addWidget(self._build_bottom_right_panel(), 2, 2, 1, 2)

        grid.setColumnStretch(0, 3); grid.setColumnStretch(1, 5)
        grid.setColumnStretch(2, 5); grid.setColumnStretch(3, 2)
        grid.setRowStretch(1, 10)
        outer.addLayout(grid, 1)

    @staticmethod
    def _action_btn(label: str, slot) -> QPushButton:
        btn = QPushButton(label); btn.setMinimumHeight(110)
        f = btn.font(); f.setPointSize(12); f.setBold(True); btn.setFont(f)
        btn.clicked.connect(slot); return btn

    # ── STEP3 ─────────────────────────────────
    def _build_step3(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2); frame.setMinimumHeight(110)
        v = QVBoxLayout(frame); v.setContentsMargins(12,12,12,12); v.setSpacing(6)
        v.addWidget(bold_label("STEP 3\n조건 입력"))
        row1 = QHBoxLayout()
        self.cb_process = QComboBox(); self.cb_process.addItems(PROCESS_CHOICES)
        self.cb_vendor  = QComboBox(); self.cb_vendor.addItems(VENDOR_CHOICES)
        row1.addWidget(QLabel("공정")); row1.addWidget(self.cb_process, 1)
        row1.addWidget(QLabel("설비사")); row1.addWidget(self.cb_vendor, 1)
        row2 = QHBoxLayout()
        self.ed_code = QLineEdit(); self.ed_code.setPlaceholderText("5D")
        self.btn_apply_map = QPushButton("코드매핑 적용 → STEP5(RACK)"); self.btn_apply_map.setEnabled(False)
        tint_button(self.btn_apply_map, "#FFE0B2")  # 연주황
        row2.addWidget(QLabel("5D")); row2.addWidget(self.ed_code, 1); row2.addWidget(self.btn_apply_map)
        v.addLayout(row1); v.addLayout(row2)
        self.cb_process.currentTextChanged.connect(self._on_step3_changed)
        self.cb_vendor.currentTextChanged.connect(self._on_step3_changed)
        self.ed_code.textChanged.connect(self._on_step3_changed)
        self.btn_apply_map.clicked.connect(self._apply_code_map)
        return frame

    # ── 의뢰파일 패널 ─────────────────────────
    def _build_req_panel(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2)
        frame.setMinimumHeight(420); frame.setFixedWidth(400)
        v = QVBoxLayout(frame); v.setContentsMargins(12,12,12,12); v.setSpacing(8)
        v.addWidget(bold_label("의뢰파일DATA", size=12))
        top = QHBoxLayout()
        self.chk_req_all = QCheckBox("전체선택"); self.chk_req_all.clicked.connect(self._on_req_select_all)
        top.addWidget(self.chk_req_all); top.addStretch(1); v.addLayout(top)
        self.req_table = _make_plain_table(4, ["선택", "설비호기(Z)", "투자정보", "수량"])
        self.req_table.setColumnWidth(0, 42)
        _rh = self.req_table.horizontalHeader()
        _rh.setSectionResizeMode(0, QHeaderView.Fixed)
        _rh.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        _rh.setSectionResizeMode(2, QHeaderView.Stretch)
        _rh.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.req_table.cellDoubleClicked.connect(self._on_req_double_click)
        self._style_req_table()
        v.addWidget(self.req_table, 1)
        return frame

    # ── STEP4 패널 ────────────────────────────
    def _build_step4_panel(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2); frame.setMinimumHeight(420)
        v = QVBoxLayout(frame); v.setContentsMargins(12,12,12,12); v.setSpacing(8)
        v.addWidget(bold_label("STEP 4\n품목(규격/단가) + 필터  (더블클릭/드래그 → STEP5)"))
        self.ed_filter = QLineEdit(); self.ed_filter.setPlaceholderText("규격 필터(실시간) - 포함 검색")
        self.ed_filter.textChanged.connect(self._apply_filter)
        self.items_table = DraggableItemsTable(0, 2)
        self.items_table.setHorizontalHeaderLabels(["규격", "단가(원)"])
        self.items_table.verticalHeader().setVisible(False)
        self.items_table.setAlternatingRowColors(True)
        self.items_table.setEditTriggers(DraggableItemsTable.NoEditTriggers)
        self.items_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.items_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.items_table.setDragEnabled(True); self.items_table.setWordWrap(False)
        self.items_table.setStyleSheet("QTableWidget::item:selected{color:black;}")
        h = self.items_table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.Stretch); h.setSectionResizeMode(1, QHeaderView.Fixed)
        self.items_table.setColumnWidth(1, 80)
        self.items_table.cellDoubleClicked.connect(self._on_item_double_click)
        v.addWidget(self.ed_filter); v.addWidget(self.items_table, 1)
        return frame

    # ── 우측 패널 (STEP5 + 버튼 + STEP7) ──────
    def _build_right_panel(self) -> QWidget:
        w = QWidget(); g = QGridLayout(w)
        g.setContentsMargins(0,0,0,0); g.setHorizontalSpacing(12)

        step5 = labeled_frame("STEP5 견적작성 대상", min_h=320)
        step5.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        sl = step5.layout()
        top5 = QHBoxLayout()
        self.chk_step5_all = QCheckBox("전체선택")
        self.chk_step5_all.clicked.connect(self._on_step5_select_all)
        top5.addWidget(self.chk_step5_all); top5.addStretch(1); sl.addLayout(top5)

        self.step5_table = DroppableQuoteTable()
        self.step5_table.setColumnCount(6)
        self.step5_table.setHorizontalHeaderLabels(["선택","분류","품목","수량","단가(원)","금액(원)"])
        self.step5_table.verticalHeader().setVisible(False)
        self.step5_table.setAlternatingRowColors(True)
        self.step5_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.step5_table.setEditTriggers(DroppableQuoteTable.NoEditTriggers)
        self.step5_table.cellDoubleClicked.connect(self._on_step5_double_click)
        self.step5_table.drop_callback = self._on_drop_to_step5
        h5 = self.step5_table.horizontalHeader()
        for i, mode in enumerate([QHeaderView.ResizeToContents]*2 + [QHeaderView.Stretch] + [QHeaderView.ResizeToContents]*3):
            h5.setSectionResizeMode(i, mode)
        self.step5_table.setColumnWidth(0, 46)
        sl.addWidget(self.step5_table, 1)

        side = QWidget(); sv = QVBoxLayout(side); sv.setContentsMargins(0,0,0,0); sv.setSpacing(12)
        sv.addWidget(self._build_generate_panel()); sv.addWidget(self._build_step7(), 1)
        g.addWidget(step5, 0, 0); g.addWidget(side, 0, 1)
        g.setColumnStretch(0, 4); g.setColumnStretch(1, 2)
        return w

    def _build_generate_panel(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2); frame.setMinimumHeight(110)
        h = QHBoxLayout(frame); h.setContentsMargins(6,6,6,6); h.setSpacing(6)
        self.btn_gen_quote = QPushButton("견적서 생성"); self.btn_gen_quote.setMinimumHeight(90)
        self.btn_gen_cover = QPushButton("갑지 생성");   self.btn_gen_cover.setMinimumHeight(90)
        tint_button(self.btn_gen_quote, "#B3E5FC")   # 연하늘
        tint_button(self.btn_gen_cover, "#F8BBD0")   # 연분홍
        for btn, slot in [(self.btn_gen_quote, self._generate_quote), (self.btn_gen_cover, self._generate_cover)]:
            f = btn.font(); f.setPointSize(12); f.setBold(True); btn.setFont(f)
            btn.clicked.connect(slot); h.addWidget(btn, 1)
        return frame

    def _build_step7(self) -> QFrame:
        frame = labeled_frame("견적서LIST", min_h=260)
        self.list_done = QListWidget()
        self.list_done.itemDoubleClicked.connect(self._open_done)
        frame.layout().addWidget(self.list_done, 1); return frame

    def _build_worklog(self) -> QFrame:
        frame = labeled_frame("작업로그", min_h=140)
        frame.setFixedHeight(140); frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.log_view = QTextEdit(); self.log_view.setReadOnly(True)
        frame.layout().addWidget(self.log_view, 1); return frame

    def _build_bottom_right_panel(self) -> QWidget:
        w = QWidget(); w.setFixedHeight(140)
        h = QHBoxLayout(w); h.setContentsMargins(0,0,0,0); h.setSpacing(12)
        h.addWidget(self._build_options_panel(), 1)
        h.addWidget(self._build_quote_info_panel())
        return w

    def _build_options_panel(self) -> QFrame:
        frame = labeled_frame("옵션", min_h=55)
        frame.setMaximumHeight(140); frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        row = QHBoxLayout()
        self.btn_credit = QPushButton("Credit"); self.btn_credit.setMinimumHeight(32)
        self.btn_credit.clicked.connect(self._open_credit_dialog)
        self.btn_investor = QPushButton("투자자 변경"); self.btn_investor.setMinimumHeight(32)
        self.btn_investor.clicked.connect(self._change_investor)
        tint_button(self.btn_credit,   "#FFF9C4")   # 연노랑
        tint_button(self.btn_investor, "#E8EAF6")   # 연라벤더
        self.chk_qt_kr = QCheckBox("국내"); self.chk_qt_cn = QCheckBox("중국"); self.chk_qt_us = QCheckBox("미국")
        self.chk_qt_kr.setChecked(True)
        for chk, qt in [(self.chk_qt_kr,"국내"),(self.chk_qt_cn,"중국"),(self.chk_qt_us,"미국")]:
            chk.clicked.connect(lambda checked, t=qt: self._on_qt_checked(t, checked))
        row.addWidget(self.btn_credit); row.addWidget(self.btn_investor); row.addSpacing(10)
        row.addWidget(QLabel("견적서 타입"))
        row.addWidget(self.chk_qt_kr); row.addWidget(self.chk_qt_cn); row.addWidget(self.chk_qt_us)
        row.addStretch(1); frame.layout().addLayout(row); return frame

    def _build_quote_info_panel(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2)
        frame.setFixedHeight(140); frame.setFixedWidth(260)
        v = QVBoxLayout(frame); v.setContentsMargins(12,6,12,6); v.setSpacing(4)
        v.addWidget(bold_label("견적서정보", size=9))
        self.lbl_pr = info_label("PR: -"); self.lbl_item = info_label("항번: -")
        self.lbl_line = info_label("라인공정: -"); self.lbl_investor = info_label("투자자: -")
        self.lbl_more = info_label("")
        for lbl in (self.lbl_pr, self.lbl_item, self.lbl_line, self.lbl_investor): v.addWidget(lbl)
        v.addStretch(1); v.addWidget(self.lbl_more); return frame

    def _style_req_table(self) -> None:
        self.req_table.setAlternatingRowColors(False)
        # item background 는 지정하지 않아야 _mark_req_done 의 setBackground 가 반영된다
        self.req_table.setStyleSheet(
            "QTableWidget{background:transparent;border:none;}"
            "QTableWidget::item:selected{color:black;background:rgba(180,200,230,140);}")

    # ══════════════════════════════════════════
    # STEP1
    # ══════════════════════════════════════════

    def _load_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "통합양식 선택", "", "Excel Files (*.xlsx)")
        if not path:
            return
        # 시트명 사전 검증 (빠름 — 메타만 읽음)
        try:
            from openpyxl import load_workbook
            wb_meta = load_workbook(path, read_only=True, data_only=False)
            missing = [n for n in SheetName.REQUIRED if n not in wb_meta.sheetnames]
            wb_meta.close()
        except Exception as e:
            QMessageBox.critical(self, "오류", f"파일을 열 수 없습니다.\n{e}")
            return
        if missing:
            QMessageBox.warning(self, "통합양식 오류", "필수 시트 없음:\n" + "\n".join(missing))
            return

        self.btn_step1.setEnabled(False)
        self.btn_step1.setText("STEP 1\n로딩 중…")
        self._tpl_thread = _TemplateLoaderThread(path, self)
        self._tpl_thread.done.connect(lambda result: self._on_template_loaded(path, result))
        self._tpl_thread.start()

    def _on_template_loaded(self, path: str, result) -> None:
        self.btn_step1.setEnabled(True)
        self.btn_step1.setText("STEP 1\n통합양식 LOAD")
        if isinstance(result, Exception):
            logger.error("통합양식 파싱 오류", exc_info=result)
            QMessageBox.critical(self, "파싱 오류", str(result))
            return
        rows, by_class, price_by_spec, order, cmap = result
        self.state.template_path = path
        self.state.items_rows = rows
        self.state.items_by_class = by_class
        self.state.price_by_spec = price_by_spec
        self.state.code_map = cmap
        self._spec_order = order
        self._apply_filter()
        self.btn_step2.setEnabled(True)
        self._step3_frame.setEnabled(True)
        self._log(f"STEP1 완료: {path}  (품목 {len(rows)}건, 코드매핑 키 {len(cmap)}개)")
        QMessageBox.information(self, "완료", "통합양식 LOAD 완료")

    # ══════════════════════════════════════════
    # STEP2
    # ══════════════════════════════════════════

    def _load_request(self) -> None:
        if not self.state.template_path:
            QMessageBox.warning(self, "순서 오류", "STEP1을 먼저 완료하세요."); return
        path, _ = QFileDialog.getOpenFileName(self, "의뢰파일 선택", "", "Excel Files (*.xlsx)")
        if not path: return
        try:
            sheet_name, rows = excel_io.parse_request_xlsx(path)
        except Exception as e:
            logger.error("의뢰파일 파싱 오류", exc_info=True)
            QMessageBox.critical(self, "오류", f"파일을 읽을 수 없습니다.\n{e}"); return
        if not rows:
            QMessageBox.warning(self, "의뢰파일", "읽을 데이터가 없습니다."); return
        self.state.request_path = path; self.state.request_sheet_name = sheet_name
        self.state.request_rows = rows
        self._fill_req_table(rows); self._step3_frame.setEnabled(True)
        self._log(f"STEP2 완료: {path}  ({len(rows)}건, 시트={sheet_name})")
        QMessageBox.information(self, "완료", "의뢰파일 LOAD 완료")

    def _fill_req_table(self, rows: List[Dict[str, Any]]) -> None:
        self._done_req_indices = set()   # 새 의뢰파일 로드 시 완료 이력 초기화
        self.req_table.setRowCount(0)
        self.chk_req_all.blockSignals(True); self.chk_req_all.setCheckState(Qt.Unchecked); self.chk_req_all.blockSignals(False)
        for i, rd in enumerate(rows):
            self.req_table.insertRow(i)
            host = centered_checkbox(lambda _s, _r=i: self._on_req_check())
            self.req_table.setCellWidget(i, 0, host)
            invest_info = parse_invest_info(rd.get("G"))
            h_val = rd.get("H")
            qty_str = fmt_qty(to_float(h_val)) if h_val is not None else ""
            for col, val in [(1, s(rd.get("Z"))), (2, invest_info), (3, qty_str)]:
                it = QTableWidgetItem(val); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                self.req_table.setItem(i, col, it)
        self.req_table.setColumnWidth(0, 42)
        self.req_table.resizeColumnToContents(1)
        self.req_table.resizeColumnToContents(3)
        self._style_req_table(); self._update_quote_info()

    def _is_req_checked(self, row: int) -> bool:
        cb = get_checkbox_from_cell(self.req_table.cellWidget(row, 0))
        return bool(cb and cb.isChecked())

    def _checked_req_rows(self) -> List[int]:
        return [r for r in range(self.req_table.rowCount()) if self._is_req_checked(r)]

    def _on_req_check(self) -> None:
        n = sum(1 for r in range(self.req_table.rowCount()) if self._is_req_checked(r))
        total = self.req_table.rowCount()
        self.chk_req_all.blockSignals(True)
        if   n == 0:     self.chk_req_all.setCheckState(Qt.Unchecked)
        elif n == total: self.chk_req_all.setCheckState(Qt.Checked)
        else:            self.chk_req_all.setCheckState(Qt.PartiallyChecked)
        self.chk_req_all.blockSignals(False); self._update_quote_info()

    def _on_req_select_all(self, checked: bool) -> None:
        for r in range(self.req_table.rowCount()):
            cb = get_checkbox_from_cell(self.req_table.cellWidget(r, 0))
            if cb: cb.blockSignals(True); cb.setChecked(checked); cb.blockSignals(False)
        self._on_req_check()

    def _update_quote_info(self) -> None:
        checked = self._checked_req_rows()
        if not checked or not self.state.request_rows:
            self.lbl_pr.setText("PR: -"); self.lbl_item.setText("항번: -")
            self.lbl_line.setText("라인공정: -"); self.lbl_investor.setText("투자자: -")
            self.lbl_more.setText(""); return
        first = self.state.request_rows[checked[0]]
        self.lbl_pr.setText(f"PR: {s(first.get('D')) or '-'}")
        self.lbl_item.setText(f"항번: {s(first.get('E')) or '-'}")
        self.lbl_line.setText(f"라인공정: {s(first.get('K')) or '-'}")
        self.lbl_investor.setText(f"투자자: {s(first.get('J')) or '-'}")
        self.lbl_more.setText(f"외 {len(checked)-1}건" if len(checked) > 1 else "")

    def _on_req_double_click(self, row: int, _col: int) -> None:
        if row >= len(self.state.request_rows): return
        rd = self.state.request_rows[row]

        def _match(text: str, choices: List[str]) -> Optional[str]:
            t = normalize_token(text)
            if not t: return None
            for c in sorted((x for x in choices if x), key=len, reverse=True):
                if normalize_token(c) in t: return c
            return None

        p = _match(s(rd.get("K")), PROCESS_CHOICES)
        if p:
            idx = self.cb_process.findText(p)
            if idx >= 0: self.cb_process.setCurrentIndex(idx)
        v = _match(s(rd.get("X")), VENDOR_CHOICES)
        if v:
            idx = self.cb_vendor.findText(v)
            if idx >= 0: self.cb_vendor.setCurrentIndex(idx)
        vcode = s(rd.get("V"))
        if vcode: self.ed_code.setText(vcode)

        f_class = s(rd.get("F")); qty = to_float(rd.get("H"))
        if f_class:
            matches = self.state.items_by_class.get(f_class, [])
            for row_data in matches:
                spec = s(row_data.get("B")); price = float(row_data.get("C", 0))
                if spec: self._add_row("PUMP", spec, qty, price)
            if matches:
                self._log(f"PUMP 자동: 분류={f_class}, {len(matches)}개, 수량={fmt_qty(qty)}")
                self._sync_step4_highlight()
        self._log(f"더블클릭 → 공정={self.cb_process.currentText()} / 설비사={self.cb_vendor.currentText()} / 5D={self.ed_code.text()}")

    # ══════════════════════════════════════════
    # STEP3
    # ══════════════════════════════════════════

    def _on_step3_changed(self) -> None:
        self.state.process = s(self.cb_process.currentText()) or None
        self.state.vendor  = s(self.cb_vendor.currentText())  or None
        self.state.code_5d = s(self.ed_code.text())           or None
        ready = bool(self.state.process and self.state.vendor and self.state.code_5d)
        self.btn_apply_map.setEnabled(ready and bool(self.state.code_map))

    def _apply_code_map(self) -> None:
        key = (self.state.process, self.state.vendor, self.state.code_5d)
        entries = self.state.code_map.get(key)
        if not entries:
            near = sum(1 for k in self.state.code_map if k[:2] == key[:2])
            QMessageBox.warning(self, "코드매핑 없음", f"키: {key}\n동일 공정/설비사 키 수: {near}"); return
        for acc, qty in entries:
            price = float(self.state.price_by_spec.get(acc, 0))
            self._add_row("RACK", acc, qty, price)
        self._log(f"코드매핑 적용 {key} → {len(entries)}개"); self._sync_step4_highlight()

    # ══════════════════════════════════════════
    # STEP4
    # ══════════════════════════════════════════

    def _apply_filter(self) -> None:
        self._filter_timer.start(200)

    def _do_filter(self) -> None:
        key = s(self.ed_filter.text()).lower()
        self._filtered_items = [
            (s(r.get("B")), float(r.get("C", 0)))
            for r in self.state.items_rows
            if s(r.get("B")) and key in s(r.get("B")).lower()
        ]
        self._render_items_table(scroll_top=False)

    def _render_items_table(self, scroll_top: bool = True) -> None:
        """Step5Manager 추상 메서드 구현: STEP4 테이블 재렌더링."""
        highlighted = self._step4_highlight

        def order(spec: str) -> int:
            return self._spec_order.get(spec, 10**9)

        sorted_rows = sorted(
            self._filtered_items,
            key=lambda x: (0 if x[0] in highlighted else 1, order(x[0])))

        self.items_table.blockSignals(True)
        self.items_table.setRowCount(len(sorted_rows))
        for i, (spec, price) in enumerate(sorted_rows):
            for col, val in [(0, spec), (1, fmt_krw(price))]:
                it = QTableWidgetItem(val); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                if spec in highlighted: it.setBackground(QBrush(QColor(220, 220, 220)))
                self.items_table.setItem(i, col, it)
        self.items_table.blockSignals(False)
        self.items_table.setColumnWidth(1, 80)
        if scroll_top: self.items_table.scrollToTop()

    def _on_item_double_click(self, row: int, _col: int) -> None:
        spec_it  = self.items_table.item(row, 0); price_it = self.items_table.item(row, 1)
        if not spec_it: return
        self._ask_qty_add(spec_it.text().strip(), to_float(price_it.text() if price_it else "0"))

    def _on_drop_to_step5(self, spec: str, price: float) -> None:
        self._ask_qty_add(spec, price)

    def _ask_qty_add(self, spec: str, unit_price: float) -> None:
        qty, ok = QInputDialog.getDouble(self, "수량 입력", f"수량 ({spec})", 1.0, 0, 1_000_000, 3)
        if ok: self._add_row("RACK", spec, qty, unit_price); self._sync_step4_highlight()

    # ══════════════════════════════════════════
    # STEP5 이벤트
    # ══════════════════════════════════════════

    def _on_step5_double_click(self, row: int, col: int) -> None:
        if self._is_total(row) or self._is_credit(row): return
        if col == 1:
            it = self.step5_table.item(row, 1)
            if it:
                it.setText("RACK" if it.text().strip() == "PUMP" else "PUMP")
                self._sort_step5_items()
                self._recalc_totals(); self._sync_step4_highlight(scroll_top=False)
        elif col == 3:
            cur = self._get_float(row, 3)
            val, ok = QInputDialog.getDouble(self, "수량 수정", "수량:", cur, 0, 1_000_000, 3)
            if ok: self._set_qty(row, val); self._recalc_row(row); self._recalc_totals(); self._sync_step4_highlight()

    def _delete_checked_rows(self) -> None:
        to_del = [
            r for r in range(self.step5_table.rowCount() - 1)
            if not self._is_total(r) and not self._is_credit(r)
            and isinstance(self.step5_table.cellWidget(r, 0), QCheckBox)
            and self.step5_table.cellWidget(r, 0).isChecked()
        ]
        if not to_del:
            sel = self.step5_table.selectionModel().selectedRows()
            if sel:
                r = sel[0].row()
                if not self._is_total(r) and not self._is_credit(r): to_del = [r]
        if not to_del: return
        reply = QMessageBox.question(self, "삭제 확인", f"{len(to_del)}개 품목을 삭제할까요?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes: return
        for r in sorted(to_del, reverse=True): self.step5_table.removeRow(r)
        self._ensure_total(); self._refresh_credits(); self._sort_step5_items(); self._recalc_totals()
        self.chk_step5_all.blockSignals(True); self.chk_step5_all.setCheckState(Qt.Unchecked); self.chk_step5_all.blockSignals(False)
        self._sync_step4_highlight()

    # ══════════════════════════════════════════
    # 옵션
    # ══════════════════════════════════════════

    def _on_qt_checked(self, selected: str, checked: bool) -> None:
        chk_map = {"국내": self.chk_qt_kr, "중국": self.chk_qt_cn, "미국": self.chk_qt_us}
        if not checked:
            # 이미 선택된 항목 재클릭 → 체크 복원 후 옵션창
            chk = chk_map[selected]; chk.blockSignals(True); chk.setChecked(True); chk.blockSignals(False)
            if selected in ("중국", "미국"):
                self._open_options(focus={"중국": "cn", "미국": "us"}[selected])
            return
        for name, chk in chk_map.items():
            chk.blockSignals(True); chk.setChecked(name == selected); chk.blockSignals(False)
        self.state.quote_type = selected; self._log(f"견적서 타입 = {selected}")
        if selected in ("중국", "미국"): self._open_options(focus={"중국": "cn", "미국": "us"}[selected])

    def _open_options(self, focus: str = "credit") -> None:
        if focus == "cn":
            dlg = ChinaOptionsDialog(self, self.state)
        elif focus == "us":
            dlg = USOptionsDialog(self, self.state)
        else:
            return
        if dlg.exec() == QDialog.Accepted:
            dlg.apply(); self._refresh_credits(); self._recalc_totals()
            self._sync_step4_highlight(scroll_top=False); self._log("[옵션] 입력값 반영")

    def _change_investor(self) -> None:
        name, ok = QInputDialog.getText(
            self, "투자자 변경", "투자자 이름:",
            text=self.state.investor_name or "채승철",
        )
        if ok:
            self.state.investor_name = name.strip() or "채승철"
            self._log(f"투자자: {self.state.investor_name}")

    def _open_credit_dialog(self) -> None:
        try:
            self._open_credit_dialog_impl()
        except Exception:
            QMessageBox.critical(self, "Credit 오류", traceback.format_exc())

    def _open_credit_dialog_impl(self) -> None:
        dlg = QDialog(self); dlg.setWindowTitle("Credit 입력"); dlg.setFixedWidth(500)
        v = QVBoxLayout(dlg); v.setSpacing(10)

        h_val = to_float(self._pick_rd().get("H")) if self.state.request_rows else 0.0

        # ── 카테고리별 품목 합계 ──────────────────────────────────
        def _cat_sum(is_pump: bool) -> float:
            total = 0.0
            for r in range(self.step5_table.rowCount()):
                if self._is_item(r):
                    cat_it = self.step5_table.item(r, 1)
                    if cat_it and (s(cat_it.text()) == "PUMP") == is_pump:
                        total += self._get_float(r, 5)
            return total

        pump_sum = _cat_sum(True)
        rack_sum = _cat_sum(False)

        # ── 섹션 빌더 ────────────────────────────────────────────
        # 반환: (ed_credit, ed_target, ed_ch)
        def _make_section(label: str, item_sum: float, init_credit: float):
            frame = QFrame(); frame.setFrameShape(QFrame.StyledPanel)
            fv    = QVBoxLayout(frame); fv.setContentsMargins(8, 6, 8, 6); fv.setSpacing(6)
            hdr = QLabel(f"<b>{label}</b>  (품목 합계: {fmt_krw(item_sum)} 원)")
            fv.addWidget(hdr)

            form = QFormLayout(); form.setVerticalSpacing(6)
            fv.addLayout(form)

            ed_credit = QLineEdit(); ed_credit.setPlaceholderText("0")
            ed_target = QLineEdit(); ed_target.setPlaceholderText(fmt_krw(item_sum))
            ed_ch     = QLineEdit(); ed_ch.setPlaceholderText(
                fmt_krw(item_sum / h_val) if h_val > 0 else "CH 값 없음"
            )
            ed_ch.setEnabled(h_val > 0)

            form.addRow("Credit 금액(원):", ed_credit)
            form.addRow("목표가(원):",      ed_target)
            form.addRow("CH당 단가(원/CH):", ed_ch)

            if init_credit:
                ed_credit.setText(fmt_krw(init_credit))

            updating = [False]

            def _from_credit():
                if updating[0]: return
                updating[0] = True
                try:
                    c = to_float(ed_credit.text())
                    t = item_sum - c
                    ed_target.setText(fmt_krw(max(0.0, t)))
                    if h_val > 0:
                        ed_ch.setText(fmt_krw(max(0.0, t) / h_val))
                finally:
                    updating[0] = False

            def _from_target():
                if updating[0]: return
                updating[0] = True
                try:
                    t = to_float(ed_target.text())
                    c = max(0.0, item_sum - t)
                    ed_credit.setText(fmt_krw(c))
                    if h_val > 0:
                        ed_ch.setText(fmt_krw(t / h_val))
                finally:
                    updating[0] = False

            def _from_ch():
                if updating[0]: return
                if h_val <= 0: return
                updating[0] = True
                try:
                    ch = to_float(ed_ch.text())
                    t  = ch * h_val
                    c  = max(0.0, item_sum - t)
                    ed_credit.setText(fmt_krw(c))
                    ed_target.setText(fmt_krw(max(0.0, t)))
                finally:
                    updating[0] = False

            ed_credit.textEdited.connect(_from_credit)
            ed_target.textEdited.connect(_from_target)
            ed_ch.textEdited.connect(_from_ch)

            return frame, ed_credit, ed_target, ed_ch

        grp_pump, ed_pump_c, ed_pump_t, ed_pump_ch = _make_section(
            "PUMP", pump_sum, self.state.pump_credit)
        grp_rack, ed_rack_c, ed_rack_t, ed_rack_ch = _make_section(
            "RACK", rack_sum, self.state.rack_credit)

        v.addWidget(grp_pump)
        v.addWidget(grp_rack)

        # ── 결과 요약 ─────────────────────────────────────────────
        sep = QFrame(); sep.setFrameShape(QFrame.HLine); sep.setFrameShadow(QFrame.Sunken)
        v.addWidget(sep)
        lbl_summary = QLabel(); lbl_summary.setStyleSheet("font-weight:bold;")
        lbl_ch_total = QLabel()
        v.addWidget(lbl_summary); v.addWidget(lbl_ch_total)

        def _refresh_summary() -> None:
            pc    = to_float(ed_pump_c.text())
            rc    = to_float(ed_rack_c.text())
            after = pump_sum + rack_sum - pc - rc
            lbl_summary.setText(
                f"Credit 반영 후 총액: {fmt_krw(after)} 원"
                f"  (감액 Pump {fmt_krw(pc)} + Rack {fmt_krw(rc)} = {fmt_krw(pc + rc)} 원)"
            )
            if h_val > 0:
                lbl_ch_total.setText(
                    f"전체 CH당 단가: {fmt_krw(after / h_val)} 원/CH  (H: {fmt_qty(h_val)} CH)"
                )
            else:
                lbl_ch_total.setText("전체 CH당 단가: H열 값 없음")

        for ed in (ed_pump_c, ed_pump_t, ed_pump_ch, ed_rack_c, ed_rack_t, ed_rack_ch):
            ed.textEdited.connect(_refresh_summary)

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel); v.addWidget(bb)

        def _ok() -> None:
            self.state.pump_credit = to_float(ed_pump_c.text())
            self.state.rack_credit = to_float(ed_rack_c.text())
            self._refresh_credits(); self._recalc_totals(); dlg.accept()

        bb.accepted.connect(_ok); bb.rejected.connect(dlg.reject)
        _refresh_summary(); dlg.exec()

    # ══════════════════════════════════════════
    # STEP6 – 생성
    # ══════════════════════════════════════════

    def _set_gen_buttons(self, enabled: bool) -> None:
        self.btn_gen_quote.setEnabled(enabled)
        self.btn_gen_cover.setEnabled(enabled)

    def _generate_quote(self) -> None:
        if not self.state.template_path:
            QMessageBox.warning(self, "오류", "STEP1 통합양식을 먼저 LOAD 하세요."); return
        if self._gen_qt_thread and self._gen_qt_thread.isRunning():
            QMessageBox.warning(self, "진행 중", "이미 생성 작업이 진행 중입니다."); return
        self._refresh_credits(); self._recalc_totals()
        items = self._snapshot_items(); qtype = self.state.quote_type or "국내"
        if qtype == "중국" and not self.state.cn_info.get("tool"): self._open_options("cn")
        if qtype == "미국"  and not self.state.us_info.get("tool"): self._open_options("us")
        checked = self._checked_req_rows()
        if qtype == "국내" and self.state.request_rows and len(checked) >= 2:
            valid = [i for i in checked if i < len(self.state.request_rows)]
            rds = [self.state.request_rows[i] for i in valid]
            self._pending_req_indices = valid
        else:
            idx = self._pick_rd_index()
            rds = [self.state.request_rows[idx]] if (idx >= 0 and self.state.request_rows) else [{}]
            self._pending_req_indices = [idx] if idx >= 0 else []
        self._log(f"견적서 생성 시작 ({qtype}, {len(rds)}건) …")
        self._set_gen_buttons(False)
        self._gen_qt_thread = _GenQuoteThread(self.state, qtype, items, rds, self)
        self._gen_qt_thread.progress.connect(self._log)
        self._gen_qt_thread.done.connect(self._on_gen_quote_done)
        self._gen_qt_thread.start()

    def _on_gen_quote_done(self, result) -> None:
        self._set_gen_buttons(True)
        if isinstance(result, Exception):
            logger.error("견적서 생성 실패", exc_info=result)
            QMessageBox.critical(self, "생성 오류", f"{result}\n\n{traceback.format_exc()}")
            return
        for i, (rd, path) in enumerate(result):
            if path:
                self._add_done(path)
                self.state.last_output_dir = os.path.dirname(path)
                if i < len(self._pending_req_indices):
                    idx = self._pending_req_indices[i]
                    self._done_req_indices.add(idx)
                    self._mark_req_done(idx)
        ok = sum(1 for _, p in result if p)
        self._log(f"견적서 생성 완료: {ok}/{len(result)}건")
        QMessageBox.information(self, "완료", f"견적서 {ok}건 생성 완료")

    def _generate_cover(self) -> None:
        if not self.state.template_path:
            QMessageBox.warning(self, "오류", "STEP1 통합양식을 먼저 LOAD 하세요."); return
        if self._gen_cv_thread and self._gen_cv_thread.isRunning():
            QMessageBox.warning(self, "진행 중", "이미 갑지 생성 작업이 진행 중입니다."); return
        start = self.state.last_output_dir or exe_dir()
        paths, _ = QFileDialog.getOpenFileNames(self, "갑지 생성 대상 엑셀파일 (다중)", start, "Excel Files (*.xlsx)")
        if not paths: return
        folder = os.path.commonpath(paths)
        if os.path.isfile(folder): folder = os.path.dirname(folder)
        self._log(f"갑지 생성 시작 ({len(paths)}건) …")
        self._set_gen_buttons(False)
        self._gen_cv_thread = _GenCoverThread(
            self.state.template_path, folder, paths, self.state.investor_name, self)
        self._gen_cv_thread.progress.connect(self._log)
        self._gen_cv_thread.done.connect(self._on_gen_cover_done)
        self._gen_cv_thread.start()

    def _on_gen_cover_done(self, result) -> None:
        self._set_gen_buttons(True)
        if isinstance(result, Exception):
            logger.error("갑지 생성 실패", exc_info=result)
            QMessageBox.critical(self, "오류", f"{result}\n\n{traceback.format_exc()}")
            return
        out = result
        self.state.last_output_dir = os.path.dirname(out)
        self._add_done_cover(out)
        self._log(f"갑지 생성 완료: {out}")
        QMessageBox.information(self, "완료", f"갑지 생성 완료\n{os.path.basename(out)}")

    def _pick_rd_index(self) -> int:
        """클릭(선택)된 행 → 체크된 첫 번째 행 → 0 순으로 의뢰행 인덱스 반환."""
        cur = self.req_table.currentRow()
        if 0 <= cur < len(self.state.request_rows):
            return cur
        checked = self._checked_req_rows()
        if checked and checked[0] < len(self.state.request_rows):
            return checked[0]
        return 0 if self.state.request_rows else -1

    def _pick_rd(self) -> Dict[str, Any]:
        idx = self._pick_rd_index()
        if idx >= 0 and idx < len(self.state.request_rows):
            return self.state.request_rows[idx]
        return {}

    def _mark_req_done(self, row: int) -> None:
        """견적 완료된 의뢰행을 연두색으로 하이라이트."""
        color = QBrush(QColor(160, 220, 160))
        for col in range(1, self.req_table.columnCount()):
            it = self.req_table.item(row, col)
            if it:
                it.setBackground(color)

    def _add_done(self, path: str) -> None:
        it = QListWidgetItem(os.path.basename(path)); it.setData(Qt.UserRole, path); self.list_done.addItem(it)

    def _add_done_cover(self, path: str) -> None:
        it = QListWidgetItem(os.path.basename(path))
        it.setData(Qt.UserRole, path)
        it.setBackground(QBrush(QColor(255, 180, 180)))
        self.list_done.addItem(it)
        self.list_done.scrollToItem(it)

    def _open_done(self, item: QListWidgetItem) -> None:
        path = item.data(Qt.UserRole)
        if path:
            try: os.startfile(path)
            except Exception as e: QMessageBox.critical(self, "오류", str(e))

    def _log(self, text: str) -> None:
        self.log_view.append(text); logger.debug(text)


# ══════════════════════════════════════════════
# RACK 구매요청서 작성 페이지
# ══════════════════════════════════════════════

class _NoScrollComboBox(QComboBox):
    """마우스 스크롤로 값이 바뀌지 않는 QComboBox."""
    def wheelEvent(self, e):
        e.ignore()


class RackPurchaseRequestPage(QWidget):
    """이미지 양식 기반 RACK 구매요청서 작성/엑셀 생성 위젯."""

    COL_LABEL = 0
    COL_ITEM = 1
    COL_QTY = 2
    COL_REF_ITEM = 3
    COL_REF_QTY = 4
    COL_REF_APPLY = 5
    COL_REQUEST = 6

    # (rack_0idx_autocomplete, rack_1idx_item_col, rack_1idx_qty_col)
    # row 2 = 인보이스 필요여부 (자동계산), row 3~ = 기존 데이터 행
    _ROW_RACK_MAP: Dict[int, Tuple] = {
        3:  (2,  3,  None),
        4:  (5,  6,  None),
        5:  (8,  9,  None),
        6:  (10, 11, None),
        7:  (22, 23, None),
        8:  (23, 24, None),
        9:  (24, 25, None),
        10: (25, 26, None),
        11: (26, 27, None),
        12: (None,28, None),
        13: (28, 29, None),
        14: (29, 30, None),
        15: (30, 31, None),
        16: (18, 19, None),
        17: (32, 33, 34),
        18: (47, 48, 49),
        19: (81, 82, 83),
        20: (52, 53, 54),
        21: (56, 57, 58),
        22: (60, 61, 62),
        23: (70, 71, 72),
        24: (None, None, 92),
        25: (94, 95, None),
        26: (96, 97, None),
        27: (None, 98, None),
        28: (99, 100, 101),
        29: (101, 102, 103),
        30: (103, 104, 105),
        31: (105, 106, 107),
    }

    def __init__(self) -> None:
        super().__init__()
        self.rack_template_path: Optional[str] = None
        self.rack_template_sheets: Dict[str, List[List[Any]]] = {}
        self.request_rows: List[Dict[str, Any]] = []
        self._last_generated_folder: Optional[str] = None
        self._matched_rack_row: Optional[List[Any]] = None
        self._build_ui()
        self._populate_table()

    def _build_ui(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(14, 14, 14, 14)
        outer.setSpacing(10)

        title_row = QHBoxLayout()
        title = bold_label("RACK구매요청서 작성", size=15)
        guide = QLabel("이미지 양식과 동일한 항목 틀에서 값을 입력한 뒤 Excel 요청서를 생성합니다.")
        guide.setStyleSheet("color:#555;")
        title_row.addWidget(title)
        title_row.addWidget(guide, 1)
        outer.addLayout(title_row)

        btn_row = QHBoxLayout()
        self.btn_load_template = QPushButton("통합양식 불러오기")
        self.btn_load_quote = QPushButton("견적의뢰DATA 불러오기")
        self.btn_generate = QPushButton("요청서 생성")
        self.btn_generate_overseas = QPushButton("국외요청서 생성")
        self.btn_generate_approval = QPushButton("결재상신용 생성")
        self.btn_upload_history = QPushButton("발주이력 업로드")
        for btn, color in [
            (self.btn_load_template, "#C8E6C9"),
            (self.btn_load_quote, "#BBDEFB"),
            (self.btn_generate, "#FFF176"),
            (self.btn_generate_overseas, "#B3E5FC"),
            (self.btn_generate_approval, "#FFE0B2"),
            (self.btn_upload_history, "#E1BEE7"),
        ]:
            btn.setMinimumHeight(42)
            f = btn.font(); f.setPointSize(11); f.setBold(True); btn.setFont(f)
            tint_button(btn, color)
            btn_row.addWidget(btn)
        btn_row.addStretch(1)
        outer.addLayout(btn_row)

        self.lbl_template_status = info_label("통합양식: 미로드")
        outer.addWidget(self.lbl_template_status)

        # ── 좌우 분할: 좌=의뢰DATA+테이블 / 우=생성된 요청서 목록(420px) ──
        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)

        # 좌측: 기존 의뢰파일DATA 패널 + 메인 테이블
        left_w = QWidget()
        left_v = QVBoxLayout(left_w)
        left_v.setContentsMargins(0, 0, 0, 0)
        left_v.setSpacing(6)
        left_v.addWidget(self._build_request_data_panel())

        self.table = QTableWidget(32, 7)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setVisible(False)
        self.table.setAlternatingRowColors(False)
        self.table.setWordWrap(False)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setShowGrid(True)
        self.table.setSpan(0, self.COL_ITEM, 1, 2)
        self.table.setSpan(0, self.COL_REF_ITEM, 1, 2)
        self.table.setColumnWidth(self.COL_LABEL, 110)
        self.table.setColumnWidth(self.COL_ITEM, 300)
        self.table.setColumnWidth(self.COL_QTY, 70)
        self.table.setColumnWidth(self.COL_REF_ITEM, 300)
        self.table.setColumnWidth(self.COL_REF_QTY, 70)
        self.table.setColumnWidth(self.COL_REF_APPLY, 60)
        self.table.setColumnWidth(self.COL_REQUEST, 150)
        self.table.horizontalHeader().setSectionResizeMode(self.COL_ITEM, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(self.COL_REF_ITEM, QHeaderView.Stretch)
        for r in range(self.table.rowCount()):
            self.table.setRowHeight(r, 24)
        self.table.setRowHeight(0, 30)
        left_v.addWidget(self.table, 1)
        splitter.addWidget(left_w)

        # 우측: 생성된 요청서 목록 (420px 고정)
        right_w = QWidget()
        right_w.setFixedWidth(420)
        right_v = QVBoxLayout(right_w)
        right_v.setContentsMargins(6, 0, 0, 0)
        right_v.setSpacing(4)

        req_title_row = QHBoxLayout()
        req_title_lbl = bold_label("생성된 요청서", size=11)
        self.btn_refresh_req_list = QPushButton("새로고침")
        self.btn_refresh_req_list.setFixedHeight(24)
        self.btn_refresh_req_list.clicked.connect(self._refresh_req_list)
        req_title_row.addWidget(req_title_lbl)
        req_title_row.addStretch(1)
        req_title_row.addWidget(self.btn_refresh_req_list)
        right_v.addLayout(req_title_row)

        self.req_list = QListWidget()
        self.req_list.setToolTip("더블클릭하면 파일을 열 수 있습니다.")
        self.req_list.itemDoubleClicked.connect(self._open_req_file)
        right_v.addWidget(self.req_list, 1)
        splitter.addWidget(right_w)

        outer.addWidget(splitter, 1)

        self.btn_load_template.clicked.connect(self._load_rack_template)
        self.btn_generate.clicked.connect(self._generate_request)
        self.btn_generate_overseas.clicked.connect(self._generate_overseas_request)
        self.btn_load_quote.clicked.connect(self._load_quote_data)
        self.btn_generate_approval.clicked.connect(self._generate_approval_doc)
        self.btn_upload_history.clicked.connect(self._upload_order_history)

        # UI 렌더링 완료 후 파일 스캔 (블로킹 방지)
        QTimer.singleShot(0, self._refresh_req_list)

    def _build_request_data_panel(self) -> QFrame:
        """불러온 견적의뢰DATA를 표시하는 붉은 박스 영역."""
        frame = QFrame()
        frame.setFrameShape(QFrame.Box)
        frame.setLineWidth(2)
        frame.setStyleSheet("QFrame { border: 2px solid #D32F2F; }")
        frame.setFixedHeight(150)
        v = QVBoxLayout(frame)
        v.setContentsMargins(8, 6, 8, 8)
        v.setSpacing(6)
        title = bold_label("의뢰파일DATA", size=11)
        title.setStyleSheet("color:#B71C1C; border:none;")
        v.addWidget(title)
        self.request_data_table = QTableWidget(0, 9)
        self.request_data_table.setHorizontalHeaderLabels([
            "선택", "PR NO.", "항번", "투자정보", "수량(CH)", "공정", "5D", "FSC", "설비호기"
        ])
        self.request_data_table.verticalHeader().setVisible(False)
        self.request_data_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.request_data_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.request_data_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.request_data_table.setStyleSheet("QTableWidget { border: 1px solid #D32F2F; }")
        self.request_data_table.cellClicked.connect(self._on_request_row_clicked)
        self.request_data_table.cellDoubleClicked.connect(lambda row, _col: self._apply_request_row(row))
        h = self.request_data_table.horizontalHeader()
        for col in range(self.request_data_table.columnCount()):
            h.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        v.addWidget(self.request_data_table, 1)
        return frame

    def _populate_table(self) -> None:
        self.table.setUpdatesEnabled(False)
        try:
            self._populate_table_inner()
        finally:
            self.table.setUpdatesEnabled(True)

    def _populate_table_inner(self) -> None:
        self._set_item(0, self.COL_LABEL, "", "#DDE2E6", bold=True, editable=False)
        self._set_item(0, self.COL_ITEM, "RACK구매요청서", "#C7EAF4", bold=True, editable=False)
        self._set_item(0, self.COL_REF_ITEM, "Ref.", "#DDE2E6", bold=True, editable=False)
        self._set_item(0, self.COL_REF_APPLY, "", "#00B050", bold=True, editable=False)
        self.btn_apply_ref = QPushButton("적용")
        self.btn_apply_ref.setStyleSheet(
            "QPushButton {"
            "  background-color: #00E676; color: #000; border: 2px solid #00C853;"
            "  border-radius: 4px; font-weight: bold; padding: 2px 8px;"
            "}"
            "QPushButton:hover { background-color: #69F0AE; border-color: #00E676; }"
            "QPushButton:pressed { background-color: #00C853; }"
        )
        self.btn_apply_ref.clicked.connect(self._apply_ref_values)
        self.table.setCellWidget(0, self.COL_REF_APPLY, self.btn_apply_ref)
        self._set_item(0, self.COL_REQUEST, "비고", "#FFFFFF", bold=True, editable=False)

        for col, text, color in [
            (self.COL_ITEM, "항목", "#C7EAF4"), (self.COL_QTY, "수량", "#C7EAF4"),
            (self.COL_REF_ITEM, "항목", "#DDE2E6"), (self.COL_REF_QTY, "수량", "#DDE2E6"),
        ]:
            self._set_item(1, col, text, color, bold=True, editable=False)
        self._set_item(1, self.COL_LABEL, "", "#DDE2E6", editable=False)
        self._set_item(1, self.COL_REF_APPLY, "", "#DDE2E6", editable=False)
        self._set_item(1, self.COL_REQUEST, "", "#FFFFFF", editable=False)

        # Row 2: 인보이스 필요여부 (투자정보 기반 자동계산, 접수일자 위쪽)
        self._set_item(2, self.COL_LABEL, "인보이스 필요여부", "#EAC3E7", bold=True, editable=False)
        self._set_item(2, self.COL_ITEM, "", "#C7EAF4", editable=False)
        self._set_item(2, self.COL_QTY, "-", "#C7EAF4", editable=False)
        self._set_item(2, self.COL_REF_ITEM, "", "#DDE2E6", editable=False)
        self._set_item(2, self.COL_REF_QTY, "-", "#DDE2E6", editable=False)
        self._set_item(2, self.COL_REF_APPLY, "-", "#DDE2E6", editable=False)
        self._set_item(2, self.COL_REQUEST, "KIT투자 시 인보이스필요", "#FFFFFF", editable=False)

        # 기존 데이터 행: start=3 (row 2가 새 인보이스 행이므로)
        for i, label in enumerate(RACK_REQUEST_LEFT_LABELS, start=3):
            left_color = "#EAC3E7" if i <= 16 else "#F4DDCF"
            self._set_item(i, self.COL_LABEL, label, left_color, bold=True, editable=False)
            self._set_item(i, self.COL_ITEM, str(i - 2), "#C7EAF4")
            self._set_item(i, self.COL_QTY, str(i + 27), "#C7EAF4")
            self._set_item(i, self.COL_REF_ITEM, str(i + 56), "#DDE2E6")
            self._set_item(i, self.COL_REF_QTY, str(i + 85), "#DDE2E6")
            if i <= 16:
                self._set_item(i, self.COL_REF_APPLY, "-", "#DDE2E6", editable=False)
            else:
                self._set_item(i, self.COL_REF_APPLY, "", "#DDE2E6", editable=False)
                host = centered_checkbox(lambda _state, _row=i: self._on_ref_apply_changed(_row))
                cb = get_checkbox_from_cell(host)
                if cb:
                    cb.setChecked(True)
                self.table.setCellWidget(i, self.COL_REF_APPLY, host)
            self._set_item(i, self.COL_REQUEST, RACK_REQUEST_REF_NOTES[i - 3], "#FFFFFF", editable=False)

        # COL_ITEM 자동완성 콤보박스 (row 6=담당자 제외, row 27=간섭여부 별도 처리)
        for _row in range(3, 32):
            if _row == 6:
                continue  # 담당자: 뒤에서 manager combobox로 덮어씀
            if _row == 27:
                _cb27 = _NoScrollComboBox()
                _cb27.setEditable(True)
                _cb27.lineEdit().setAlignment(Qt.AlignCenter)
                _cb27.lineEdit().setReadOnly(True)
                _cb27.addItems(["", "간섭없음", "좌측간섭", "우측간섭", "좌/우간섭"])
                _cb27.setStyleSheet("QComboBox { background-color: #C7EAF4; }")
                self.table.setCellWidget(_row, self.COL_ITEM, _cb27)
                continue
            _cb = _NoScrollComboBox()
            _cb.setEditable(True)
            _cb.setInsertPolicy(QComboBox.NoInsert)
            _cb.lineEdit().setAlignment(Qt.AlignCenter)
            _cb.setStyleSheet("QComboBox { background-color: #C7EAF4; }")
            # FSC(16): 5D 자동 매칭 전용 — 수기 입력 불가
            if _row == 16:
                _cb.lineEdit().setReadOnly(True)
                _cb.setStyleSheet("QComboBox { background-color: #DDE8F0; }")
            else:
                _cmpl = _cb.completer()
                if _cmpl:
                    _cmpl.setFilterMode(Qt.MatchContains)
                    _cmpl.setCompletionMode(QCompleter.PopupCompletion)
            self.table.setCellWidget(_row, self.COL_ITEM, _cb)
            # 공정(7), 설비(10), 5D(15) 변경 시 실시간 Ref. 매칭
            if _row in (7, 10, 15):
                _cb.currentTextChanged.connect(self._on_key_field_changed)
            # 5D(15) 변경 시 FSC 자동 매칭
            if _row == 15:
                _cb.currentTextChanged.connect(self._auto_match_fsc)

        manager_row, manager_col = self._slot_pos(4)
        manager = _NoScrollComboBox()
        manager.setEditable(True)
        manager.lineEdit().setAlignment(Qt.AlignCenter)
        manager.addItems(RACK_REQUEST_MANAGERS)
        manager.setCurrentText(RACK_REQUEST_MANAGERS[0])
        manager.setStyleSheet("QComboBox { background-color: #C7EAF4; }")
        self.table.setCellWidget(manager_row, manager_col, manager)

        # RACK발주 매칭 REF 슬롯 초기화 (rows 17-31)
        for row in range(17, 32):
            item = self.table.item(row, self.COL_REF_ITEM)
            if item:
                item.setText("")
            item = self.table.item(row, self.COL_REF_QTY)
            if item:
                item.setText("")

        # COL_QTY 고정 "-" (rows 2-16, 인보이스 포함)
        for row in range(2, 17):
            self._set_item(row, self.COL_QTY, "-", "#C7EAF4", editable=False)

        # COL_REF_QTY 고정 "-" (rows 2-16)
        for row in range(2, 17):
            self._set_item(row, self.COL_REF_QTY, "-", "#DDE2E6", editable=False)

        # COL_REF_ITEM 초기값 "" (rows 3-16, 접수일자~RACK CH)
        for row in range(3, 17):
            item = self.table.item(row, self.COL_REF_ITEM)
            if item:
                item.setText("")

        # slot 80 (row 24, COL_REF_ITEM): 고정 "-"
        self._set_item(24, self.COL_REF_ITEM, "-", "#DDE2E6", editable=False)

        # slots 110-112 (rows 25-27, COL_REF_QTY): 고정 "-"
        for row in (25, 26, 27):
            self._set_item(row, self.COL_REF_QTY, "-", "#DDE2E6", editable=False)

        # slot 83 (row 27, COL_REF_ITEM): 간섭여부 드롭다운 (가운데 정렬)
        combo_83 = _NoScrollComboBox()
        combo_83.setEditable(True)
        combo_83.lineEdit().setAlignment(Qt.AlignCenter)
        combo_83.lineEdit().setReadOnly(True)
        combo_83.addItems(["", "간섭없음", "좌측간섭", "우측간섭", "좌/우간섭"])
        combo_83.setStyleSheet("QComboBox { background-color: #DDE2E6; }")
        self.table.setCellWidget(27, self.COL_REF_ITEM, combo_83)

        self._set_default_values()
        # _populate_table_inner 끝 — setUpdatesEnabled(True)는 _populate_table에서 처리

    def _set_default_values(self) -> None:
        """견적의뢰DATA와 무관한 기본값 설정 (통합양식 로드 후에도 재적용)."""
        inv_item = self.table.item(2, self.COL_ITEM)
        if inv_item:
            inv_item.setText("투자미확인")
        self._set_slot_value(1, datetime.now().strftime("%Y-%m-%d"))
        self._set_slot_value(2, "입력필요")
        self._set_slot_value(3, "구두발주")
        # 담당자/공정/세부공정/라인/설비/설비MODEL/설비호기/PUMP MODEL/수량(CH)/5D: 공란
        for slot in (4, 5, 6, 7, 8, 9, 10, 11, 12, 13):
            self._set_slot_value(slot, "")
        # FSC 공란 (5D 입력 시 자동 채워짐)
        fsc_widget = self.table.cellWidget(16, self.COL_ITEM)
        if isinstance(fsc_widget, QComboBox):
            fsc_widget.setCurrentText("")

    def _set_item(self, row: int, col: int, text: str, color: str, *, bold: bool = False,
                  editable: bool = True, size: Optional[int] = None) -> None:
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        item.setBackground(QBrush(QColor(color)))
        flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable
        if editable:
            flags |= Qt.ItemIsEditable
        item.setFlags(flags)
        font = item.font()
        font.setBold(bold)
        if size:
            font.setPointSize(size)
        item.setFont(font)
        self.table.setItem(row, col, item)

    def reset_page(self) -> None:
        """좌측 하단 초기화 버튼에서 호출되는 현재 페이지 초기화 진입점."""
        self.rack_template_path = None
        self.rack_template_sheets = {}
        self.request_rows = []
        self.lbl_template_status.setText("통합양식: 미로드")
        self.request_data_table.setRowCount(0)
        self._populate_table()

    def _is_ref_apply_checked(self, row: int) -> bool:
        cb = get_checkbox_from_cell(self.table.cellWidget(row, self.COL_REF_APPLY))
        return bool(cb and cb.isChecked())

    def _on_ref_apply_changed(self, row: int) -> None:
        """COL_REF_APPLY 체크박스 토글 시 해당 행의 Ref 값을 항목/수량에 즉시 반영."""
        if not self._is_ref_apply_checked(row):
            return
        ref_item = self._text(row, self.COL_REF_ITEM)
        ref_qty  = self._text(row, self.COL_REF_QTY)
        w = self.table.cellWidget(row, self.COL_ITEM)
        if isinstance(w, QComboBox):
            w.setCurrentText(ref_item)
        else:
            it = self.table.item(row, self.COL_ITEM)
            if it:
                it.setText(ref_item)
        it_qty = self.table.item(row, self.COL_QTY)
        if it_qty:
            it_qty.setText(ref_qty)

    def _clear_lower_rows(self) -> None:
        """rows 17-31(RACK CH~3단 모니터링 Cable) 항목/수량 값 공란으로 초기화."""
        for row in range(17, 32):
            w = self.table.cellWidget(row, self.COL_ITEM)
            if isinstance(w, QComboBox):
                w.setCurrentText("")
            else:
                it = self.table.item(row, self.COL_ITEM)
                if it:
                    it.setText("")
            it_qty = self.table.item(row, self.COL_QTY)
            if it_qty:
                it_qty.setText("")

    def _apply_all_checked_refs(self) -> None:
        """rows 17-31 중 체크된 행의 Ref 항목/수량을 COL_ITEM/COL_QTY에 즉시 반영."""
        for row in range(17, 32):
            self._on_ref_apply_changed(row)

    def _slot_pos(self, slot: int) -> Tuple[int, int]:
        """이미지 양식의 1~58 입력칸 번호를 테이블 row/col로 변환.
        row 2 = 인보이스 필요여부(자동계산), row 3~ = slot 1~ 대응."""
        if 1 <= slot <= 29:
            return slot + 2, self.COL_ITEM
        if 30 <= slot <= 58:
            return slot - 27, self.COL_QTY
        raise ValueError(f"지원하지 않는 RACK 요청서 칸 번호: {slot}")

    def _set_slot_value(self, slot: int, value: Any) -> None:
        row, col = self._slot_pos(slot)
        widget = self.table.cellWidget(row, col)
        if isinstance(widget, QComboBox):
            widget.setCurrentText(s(value))
            return
        item = self.table.item(row, col)
        if item:
            item.setText(s(value))

    def _reset_slot_values(self, slots: List[int]) -> None:
        for slot in slots:
            self._set_slot_value(slot, "")

    def _date_yyyy_mm_dd(self, value: Any) -> str:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        text = s(value)
        if not text:
            return ""
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
            try:
                return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
        if "T" in text:
            text = text.split("T", 1)[0]
        if " " in text:
            text = text.split(" ", 1)[0]
        return text.replace("/", "-")

    def _split_line_process(self, value: Any) -> Tuple[str, str]:
        text = s(value)
        if "_" not in text:
            return text, ""
        line, process = text.split("_", 1)
        return line, process

    def _pump_model_from_g(self, value: Any) -> str:
        parts = [part.strip() for part in s(value).split(",") if part.strip()]
        for idx, part in enumerate(parts):
            if part.upper() == "LOT" and idx + 1 < len(parts):
                return parts[idx + 1]
        return parts[-1] if parts else ""

    def _rack_count_from_v(self, value: Any) -> str:
        match = re.search(r"(\d+(?:\.\d+)?)\s*RACK", s(value), re.IGNORECASE)
        return match.group(1) if match else ""

    def _populate_item_combos(self) -> None:
        """통합양식 로드 후 COL_ITEM 콤보박스에 RACK발주 열 고유값을 채운다."""
        rack_rows = self.rack_template_sheets.get("RACK발주", [])
        for row, info in self._ROW_RACK_MAP.items():
            col_0idx = info[0]
            if col_0idx is None:
                continue
            widget = self.table.cellWidget(row, self.COL_ITEM)
            if not isinstance(widget, QComboBox):
                continue
            seen: set = {widget.itemText(i) for i in range(widget.count())}
            widget.blockSignals(True)
            for r in rack_rows:
                if col_0idx < len(r):
                    val = s(r[col_0idx])
                    if val and val not in seen:
                        widget.addItem(val)
                        seen.add(val)
            widget.blockSignals(False)

    def _lookup_fsc_from_template(self, value: Any) -> str:
        key = s(value)
        if not key:
            return ""
        for row in self.rack_template_sheets.get("FSC All List", []):
            if len(row) >= 8 and key in s(row[7]):
                return s(row[1]) if len(row) >= 2 else ""
        return ""

    def _apply_ref_values(self) -> None:
        """Ref. 적용 체크 행의 Ref 항목/수량을 좌측 수기입력 영역으로 복사."""
        applied = 0
        for row in range(2, self.table.rowCount()):
            if not self._is_ref_apply_checked(row):
                continue
            ref_item = self._text(row, self.COL_REF_ITEM)
            ref_qty  = self._text(row, self.COL_REF_QTY)
            w = self.table.cellWidget(row, self.COL_ITEM)
            if isinstance(w, QComboBox):
                w.setCurrentText(ref_item)
            else:
                it = self.table.item(row, self.COL_ITEM)
                if it:
                    it.setText(ref_item)
            it_qty = self.table.item(row, self.COL_QTY)
            if it_qty:
                it_qty.setText(ref_qty)
            applied += 1
        QMessageBox.information(self, "Ref. 적용", f"Ref. 항목/수량 {applied}건을 수기입력 칸에 반영했습니다.")

    def _load_rack_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "통합양식 선택", "", "Excel Files (*.xlsx)")
        if not path:
            return

        def _do_load(p):
            from openpyxl import load_workbook
            required = ["RACK발주", "RACK발주양식", "FSC All List"]
            wb = load_workbook(p, data_only=True)
            missing = [name for name in required if name not in wb.sheetnames]
            if missing:
                wb.close()
                raise ValueError("필수 시트 없음:\n" + "\n".join(missing))
            loaded: Dict[str, List[List[Any]]] = {}
            counts: List[str] = []
            for name in required:
                ws = wb[name]
                rows: List[List[Any]] = []
                for row in ws.iter_rows(values_only=True):
                    values = list(row)
                    if any(s(cell) for cell in values):
                        rows.append(values)
                loaded[name] = rows
                counts.append(f"{name} {len(rows)}행")
            wb.close()
            return loaded, counts

        def _on_result(res):
            loaded, counts = res
            self.rack_template_path = path
            self.rack_template_sheets = loaded
            self._populate_item_combos()
            self._set_default_values()
            self.lbl_template_status.setText(f"통합양식: {os.path.basename(path)} ({', '.join(counts)})")
            QMessageBox.information(self, "완료", "통합양식 LOAD 완료\n" + "\n".join(counts))

        def _on_error(e):
            if "필수 시트 없음" in e:
                QMessageBox.warning(self, "통합양식 오류", e.split("\n\n")[0])
            else:
                QMessageBox.critical(self, "불러오기 오류", f"통합양식을 읽을 수 없습니다.\n{e}")

        _BgWorker.run_with_progress(self, "통합양식 읽는 중...", _do_load, path,
                                    on_result=_on_result, on_error=_on_error)

    def _load_quote_data(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "견적의뢰DATA 선택", "", "Excel Files (*.xlsx)")
        if not path:
            return

        def _on_result(res):
            sheet_name, rows = res
            if not rows:
                QMessageBox.warning(self, "견적의뢰DATA", "읽을 데이터가 없습니다.")
                return
            self.request_rows = rows
            self._fill_request_data_table(rows)
            self._apply_request_row(0, show_message=False)
            QMessageBox.information(self, "완료",
                f"견적의뢰DATA {len(rows)}건을 붉은 박스 영역에 표시했습니다.\n시트: {sheet_name}")

        _BgWorker.run_with_progress(self, "견적의뢰DATA 읽는 중...",
                                    excel_io.parse_request_xlsx, path,
                                    on_result=_on_result)

    def _fill_request_data_table(self, rows: List[Dict[str, Any]]) -> None:
        self.request_data_table.setRowCount(0)
        for i, rd in enumerate(rows):
            self.request_data_table.insertRow(i)
            self.request_data_table.setCellWidget(
                i, 0, centered_checkbox(lambda _state, _r=i: self._on_request_target_changed(_r))
            )
            values = [
                s(rd.get("D")),
                s(rd.get("E")),
                parse_invest_info(rd.get("G")),
                fmt_qty(to_float(rd.get("H"))) if rd.get("H") is not None else "",
                s(rd.get("K")),
                s(rd.get("V")),
                s(rd.get("X")),
                s(rd.get("Z")),
            ]
            for col, value in enumerate(values, start=1):
                item = QTableWidgetItem(value)
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                item.setTextAlignment(Qt.AlignCenter)
                self.request_data_table.setItem(i, col, item)
            self.request_data_table.resizeRowToContents(i)
        self.request_data_table.resizeColumnsToContents()
        if rows:
            self._set_request_checked(0, True)
            self.request_data_table.selectRow(0)

    def _set_request_checked(self, row: int, checked: bool) -> None:
        cb = get_checkbox_from_cell(self.request_data_table.cellWidget(row, 0))
        if cb:
            cb.blockSignals(True)
            cb.setChecked(checked)
            cb.blockSignals(False)

    def _is_request_checked(self, row: int) -> bool:
        cb = get_checkbox_from_cell(self.request_data_table.cellWidget(row, 0))
        return bool(cb and cb.isChecked())

    def _checked_request_rows(self) -> List[int]:
        return [r for r in range(self.request_data_table.rowCount()) if self._is_request_checked(r)]

    def _highlight_request_row(self, row: int, color: str) -> None:
        """request_data_table 특정 행의 배경색 설정 (color="" 이면 초기화)."""
        for col in range(1, self.request_data_table.columnCount()):
            item = self.request_data_table.item(row, col)
            if item:
                if color:
                    item.setBackground(QBrush(QColor(color)))
                else:
                    item.setBackground(QBrush())

    def _mark_request_generated(self, row: int) -> None:
        """생성 완료 행을 붉은 파스텔 음영으로 표시."""
        self._highlight_request_row(row, "#FFCDD2")

    def _on_request_row_clicked(self, row: int, _col: int) -> None:
        """클릭한 행만 파스텔 파랑; 이전 클릭 행 색 해제 (생성완료·체크 행 제외)."""
        n = self.request_data_table.rowCount()
        for r in range(n):
            item1 = self.request_data_table.item(r, 1)
            if item1 is None:
                continue
            bg = item1.background().color().name()
            if r == row:
                if bg != "#ffcdd2":
                    self._highlight_request_row(r, "#E3F2FD")
            else:
                if bg == "#e3f2fd" and not self._is_request_checked(r):
                    self._highlight_request_row(r, "")

    def _on_request_target_changed(self, row: int) -> None:
        if self._is_request_checked(row):
            self._highlight_request_row(row, "#E3F2FD")
            self._apply_request_row(row, show_message=False)
        else:
            self._highlight_request_row(row, "")
            self._reset_slot_values(list(range(1, 15)))
            self._fill_rack_ref_slots(None)
            self._clear_lower_rows()

    def _apply_request_row(self, row: int, show_message: bool = True) -> None:
        if row < 0 or row >= len(self.request_rows):
            return
        rd = self.request_rows[row]
        line, process = self._split_line_process(rd.get("K"))
        defaults = {
            1: datetime.now().strftime("%Y-%m-%d"),
            2: self._date_yyyy_mm_dd(rd.get("AA")),
            3: rd.get("D"),
            5: process,
            6: rd.get("N"),
            7: line,
            8: rd.get("X"),
            9: rd.get("Y"),
            10: rd.get("Z"),
            11: self._pump_model_from_g(rd.get("G")),
            12: self._rack_count_from_v(rd.get("V")),
            13: rd.get("V"),
            14: self._lookup_fsc_from_template(rd.get("V")),
        }
        for slot, value in defaults.items():
            self._set_slot_value(slot, value)
        self._set_request_checked(row, True)
        self.request_data_table.selectRow(row)

        # RACK발주 시트에서 공정/설비/5D로 매칭 행 조회 후 REF 슬롯 채우기
        # 공정: K열을 "_"로 split한 process 부분 사용 (라인_공정 → 공정만)
        rack_process = process
        rack_vendor  = s(rd.get("X"))
        rack_5d      = s(rd.get("V"))
        rack_row = self._lookup_rack_row(rack_process, rack_vendor, rack_5d)
        self._fill_rack_ref_slots(rack_row)
        self._apply_all_checked_refs()

        invest_info = parse_invest_info(rd.get("G"))
        invoice_text = "인보이스 필요" if "KIT" in invest_info.upper() else "해당 없음"
        inv_item = self.table.item(2, self.COL_ITEM)
        if inv_item:
            inv_item.setText(invoice_text)

        if show_message:
            if rack_row is not None:
                msg = f"{row + 1}번째 의뢰파일DATA를 반영했습니다.\n[RACK발주 매칭 성공] 공정={rack_process} / 설비={rack_vendor} / 5D={rack_5d}"
            else:
                msg = (f"{row + 1}번째 의뢰파일DATA를 반영했습니다.\n"
                       f"[RACK발주 매칭 없음] 공정={rack_process} / 설비={rack_vendor} / 5D={rack_5d}\n"
                       f"통합양식 RACK발주 시트에서 일치하는 행이 없습니다.")
            QMessageBox.information(self, "의뢰파일DATA 적용", msg)

    def _lookup_rack_row(self, process: str, vendor: str, code_5d: str) -> Optional[List[Any]]:
        """RACK발주 시트에서 공정/설비/5D가 일치하는 가장 마지막 행을 반환 (역순 탐색).

        비교 시 앞뒤 공백 제거 + 대소문자 무시.
        W열(col 23) = 공정, Z열(col 26) = 설비, AE열(col 31) = 5D.
        """
        rows = self.rack_template_sheets.get("RACK발주", [])
        proc = s(process).upper().replace(" ", "")
        vend = s(vendor).upper().replace(" ", "")
        d5   = s(code_5d).upper().replace(" ", "")
        if not (proc and vend and d5):
            logger.debug("RACK발주 조회 스킵: 공정=%r 설비=%r 5D=%r", process, vendor, code_5d)
            return None
        logger.debug("RACK발주 조회: 공정=%r 설비=%r 5D=%r (총 %d행)", proc, vend, d5, len(rows))
        for row in reversed(rows):
            if len(row) <= 30:
                continue
            w  = s(row[22]).upper().replace(" ", "")   # W열
            z  = s(row[25]).upper().replace(" ", "")   # Z열
            ae = s(row[30]).upper().replace(" ", "")   # AE열
            if w == proc and z == vend and ae == d5:
                logger.debug("RACK발주 매칭 성공: W=%r Z=%r AE=%r", row[22], row[25], row[30])
                return row
        logger.debug("RACK발주 매칭 없음: 공정=%r 설비=%r 5D=%r", proc, vend, d5)
        return None

    def _fill_rack_ref_slots(self, rack_row: Optional[List[Any]]) -> None:
        """찾은 RACK발주 행의 값으로 REF 슬롯(59~87, 102~116)을 채운다."""
        def _get(col_0idx: int) -> str:
            if rack_row is None or col_0idx >= len(rack_row):
                return ""
            val = rack_row[col_0idx]
            if isinstance(val, (datetime, date)):
                return val.strftime("%Y-%m-%d")
            return s(val)

        # COL_REF_ITEM (rows 3-16, slots 59-72): RACK발주 기본 정보 열
        basic_ref_cols = {
            3:  2,   # C열  (접수 일자)
            4:  5,   # F열  (납품 요청)
            5:  8,   # I열  (PR NO.)
            6:  10,  # K열  (담당자)
            7:  22,  # W열  (공 정)
            8:  23,  # X열  (세부 공정)
            9:  24,  # Y열  (라인)
            10: 25,  # Z열  (설 비)
            11: 26,  # AA열 (설비MODEL)
            12: 27,  # AB열 (설비호기)
            13: 28,  # AC열 (PUMP MODEL)
            14: 29,  # AD열 (수량(CH))
            15: 30,  # AE열 (5D)
            16: 18,  # S열  (FSC)
        }
        for row, col_0idx in basic_ref_cols.items():
            item = self.table.item(row, self.COL_REF_ITEM)
            if item:
                item.setText(_get(col_0idx))

        # COL_REF_ITEM (rows 17-31, slots 73-87)
        ref_item_cols = {
            17: 32,   # AG(33)
            18: 47,   # AV(48)
            19: 81,   # CD(82)
            20: 52,   # BA(53)
            21: 56,   # BE(57)
            22: 60,   # BI(61)
            23: 70,   # BS(71)
            # 24: slot 80 → 고정 "-"
            25: 94,   # CQ(95)
            26: 96,   # CS(97)
            # 27: 간섭여부 → QComboBox (CT열)
            28: 99,   # CV(100)
            29: 101,  # CX(102)
            30: 103,  # CZ(104)
            31: 105,  # DB(106)
        }
        for row, col_0idx in ref_item_cols.items():
            item = self.table.item(row, self.COL_REF_ITEM)
            if item:
                item.setText(_get(col_0idx))

        # slot 80 (row 24): 항상 "-"
        item = self.table.item(24, self.COL_REF_ITEM)
        if item:
            item.setText("-")

        # row 27 (간섭여부) COL_REF_ITEM: QComboBox → CT열(0-indexed 97)
        combo_27 = self.table.cellWidget(27, self.COL_REF_ITEM)
        if isinstance(combo_27, QComboBox):
            combo_27.setCurrentText(_get(97))

        # COL_REF_QTY (rows 17-31, slots 102-116)
        ref_qty_cols = {
            17: 33,   # AH(34)
            18: 48,   # AW(49)
            19: 82,   # CE(83)
            20: 53,   # BB(54)
            21: 57,   # BF(58)
            22: 61,   # BJ(62)
            23: 71,   # BT(72)
            24: 91,   # CN(92)
            # 25, 26, 27: slots 110-112 → 고정 "-"
            28: 100,  # CW(101)
            29: 102,  # CY(103)
            30: 104,  # DA(105)
            31: 106,  # DC(107)
        }
        for row, col_0idx in ref_qty_cols.items():
            item = self.table.item(row, self.COL_REF_QTY)
            if item:
                item.setText(_get(col_0idx))

        # slots 110-112 (rows 25-27): 항상 "-"
        for row in (25, 26, 27):
            item = self.table.item(row, self.COL_REF_QTY)
            if item:
                item.setText("-")

    def _text(self, row: int, col: int) -> str:
        widget = self.table.cellWidget(row, col)
        if isinstance(widget, QComboBox):
            return widget.currentText().strip()
        cb = get_checkbox_from_cell(widget)
        if cb:
            return "☑" if cb.isChecked() else ""
        item = self.table.item(row, col)
        return item.text().strip() if item else ""

    def _on_key_field_changed(self, _text: str = "") -> None:
        """공정/설비/5D 수기 수정 시 실시간 Ref. 매칭."""
        if not self.rack_template_sheets.get("RACK발주"):
            return
        process = self._text(7,  self.COL_ITEM)
        vendor  = self._text(10, self.COL_ITEM)
        code_5d = self._text(15, self.COL_ITEM)
        rack_row = self._lookup_rack_row(process, vendor, code_5d)
        self._fill_rack_ref_slots(rack_row)
        self._apply_all_checked_refs()

    def _auto_match_fsc(self, _text: str = "") -> None:
        """5D 입력 시 FSC All List에서 FSC를 실시간 자동 매칭."""
        code_5d = self._text(15, self.COL_ITEM)
        fsc = self._lookup_fsc_from_template(code_5d)
        if not fsc:
            return
        w = self.table.cellWidget(16, self.COL_ITEM)
        if isinstance(w, QComboBox):
            w.blockSignals(True)
            w.setCurrentText(fsc)
            w.blockSignals(False)
        else:
            it = self.table.item(16, self.COL_ITEM)
            if it:
                it.setText(fsc)

    def _upload_order_history(self) -> None:
        """발주이력 업로드: 선택한 요청서 xlsx의 RACK발주양식 2행을 통합양식 RACK발주 시트에 추가."""
        if not self.rack_template_path:
            QMessageBox.warning(self, "오류", "통합양식을 먼저 불러오세요.")
            return
        if not self._is_file_writable(self.rack_template_path):
            QMessageBox.warning(self, "파일 잠금",
                "통합양식 파일이 다른 프로그램(Excel)에서 열려 있어 수정할 수 없습니다.\n"
                "Excel에서 통합양식 파일을 닫은 후 다시 실행해 주세요.")
            return

        paths, _ = QFileDialog.getOpenFileNames(
            self, "발주이력 요청서 선택", "", "Excel Files (*.xlsx)")
        if not paths:
            return

        tmpl_path = self.rack_template_path

        def _do_upload(tmpl, src_paths):
            from openpyxl import load_workbook as _lw
            tmpl_wb = _lw(tmpl)
            if "RACK발주" not in tmpl_wb.sheetnames:
                tmpl_wb.close()
                raise ValueError("통합양식에 'RACK발주' 시트가 없습니다.")
            ws = tmpl_wb["RACK발주"]

            first_empty = ws.max_row + 1
            for r in range(1, ws.max_row + 2):
                if ws.cell(r, 3).value is None or str(ws.cell(r, 3).value).strip() == "":
                    first_empty = r
                    break

            existing_rows: List[Dict] = []
            for r in range(1, first_empty):
                if ws.cell(r, 3).value is not None and str(ws.cell(r, 3).value).strip():
                    rd = {c: ws.cell(r, c).value
                          for c in range(1, (ws.max_column or 0) + 1)
                          if ws.cell(r, c).value is not None}
                    existing_rows.append(rd)

            added = 0
            duplicates: List[str] = []
            errors: List[str] = []
            for path in src_paths:
                try:
                    src_wb = _lw(path, data_only=True)
                    if "RACK발주양식" not in src_wb.sheetnames:
                        errors.append(f"{os.path.basename(path)}: 'RACK발주양식' 시트 없음")
                        src_wb.close()
                        continue
                    src_ws = src_wb["RACK발주양식"]
                    new_row: Dict = {cell.column: cell.value
                                     for cell in src_ws[2] if cell.value is not None}
                    src_wb.close()

                    if any(ex == new_row for ex in existing_rows):
                        duplicates.append(os.path.basename(path))
                        continue

                    for col, val in new_row.items():
                        ws.cell(first_empty, col, val)
                    existing_rows.append(new_row)
                    first_empty += 1
                    added += 1
                except Exception as e:
                    errors.append(f"{os.path.basename(path)}: {e}")

            tmpl_wb.save(tmpl)
            tmpl_wb.close()
            return added, duplicates, errors

        def _on_result(res):
            added, duplicates, errors = res
            msg = f"{added}건 발주이력을 RACK발주 시트에 추가했습니다."
            if duplicates:
                msg += "\n\n중복 항목 발견 (등록 제외됨):\n" + "\n".join(duplicates)
            if errors:
                msg += "\n\n오류:\n" + "\n".join(errors)
            QMessageBox.information(self, "발주이력 업로드 완료", msg)

        _BgWorker.run_with_progress(self, "발주이력 업로드 중...", _do_upload, tmpl_path, paths,
                                    on_result=_on_result)

    def _refresh_req_list(self) -> None:
        """RACK구매요청서 폴더의 xlsx 파일 목록을 우측 패널에 갱신."""
        root = os.path.join(exe_dir(), "RACK구매요청서")
        self.req_list.clear()
        if not os.path.isdir(root):
            return
        files = []
        for dirpath, _, filenames in os.walk(root):
            for fn in filenames:
                if fn.endswith(".xlsx") and not fn.startswith("~$"):
                    full = os.path.join(dirpath, fn)
                    files.append((os.path.getmtime(full), full))
        files.sort(reverse=True)  # 최신순
        for _, full in files:
            rel = os.path.relpath(full, root)
            item = QListWidgetItem(rel)
            item.setData(Qt.UserRole, full)
            item.setToolTip(full)
            self.req_list.addItem(item)

    def _open_req_file(self, item: QListWidgetItem) -> None:
        """더블클릭한 요청서 파일을 기본 프로그램으로 열기."""
        path = item.data(Qt.UserRole)
        if path and os.path.exists(path):
            os.startfile(path)
        else:
            QMessageBox.warning(self, "파일 없음", f"파일을 찾을 수 없습니다:\n{path}")

    @staticmethod
    def _is_file_writable(path: str) -> bool:
        """파일이 쓰기 가능한지 확인 (다른 프로세스에 잠겨 있으면 False)."""
        try:
            with open(path, 'r+b'):
                return True
        except (IOError, PermissionError, OSError):
            return False

    def _do_generate_one(self) -> str:
        """현재 self.table 상태로 파일 1개 생성 후 경로 반환."""
        ymd      = datetime.now().strftime("%y%m%d")
        process  = safe_filename(self._text(7,  self.COL_ITEM)) or "RACK"
        line     = safe_filename(self._text(9,  self.COL_ITEM)) or ""
        equip_no = safe_filename(self._text(12, self.COL_ITEM)) or "설비호기"
        folder_mid = f"{ymd}_{line}_{process}" if line else f"{ymd}_{process}"
        folder = ensure_dir(os.path.join(
            exe_dir(), "RACK구매요청서", f"{folder_mid}_RACK구매요청서"))
        out_path = unique_path(os.path.join(folder, f"{equip_no}_RACK구매요청서.xlsx"))
        shutil.copy2(self.rack_template_path, out_path)
        self._direct_patch_xlsx(out_path)
        self._last_generated_folder = folder
        return out_path

    def _generate_request(self) -> None:
        if not self.rack_template_path:
            QMessageBox.warning(self, "오류", "통합양식을 먼저 불러오세요.")
            return

        checked_rows = self._checked_request_rows()

        if not checked_rows or len(checked_rows) == 1:
            # 체크 없음 또는 1개 → 현재 테이블 상태 그대로 생성 (수기 수정 반영)
            try:
                out_path = self._do_generate_one()
                QMessageBox.information(self, "완료",
                    f"RACK 구매요청서 생성 완료\n{os.path.basename(out_path)}")
                if checked_rows:
                    self._mark_request_generated(checked_rows[0])
            except Exception as e:
                QMessageBox.critical(self, "생성 오류", f"요청서 생성 중 오류:\n{e}")
                logger.error("요청서 생성 오류", exc_info=True)
        else:
            # 체크된 행 여러 개 → 각 행 적용 후 순서대로 생성
            # rows 17-31(RACK CH~3단모니터링Cable) 현재 값을 저장해 각 파일에 그대로 반영
            saved_lower: Dict[Tuple[int, int], str] = {}
            for r in range(17, 32):
                w = self.table.cellWidget(r, self.COL_ITEM)
                saved_lower[(r, self.COL_ITEM)] = (
                    w.currentText() if isinstance(w, QComboBox)
                    else (self.table.item(r, self.COL_ITEM).text()
                          if self.table.item(r, self.COL_ITEM) else ""))
                it_q = self.table.item(r, self.COL_QTY)
                saved_lower[(r, self.COL_QTY)] = it_q.text() if it_q else ""

            success, errors = [], []
            for row in checked_rows:
                try:
                    self._apply_request_row(row, show_message=False)
                    # rows 17-31 복원 (의뢰DATA 적용이 덮어쓴 값 되돌리기)
                    for r in range(17, 32):
                        item_val = saved_lower.get((r, self.COL_ITEM), "")
                        w = self.table.cellWidget(r, self.COL_ITEM)
                        if isinstance(w, QComboBox):
                            w.setCurrentText(item_val)
                        else:
                            it = self.table.item(r, self.COL_ITEM)
                            if it:
                                it.setText(item_val)
                        qty_val = saved_lower.get((r, self.COL_QTY), "")
                        it_q = self.table.item(r, self.COL_QTY)
                        if it_q:
                            it_q.setText(qty_val)
                    out_path = self._do_generate_one()
                    success.append(os.path.basename(out_path))
                    self._mark_request_generated(row)
                except Exception as e:
                    errors.append(f"행 {row + 1}: {e}")
                    logger.error("멀티 요청서 생성 오류 (행 %d)", row, exc_info=True)
            msg = f"총 {len(success)}개 생성 완료:\n" + "\n".join(success)
            if errors:
                msg += f"\n\n실패 {len(errors)}건:\n" + "\n".join(errors)
            QMessageBox.information(self, "생성 완료", msg)
        self._refresh_req_list()

    def _generate_overseas_request(self) -> None:
        """국외요청서 생성: 의뢰파일DATA 체크 여부와 무관하게 현재 입력값으로 요청서 1개 생성."""
        if not self.rack_template_path:
            QMessageBox.warning(self, "오류", "통합양식을 먼저 불러오세요.")
            return
        try:
            out_path = self._do_generate_one()
            QMessageBox.information(self, "완료",
                f"국외 RACK 구매요청서 생성 완료\n{os.path.basename(out_path)}")
        except Exception as e:
            QMessageBox.critical(self, "생성 오류", f"요청서 생성 중 오류:\n{e}")
            logger.error("국외요청서 생성 오류", exc_info=True)
        self._refresh_req_list()

    # ── xlsx 직접 조작 helper (ET 미사용, regex/raw bytes) ────────────────

    @staticmethod
    def _xlsx_find_sheet_file(wb_bytes: bytes, rels_bytes: bytes, sheet_name: str) -> Optional[str]:
        """workbook.xml / .rels에서 시트 파일의 zip 내부 경로 반환."""
        wb_str   = wb_bytes.decode('utf-8', errors='replace')
        rels_str = rels_bytes.decode('utf-8', errors='replace')
        rid = None
        for m in re.finditer(r'<sheet\b([^>]*/?>)', wb_str):
            attrs = m.group(1)
            nm_m  = re.search(r'\bname="([^"]*)"', attrs)
            rid_m = re.search(r'\br:id="([^"]*)"', attrs)
            if nm_m and rid_m and nm_m.group(1) == sheet_name:
                rid = rid_m.group(1)
                break
        if not rid:
            return None
        for m in re.finditer(r'<Relationship\b([^>]*/?>)', rels_str):
            attrs = m.group(1)
            id_m  = re.search(r'\bId="([^"]*)"', attrs)
            tgt_m = re.search(r'\bTarget="([^"]*)"', attrs)
            if id_m and id_m.group(1) == rid and tgt_m:
                tgt = tgt_m.group(1).lstrip('/')
                return tgt if tgt.startswith('xl/') else f'xl/{tgt}'
        return None

    @staticmethod
    def _xlsx_patch_wb_visibility(wb_bytes: bytes, visible_sheet: str) -> bytes:
        """workbook.xml에서 visible_sheet 외 모든 시트 veryHidden 처리."""
        xml = wb_bytes.decode('utf-8', errors='replace')

        def repl(m: re.Match) -> str:
            tag = m.group(0)
            nm_m = re.search(r'\bname="([^"]*)"', tag)
            nm = nm_m.group(1) if nm_m else ''
            tag = re.sub(r'\s+state=(?:"[^"]*"|\'[^\']*\')', '', tag)
            if nm and nm != visible_sheet:
                tag = re.sub(r'(\s*/?>)$', r' state="veryHidden"\1', tag)
            return tag

        xml = re.sub(r'<sheet\b[^>]*/>', repl, xml)

        if 'fullCalcOnLoad' not in xml:
            if '<calcPr' in xml:
                xml = re.sub(r'(<calcPr\b[^/]*?)(/?>)',
                             lambda m: m.group(1) + ' fullCalcOnLoad="1"' + m.group(2), xml)
            else:
                xml = xml.replace('</workbook>', '<calcPr fullCalcOnLoad="1"/></workbook>', 1)

        return xml.encode('utf-8')

    @staticmethod
    def _xlsx_patch_sheet_row(sheet_bytes: bytes, row_num: int,
                               col_vals: Dict[int, Any], get_col_letter,
                               cell_styles: Optional[Dict[str, str]] = None) -> bytes:
        """sheet XML 특정 행을 raw 문자열로 삽입/교체 (ET 미사용)."""
        xml = sheet_bytes.decode('utf-8', errors='replace')
        if not col_vals:
            return sheet_bytes

        styles = cell_styles or {}

        def _col_1idx(letters: str) -> int:
            n = 0
            for ch in letters:
                n = n * 26 + (ord(ch) - 64)
            return n

        all_cols = set(col_vals.keys()) | {_col_1idx(l) for l in styles}
        cells_xml = ''
        for col_idx in sorted(all_cols):
            col_l = get_col_letter(col_idx)
            ref   = f'{col_l}{row_num}'
            s_attr = f' s="{styles[col_l]}"' if col_l in styles else ''
            val   = col_vals.get(col_idx)
            if val is None:
                cells_xml += f'<c r="{ref}"{s_attr}/>'
            elif isinstance(val, (int, float)):
                num_s = str(int(val)) if float(val) == int(val) else str(val)
                cells_xml += f'<c r="{ref}"{s_attr}><v>{num_s}</v></c>'
            elif isinstance(val, datetime):
                serial = (val.date() - date(1899, 12, 30)).days
                cells_xml += f'<c r="{ref}"{s_attr}><v>{serial}</v></c>'
            elif isinstance(val, date):
                serial = (val - date(1899, 12, 30)).days
                cells_xml += f'<c r="{ref}"{s_attr}><v>{serial}</v></c>'
            else:
                sv = (str(val).replace('&', '&amp;')
                      .replace('<', '&lt;').replace('>', '&gt;'))
                if sv:
                    cells_xml += f'<c r="{ref}"{s_attr} t="inlineStr"><is><t>{sv}</t></is></c>'
                else:
                    cells_xml += f'<c r="{ref}"{s_attr}/>'

        row_xml = f'<row r="{row_num}">' + cells_xml + '</row>'
        rn = str(row_num)

        # 기존 같은 번호의 행 제거
        xml = re.sub(rf'<row\b[^>]*\br="{rn}"[^>]*/>', '', xml)
        xml = re.sub(rf'<row\b[^>]*\br="{rn}"[^>]*>.*?</row>', '', xml, flags=re.DOTALL)

        # 삽입 위치: row_num보다 큰 첫 번째 행 앞, 없으면 </sheetData> 앞
        insert_pos: Optional[int] = None
        for mm in re.finditer(r'<row\b[^>]*\br="(\d+)"', xml):
            if int(mm.group(1)) > row_num:
                insert_pos = mm.start()
                break

        if insert_pos is not None:
            xml = xml[:insert_pos] + row_xml + xml[insert_pos:]
        elif '</sheetData>' in xml:
            xml = xml.replace('</sheetData>', row_xml + '</sheetData>', 1)
        else:
            xml = re.sub(r'<sheetData\s*/>', f'<sheetData>{row_xml}</sheetData>', xml, count=1)

        return xml.encode('utf-8')

    @staticmethod
    def _xlsx_remove_external_links(files: dict, names: list) -> tuple:
        """zip에서 externalLinks 파일 및 모든 참조를 제거."""
        names = [n for n in names if not n.startswith('xl/externalLinks/')]
        for key in [k for k in files if k.startswith('xl/externalLinks/')]:
            del files[key]
        rels_key = 'xl/_rels/workbook.xml.rels'
        if rels_key in files:
            rs = files[rels_key].decode('utf-8', errors='replace')
            rs = re.sub(r'<Relationship\b[^>]*Type="[^"]*externalLink[^"]*"[^>]*/>', '', rs)
            files[rels_key] = rs.encode('utf-8')
        wb_key = 'xl/workbook.xml'
        if wb_key in files:
            ws = files[wb_key].decode('utf-8', errors='replace')
            ws = re.sub(r'<externalReferences\b[^>]*>.*?</externalReferences>', '', ws, flags=re.DOTALL)
            ws = re.sub(r'<externalReferences\s*/>', '', ws)
            files[wb_key] = ws.encode('utf-8')
        # [Content_Types].xml에서 externalLink Override 항목 제거
        ct_key = '[Content_Types].xml'
        if ct_key in files:
            ct = files[ct_key].decode('utf-8', errors='replace')
            ct = re.sub(r'<Override\b[^>]*PartName="[^"]*externalLink[^"]*"[^>]*/>', '', ct)
            files[ct_key] = ct.encode('utf-8')
        return files, names

    @staticmethod
    def _xlsx_strip_external_ref_formulas(files: dict) -> dict:
        """워크시트에서 [N] 외부링크 참조 수식의 <f> 태그와 고아 공유수식 인스턴스를 제거."""
        for name in list(files.keys()):
            if not re.match(r'xl/worksheets/sheet\d+\.xml$', name):
                continue
            xml = files[name].decode('utf-8', errors='replace')

            # Step 1: 외부링크 공유수식 마스터의 si 인덱스 수집
            dead_si: set = set()
            for m in re.finditer(r'<f\b([^>]*)>(.*?)</f>', xml, flags=re.DOTALL):
                if re.search(r'\[\d+\]', m.group(2)):
                    si_m = re.search(r'\bsi="(\d+)"', m.group(1))
                    if si_m:
                        dead_si.add(si_m.group(1))

            # Step 2: 외부링크 포함 <f>...</f> 제거
            def _drop(m: re.Match) -> str:
                return '' if re.search(r'\[\d+\]', m.group(0)) else m.group(0)
            new_xml = re.sub(r'<f(?:\s[^>]*)?>.*?</f>', _drop, xml, flags=re.DOTALL)

            # Step 3: 고아 공유수식 인스턴스 (자체종결 <f si="N"/>) 제거
            for si in dead_si:
                new_xml = re.sub(rf'<f\b[^>]*\bsi="{re.escape(si)}"[^>]*/>', '', new_xml)

            if new_xml != xml:
                files[name] = new_xml.encode('utf-8')
        return files

    @staticmethod
    def _xlsx_strip_all_formulas(files: dict) -> dict:
        """모든 워크시트의 수식을 제거하고 셀 타입을 정규화 (복구 메시지 완전 차단).

        모든 <c> 셀을 단일 패스로 처리하여 엣지케이스 누락 방지:
        - 수식 포함 셀: 수식 제거 후 t 속성을 값 타입에 맞게 재설정
        - t="str" 고아 셀(수식 없음): inlineStr 변환 또는 빈 셀
        - 자체종결 t="str" 셀: t 속성 제거
        """
        for name in list(files.keys()):
            if not re.match(r'xl/worksheets/sheet\d+\.xml$', name):
                continue
            xml = files[name].decode('utf-8', errors='replace')

            # ── 비자체종결 셀 처리 ───────────────────────────────────────────
            def _fix_cell(m: re.Match) -> str:
                full  = m.group(0)
                gt    = full.index('>')
                attrs = full[2:gt]        # <c 와 > 사이의 속성 문자열
                body  = full[gt + 1:-4]  # > 와 </c> 사이의 본문

                has_f    = '<f' in body
                has_tstr = 't="str"' in attrs

                if not has_f and not has_tstr:
                    return full  # 변경 불필요

                # 수식 태그 제거
                if has_f:
                    body = re.sub(r'<f(?:\s[^>]*)?>.*?</f>', '', body, flags=re.DOTALL)
                    body = re.sub(r'<f\b[^>]*/>', '', body)

                # t="str" 제거 (수식 없으면 OOXML 규격 위반)
                clean = re.sub(r'\s*\bt="str"', '', attrs)  # 뒤 \b 생략 (""는 word char 아님)

                vm  = re.search(r'<v>(.*?)</v>', body, re.DOTALL)
                val = vm.group(1) if vm else None

                if val is not None:
                    if has_tstr:
                        # 문자열 수식 결과 → inlineStr로 변환
                        body = re.sub(r'<v>.*?</v>',
                                      f'<is><t>{val}</t></is>', body, flags=re.DOTALL)
                        return f'<c{clean} t="inlineStr">{body}</c>'
                    else:
                        # 숫자/불리언 등 → 타입 없는 숫자 셀 유지
                        return f'<c{clean}>{body}</c>'
                else:
                    # 캐시 값 없음 → 빈 셀
                    stripped = body.strip()
                    if stripped:
                        return f'<c{clean}>{stripped}</c>'
                    return f'<c{clean.rstrip()}/>'

            xml = re.sub(r'<c\b[^>]*>.*?</c>', _fix_cell, xml, flags=re.DOTALL)

            # ── 자체종결 셀 안전망 ──────────────────────────────────────────
            # <c ... t="str" ... /> 에서 t="str" 제거 (빈 숫자형 셀)
            xml = re.sub(r'(<c\b[^>]*?)\s+t="str"([^>]*/>)', r'\1\2', xml)

            files[name] = xml.encode('utf-8')
        return files

    @staticmethod
    def _xlsx_delete_leading_cols(sheet_bytes: bytes, delete_count: int) -> bytes:
        """sheet XML에서 앞 delete_count개 열 삭제 후 나머지 열을 왼쪽으로 이동."""
        from openpyxl.utils import column_index_from_string as _ci, get_column_letter as _cl
        xml = sheet_bytes.decode('utf-8', errors='replace')
        del_letters = [_cl(i) for i in range(1, delete_count + 1)]
        del_pat = '|'.join(re.escape(l) for l in del_letters)

        # ── 셀 데이터 제거 ──────────────────────────────────────────────────
        xml = re.sub(rf'<c\b[^>]*\br="(?:{del_pat})\d+"[^>]*/>', '', xml)
        xml = re.sub(rf'<c\b[^>]*\br="(?:{del_pat})\d+"[^>]*>.*?</c>', '', xml, flags=re.DOTALL)

        def _shift_col(col_idx: int) -> int:
            return max(1, col_idx - delete_count)

        def _shift_cell(ref: str):
            """(new_ref, was_deleted) 반환."""
            rm = re.match(r'([A-Z]+)(\d+)', ref)
            if not rm:
                return ref, False
            ci = _ci(rm.group(1))
            if ci <= delete_count:
                return None, True
            return f'{_cl(ci - delete_count)}{rm.group(2)}', False

        # ── r="XN" 셀 참조 시프트 (cell 태그) ──────────────────────────────
        def shift_ref(m: re.Match) -> str:
            rm = re.match(r'([A-Z]+)(\d+)', m.group(1))
            if not rm:
                return m.group(0)
            ci = _ci(rm.group(1))
            if ci <= delete_count:
                return m.group(0)
            return f'r="{_cl(ci - delete_count)}{rm.group(2)}"'
        xml = re.sub(r'\br="([A-Z]+\d+)"', shift_ref, xml)

        # ── mergeCells 처리 ─────────────────────────────────────────────────
        def shift_merge(m: re.Match) -> str:
            ref = m.group(1)
            if ':' not in ref:
                return ''  # 단일 셀 병합 → 제거
            from_r, to_r = ref.split(':', 1)
            fm = re.match(r'([A-Z]+)(\d+)', from_r)
            tm = re.match(r'([A-Z]+)(\d+)', to_r)
            if not fm or not tm:
                return m.group(0)
            fc, frow = _ci(fm.group(1)), fm.group(2)
            tc, trow = _ci(tm.group(1)), tm.group(2)
            if tc <= delete_count:
                return ''  # 완전히 삭제 범위 → 제거
            new_fc = 1 if fc <= delete_count else fc - delete_count
            new_tc = tc - delete_count
            if new_fc == new_tc and frow == trow:
                return ''  # 단일 셀이 됨 → 병합 불필요
            return f'<mergeCell ref="{_cl(new_fc)}{frow}:{_cl(new_tc)}{trow}"/>'
        xml = re.sub(r'<mergeCell\b[^>]*\bref="([^"]+)"[^>]*/>', shift_merge, xml)
        xml = re.sub(r'<mergeCells[^>]*>\s*</mergeCells>', '', xml)

        # ── sqref 속성 처리 (dataValidation / conditionalFormatting) ────────
        def shift_sqref(sqref_val: str) -> str:
            new_parts = []
            for part in sqref_val.split():
                if ':' in part:
                    fr, tr = part.split(':', 1)
                    fm = re.match(r'([A-Z]+)(\d+)', fr)
                    tm = re.match(r'([A-Z]+)(\d+)', tr)
                    if fm and tm:
                        fc, frow = _ci(fm.group(1)), fm.group(2)
                        tc, trow = _ci(tm.group(1)), tm.group(2)
                        if tc <= delete_count:
                            continue
                        new_fc = 1 if fc <= delete_count else fc - delete_count
                        new_parts.append(f'{_cl(new_fc)}{frow}:{_cl(tc - delete_count)}{trow}')
                    else:
                        new_parts.append(part)
                else:
                    mm = re.match(r'([A-Z]+)(\d+)', part)
                    if mm:
                        ci = _ci(mm.group(1))
                        if ci <= delete_count:
                            continue
                        new_parts.append(f'{_cl(ci - delete_count)}{mm.group(2)}')
                    else:
                        new_parts.append(part)
            return ' '.join(new_parts)

        def repl_sqref(m: re.Match) -> str:
            new_val = shift_sqref(m.group(1))
            return f'sqref="{new_val}"' if new_val else 'sqref=""'
        xml = re.sub(r'\bsqref="([^"]+)"', repl_sqref, xml)
        # sqref가 빈 조건부서식/유효성 블록 제거
        xml = re.sub(r'<conditionalFormatting\b[^>]*sqref=""[^>]*/>', '', xml)
        xml = re.sub(r'<conditionalFormatting\b[^>]*sqref=""[^>]*>.*?</conditionalFormatting>',
                     '', xml, flags=re.DOTALL)
        xml = re.sub(r'<dataValidation\b[^>]*sqref=""[^>]*/>', '', xml)
        xml = re.sub(r'<dataValidation\b[^>]*sqref=""[^>]*>.*?</dataValidation>',
                     '', xml, flags=re.DOTALL)

        # ── sheetView 내 activeCell / topLeftCell 시프트 ─────────────────────
        def shift_attr_cell(m: re.Match) -> str:
            attr, cell_ref = m.group(1), m.group(2)
            rm = re.match(r'([A-Z]+)(\d+)', cell_ref)
            if not rm:
                return m.group(0)
            ci = _ci(rm.group(1))
            if ci <= delete_count:
                return f'{attr}="A{rm.group(2)}"'  # 삭제된 열 → A열로 리셋
            return f'{attr}="{_cl(ci - delete_count)}{rm.group(2)}"'
        xml = re.sub(r'\b(activeCell|topLeftCell)="([^"]+)"', shift_attr_cell, xml)

        return xml.encode('utf-8')

    @staticmethod
    def _xlsx_shift_drawing_cols(files: dict, sheet_zip_path: str, delete_count: int) -> dict:
        """sheet와 연결된 drawing XML의 열 앵커를 왼쪽으로 보정 (_xlsx_delete_leading_cols 후 호출)."""
        # 시트 rels 경로: xl/worksheets/sheet1.xml → xl/worksheets/_rels/sheet1.xml.rels
        base, fname = sheet_zip_path.rsplit('/', 1) if '/' in sheet_zip_path else ('', sheet_zip_path)
        rels_path = f"{base}/_rels/{fname}.rels" if base else f"_rels/{fname}.rels"
        if rels_path not in files:
            return files

        rels_xml = files[rels_path].decode('utf-8', errors='replace')
        for m in re.finditer(r'<Relationship\b([^>]*/?>)', rels_xml):
            attrs  = m.group(1)
            type_m = re.search(r'\bType="([^"]*)"', attrs)
            tgt_m  = re.search(r'\bTarget="([^"]*)"', attrs)
            if not (type_m and tgt_m and '/drawing' in type_m.group(1)):
                continue

            tgt = tgt_m.group(1)
            # "../drawings/drawing1.xml" → "xl/drawings/drawing1.xml"
            if tgt.startswith('../'):
                drawing_path = base.rsplit('/', 1)[0] + '/' + tgt[3:] if '/' in base else tgt[3:]
            elif tgt.startswith('/'):
                drawing_path = tgt.lstrip('/')
            else:
                drawing_path = (base + '/' + tgt) if base else tgt

            if drawing_path not in files:
                continue

            draw_xml = files[drawing_path].decode('utf-8', errors='replace')

            def _shift(cm: re.Match) -> str:
                return f'<xdr:col>{max(0, int(cm.group(1)) - delete_count)}</xdr:col>'

            files[drawing_path] = re.sub(
                r'<xdr:col>(\d+)</xdr:col>', _shift, draw_xml).encode('utf-8')

        return files

    # ── xlsx 생성 / 통합양식 업데이트 ──────────────────────────────────────

    @staticmethod
    def _xlsx_load_and_clean(path: str, *, strip_all_formulas: bool = False) -> Tuple[dict, list]:
        """xlsx ZIP을 읽어 calcChain·externalLinks를 제거한 (files, names) 반환.

        strip_all_formulas=True: 모든 수식 제거 (출력 파일용 — 복구 메시지 원천 차단)
        strip_all_formulas=False: 외부링크+고아 공유수식만 제거 (템플릿 업데이트용)
        """
        import zipfile as _zf
        with _zf.ZipFile(path, 'r') as zf:
            names = zf.namelist()
            files = {n: zf.read(n) for n in names}
        files.pop('xl/calcChain.xml', None)
        names = [n for n in names if n != 'xl/calcChain.xml']
        files, names = RackPurchaseRequestPage._xlsx_remove_external_links(files, names)
        if strip_all_formulas:
            files = RackPurchaseRequestPage._xlsx_strip_all_formulas(files)
        else:
            files = RackPurchaseRequestPage._xlsx_strip_external_ref_formulas(files)
        return files, names

    @staticmethod
    def _xlsx_save(path: str, files: dict, names: list) -> None:
        """(files, names)를 xlsx ZIP으로 저장."""
        import zipfile as _zf
        with _zf.ZipFile(path, 'w', _zf.ZIP_DEFLATED) as zf:
            for name in names:
                if name in files:
                    zf.writestr(name, files[name])

    @staticmethod
    def _xlsx_collect_row_styles(sheet_xml: str, row_num: int) -> Dict[str, str]:
        """sheet XML에서 지정 행의 열별 s 속성(스타일 인덱스) 수집."""
        styles: Dict[str, str] = {}
        for row_m in re.finditer(
                rf'<row\b[^>]*\br="{row_num}"[^>]*>(.*?)</row>', sheet_xml, flags=re.DOTALL):
            for cell_m in re.finditer(r'<c\b([^>]*)>', row_m.group(1)):
                attrs = cell_m.group(1)
                ref_m = re.search(r'\br="([A-Z]+)\d+"', attrs)
                s_m   = re.search(r'\bs="(\d+)"', attrs)
                if ref_m and s_m:
                    styles[ref_m.group(1)] = s_m.group(1)
            break
        return styles

    def _direct_patch_xlsx(self, path: str) -> None:
        """drawings 손상·calcChain 오류 없이 xlsx를 직접 수정 (regex 방식)."""
        from openpyxl.utils import get_column_letter as _gcl

        files, names = self._xlsx_load_and_clean(path, strip_all_formulas=True)

        wb_bytes   = files.get('xl/workbook.xml', b'')
        rels_bytes = files.get('xl/_rels/workbook.xml.rels', b'')
        files['xl/workbook.xml'] = self._xlsx_patch_wb_visibility(wb_bytes, 'RACK발주양식')
        rack_file = self._xlsx_find_sheet_file(wb_bytes, rels_bytes, 'RACK발주양식')

        row2_vals: Dict[int, Any] = {}
        for row, (_, item_col, qty_col) in self._ROW_RACK_MAP.items():
            for col, tbl_col in ((item_col, self.COL_ITEM), (qty_col, self.COL_QTY)):
                if col is None:
                    continue
                txt = self._text(row, tbl_col)
                if txt == '':
                    continue
                try:
                    v: Any = int(txt) if '.' not in txt else float(txt)
                except (ValueError, TypeError):
                    v = txt
                row2_vals[col] = v

        rack_ch_item = self._text(17, self.COL_ITEM)
        rack_ch_qty  = self._text(17, self.COL_QTY)
        if rack_ch_item:
            row2_vals[41] = rack_ch_item  # AO열(41): RACK CH 항목
        if rack_ch_qty:
            try:
                rack_ch_qty_v: Any = int(rack_ch_qty) if '.' not in rack_ch_qty else float(rack_ch_qty)
            except (ValueError, TypeError):
                rack_ch_qty_v = rack_ch_qty
            row2_vals[42] = rack_ch_qty_v  # AP열(42): RACK CH 수량

        row2_vals[88] = "전기공통"  # CJ열(88) 항상 기입

        invoice_text = self._text(2, self.COL_ITEM)
        if invoice_text:
            row2_vals[2] = invoice_text  # B열(2): 인보이스 필요여부
        if invoice_text == "인보이스 필요":
            row2_vals[95] = "KIT투자 시 인보이스필요"  # CQ열(95): 비고

        # 데이터 입력된 열 중 비어 있는 셀은 "-"로 채움
        all_data_cols: set = set()
        for _, item_col, qty_col in self._ROW_RACK_MAP.values():
            if item_col is not None:
                all_data_cols.add(item_col)
            if qty_col is not None:
                all_data_cols.add(qty_col)
        for col in all_data_cols:
            if col not in row2_vals:
                row2_vals[col] = "-"

        if rack_file and rack_file in files:
            try:
                sheet_xml = files[rack_file].decode('utf-8', errors='replace')
                row2_styles = self._xlsx_collect_row_styles(sheet_xml, 2)
                files[rack_file] = self._xlsx_patch_sheet_row(
                    files[rack_file], 2, row2_vals, _gcl, row2_styles)
            except Exception as e:
                logger.warning("RACK발주양식 시트 패치 실패: %s", e)

        self._xlsx_save(path, files, names)


    def _generate_approval_doc(self) -> None:
        """결재상신용: 여러 요청서 RACK발주양식 2행을 합쳐 A·B열 삭제 후 저장 (zipfile 방식)."""
        from openpyxl.utils import get_column_letter as _gcl

        start_dir = self._last_generated_folder or exe_dir()
        paths, _ = QFileDialog.getOpenFileNames(
            self, "RACK 구매요청서 선택 (2개 이상)", start_dir, "Excel Files (*.xlsx)"
        )
        if not paths:
            return
        if len(paths) < 2:
            QMessageBox.warning(self, "선택 오류", "2개 이상의 파일을 선택하세요.")
            return

        folder = os.path.dirname(paths[0])
        folder_name = os.path.basename(folder)
        process = folder_name
        if "_RACK구매요청서" in folder_name:
            base = folder_name.split("_RACK구매요청서")[0]
            parts = base.split("_")
            process = "_".join(parts[1:]) if len(parts) > 1 else base

        today    = datetime.now().strftime("%Y-%m-%d")
        out_path = unique_path(os.path.join(
            folder, f"{process} Rack 구매요청서 ({today}).xlsx"))

        try:
            all_row_vals: List[Dict[int, Any]] = []
            for path in paths:
                try:
                    from openpyxl import load_workbook
                    src_wb = load_workbook(path, data_only=True)
                    ws_r   = src_wb["RACK발주양식"]
                    rv = {c: ws_r.cell(2, c).value
                          for c in range(1, ws_r.max_column + 1)
                          if ws_r.cell(2, c).value is not None}
                    src_wb.close()
                    all_row_vals.append(rv)
                except Exception as e:
                    logger.warning("결재상신용 파일 읽기 실패 (%s): %s", path, e)
                    all_row_vals.append({})

            shutil.copy2(paths[0], out_path)
            files, names = self._xlsx_load_and_clean(out_path, strip_all_formulas=True)

            wb_bytes   = files.get('xl/workbook.xml', b'')
            rels_bytes = files.get('xl/_rels/workbook.xml.rels', b'')
            rack_file  = self._xlsx_find_sheet_file(wb_bytes, rels_bytes, 'RACK발주양식')

            if not rack_file or rack_file not in files:
                QMessageBox.critical(self, "오류", "RACK발주양식 시트를 찾을 수 없습니다.")
                return

            row2_styles = self._xlsx_collect_row_styles(
                files[rack_file].decode('utf-8', errors='replace'), 2)

            for i, rv in enumerate(all_row_vals[1:], start=3):
                try:
                    files[rack_file] = self._xlsx_patch_sheet_row(
                        files[rack_file], i, rv, _gcl, row2_styles)
                except Exception as e:
                    logger.warning("결재상신용 행 추가 실패 (row=%d): %s", i, e)

            files[rack_file] = self._xlsx_delete_leading_cols(files[rack_file], 2)
            files = self._xlsx_shift_drawing_cols(files, rack_file, 2)

            # 열 숨기기 전체 해제 (<col hidden="1"> 속성 제거)
            sheet_str = files[rack_file].decode('utf-8', errors='replace')
            sheet_str = re.sub(r'(<col\b[^>]*?)\s+hidden="1"', r'\1', sheet_str)
            files[rack_file] = sheet_str.encode('utf-8')

            self._xlsx_save(out_path, files, names)

            QMessageBox.information(self, "완료",
                f"결재상신용 생성 완료\n{os.path.basename(out_path)}")
        except Exception as e:
            QMessageBox.critical(self, "생성 오류", f"결재상신용 생성 중 오류:\n{e}")
            logger.error("결재상신용 생성 오류", exc_info=True)



# ══════════════════════════════════════════════
# 전자서명 페이지
# ══════════════════════════════════════════════

class ESignPage(QWidget):

    SIGN_W = 170
    SIGN_H = 40

    def __init__(self) -> None:
        super().__init__()
        self._code        : str  = ""
        self._signs       : List = []
        self._files       : List[str] = []
        self._base_folder : str  = ""
        self._sheet_pngs  : List[List[str]] = []   # 파일별 시트 PNG 경로 리스트
        self._cur_pngs    : List[str] = []          # 현재 파일 PNG 경로들
        self._cur_file    : int  = 0
        self._cur_page    : int  = 0
        self._sign_items    : Dict[Tuple[int,int], List[SignatureItem]] = {}
        self._render_sz     : Dict[Tuple[int,int], Tuple[int,int]]     = {}
        self._bg_item              = None
        self._shown_key     : Optional[Tuple[int,int]] = None
        self._loader_thread : Optional[_ExcelLoaderThread] = None
        self._load_progress : Optional[QProgressDialog]   = None
        self._tmp_dir       : Optional[str]               = None
        self._build_ui()

    def _build_ui(self) -> None:
        outer = QVBoxLayout(self); outer.setContentsMargins(14,14,14,14); outer.setSpacing(10)
        top = QHBoxLayout()
        self.btn_code   = QPushButton("승인코드 LOAD")
        self.btn_excel  = QPushButton("엑셀 LOAD")
        self.btn_save   = QPushButton("PDF 저장")
        tint_button(self.btn_code,  "#DCEDC8")   # 연초록
        tint_button(self.btn_excel, "#B3E5FC")   # 연하늘
        tint_button(self.btn_save,  "#FFE0B2")   # 연주황
        self.lbl_status = QLabel("준비")
        for b in (self.btn_code, self.btn_excel, self.btn_save): top.addWidget(b); top.addSpacing(8)
        top.addStretch(1); top.addWidget(self.lbl_status); outer.addLayout(top)
        mid = QHBoxLayout()
        self.file_list = QListWidget(); self.file_list.setFixedWidth(360); mid.addWidget(self.file_list)
        self.scene = QGraphicsScene(self)
        self.view  = PdfView(self); self.view.setScene(self.scene); self.view.setAlignment(Qt.AlignCenter)
        self.view.setFocusPolicy(Qt.StrongFocus); self.view.setFocus(); mid.addWidget(self.view, 1)
        outer.addLayout(mid, 1)
        self.btn_code.clicked.connect(self._load_code)
        self.btn_excel.clicked.connect(self._load_excels)
        self.btn_save.clicked.connect(self._save_pdf)
        self.file_list.currentRowChanged.connect(self._on_select_file)
        self.view.on_prev = self._prev_page; self.view.on_next = self._next_page
        self.view.on_double_click = self._add_sign

    def _load_code(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "승인코드 TXT", "", "Text Files (*.txt)")
        if not path: return
        try:
            with open(path, encoding="utf-8") as f:
                self._code = f.read().strip()
        except Exception as e:
            QMessageBox.critical(self, "오류", str(e))
            return
        folder = os.path.dirname(path); exts = (".png",".jpg",".jpeg",".bmp",".webp")
        imgs = sorted([os.path.join(folder,f) for f in os.listdir(folder) if f.lower().endswith(exts)],
                       key=lambda p: os.path.basename(p).lower())
        def _pick(kws):
            for kw in kws:
                for p in imgs:
                    if kw in os.path.basename(p).lower(): return p
            return None
        p1 = _pick(["서명1","sign1","signature1","stamp1"]); p2 = _pick(["서명2","sign2","signature2","stamp2"])
        if not p1 or not p2:
            p1 = p1 or (imgs[0] if imgs else None); p2 = p2 or (imgs[1] if len(imgs)>1 else None)
        if not p1 or not p2:
            QMessageBox.warning(self, "안내", "서명 이미지 2개를 찾지 못했습니다.")
            self._signs = []; self.lbl_status.setText("승인코드 OK, 서명 없음"); return
        pm1 = QPixmap(p1).scaled(self.SIGN_W,self.SIGN_H,Qt.IgnoreAspectRatio,Qt.SmoothTransformation)
        pm2 = QPixmap(p2).scaled(self.SIGN_W,self.SIGN_H,Qt.IgnoreAspectRatio,Qt.SmoothTransformation)
        if pm1.isNull() or pm2.isNull(): QMessageBox.critical(self,"오류","서명 이미지 로드 실패"); return
        self._signs = [pm1, pm2]
        self.lbl_status.setText(f"승인코드 OK / {os.path.basename(p1)}, {os.path.basename(p2)}")
        QMessageBox.information(self, "완료", "승인코드 LOAD 완료\n전자서명 ON")

    def _load_excels(self) -> None:
        if not excel_io.COM_AVAILABLE:
            QMessageBox.critical(self, "오류", "Excel COM이 없습니다.")
            return
        if self._loader_thread and self._loader_thread.isRunning():
            QMessageBox.information(self, "안내", "이미 로딩 중입니다.")
            return
        paths, _ = QFileDialog.getOpenFileNames(self, "엑셀 선택(다중)", "", "Excel Files (*.xlsx)")
        if not paths:
            return
        paths = sorted(paths, key=lambda p: (0 if "갑지" in os.path.basename(p).lower() else 1, os.path.basename(p).lower()))
        base = os.path.commonpath(paths)
        if os.path.isfile(base):
            base = os.path.dirname(base)
        self._base_folder = base
        self._files = paths
        self._sheet_pngs = []
        self._cur_pngs   = []
        self._sign_items.clear()

        self.file_list.blockSignals(True)
        self.file_list.clear()
        for p in paths:
            it = QListWidgetItem(os.path.basename(p))
            it.setData(Qt.UserRole, p)
            self.file_list.addItem(it)
        self.file_list.blockSignals(False)

        self._cleanup_tmp()
        tmp = ensure_dir(os.path.join(base, "_esign_tmp_pdf"))
        self._tmp_dir = tmp

        self._load_progress = QProgressDialog("변환 준비 중...", "취소", 0, len(paths), self)
        self._load_progress.setWindowTitle("엑셀 → PDF 변환")
        self._load_progress.setWindowModality(Qt.WindowModal)
        self._load_progress.setMinimumDuration(0)
        self._load_progress.setValue(0)

        self._loader_thread = _ExcelLoaderThread(paths, tmp, self)
        self._loader_thread.progress.connect(self._on_load_progress)
        self._loader_thread.finished.connect(self._on_load_finished)
        self._load_progress.canceled.connect(self._loader_thread.cancel)

        self.btn_excel.setEnabled(False)
        self._loader_thread.start()

    def _on_load_progress(self, done: int, total: int, fname: str) -> None:
        if not self._load_progress:
            return
        self._load_progress.setValue(done)
        if done < total:
            self._load_progress.setLabelText(f"변환 중 ({done + 1}/{total}): {fname}")
        else:
            self._load_progress.setLabelText("변환 완료")

    def _on_load_finished(self) -> None:
        if self._load_progress:
            self._load_progress.close()
            self._load_progress = None
        self._sheet_pngs = self._loader_thread.sheet_pngs
        while len(self._sheet_pngs) < len(self._files):
            self._sheet_pngs.append([])
        self.btn_excel.setEnabled(True)
        self.lbl_status.setText(f"{len(self._files)}개 로드 완료")
        if self.file_list.count() > 0:
            self.file_list.setCurrentRow(0)

    def _on_select_file(self, row: int) -> None:
        if row < 0 or row >= len(self._files): return
        self._cur_file = row; self._cur_page = 0; self._load_sheets(); self._render()

    def _load_sheets(self) -> None:
        self._cur_pngs = (self._sheet_pngs[self._cur_file]
                          if self._cur_file < len(self._sheet_pngs) else [])
        if not self._cur_pngs:
            self.lbl_status.setText("표시할 시트 없음(스킵)"); self.scene.clear()

    def _render(self) -> None:
        if not self._cur_pngs: return
        self._cur_page = max(0, min(self._cur_page, len(self._cur_pngs) - 1))
        pm = QPixmap(self._cur_pngs[self._cur_page])
        if pm.isNull(): return
        self._render_sz[(self._cur_file, self._cur_page)] = (pm.width(), pm.height())
        if self._shown_key is not None:
            for it in list(self._sign_items.get(self._shown_key, [])):
                try:
                    if it.scene() is self.scene: self.scene.removeItem(it)
                except RuntimeError: pass
        if self._bg_item is not None:
            try:
                if self._bg_item.scene() is self.scene: self.scene.removeItem(self._bg_item)
            except Exception: pass
        bg = QGraphicsPixmapItem(pm); bg.setZValue(0); bg.setAcceptedMouseButtons(Qt.NoButton)
        self.scene.addItem(bg); self._bg_item = bg; self._shown_key = (self._cur_file, self._cur_page)
        for it in list(self._sign_items.get(self._shown_key, [])):
            try: self.scene.addItem(it); it.setZValue(10)
            except RuntimeError: pass
        self.scene.setSceneRect(bg.boundingRect()); self.view.resetTransform()
        vp_w = max(1, self.view.viewport().width())
        scale = vp_w / max(1, pm.width()); self.view.scale(scale, scale); self.view.setFocus()
        self.view.verticalScrollBar().setValue(self.view.verticalScrollBar().minimum())
        self.lbl_status.setText(f"파일 {self._cur_file+1}/{len(self._files)} / 시트 {self._cur_page+1}/{len(self._cur_pngs)}")

    def _next_page(self) -> None:
        if not self._cur_pngs: return
        if self._cur_page+1 < len(self._cur_pngs): self._cur_page += 1; self._render()
        elif self._cur_file+1 < len(self._files): self.file_list.setCurrentRow(self._cur_file+1)

    def _prev_page(self) -> None:
        if not self._cur_pngs: return
        if self._cur_page-1 >= 0: self._cur_page -= 1; self._render()
        elif self._cur_file-1 >= 0:
            self.file_list.setCurrentRow(self._cur_file-1)
            if self._cur_pngs: self._cur_page = max(0, len(self._cur_pngs)-1); self._render()

    def _add_sign(self, scene_pos: QPointF) -> None:
        if not self._signs:
            QMessageBox.information(self, "안내", "승인코드 LOAD 후 서명 이미지가 필요합니다."); return
        if not self._cur_pngs: return
        dlg = PasswordDialog(self, self._code)
        if dlg.exec() != QDialog.Accepted or not dlg.verified: return
        idx = min(1 if (QApplication.keyboardModifiers() & Qt.ShiftModifier) else 0, len(self._signs)-1)
        item = SignatureItem(self._signs[idx], self._cur_page); item.setZValue(10)
        item.setPos(QPointF(scene_pos.x()-self.SIGN_W/2, scene_pos.y()-self.SIGN_H/2))
        self.scene.addItem(item)
        self._sign_items.setdefault((self._cur_file, self._cur_page), []).append(item)
        self.scene.update()

    def _cleanup_tmp(self) -> None:
        self._cur_pngs = []
        if self._tmp_dir and os.path.isdir(self._tmp_dir):
            try:
                shutil.rmtree(self._tmp_dir)
            except Exception as e:
                logger.warning("tmp 폴더 삭제 실패: %s", e)
            self._tmp_dir = None

    def _save_pdf(self) -> None:
        if not any(self._sheet_pngs) or not self._files:
            QMessageBox.information(self, "안내", "먼저 엑셀을 LOAD 하세요."); return
        folder_name = os.path.basename(self._base_folder.rstrip("\\/"))
        out = unique_path(os.path.join(self._base_folder, f"{folder_name}.pdf"))
        try:
            self._build_pdf(out)
            self._cleanup_tmp()
            QMessageBox.information(self, "완료", f"저장 완료:\n{out}"); self.lbl_status.setText("PDF 저장 완료")
        except Exception as e:
            logger.error("PDF 저장 실패", exc_info=True)
            QMessageBox.critical(self, "오류", f"{e}\n\n{traceback.format_exc()}")

    def _build_pdf(self, out: str) -> None:
        import fitz
        from PySide6.QtGui import QPainter, QImage
        A4_W, A4_H = 595.0, 842.0       # A4 in points (1pt = 1/72 inch)
        MAX_PX = int(A4_W * 150 / 72)   # 방안A: A4 150 DPI 기준 최대 너비 ≈ 1240px
        final = fitz.open()

        for fi, pngs in enumerate(self._sheet_pngs):
            for pno, png_path in enumerate(pngs):
                if not os.path.exists(png_path):
                    continue
                base_pm = QPixmap(png_path)
                if base_pm.isNull():
                    continue
                iw, ih = base_pm.width(), base_pm.height()
                signs = self._sign_items.get((fi, pno), [])

                if signs:
                    composed = QPixmap(iw, ih)
                    composed.fill(Qt.white)
                    painter = QPainter(composed)
                    painter.drawPixmap(0, 0, base_pm)
                    for it in list(signs):
                        try:
                            x, y = float(it.pos().x()), float(it.pos().y())
                        except RuntimeError:
                            continue
                        painter.drawPixmap(int(x), int(y), it.pixmap())
                    painter.end()
                    final_pm = composed
                else:
                    final_pm = base_pm

                # 방안A: 해상도 상한 (너비 MAX_PX 초과 시 축소)
                if final_pm.width() > MAX_PX:
                    final_pm = final_pm.scaled(
                        MAX_PX, MAX_PX * 10,
                        Qt.KeepAspectRatio, Qt.SmoothTransformation)

                # 방안B: 입고검수확인서만 그레이스케일 변환
                if "입고검수확인서" in png_path:
                    gray_img = final_pm.toImage().convertToFormat(
                        QImage.Format_Grayscale8)
                    final_pm = QPixmap.fromImage(gray_img)

                iw, ih = final_pm.width(), final_pm.height()

                # QPixmap → JPEG bytes
                ba = QByteArray(); buf = QBuffer(ba); buf.open(QIODevice.WriteOnly)
                final_pm.save(buf, "JPEG", 45); buf.close()

                # 고정 A4 페이지, 이미지를 비율 유지하며 중앙 배치
                page = final.new_page(width=A4_W, height=A4_H)
                scale = min(A4_W / max(1, iw), A4_H / max(1, ih))
                pw, ph = iw * scale, ih * scale
                x0 = (A4_W - pw) / 2
                y0 = (A4_H - ph) / 2
                page.insert_image(fitz.Rect(x0, y0, x0 + pw, y0 + ph), stream=bytes(ba))

        final.save(out, deflate=True, garbage=4); final.close()
