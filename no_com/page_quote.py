"""
page_quote.py
=============
견적서 작성 페이지와 관련 스레드/다이얼로그.
"""
from __future__ import annotations

import logging
import os
import traceback
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from PySide6.QtCore import Qt, QEvent, QThread, QTimer, Signal
from PySide6.QtGui import QBrush, QColor, QKeySequence, QPixmap, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView, QCheckBox, QComboBox, QCompleter, QDialog,
    QDialogButtonBox, QDoubleSpinBox, QFileDialog,
    QFormLayout, QFrame, QGridLayout, QHBoxLayout,
    QHeaderView, QInputDialog, QLabel, QLineEdit,
    QListWidget, QListWidgetItem, QMessageBox,
    QProgressDialog, QPushButton, QSizePolicy, QSpinBox, QSplitter, QTabWidget,
    QTableWidget, QTableWidgetItem, QTextEdit,
    QVBoxLayout, QWidget,
)

from core import (
    QuoteState, SheetName,
    exe_dir, fmt_krw, fmt_qty,
    normalize_token, parse_invest_info, s, to_float, unique_path,
)
from widgets import (
    DraggableItemsTable, DroppableQuoteTable,
    bold_label, centered_checkbox,
    get_checkbox_from_cell, info_label, labeled_frame, tint_button,
)
from page_common import _friendly_error_msg, _make_plain_table, _ScrollableErrorDialog

logger = logging.getLogger("QuoteApp")


PROCESS_CHOICES = ["", "CVD", "DIFF", "IMP", "ETCH", "METAL", "CLEAN", "GCS"]
VENDOR_CHOICES  = [
    "", "AMAT", "ASM", "AXCELIS", "EUGENETECH", "WONIK_IPS",
    "KOKUSAI", "LAM", "SEMES", "TEL", "TES", "ULVAC", "GCS",
]


# ──────────────────────────────────────────────
# 백그라운드 Excel → PDF 변환 스레드
# ──────────────────────────────────────────────
class _TemplateLoaderThread(QThread):
    """STEP1용: 통합양식 Excel을 백그라운드에서 파싱."""
    done    = Signal(object)   # (rows, by_class, price_by_spec, order, cmap) 또는 Exception
    def __init__(self, path: str, parent=None) -> None:
        super().__init__(parent)
        self.path = path
    def run(self) -> None:
        try:
            import excel_io
            from openpyxl import load_workbook
            wb = load_workbook(self.path, data_only=False)
            rows, by_class, price_by_spec, order = excel_io.parse_items_sheet(wb[SheetName.ITEMS])
            cmap = excel_io.parse_code_map_sheet(wb[SheetName.CODE_MAP])
            self.done.emit((rows, by_class, price_by_spec, order, cmap))
        except Exception as e:
            self.done.emit(e)


class _RequestLoaderThread(QThread):
    """STEP2용: 의뢰파일 Excel을 백그라운드에서 파싱 (메인 스레드 동결 방지)."""
    done = Signal(object)  # (sheet_name, rows) tuple 또는 Exception
    def __init__(self, path: str, parent=None) -> None:
        super().__init__(parent)
        self.path = path
    def run(self) -> None:
        try:
            import excel_io
            sheet_name, rows = excel_io.parse_request_xlsx(self.path)
            self.done.emit((sheet_name, rows))
        except Exception as e:
            self.done.emit(e)


class _GenQuoteThread(QThread):
    """견적서 생성을 백그라운드에서 실행."""
    progress = Signal(str)          # 로그 메시지
    done     = Signal(object)       # List[(rd, path)] 또는 Exception

    def __init__(self, state, qtype: str, items, rds: list, parent=None) -> None:
        super().__init__(parent)
        self.state = state; self.qtype = qtype
        self.items = items; self.rds   = rds

    def run(self) -> None:
        try:
            import excel_io
            if self.qtype == "국내" and len(self.rds) >= 2:
                def _cb(done, total, name):
                    self.progress.emit(f"[{done}/{total}] 저장: {name}")
                results = excel_io.generate_quote_multi(
                    self.state, self.items, self.rds, progress_cb=_cb)
            else:
                rd   = self.rds[0] if self.rds else {}
                path = excel_io.generate_quote(self.state, self.qtype, self.items, rd)
                self.progress.emit(f"저장: {os.path.basename(path)}")
                results = [(rd, path)]
            self.done.emit(results)
        except Exception as e:
            self.done.emit(e)


class _GenCoverThread(QThread):
    """갑지 생성을 백그라운드에서 실행."""
    progress = Signal(str)        # 로그 메시지
    step     = Signal(int, int)   # (done, total) — 진행률 다이얼로그용
    done     = Signal(object)     # 저장 경로(str) 또는 Exception

    def __init__(self, template_path: str, folder: str, paths: list,
                 investor_name: str, warranty_years: int = 2, parent=None) -> None:
        super().__init__(parent)
        self.template_path = template_path; self.folder = folder
        self.paths = paths; self.investor_name = investor_name
        self.warranty_years = warranty_years

    def run(self) -> None:
        try:
            import excel_io
            def _cb(done, total, name):
                self.progress.emit(f"[{done}/{total}] 읽는 중: {name}")
                self.step.emit(done, total)
            out = excel_io.generate_cover(
                self.template_path, self.folder, self.paths,
                investor_name=self.investor_name, progress_cb=_cb,
                warranty_years=self.warranty_years)
            self.done.emit(out)
        except Exception as e:
            self.done.emit(e)


# ══════════════════════════════════════════════
# ══════════════════════════════════════════════
# 중국 옵션 다이얼로그
# ══════════════════════════════════════════════

class ChinaOptionsDialog(QDialog):
    def __init__(self, parent, state: QuoteState) -> None:
        super().__init__(parent)
        self.setWindowTitle("중국 옵션 입력")
        self.state = state
        self.setModal(True)
        self.setFixedWidth(380)

        cn = state.cn_info or {}
        root = QVBoxLayout(self)
        form = QFormLayout()
        self._maker   = QLineEdit(cn.get("maker", ""))
        self._process = QLineEdit(cn.get("process", ""))
        self._tool    = QLineEdit(cn.get("tool", ""))
        self._line    = QLineEdit(cn.get("line", ""))
        form.addRow("Maker",    self._maker)
        form.addRow("Process",  self._process)
        form.addRow("설비호기", self._tool)
        form.addRow("라인",     self._line)
        root.addLayout(form)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self.accept); bb.rejected.connect(self.reject)
        root.addWidget(bb)

    def apply(self) -> None:
        self.state.cn_info = {
            "maker": self._maker.text().strip(), "process": self._process.text().strip(),
            "tool":  self._tool.text().strip(),  "line":    self._line.text().strip(),
        }


# ══════════════════════════════════════════════
# 미국 옵션 다이얼로그
# ══════════════════════════════════════════════

class USOptionsDialog(QDialog):
    def __init__(self, parent, state: QuoteState) -> None:
        super().__init__(parent)
        self.setWindowTitle("미국 옵션 입력")
        self.state = state
        self.setModal(True)
        self.setFixedWidth(420)

        us = state.us_info or {}
        root = QVBoxLayout(self)
        form = QFormLayout()
        self._maker   = QLineEdit(us.get("maker", ""))
        self._process = QLineEdit(us.get("process", ""))
        self._tool    = QLineEdit(us.get("tool", ""))
        self._exchange = QDoubleSpinBox()
        self._exchange.setRange(0, 1e9); self._exchange.setDecimals(2)
        self._exchange.setValue(float(us.get("exchange", 0) or 0))
        now = datetime.now()
        y, m = us.get("base_ym", (now.year, now.month))
        self._year  = QSpinBox(); self._year.setRange(2000, 2100); self._year.setValue(int(y))
        self._month = QSpinBox(); self._month.setRange(1, 12);      self._month.setValue(int(m))
        self._site  = QComboBox(); self._site.addItems(["Taylor", "Austin"])
        site = us.get("site", "Taylor")
        self._site.setCurrentText(site if site in ("Taylor", "Austin") else "Taylor")
        ym_row = QWidget()
        ym_h = QHBoxLayout(ym_row); ym_h.setContentsMargins(0,0,0,0)
        ym_h.addWidget(self._year); ym_h.addWidget(QLabel("년"))
        ym_h.addWidget(self._month); ym_h.addWidget(QLabel("월"))
        ym_h.addStretch(1)
        form.addRow("Maker",         self._maker)
        form.addRow("Process",       self._process)
        form.addRow("설비호기",      self._tool)
        form.addRow("환율",          self._exchange)
        form.addRow("기준일(년/월)", ym_row)
        form.addRow("Site",          self._site)
        root.addLayout(form)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self.accept); bb.rejected.connect(self.reject)
        root.addWidget(bb)

    def apply(self) -> None:
        self.state.us_info = {
            "maker": self._maker.text().strip(), "process": self._process.text().strip(),
            "tool":  self._tool.text().strip(),  "exchange": float(self._exchange.value()),
            "base_ym": (int(self._year.value()), int(self._month.value())),
            "site": self._site.currentText(),
        }


# ══════════════════════════════════════════════
# STEP5 테이블 관리 믹스인
# ══════════════════════════════════════════════

class Step5Manager:
    """
    STEP5 테이블의 행 CRUD + TOTAL/CREDIT 행 관리.
    QuoteBuilderPage 가 다중 상속하여 사용한다.

    하위 클래스가 반드시 제공해야 하는 속성:
        self.step5_table      : DroppableQuoteTable
        self.chk_step5_all    : QCheckBox
        self.state            : QuoteState
        self._step4_highlight : set
    하위 클래스가 구현해야 하는 메서드:
        self._render_items_table(scroll_top: bool)
    """

    # ── 행 종류 판별 ────────────────────────────
    def _row_role(self, row: int) -> str:
        it = self.step5_table.item(row, 1)
        return s(it.data(Qt.UserRole)) if it else ""

    def _is_total(self, row: int)  -> bool: return self._row_role(row) == "TOTAL"
    def _is_credit(self, row: int) -> bool: return self._row_role(row) == "CREDIT"
    def _is_item(self, row: int)   -> bool: return self._row_role(row) == "ITEM"

    # ── 셀 setter ───────────────────────────────
    def _set_qty(self, row: int, qty: float, blank: bool = False) -> None:
        it = QTableWidgetItem("" if blank else fmt_qty(qty))
        it.setData(Qt.UserRole, qty); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(row, 3, it)

    def _set_price(self, row: int, price: float, blank: bool = False) -> None:
        it = QTableWidgetItem("" if blank else fmt_krw(price))
        it.setData(Qt.UserRole, price); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(row, 4, it)

    def _set_amt(self, row: int, amt: float) -> None:
        it = QTableWidgetItem(fmt_krw(amt))
        it.setData(Qt.UserRole, amt); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(row, 5, it)

    def _get_float(self, row: int, col: int) -> float:
        it = self.step5_table.item(row, col)
        return float(it.data(Qt.UserRole) or 0.0) if it else 0.0

    # ── TOTAL 행 보장 ────────────────────────────
    def _ensure_total(self) -> None:
        n = self.step5_table.rowCount()
        if n == 0 or not self._is_total(n - 1):
            self._append_total()

    def _append_total(self) -> None:
        r = self.step5_table.rowCount()
        self.step5_table.insertRow(r)
        it_cat = QTableWidgetItem("TOTAL")
        it_cat.setData(Qt.UserRole, "TOTAL"); it_cat.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(r, 1, it_cat)
        it_spec = QTableWidgetItem("총 합계")
        it_spec.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(r, 2, it_spec)
        self._set_qty(r, 0, blank=True); self._set_price(r, 0, blank=True); self._set_amt(r, 0)

    # ── 합계 재계산 ──────────────────────────────
    def _recalc_totals(self) -> None:
        self._ensure_total()
        total_qty = total_amt = 0.0
        last = self.step5_table.rowCount() - 1
        for r in range(last):
            if self._is_total(r): continue
            total_amt += self._get_float(r, 5)
            if not self._is_credit(r):
                total_qty += self._get_float(r, 3)
        self._set_qty(last, total_qty, blank=True)
        self._set_amt(last, total_amt)

    def _recalc_row(self, row: int) -> None:
        self._set_amt(row, self._get_float(row, 3) * self._get_float(row, 4))

    # ── 행 추가 ─────────────────────────────────
    def _add_row(self, category: str, spec: str, qty: float, unit_price: float) -> None:
        """같은 spec 이면 수량 누적, 없으면 TOTAL 행 위에 새 행 삽입."""
        self._ensure_total()
        last = self.step5_table.rowCount() - 1
        for r in range(last):
            if self._is_total(r) or self._is_credit(r): continue
            it = self.step5_table.item(r, 2)
            if it and it.text().strip() == spec.strip():
                self._set_qty(r, self._get_float(r, 3) + qty)
                self._recalc_row(r); self._recalc_totals(); return

        ins = last
        self.step5_table.insertRow(ins)
        cb = QCheckBox()
        cb.stateChanged.connect(self._on_step5_check_changed)
        self.step5_table.setCellWidget(ins, 0, cb)
        it_cat = QTableWidgetItem(category)
        it_cat.setData(Qt.UserRole, "ITEM"); it_cat.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(ins, 1, it_cat)
        it_spec = QTableWidgetItem(spec)
        it_spec.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(ins, 2, it_spec)
        self._set_qty(ins, qty); self._set_price(ins, unit_price)
        self._set_amt(ins, qty * unit_price); self._recalc_totals()
        self._sort_step5_items()

    # ── Credit 행 ────────────────────────────────
    def _clear_credit_rows(self) -> None:
        for r in range(self.step5_table.rowCount() - 1, -1, -1):
            if self._is_credit(r): self.step5_table.removeRow(r)
        self._ensure_total()

    def _add_credit_row(self, label: str, amount: float) -> None:
        self._ensure_total()
        ins = self.step5_table.rowCount() - 1
        self.step5_table.insertRow(ins)
        red = QBrush(QColor(200, 0, 0))
        it_cat = QTableWidgetItem("CREDIT")
        it_cat.setData(Qt.UserRole, "CREDIT"); it_cat.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(ins, 1, it_cat)
        it_spec = QTableWidgetItem(label)
        it_spec.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable); it_spec.setForeground(red)
        self.step5_table.setItem(ins, 2, it_spec)
        self.step5_table.setItem(ins, 3, QTableWidgetItem(""))
        self.step5_table.setItem(ins, 4, QTableWidgetItem(""))
        it_amt = QTableWidgetItem(fmt_krw(amount))
        it_amt.setData(Qt.UserRole, float(amount)); it_amt.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        it_amt.setForeground(red); self.step5_table.setItem(ins, 5, it_amt)

    def _refresh_credits(self) -> None:
        self._clear_credit_rows()
        if self.state.pump_credit: self._add_credit_row("Pump Credit", -abs(self.state.pump_credit))
        if self.state.rack_credit: self._add_credit_row("Rack Credit", -abs(self.state.rack_credit))

    # ── STEP5 정렬 (STEP4 품목 순서 기준) ────────
    def _sort_step5_items(self) -> None:
        """ITEM 행을 self._spec_order 순서에 따라 제자리 정렬."""
        spec_order: Dict[str, int] = getattr(self, "_spec_order", {})
        if not spec_order:
            return
        item_rows = [r for r in range(self.step5_table.rowCount()) if self._is_item(r)]
        if len(item_rows) <= 1:
            return
        # 각 행 데이터 스냅샷
        snaps = []
        for r in item_rows:
            cb = self.step5_table.cellWidget(r, 0)
            cat_it  = self.step5_table.item(r, 1)
            spec_it = self.step5_table.item(r, 2)
            snaps.append({
                "checked": cb.isChecked() if isinstance(cb, QCheckBox) else False,
                "cat":     s(cat_it.text())  if cat_it  else "",
                "spec":    s(spec_it.text()) if spec_it else "",
                "qty":     self._get_float(r, 3),
                "price":   self._get_float(r, 4),
                "amt":     self._get_float(r, 5),
            })
        snaps.sort(key=lambda x: (0 if x["cat"] == "PUMP" else 1, spec_order.get(x["spec"], 10**9)))
        # 정렬된 데이터를 같은 위치에 재기입 (행 삽입/삭제 없이)
        self.step5_table.blockSignals(True)
        for r, d in zip(item_rows, snaps):
            cb = self.step5_table.cellWidget(r, 0)
            if isinstance(cb, QCheckBox):
                cb.blockSignals(True); cb.setChecked(d["checked"]); cb.blockSignals(False)
            cat_it  = self.step5_table.item(r, 1)
            spec_it = self.step5_table.item(r, 2)
            if cat_it:  cat_it.setText(d["cat"])
            if spec_it: spec_it.setText(d["spec"])
            self._set_qty(r, d["qty"])
            self._set_price(r, d["price"])
            self._set_amt(r, d["amt"])
        self.step5_table.blockSignals(False)

    # ── STEP5 스냅샷 ─────────────────────────────
    def _snapshot_items(self) -> List[Dict[str, Any]]:
        rows = []
        for r in range(self.step5_table.rowCount()):
            if self._is_total(r): continue
            cat_it  = self.step5_table.item(r, 1)
            spec_it = self.step5_table.item(r, 2)
            rows.append({
                "role":  s(cat_it.data(Qt.UserRole)) if cat_it else "",
                "cat":   s(cat_it.text())             if cat_it else "",
                "spec":  s(spec_it.text())            if spec_it else "",
                "qty":   self._get_float(r, 3),
                "price": self._get_float(r, 4),
                "amt":   self._get_float(r, 5),
            })
        return rows

    # ── 전체선택 동기화 ──────────────────────────
    def _sync_step5_all_chk(self) -> None:
        selectable = checked = 0
        last = self.step5_table.rowCount() - 1
        for r in range(last + 1):
            if self._is_total(r) or self._is_credit(r): continue
            cb = self.step5_table.cellWidget(r, 0)
            if isinstance(cb, QCheckBox):
                selectable += 1
                if cb.isChecked(): checked += 1
        self.chk_step5_all.blockSignals(True)
        if   selectable == 0 or checked == 0: self.chk_step5_all.setCheckState(Qt.Unchecked)
        elif checked == selectable:            self.chk_step5_all.setCheckState(Qt.Checked)
        else:                                  self.chk_step5_all.setCheckState(Qt.PartiallyChecked)
        self.chk_step5_all.blockSignals(False)

    def _on_step5_check_changed(self, _state: int) -> None:
        """개별 체크박스 변경 → 전체선택 상태 동기화."""
        self._sync_step5_all_chk()

    def _on_step5_select_all(self, checked: bool) -> None:
        """전체선택 체크박스 핸들러."""
        last = self.step5_table.rowCount() - 1
        for r in range(last + 1):
            if self._is_total(r) or self._is_credit(r): continue
            cb = self.step5_table.cellWidget(r, 0)
            if isinstance(cb, QCheckBox):
                cb.blockSignals(True); cb.setChecked(checked); cb.blockSignals(False)
        self._sync_step5_all_chk()

    # ── STEP4 하이라이트 동기화 ──────────────────
    def _sync_step4_highlight(self, scroll_top: bool = True) -> None:
        transferred: set = set()
        for r in range(self.step5_table.rowCount()):
            if self._is_total(r) or self._is_credit(r): continue
            it = self.step5_table.item(r, 2)
            if it and it.text().strip():
                transferred.add(it.text().strip())
        self._step4_highlight = transferred
        self._render_items_table(scroll_top=scroll_top)

    def _render_items_table(self, scroll_top: bool = True) -> None:
        """하위 클래스(QuoteBuilderPage)에서 구현."""
        raise NotImplementedError("_render_items_table must be implemented by subclass")


# ══════════════════════════════════════════════
# 견적서 LIST 독립 창
# ══════════════════════════════════════════════

class _QuoteListWindow(QWidget):
    """견적서 LIST를 보여주는 독립(플로팅) 창."""

    def __init__(self, parent=None):
        super().__init__(parent, Qt.Window)
        self.setWindowTitle("견적서 LIST")
        self.resize(480, 620)
        # (tool, label, full_path) — 추가 시점에 한 번만 파싱해 캐시
        self._session_entries: List[Tuple[str, str, str]] = []
        self._build()

    def _build(self) -> None:
        v = QVBoxLayout(self)
        v.setContentsMargins(12, 12, 12, 12); v.setSpacing(6)

        title_row = QHBoxLayout()
        lbl = QLabel("견적서 LIST")
        f = lbl.font(); f.setPointSize(12); f.setBold(True); lbl.setFont(f)
        title_row.addWidget(lbl)
        v.addLayout(title_row)

        filter_row = QHBoxLayout(); filter_row.setSpacing(6)
        lbl_f = QLabel("설비호기:"); lbl_f.setFixedWidth(52)
        self.edit_filter = QLineEdit()
        self.edit_filter.setPlaceholderText("설비호기 입력 → 실시간 필터")
        self.edit_filter.setFixedHeight(24)
        self.edit_filter.setClearButtonEnabled(True)
        self.edit_filter.textChanged.connect(self._apply_filter)
        filter_row.addWidget(lbl_f); filter_row.addWidget(self.edit_filter)
        v.addLayout(filter_row)

        self.list_widget = QListWidget()
        self.list_widget.itemDoubleClicked.connect(self._open_item)
        v.addWidget(self.list_widget, 1)

    def add_item(self, path: str) -> None:
        entry = QuoteBuilderPage._parse_quote_entry(path)
        if entry is not None:
            self._session_entries.insert(0, entry)
        self._apply_filter()

    def refresh_session(self, files: List[str]) -> None:
        self._session_entries = [
            e for e in (QuoteBuilderPage._parse_quote_entry(f) for f in files)
            if e is not None
        ]
        self._apply_filter()

    def _apply_filter(self) -> None:
        keyword = self.edit_filter.text().strip().lower()
        self.list_widget.clear()

        filtered = [
            (tool, label, full)
            for tool, label, full in self._session_entries
            if not keyword or keyword in label.lower()
        ]

        # 중복 레이블 카운트: 두 번째부터 (2), (3) … 표시
        label_counts: Dict[str, int] = {}
        for _, label, _ in filtered:
            label_counts[label] = label_counts.get(label, 0) + 1
        label_seen: Dict[str, int] = {}

        for tool, label, full in filtered:
            if label_counts[label] > 1:
                label_seen[label] = label_seen.get(label, 0) + 1
                n = label_seen[label]
                display = label if n == 1 else f"{label} ({n})"
            else:
                display = label
            item = QListWidgetItem(display)
            item.setData(Qt.UserRole, full)
            item.setToolTip(full)
            if tool == "":
                item.setBackground(QBrush(QColor(255, 153, 153)))
            self.list_widget.addItem(item)

    def _open_item(self, item: QListWidgetItem) -> None:
        path = item.data(Qt.UserRole)
        if path:
            try: os.startfile(path)
            except Exception as e: QMessageBox.critical(self, "오류", str(e))

    def show_and_raise(self) -> None:
        self.show()
        self.raise_()
        self.activateWindow()
        if self.list_widget.count() > 0:
            self.list_widget.scrollToTop()


# ══════════════════════════════════════════════
# 견적서 작성 페이지
# ══════════════════════════════════════════════

class QuoteBuilderPage(Step5Manager, QWidget):

    def __init__(self) -> None:
        QWidget.__init__(self)
        self.state = QuoteState()

        self._step4_highlight     : set                     = set()
        self._filtered_items      : List[Tuple[str, float]] = []
        self._spec_order          : Dict[str, int]          = {}
        self._done_req_indices    : set                     = set()   # 견적 완료된 의뢰행 인덱스
        self._pending_req_indices : List[int]               = []      # 현재 생성 중인 의뢰행 인덱스

        self._filter_timer  : QTimer                          = QTimer(self)
        self._filter_timer.setSingleShot(True)
        self._filter_timer.timeout.connect(self._do_filter)
        self._tpl_thread    : Optional[_TemplateLoaderThread]  = None
        self._req_thread    : Optional[_RequestLoaderThread]   = None
        self._gen_qt_thread : Optional[_GenQuoteThread]        = None
        self._gen_cv_thread  : Optional[_GenCoverThread]        = None
        self._cover_progress : Optional[QProgressDialog]      = None
        self._list_all_files : List[str]                       = []
        self._quote_list_window: Optional[_QuoteListWindow]    = None

        self._build_ui()
        self._ensure_total()
        self._log("준비: STEP1 → STEP2 → STEP4/드래그 → STEP5 → 생성")

        sc = QShortcut(QKeySequence.Delete, self)
        sc.activated.connect(self._delete_checked_rows)

    # ══════════════════════════════════════════
    # UI 구성
    # ══════════════════════════════════════════

    def _build_ui(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(14, 14, 14, 14)
        outer.setSpacing(8)

        # ── 툴바 (STEP1·2 LOAD + 생성버튼 + LIST 버튼) ──────────────
        outer.addWidget(self._build_toolbar())

        # ── 메인 콘텐츠 (STEP3+의뢰 | STEP4|STEP5) ──────────────────
        content = QWidget()
        ch = QHBoxLayout(content)
        ch.setContentsMargins(0, 0, 0, 0); ch.setSpacing(12)

        # 왼쪽: STEP3 (위) + 의뢰파일DATA (아래)
        left = QWidget(); left.setFixedWidth(630)
        lv = QVBoxLayout(left)
        lv.setContentsMargins(0, 0, 0, 0); lv.setSpacing(8)
        self._step3_frame = self._build_step3()
        self._step3_frame.setEnabled(False)
        lv.addWidget(self._step3_frame)
        lv.addWidget(self._build_req_panel(), 1)
        ch.addWidget(left)

        # 오른쪽: STEP4 (왼) | STEP5 (오) 가로 배치
        ch.addWidget(self._build_right_panel(), 1)

        outer.addWidget(content, 1)

        # ── 하단: 로그 + 옵션 ────────────────────────────────────────
        bottom = QWidget(); bottom.setFixedHeight(140)
        bh = QHBoxLayout(bottom); bh.setContentsMargins(0, 0, 0, 0); bh.setSpacing(12)
        bh.addWidget(self._build_worklog(), 1)
        bh.addWidget(self._build_bottom_right_panel())
        outer.addWidget(bottom)


    @staticmethod
    def _action_btn(label: str, slot) -> QPushButton:
        btn = QPushButton(label); btn.setFixedHeight(52)
        f = btn.font(); f.setPointSize(10); f.setBold(True); btn.setFont(f)
        btn.clicked.connect(slot); return btn

    # ── 툴바 ──────────────────────────────────
    def _build_toolbar(self) -> QWidget:
        bar = QWidget(); bar.setFixedHeight(58)
        h = QHBoxLayout(bar); h.setContentsMargins(0, 0, 0, 0); h.setSpacing(6)

        self.btn_step1 = self._action_btn("STEP 1\n통합양식 LOAD", self._load_template)
        self.btn_step2 = self._action_btn("STEP 2\n의뢰파일 LOAD", self._load_request)
        tint_button(self.btn_step1, "#C8E6C9")
        tint_button(self.btn_step2, "#BBDEFB")
        self.btn_step2.setEnabled(False)

        self.btn_gen_quote = self._action_btn("견적서 생성", self._generate_quote)
        self.btn_gen_cover = self._action_btn("갑지 생성",   self._generate_cover)
        tint_button(self.btn_gen_quote, "#B3E5FC")
        tint_button(self.btn_gen_cover, "#F8BBD0")

        btn_list = QPushButton("견적서 LIST"); btn_list.setFixedHeight(52)
        f = btn_list.font(); f.setPointSize(10); f.setBold(True); btn_list.setFont(f)
        tint_button(btn_list, "#E8EAF6")
        btn_list.clicked.connect(self._open_quote_list_window)

        sep = QFrame(); sep.setFrameShape(QFrame.VLine); sep.setFrameShadow(QFrame.Sunken)

        h.addWidget(self.btn_step1)
        h.addWidget(self.btn_step2)
        h.addWidget(sep)
        h.addWidget(self.btn_gen_quote)
        h.addWidget(self.btn_gen_cover)
        h.addWidget(btn_list)
        h.addStretch(1)
        return bar

    # ── STEP3 ─────────────────────────────────
    def _build_step3(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2); frame.setMinimumHeight(110)
        v = QVBoxLayout(frame); v.setContentsMargins(12,12,12,12); v.setSpacing(6)
        v.addWidget(bold_label("STEP 3\n조건 입력"))
        row1 = QHBoxLayout()
        self.cb_process = QComboBox(); self.cb_process.addItems(PROCESS_CHOICES)
        self.cb_vendor  = QComboBox(); self.cb_vendor.addItems(VENDOR_CHOICES)
        row1.addWidget(QLabel("공정")); row1.addWidget(self.cb_process, 1)
        row1.addWidget(QLabel("설비사")); row1.addWidget(self.cb_vendor, 1)
        row2 = QHBoxLayout()
        self.ed_code = QLineEdit(); self.ed_code.setPlaceholderText("5D")
        self.btn_apply_map = QPushButton("코드매핑 적용 → STEP5(RACK)"); self.btn_apply_map.setEnabled(False)
        tint_button(self.btn_apply_map, "#FFE0B2")  # 연주황
        row2.addWidget(QLabel("5D")); row2.addWidget(self.ed_code, 1); row2.addWidget(self.btn_apply_map)
        v.addLayout(row1); v.addLayout(row2)
        self.cb_process.currentTextChanged.connect(self._on_step3_changed)
        self.cb_vendor.currentTextChanged.connect(self._on_step3_changed)
        self.ed_code.textChanged.connect(self._on_step3_changed)
        self.btn_apply_map.clicked.connect(self._apply_code_map)
        return frame

    # ── 의뢰파일 패널 ─────────────────────────
    def _build_req_panel(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2)
        frame.setMinimumHeight(180)
        v = QVBoxLayout(frame); v.setContentsMargins(12,12,12,12); v.setSpacing(8)
        v.addWidget(bold_label("의뢰파일DATA", size=12))
        top = QHBoxLayout()
        self.chk_req_all = QCheckBox("전체선택"); self.chk_req_all.clicked.connect(self._on_req_select_all)
        top.addWidget(self.chk_req_all); top.addStretch(1); v.addLayout(top)
        self.req_table = _make_plain_table(
            8, ["선택", "설비호기(Z)", "투자정보", "수량", "Rack", "Maker", "설비", "5D"])
        self.req_table.setColumnWidth(0, 42)
        _rh = self.req_table.horizontalHeader()
        _rh.setSectionResizeMode(0, QHeaderView.Fixed)
        for _ci in range(1, 8):
            _rh.setSectionResizeMode(_ci, QHeaderView.Interactive)
        self.req_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.req_table.cellDoubleClicked.connect(self._on_req_double_click)
        self._style_req_table()
        v.addWidget(self.req_table, 1)
        return frame

    # ── STEP4 패널 ────────────────────────────
    def _build_step4_panel(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2); frame.setMinimumHeight(150)
        v = QVBoxLayout(frame); v.setContentsMargins(12,12,12,12); v.setSpacing(8)
        v.addWidget(bold_label("STEP 4\n품목(규격/단가) + 필터  (더블클릭/드래그 → STEP5)"))
        self.ed_filter = QLineEdit(); self.ed_filter.setPlaceholderText("규격 필터(실시간) - 포함 검색")
        self.ed_filter.textChanged.connect(self._apply_filter)
        self.items_table = DraggableItemsTable(0, 2)
        self.items_table.setHorizontalHeaderLabels(["규격", "단가(원)"])
        self.items_table.verticalHeader().setVisible(False)
        self.items_table.setAlternatingRowColors(True)
        self.items_table.setEditTriggers(DraggableItemsTable.NoEditTriggers)
        self.items_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.items_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.items_table.setDragEnabled(True); self.items_table.setWordWrap(False)
        self.items_table.setStyleSheet("QTableWidget::item:selected{color:black;}")
        h = self.items_table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.Stretch); h.setSectionResizeMode(1, QHeaderView.Fixed)
        self.items_table.setColumnWidth(1, 80)
        self.items_table.cellDoubleClicked.connect(self._on_item_double_click)
        v.addWidget(self.ed_filter); v.addWidget(self.items_table, 1)
        return frame

    # ── 우측 패널 (STEP4 왼 | STEP5 오 가로 배치) ─
    def _build_right_panel(self) -> QWidget:
        splitter = QSplitter(Qt.Horizontal)

        splitter.addWidget(self._build_step4_panel())
        splitter.addWidget(self._build_step5_panel())
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 3)
        splitter.setSizes([400, 600])
        return splitter

    def _build_step5_panel(self) -> QFrame:
        step5 = labeled_frame("STEP5 견적작성 대상", min_h=150)
        step5.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        sl = step5.layout()
        top5 = QHBoxLayout()
        self.chk_step5_all = QCheckBox("전체선택")
        self.chk_step5_all.clicked.connect(self._on_step5_select_all)
        top5.addWidget(self.chk_step5_all); top5.addStretch(1); sl.addLayout(top5)

        self.step5_table = DroppableQuoteTable()
        self.step5_table.setColumnCount(6)
        self.step5_table.setHorizontalHeaderLabels(["선택","분류","품목","수량","단가(원)","금액(원)"])
        self.step5_table.verticalHeader().setVisible(False)
        self.step5_table.setAlternatingRowColors(True)
        self.step5_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.step5_table.setEditTriggers(DroppableQuoteTable.NoEditTriggers)
        self.step5_table.cellDoubleClicked.connect(self._on_step5_double_click)
        self.step5_table.drop_callback = self._on_drop_to_step5
        h5 = self.step5_table.horizontalHeader()
        h5.setSectionResizeMode(0, QHeaderView.Fixed)
        for i in [1, 3, 4, 5]:
            h5.setSectionResizeMode(i, QHeaderView.Fixed)
        h5.setSectionResizeMode(2, QHeaderView.Interactive)
        h5.setStretchLastSection(False)
        self.step5_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.step5_table.setColumnWidth(0, 46)
        for c in [1, 3, 4, 5]:
            self.step5_table.resizeColumnToContents(c)
        h5.sectionResized.connect(self._on_step5_col2_resized)
        self.step5_table.viewport().installEventFilter(self)
        QTimer.singleShot(0, self._step5_fit_col2)
        sl.addWidget(self.step5_table, 1)
        return step5

    def _build_worklog(self) -> QFrame:
        frame = labeled_frame("작업로그", min_h=140)
        frame.setFixedHeight(140); frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.log_view = QTextEdit(); self.log_view.setReadOnly(True)
        frame.layout().addWidget(self.log_view, 1); return frame

    def _build_bottom_right_panel(self) -> QWidget:
        w = QWidget(); w.setFixedHeight(140)
        h = QHBoxLayout(w); h.setContentsMargins(0,0,0,0); h.setSpacing(12)
        h.addWidget(self._build_options_panel(), 1)
        h.addWidget(self._build_quote_info_panel())
        return w

    def _build_options_panel(self) -> QFrame:
        frame = labeled_frame("옵션", min_h=55)
        frame.setMaximumHeight(140); frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        row = QHBoxLayout()

        self.btn_credit = QPushButton("Credit"); self.btn_credit.setMinimumHeight(32)
        self.btn_credit.clicked.connect(self._open_credit_dialog)
        tint_button(self.btn_credit, "#FFF9C4")   # 연노랑

        # 보증기간 변경 + 투자자 변경을 세로로 묶음
        self.btn_warranty = QPushButton("보증기간 변경"); self.btn_warranty.setMinimumHeight(28)
        self.btn_warranty.clicked.connect(self._change_warranty)
        tint_button(self.btn_warranty, "#E0F2F1")   # 연민트

        self.btn_investor = QPushButton("투자자 변경"); self.btn_investor.setMinimumHeight(28)
        self.btn_investor.clicked.connect(self._change_investor)
        tint_button(self.btn_investor, "#E8EAF6")   # 연라벤더

        btn_col = QVBoxLayout(); btn_col.setSpacing(4)
        btn_col.addWidget(self.btn_warranty)
        btn_col.addWidget(self.btn_investor)

        self.chk_qt_kr = QCheckBox("국내"); self.chk_qt_cn = QCheckBox("중국"); self.chk_qt_us = QCheckBox("미국")
        self.chk_qt_kr.setChecked(True)
        for chk, qt in [(self.chk_qt_kr,"국내"),(self.chk_qt_cn,"중국"),(self.chk_qt_us,"미국")]:
            chk.clicked.connect(lambda checked, t=qt: self._on_qt_checked(t, checked))

        row.addWidget(self.btn_credit)
        row.addLayout(btn_col)
        row.addSpacing(10)
        row.addWidget(QLabel("견적서 타입"))
        row.addWidget(self.chk_qt_kr); row.addWidget(self.chk_qt_cn); row.addWidget(self.chk_qt_us)
        row.addStretch(1); frame.layout().addLayout(row); return frame

    def _build_quote_info_panel(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2)
        frame.setFixedHeight(140); frame.setFixedWidth(260)
        v = QVBoxLayout(frame); v.setContentsMargins(12,6,12,6); v.setSpacing(4)
        v.addWidget(bold_label("견적서정보", size=9))
        self.lbl_pr = info_label("PR: -"); self.lbl_item = info_label("항번: -")
        self.lbl_line = info_label("라인공정: -"); self.lbl_investor = info_label("투자자: -")
        self.lbl_more = info_label("")
        for lbl in (self.lbl_pr, self.lbl_item, self.lbl_line, self.lbl_investor): v.addWidget(lbl)
        v.addStretch(1); v.addWidget(self.lbl_more); return frame

    def _style_req_table(self) -> None:
        self.req_table.setAlternatingRowColors(False)
        # item background 는 지정하지 않아야 _mark_req_done 의 setBackground 가 반영된다
        self.req_table.setStyleSheet(
            "QTableWidget{background:transparent;border:none;}"
            "QTableWidget::item:selected{color:black;background:rgba(180,200,230,140);}")

    # ══════════════════════════════════════════
    # STEP1
    # ══════════════════════════════════════════

    def _load_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "통합양식 선택", "", "Excel Files (*.xlsx)")
        if not path:
            return
        # 시트명 사전 검증 (빠름 — 메타만 읽음)
        try:
            from openpyxl import load_workbook
            wb_meta = load_workbook(path, read_only=True, data_only=False)
            missing = [n for n in SheetName.REQUIRED if n not in wb_meta.sheetnames]
            wb_meta.close()
        except Exception as e:
            QMessageBox.critical(self, "오류", f"파일을 열 수 없습니다.\n{e}")
            return
        if missing:
            QMessageBox.warning(self, "통합양식 오류", "필수 시트 없음:\n" + "\n".join(missing))
            return

        self.btn_step1.setEnabled(False)
        self.btn_step1.setText("STEP 1\n로딩 중…")
        self._tpl_thread = _TemplateLoaderThread(path, self)
        self._tpl_thread.done.connect(lambda result: self._on_template_loaded(path, result))
        self._tpl_thread.start()

    def _on_template_loaded(self, path: str, result) -> None:
        self.btn_step1.setEnabled(True)
        self.btn_step1.setText("STEP 1\n통합양식 LOAD")
        if isinstance(result, Exception):
            logger.error("통합양식 파싱 오류", exc_info=result)
            QMessageBox.critical(self, "파싱 오류", str(result))
            return
        rows, by_class, price_by_spec, order, cmap = result
        self.state.template_path = path
        self.state.items_rows = rows
        self.state.items_by_class = by_class
        self.state.price_by_spec = price_by_spec
        self.state.code_map = cmap
        self._spec_order = order
        self._apply_filter()
        self.btn_step2.setEnabled(True)
        self._step3_frame.setEnabled(True)
        self._log(f"STEP1 완료: {path}  (품목 {len(rows)}건, 코드매핑 키 {len(cmap)}개)")
        QMessageBox.information(self, "완료", "통합양식 LOAD 완료")

    # ══════════════════════════════════════════
    # STEP2
    # ══════════════════════════════════════════

    def _load_request(self) -> None:
        if not self.state.template_path:
            QMessageBox.warning(self, "순서 오류", "STEP1을 먼저 완료하세요."); return
        if self._req_thread and self._req_thread.isRunning():
            return
        path, _ = QFileDialog.getOpenFileName(self, "의뢰파일 선택", "", "Excel Files (*.xlsx)")
        if not path: return
        self._req_thread = _RequestLoaderThread(path, self)
        self._req_thread.done.connect(lambda result: self._on_request_loaded(path, result))
        self._req_thread.start()

    def _on_request_loaded(self, path: str, result) -> None:
        if isinstance(result, Exception):
            logger.error("의뢰파일 파싱 오류", exc_info=True)
            QMessageBox.critical(self, "오류", f"파일을 읽을 수 없습니다.\n{result}"); return
        sheet_name, rows = result
        if not rows:
            QMessageBox.warning(self, "의뢰파일", "읽을 데이터가 없습니다."); return
        self.state.request_path = path; self.state.request_sheet_name = sheet_name
        self.state.request_rows = rows
        self._fill_req_table(rows); self._step3_frame.setEnabled(True)
        self._log(f"STEP2 완료: {path}  ({len(rows)}건, 시트={sheet_name})")
        QMessageBox.information(self, "완료", "의뢰파일 LOAD 완료")

    def _fill_req_table(self, rows: List[Dict[str, Any]]) -> None:
        self._done_req_indices = set()   # 새 의뢰파일 로드 시 완료 이력 초기화
        self.req_table.setRowCount(0)
        self.chk_req_all.blockSignals(True); self.chk_req_all.setCheckState(Qt.Unchecked); self.chk_req_all.blockSignals(False)
        for i, rd in enumerate(rows):
            self.req_table.insertRow(i)
            host = centered_checkbox(lambda _s, _r=i: self._on_req_check())
            self.req_table.setCellWidget(i, 0, host)
            invest_info = parse_invest_info(rd.get("G"))
            h_val = rd.get("H")
            qty_str = fmt_qty(to_float(h_val)) if h_val is not None else ""
            for col, val in [
                (1, s(rd.get("Z"))),
                (2, invest_info),
                (3, qty_str),
                (4, s(rd.get("R"))),   # Rack (R열=col18)
                (5, s(rd.get("X"))),   # Maker (X열=col24)
                (6, s(rd.get("Y"))),   # 설비 (Y열=col25)
                (7, s(rd.get("V"))),   # 5D (V열=col22)
            ]:
                it = QTableWidgetItem(val)
                it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                it.setToolTip(val)   # 잘린 텍스트 전체를 마우스오버 시 표시
                self.req_table.setItem(i, col, it)
        # 데이터 기준 컬럼 폭 자동 조정 후 체크박스 열만 고정
        self.req_table.resizeColumnsToContents()
        self.req_table.setColumnWidth(0, 42)
        self._style_req_table(); self._update_quote_info()

    def _is_req_checked(self, row: int) -> bool:
        cb = get_checkbox_from_cell(self.req_table.cellWidget(row, 0))
        return bool(cb and cb.isChecked())

    def _checked_req_rows(self) -> List[int]:
        return [r for r in range(self.req_table.rowCount()) if self._is_req_checked(r)]

    def _on_req_check(self) -> None:
        n = sum(1 for r in range(self.req_table.rowCount()) if self._is_req_checked(r))
        total = self.req_table.rowCount()
        self.chk_req_all.blockSignals(True)
        if   n == 0:     self.chk_req_all.setCheckState(Qt.Unchecked)
        elif n == total: self.chk_req_all.setCheckState(Qt.Checked)
        else:            self.chk_req_all.setCheckState(Qt.PartiallyChecked)
        self.chk_req_all.blockSignals(False); self._update_quote_info()

    def _on_req_select_all(self, checked: bool) -> None:
        for r in range(self.req_table.rowCount()):
            cb = get_checkbox_from_cell(self.req_table.cellWidget(r, 0))
            if cb: cb.blockSignals(True); cb.setChecked(checked); cb.blockSignals(False)
        self._on_req_check()

    def _update_quote_info(self) -> None:
        checked = self._checked_req_rows()
        if not checked or not self.state.request_rows:
            self.lbl_pr.setText("PR: -"); self.lbl_item.setText("항번: -")
            self.lbl_line.setText("라인공정: -"); self.lbl_investor.setText("투자자: -")
            self.lbl_more.setText(""); return
        first = self.state.request_rows[checked[0]]
        self.lbl_pr.setText(f"PR: {s(first.get('D')) or '-'}")
        self.lbl_item.setText(f"항번: {s(first.get('E')) or '-'}")
        self.lbl_line.setText(f"라인공정: {s(first.get('K')) or '-'}")
        self.lbl_investor.setText(f"투자자: {s(first.get('J')) or '-'}")
        self.lbl_more.setText(f"외 {len(checked)-1}건" if len(checked) > 1 else "")

    def _on_req_double_click(self, row: int, _col: int) -> None:
        if row >= len(self.state.request_rows): return
        rd = self.state.request_rows[row]

        def _match(text: str, choices: List[str]) -> Optional[str]:
            t = normalize_token(text)
            if not t: return None
            for c in sorted((x for x in choices if x), key=len, reverse=True):
                if normalize_token(c) in t: return c
            return None

        p = _match(s(rd.get("K")), PROCESS_CHOICES)
        if p:
            idx = self.cb_process.findText(p)
            if idx >= 0: self.cb_process.setCurrentIndex(idx)
        v = _match(s(rd.get("X")), VENDOR_CHOICES)
        if v:
            idx = self.cb_vendor.findText(v)
            if idx >= 0: self.cb_vendor.setCurrentIndex(idx)
        vcode = s(rd.get("V"))
        if vcode: self.ed_code.setText(vcode)

        f_class = s(rd.get("F")); qty = to_float(rd.get("H"))
        if f_class:
            matches = self.state.items_by_class.get(f_class, [])
            for row_data in matches:
                spec = s(row_data.get("B")); price = float(row_data.get("C", 0))
                if spec: self._add_row("PUMP", spec, qty, price)
            if matches:
                self._log(f"PUMP 자동: 분류={f_class}, {len(matches)}개, 수량={fmt_qty(qty)}")
                self._sync_step4_highlight()
        self._log(f"더블클릭 → 공정={self.cb_process.currentText()} / 설비사={self.cb_vendor.currentText()} / 5D={self.ed_code.text()}")

    # ══════════════════════════════════════════
    # STEP3
    # ══════════════════════════════════════════

    def _on_step3_changed(self) -> None:
        self.state.process = s(self.cb_process.currentText()) or None
        self.state.vendor  = s(self.cb_vendor.currentText())  or None
        self.state.code_5d = s(self.ed_code.text())           or None
        ready = bool(self.state.process and self.state.vendor and self.state.code_5d)
        self.btn_apply_map.setEnabled(ready and bool(self.state.code_map))

    def _apply_code_map(self) -> None:
        key = (self.state.process, self.state.vendor, self.state.code_5d)
        entries = self.state.code_map.get(key)
        if not entries:
            near = sum(1 for k in self.state.code_map if k[:2] == key[:2])
            QMessageBox.warning(self, "코드매핑 없음", f"키: {key}\n동일 공정/설비사 키 수: {near}"); return
        for acc, qty in entries:
            price = float(self.state.price_by_spec.get(acc, 0))
            self._add_row("RACK", acc, qty, price)
        self._log(f"코드매핑 적용 {key} → {len(entries)}개"); self._sync_step4_highlight()

    # ══════════════════════════════════════════
    # STEP4
    # ══════════════════════════════════════════

    def _apply_filter(self) -> None:
        self._filter_timer.start(200)

    def _do_filter(self) -> None:
        key = s(self.ed_filter.text()).lower()
        self._filtered_items = [
            (s(r.get("B")), float(r.get("C", 0)))
            for r in self.state.items_rows
            if s(r.get("B")) and key in s(r.get("B")).lower()
        ]
        self._render_items_table(scroll_top=False)

    def _render_items_table(self, scroll_top: bool = True) -> None:
        """Step5Manager 추상 메서드 구현: STEP4 테이블 재렌더링."""
        highlighted = self._step4_highlight

        def order(spec: str) -> int:
            return self._spec_order.get(spec, 10**9)

        sorted_rows = sorted(
            self._filtered_items,
            key=lambda x: (0 if x[0] in highlighted else 1, order(x[0])))

        self.items_table.blockSignals(True)
        self.items_table.setRowCount(len(sorted_rows))
        for i, (spec, price) in enumerate(sorted_rows):
            for col, val in [(0, spec), (1, fmt_krw(price))]:
                it = QTableWidgetItem(val); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                if spec in highlighted: it.setBackground(QBrush(QColor(220, 220, 220)))
                self.items_table.setItem(i, col, it)
        self.items_table.blockSignals(False)
        self.items_table.setColumnWidth(1, 80)
        if scroll_top: self.items_table.scrollToTop()

    def _on_item_double_click(self, row: int, _col: int) -> None:
        spec_it  = self.items_table.item(row, 0); price_it = self.items_table.item(row, 1)
        if not spec_it: return
        self._ask_qty_add(spec_it.text().strip(), to_float(price_it.text() if price_it else "0"))

    def _on_drop_to_step5(self, spec: str, price: float) -> None:
        self._ask_qty_add(spec, price)

    def _ask_qty_add(self, spec: str, unit_price: float) -> None:
        qty, ok = QInputDialog.getDouble(self, "수량 입력", f"수량 ({spec})", 1.0, 0, 1_000_000, 3)
        if ok: self._add_row("RACK", spec, qty, unit_price); self._sync_step4_highlight()

    # ══════════════════════════════════════════
    # STEP5 이벤트
    # ══════════════════════════════════════════

    def _on_step5_double_click(self, row: int, col: int) -> None:
        if self._is_total(row) or self._is_credit(row): return
        if col == 1:
            it = self.step5_table.item(row, 1)
            if it:
                it.setText("RACK" if it.text().strip() == "PUMP" else "PUMP")
                self._sort_step5_items()
                self._recalc_totals(); self._sync_step4_highlight(scroll_top=False)
        elif col == 3:
            cur = self._get_float(row, 3)
            val, ok = QInputDialog.getDouble(self, "수량 수정", "수량:", cur, 0, 1_000_000, 3)
            if ok: self._set_qty(row, val); self._recalc_row(row); self._recalc_totals(); self._sync_step4_highlight()

    def _delete_checked_rows(self) -> None:
        to_del = [
            r for r in range(self.step5_table.rowCount() - 1)
            if not self._is_total(r) and not self._is_credit(r)
            and isinstance(self.step5_table.cellWidget(r, 0), QCheckBox)
            and self.step5_table.cellWidget(r, 0).isChecked()
        ]
        if not to_del:
            sel = self.step5_table.selectionModel().selectedRows()
            if sel:
                r = sel[0].row()
                if not self._is_total(r) and not self._is_credit(r): to_del = [r]
        if not to_del: return
        reply = QMessageBox.question(self, "삭제 확인", f"{len(to_del)}개 품목을 삭제할까요?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes: return
        for r in sorted(to_del, reverse=True): self.step5_table.removeRow(r)
        self._ensure_total(); self._refresh_credits(); self._sort_step5_items(); self._recalc_totals()
        self.chk_step5_all.blockSignals(True); self.chk_step5_all.setCheckState(Qt.Unchecked); self.chk_step5_all.blockSignals(False)
        self._sync_step4_highlight()

    # ══════════════════════════════════════════
    # 옵션
    # ══════════════════════════════════════════

    def _on_qt_checked(self, selected: str, checked: bool) -> None:
        chk_map = {"국내": self.chk_qt_kr, "중국": self.chk_qt_cn, "미국": self.chk_qt_us}
        if not checked:
            # 이미 선택된 항목 재클릭 → 체크 복원 후 옵션창
            chk = chk_map[selected]; chk.blockSignals(True); chk.setChecked(True); chk.blockSignals(False)
            if selected in ("중국", "미국"):
                self._open_options(focus={"중국": "cn", "미국": "us"}[selected])
            return
        for name, chk in chk_map.items():
            chk.blockSignals(True); chk.setChecked(name == selected); chk.blockSignals(False)
        self.state.quote_type = selected; self._log(f"견적서 타입 = {selected}")
        if selected in ("중국", "미국"): self._open_options(focus={"중국": "cn", "미국": "us"}[selected])

    def _open_options(self, focus: str = "credit") -> None:
        if focus == "cn":
            dlg = ChinaOptionsDialog(self, self.state)
        elif focus == "us":
            dlg = USOptionsDialog(self, self.state)
        else:
            return
        if dlg.exec() == QDialog.Accepted:
            dlg.apply(); self._refresh_credits(); self._recalc_totals()
            self._sync_step4_highlight(scroll_top=False); self._log("[옵션] 입력값 반영")

    # ── STEP5 컬럼 너비 관리 ─────────────────────────────────────────
    _S5_ADJ = [1, 3, 4, 5]   # 품목 변경 시 비례 조정되는 컬럼들
    _S5_ADJ_MIN = 28          # 비례 컬럼 최소 너비(px)
    _S5_COL2_MIN = 60         # 품목 최소 너비(px)

    def _step5_fit_col2(self) -> None:
        """초기화 시 품목(col 2) 너비를 364px로 설정하고 나머지 칼럼들을 비례 배분."""
        target    = 364
        total     = self.step5_table.viewport().width()
        col0_w    = self.step5_table.columnWidth(0)
        min_adj   = len(self._S5_ADJ) * self._S5_ADJ_MIN
        col2      = max(self._S5_COL2_MIN, min(target, total - col0_w - min_adj))
        remaining = total - col0_w - col2
        old_ws    = [self.step5_table.columnWidth(c) for c in self._S5_ADJ]
        s         = sum(old_ws) or 1
        new_ws    = [max(self._S5_ADJ_MIN, round(w / s * remaining)) for w in old_ws]
        new_ws[-1] = max(self._S5_ADJ_MIN, remaining - sum(new_ws[:-1]))
        h5 = self.step5_table.horizontalHeader()
        h5.blockSignals(True)
        for c, w in zip(self._S5_ADJ, new_ws):
            self.step5_table.setColumnWidth(c, w)
        self.step5_table.setColumnWidth(2, col2)
        h5.blockSignals(False)

    def _on_step5_col2_resized(self, logical: int, _old: int, new_size: int) -> None:
        """품목(col 2) 드래그 시 나머지 컬럼들을 현재 비율 그대로 비례 조정."""
        if logical != 2:
            return
        total    = self.step5_table.viewport().width()
        col0_w   = self.step5_table.columnWidth(0)
        min_adj  = self._S5_ADJ_MIN
        # 품목 너비 범위 제한
        max_col2 = total - col0_w - len(self._S5_ADJ) * min_adj
        clamped  = max(self._S5_COL2_MIN, min(new_size, max_col2))
        remaining = total - col0_w - clamped

        old_ws = [self.step5_table.columnWidth(c) for c in self._S5_ADJ]
        s = sum(old_ws) or 1
        new_ws = [max(min_adj, round(w / s * remaining)) for w in old_ws]
        # 마지막 컬럼이 반올림 오차 흡수
        new_ws[-1] = max(min_adj, remaining - sum(new_ws[:-1]))

        h5 = self.step5_table.horizontalHeader()
        h5.blockSignals(True)
        for c, w in zip(self._S5_ADJ, new_ws):
            self.step5_table.setColumnWidth(c, w)
        if clamped != new_size:
            self.step5_table.setColumnWidth(2, clamped)
        h5.blockSignals(False)

    def eventFilter(self, obj, event) -> bool:
        if (hasattr(self, 'step5_table')
                and obj is self.step5_table.viewport()
                and event.type() == QEvent.Resize):
            self._step5_fit_col2()
        return super().eventFilter(obj, event)

    def _change_warranty(self) -> None:
        dlg = QDialog(self)
        dlg.setWindowTitle("보증기간 변경")
        dlg.setFixedWidth(300)
        v = QVBoxLayout(dlg)
        v.setSpacing(10)

        form = QFormLayout()
        spin = QSpinBox()
        spin.setRange(1, 10)
        spin.setValue(self.state.warranty_years)
        spin.setSuffix(" Year after delivery")
        form.addRow("보증기간:", spin)
        v.addLayout(form)

        preview = QLabel(f"{self.state.warranty_years} Year after delivery")
        preview.setStyleSheet("color: #555; font-size: 11px;")
        spin.valueChanged.connect(lambda n: preview.setText(f"{n} Year after delivery"))
        v.addWidget(preview)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        v.addWidget(btns)

        if dlg.exec() == QDialog.Accepted:
            self.state.warranty_years = spin.value()
            self._log(f"보증기간: {self.state.warranty_years} Year after delivery")

    def _change_investor(self) -> None:
        name, ok = QInputDialog.getText(
            self, "투자자 변경", "투자자 이름:",
            text=self.state.investor_name or "채승철",
        )
        if ok:
            self.state.investor_name = name.strip() or "채승철"
            self._log(f"투자자: {self.state.investor_name}")

    def _open_credit_dialog(self) -> None:
        try:
            self._open_credit_dialog_impl()
        except Exception:
            QMessageBox.critical(self, "Credit 오류", traceback.format_exc())

    def _open_credit_dialog_impl(self) -> None:
        dlg = QDialog(self); dlg.setWindowTitle("Credit 입력"); dlg.setFixedWidth(500)
        v = QVBoxLayout(dlg); v.setSpacing(10)

        h_val = to_float(self._pick_rd().get("H")) if self.state.request_rows else 0.0

        # ── 카테고리별 품목 합계 ──────────────────────────────────
        def _cat_sum(is_pump: bool) -> float:
            total = 0.0
            for r in range(self.step5_table.rowCount()):
                if self._is_item(r):
                    cat_it = self.step5_table.item(r, 1)
                    if cat_it and (s(cat_it.text()) == "PUMP") == is_pump:
                        total += self._get_float(r, 5)
            return total

        pump_sum = _cat_sum(True)
        rack_sum = _cat_sum(False)

        # ── 섹션 빌더 ────────────────────────────────────────────
        # 반환: (ed_credit, ed_target, ed_ch)
        def _make_section(label: str, item_sum: float, init_credit: float):
            frame = QFrame(); frame.setFrameShape(QFrame.StyledPanel)
            fv    = QVBoxLayout(frame); fv.setContentsMargins(8, 6, 8, 6); fv.setSpacing(6)
            hdr = QLabel(f"<b>{label}</b>  (품목 합계: {fmt_krw(item_sum)} 원)")
            fv.addWidget(hdr)

            form = QFormLayout(); form.setVerticalSpacing(6)
            fv.addLayout(form)

            ed_credit = QLineEdit(); ed_credit.setPlaceholderText("0")
            ed_target = QLineEdit(); ed_target.setPlaceholderText(fmt_krw(item_sum))
            ed_ch     = QLineEdit(); ed_ch.setPlaceholderText(
                fmt_krw(item_sum / h_val) if h_val > 0 else "CH 값 없음"
            )
            ed_ch.setEnabled(h_val > 0)

            form.addRow("Credit 금액(원):", ed_credit)
            form.addRow("목표가(원):",      ed_target)
            form.addRow("CH당 단가(원/CH):", ed_ch)

            if init_credit:
                ed_credit.setText(fmt_krw(init_credit))

            updating = [False]

            def _from_credit():
                if updating[0]: return
                updating[0] = True
                try:
                    c = to_float(ed_credit.text())
                    t = item_sum - c
                    ed_target.setText(fmt_krw(max(0.0, t)))
                    if h_val > 0:
                        ed_ch.setText(fmt_krw(max(0.0, t) / h_val))
                finally:
                    updating[0] = False

            def _from_target():
                if updating[0]: return
                updating[0] = True
                try:
                    t = to_float(ed_target.text())
                    c = max(0.0, item_sum - t)
                    ed_credit.setText(fmt_krw(c))
                    if h_val > 0:
                        ed_ch.setText(fmt_krw(t / h_val))
                finally:
                    updating[0] = False

            def _from_ch():
                if updating[0]: return
                if h_val <= 0: return
                updating[0] = True
                try:
                    ch = to_float(ed_ch.text())
                    t  = ch * h_val
                    c  = max(0.0, item_sum - t)
                    ed_credit.setText(fmt_krw(c))
                    ed_target.setText(fmt_krw(max(0.0, t)))
                finally:
                    updating[0] = False

            ed_credit.textEdited.connect(_from_credit)
            ed_target.textEdited.connect(_from_target)
            ed_ch.textEdited.connect(_from_ch)

            return frame, ed_credit, ed_target, ed_ch

        grp_pump, ed_pump_c, ed_pump_t, ed_pump_ch = _make_section(
            "PUMP", pump_sum, self.state.pump_credit)
        grp_rack, ed_rack_c, ed_rack_t, ed_rack_ch = _make_section(
            "RACK", rack_sum, self.state.rack_credit)

        v.addWidget(grp_pump)
        v.addWidget(grp_rack)

        # ── 결과 요약 ─────────────────────────────────────────────
        sep = QFrame(); sep.setFrameShape(QFrame.HLine); sep.setFrameShadow(QFrame.Sunken)
        v.addWidget(sep)
        lbl_summary = QLabel(); lbl_summary.setStyleSheet("font-weight:bold;")
        lbl_ch_total = QLabel()
        v.addWidget(lbl_summary); v.addWidget(lbl_ch_total)

        def _refresh_summary() -> None:
            pc    = to_float(ed_pump_c.text())
            rc    = to_float(ed_rack_c.text())
            after = pump_sum + rack_sum - pc - rc
            lbl_summary.setText(
                f"Credit 반영 후 총액: {fmt_krw(after)} 원"
                f"  (감액 Pump {fmt_krw(pc)} + Rack {fmt_krw(rc)} = {fmt_krw(pc + rc)} 원)"
            )
            if h_val > 0:
                lbl_ch_total.setText(
                    f"전체 CH당 단가: {fmt_krw(after / h_val)} 원/CH  (H: {fmt_qty(h_val)} CH)"
                )
            else:
                lbl_ch_total.setText("전체 CH당 단가: H열 값 없음")

        for ed in (ed_pump_c, ed_pump_t, ed_pump_ch, ed_rack_c, ed_rack_t, ed_rack_ch):
            ed.textEdited.connect(_refresh_summary)

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel); v.addWidget(bb)

        def _ok() -> None:
            self.state.pump_credit = to_float(ed_pump_c.text())
            self.state.rack_credit = to_float(ed_rack_c.text())
            self._refresh_credits(); self._recalc_totals(); dlg.accept()

        bb.accepted.connect(_ok); bb.rejected.connect(dlg.reject)
        _refresh_summary(); dlg.exec()

    # ══════════════════════════════════════════
    # STEP6 – 생성
    # ══════════════════════════════════════════

    def _set_gen_buttons(self, enabled: bool) -> None:
        self.btn_gen_quote.setEnabled(enabled)
        self.btn_gen_cover.setEnabled(enabled)

    def _generate_quote(self) -> None:
        if not self.state.template_path:
            QMessageBox.warning(self, "오류", "STEP1 통합양식을 먼저 LOAD 하세요."); return
        if self._gen_qt_thread and self._gen_qt_thread.isRunning():
            QMessageBox.warning(self, "진행 중", "이미 생성 작업이 진행 중입니다."); return
        self._refresh_credits(); self._recalc_totals()
        items = self._snapshot_items(); qtype = self.state.quote_type or "국내"
        if qtype == "중국" and not self.state.cn_info.get("tool"):
            self._open_options("cn")
            if not self.state.cn_info.get("tool"): return
        if qtype == "미국" and not self.state.us_info.get("tool"):
            self._open_options("us")
            if not self.state.us_info.get("tool"): return
        checked = self._checked_req_rows()
        if qtype == "국내" and self.state.request_rows and len(checked) >= 2:
            valid = [i for i in checked if i < len(self.state.request_rows)]
            rds = [self.state.request_rows[i] for i in valid]
            self._pending_req_indices = valid
        else:
            idx = self._pick_rd_index()
            rds = [self.state.request_rows[idx]] if (idx >= 0 and self.state.request_rows) else [{}]
            self._pending_req_indices = [idx] if idx >= 0 else []
        self._log(f"견적서 생성 시작 ({qtype}, {len(rds)}건) …")
        self._set_gen_buttons(False)
        self._gen_qt_thread = _GenQuoteThread(self.state, qtype, items, rds, self)
        self._gen_qt_thread.progress.connect(self._log)
        self._gen_qt_thread.done.connect(self._on_gen_quote_done)
        self._gen_qt_thread.start()

    def _on_gen_quote_done(self, result) -> None:
        self._set_gen_buttons(True)
        if isinstance(result, Exception):
            logger.error("견적서 생성 실패", exc_info=result)
            tb = "".join(traceback.format_exception(type(result), result, result.__traceback__))
            user_msg, hint = _friendly_error_msg(result)
            _ScrollableErrorDialog(self, tb, user_msg=user_msg, hint=hint).exec()
            return
        for i, (rd, path) in enumerate(result):
            if path:
                self._add_done(path)
                self.state.last_output_dir = os.path.dirname(path)
                if i < len(self._pending_req_indices):
                    idx = self._pending_req_indices[i]
                    self._done_req_indices.add(idx)
                    self._mark_req_done(idx)
        ok = sum(1 for _, p in result if p)
        self._log(f"견적서 생성 완료: {ok}/{len(result)}건")
        QMessageBox.information(self, "완료", f"견적서 {ok}건 생성 완료")
        self._open_quote_list_window()

    def _generate_cover(self) -> None:
        if not self.state.template_path:
            QMessageBox.warning(self, "오류", "STEP1 통합양식을 먼저 LOAD 하세요."); return
        if self._gen_cv_thread and self._gen_cv_thread.isRunning():
            QMessageBox.warning(self, "진행 중", "이미 갑지 생성 작업이 진행 중입니다."); return
        start = self.state.last_output_dir or exe_dir()
        paths, _ = QFileDialog.getOpenFileNames(self, "갑지 생성 대상 엑셀파일 (다중)", start, "Excel Files (*.xlsx)")
        if not paths: return
        folder = os.path.commonpath(paths)
        if os.path.isfile(folder): folder = os.path.dirname(folder)
        self._log(f"갑지 생성 시작 ({len(paths)}건) …")
        self._set_gen_buttons(False)
        self._gen_cv_thread = _GenCoverThread(
            self.state.template_path, folder, paths, self.state.investor_name,
            self.state.warranty_years, self)
        self._gen_cv_thread.progress.connect(self._log)
        self._gen_cv_thread.done.connect(self._on_gen_cover_done)

        self._cover_progress = QProgressDialog("갑지 생성 중…", None, 0, len(paths), self)
        self._cover_progress.setWindowTitle("갑지 생성")
        self._cover_progress.setWindowModality(Qt.WindowModal)
        self._cover_progress.setMinimumDuration(500)
        self._cover_progress.setValue(0)
        self._gen_cv_thread.step.connect(
            lambda d, t: self._cover_progress.setValue(d) if self._cover_progress else None
        )

        self._gen_cv_thread.start()

    def _on_gen_cover_done(self, result) -> None:
        if self._cover_progress:
            self._cover_progress.close()
            self._cover_progress = None
        self._set_gen_buttons(True)
        if isinstance(result, Exception):
            logger.error("갑지 생성 실패", exc_info=result)
            tb = "".join(traceback.format_exception(type(result), result, result.__traceback__))
            user_msg, hint = _friendly_error_msg(result)
            _ScrollableErrorDialog(self, tb, user_msg=user_msg, hint=hint).exec()
            return
        out = result
        self.state.last_output_dir = os.path.dirname(out)
        self._add_done_cover(out)
        self._log(f"갑지 생성 완료: {out}")
        QMessageBox.information(self, "완료", f"갑지 생성 완료\n{os.path.basename(out)}")

    def _pick_rd_index(self) -> int:
        """클릭(선택)된 행 → 체크된 첫 번째 행 → 0 순으로 의뢰행 인덱스 반환."""
        cur = self.req_table.currentRow()
        if 0 <= cur < len(self.state.request_rows):
            return cur
        checked = self._checked_req_rows()
        if checked and checked[0] < len(self.state.request_rows):
            return checked[0]
        return 0 if self.state.request_rows else -1

    def _pick_rd(self) -> Dict[str, Any]:
        idx = self._pick_rd_index()
        if idx >= 0 and idx < len(self.state.request_rows):
            return self.state.request_rows[idx]
        return {}

    def _mark_req_done(self, row: int) -> None:
        """견적 완료된 의뢰행을 연두색으로 하이라이트."""
        color = QBrush(QColor(160, 220, 160))
        for col in range(1, self.req_table.columnCount()):
            it = self.req_table.item(row, col)
            if it:
                it.setBackground(color)

    def _add_done(self, path: str) -> None:
        self._list_all_files.insert(0, path)
        if self._quote_list_window:
            self._quote_list_window.add_item(path)

    def _add_done_cover(self, path: str) -> None:
        self._list_all_files.insert(0, path)
        if self._quote_list_window:
            self._quote_list_window.add_item(path)

    @staticmethod
    def _parse_quote_label(full_path: str) -> Tuple[Optional[str], str]:
        """파일 경로에서 (설비호기, 표시레이블) 추출.

        반환값:
        - (None, '')    : 패턴 불일치 → 목록에서 제외
        - ('', label)   : 갑지 파일  → label 끝이 '_갑지', 붉은색으로 표시
        - (tool, label) : 일반 견적서 → 라인_공정_설비호기 형식
        """
        import re as _re
        fn = os.path.basename(full_path)
        if not fn.endswith('.xlsx'):
            return None, ""
        stem   = fn[:-5]
        folder = os.path.basename(os.path.dirname(full_path))

        # ── 갑지: tool='' 로 반환해 붉은색 음영 처리 신호 ───────────
        if _re.search(r'_갑지(?:_\d+)?$', stem):
            m_dom = _re.match(r'^\d{6}_LOT베큠_(.+)$', folder)
            if m_dom:
                return "", f"{m_dom.group(1)}_갑지"
            m_cn = _re.match(r'^\d{6}_중국_SCS_(.+)$', folder)
            if m_cn:
                return "", f"중국_{m_cn.group(1)}_갑지"
            return "", "갑지"

        # ── 미국 Quotation: ..._{tool}_Quotation ─────────────────────
        m_us = _re.search(r'_([^_]+)_Quotation$', stem)
        if m_us:
            tool = m_us.group(1)
            mf   = _re.match(r'^\d{6}_미국_(.+)$', folder)
            return tool, f"미국_{mf.group(1) if mf else '미국'}_{tool}"

        # ── 국내·중국: 마지막 _구분자 이후 = 설비호기 ───────────────
        parts = stem.split('_')
        tool  = parts[-1] if parts else ""
        if not tool:
            return None, ""

        m_dom = _re.match(r'^\d{6}_LOT베큠_(.+)$', folder)
        if m_dom:
            lp_parts = m_dom.group(1).split('_')
            lineproc = '_'.join(lp_parts[:-1]) if len(lp_parts) >= 2 else m_dom.group(1)
            return tool, f"{lineproc}_{tool}"

        m_cn = _re.match(r'^\d{6}_중국_SCS_(.+)$', folder)
        if m_cn:
            return tool, f"중국_{m_cn.group(1)}_{tool}"

        return tool, tool

    @staticmethod
    def _parse_quote_entry(full_path: str) -> Optional[Tuple[str, str, str]]:
        """_parse_quote_label 을 호출하고 (tool, label, full_path) 튜플로 감싼다.
        패턴 불일치(tool is None)이면 None 반환."""
        tool, label = QuoteBuilderPage._parse_quote_label(full_path)
        if tool is None:
            return None
        return tool, label, full_path

    def _open_quote_list_window(self) -> None:
        """견적서 LIST 독립 창을 열거나 앞으로 가져온다."""
        if self._quote_list_window is None:
            self._quote_list_window = _QuoteListWindow(self)
            self._quote_list_window.refresh_session(self._list_all_files)
        self._quote_list_window.show_and_raise()

    def _log(self, text: str) -> None:
        self.log_view.append(text); logger.debug(text)
