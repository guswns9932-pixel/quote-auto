"""
page_rack.py
============
RACK 구매요청서 페이지 (현재 비활성화된 기능).
"""
from __future__ import annotations

import logging
import os
import re
import shutil
import traceback
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from PySide6.QtCore import Qt, QEvent, Signal
from PySide6.QtGui import QColor, QKeySequence
from PySide6.QtWidgets import (
    QAbstractItemView, QApplication, QCheckBox, QComboBox, QDialog,
    QDialogButtonBox, QFormLayout, QFrame, QGridLayout, QHBoxLayout,
    QHeaderView, QInputDialog, QLabel, QLineEdit,
    QListWidget, QListWidgetItem, QMessageBox,
    QPushButton, QSizePolicy, QSplitter, QTabWidget,
    QTableWidget, QTableWidgetItem,
    QVBoxLayout, QWidget,
)

from core import (
    ensure_dir, exe_dir, fmt_qty, parse_invest_info, s,
    safe_filename, to_float, unique_path,
)
from widgets import (
    bold_label, centered_checkbox,
    get_checkbox_from_cell, info_label, tint_button,
)
from page_common import _BgWorker

logger = logging.getLogger("QuoteApp")


RACK_REQUEST_LEFT_LABELS = [
    "접수 일자", "납품 요청", "PR NO.", "담당자", "공 정", "세부 공정", "라인", "설 비",
    "설비MODEL", "설비호기", "PUMP MODEL", "수량(CH)", "5D", "FSC", "RACK CH",
    "GATE V/V TYPE", "METAL BELLOWS", "HOT-N2 TYPE", "IMS 중계기", "DA / LEP", "INVERTER",
    "ANGLE Valve", "Option", "비 고", "간섭여부", "Interface Cable", "Dual Cable",
    "통신모듈 Cable", "3단 모니터링 Cable",
]
RACK_REQUEST_REF_NOTES = [
    "작성일", "Main설비반입일 D-28", "견적의뢰DATA", "수기입력", "견적의뢰DATA", "견적의뢰DATA",
    "견적의뢰DATA", "견적의뢰DATA", "견적의뢰DATA", "견적의뢰DATA", "견적의뢰DATA", "견적의뢰DATA",
    "견적의뢰DATA", "5D 매칭",
] + ["공정+설비+5D 매칭"] * 15
RACK_REQUEST_MANAGERS = ["김종균", "박현준", "이기웅", "정상철", "황명선"]

class _NoScrollComboBox(QComboBox):
    """마우스 스크롤로 값이 바뀌지 않는 QComboBox."""
    def wheelEvent(self, e):
        e.ignore()


class _GroupwareLoginDialog(QDialog):
    """그룹웨어 로그인 계정/비밀번호 입력 다이얼로그."""

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("그룹웨어 로그인")
        self.setFixedWidth(340)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        layout = QFormLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(24, 20, 24, 20)

        self.id_edit = QLineEdit()
        self.id_edit.setPlaceholderText("계정")
        self.id_edit.setMinimumHeight(36)

        self.pw_edit = QLineEdit()
        self.pw_edit.setPlaceholderText("비밀번호")
        self.pw_edit.setEchoMode(QLineEdit.Password)
        self.pw_edit.setMinimumHeight(36)

        layout.addRow("계정", self.id_edit)
        layout.addRow("비밀번호", self.pw_edit)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText("로그인")
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        layout.addRow(btns)

        self.pw_edit.returnPressed.connect(btns.button(QDialogButtonBox.Ok).click)

    def _on_accept(self) -> None:
        if not self.id_edit.text().strip() or not self.pw_edit.text():
            QMessageBox.warning(self, "입력 오류", "계정과 비밀번호를 모두 입력하세요.")
            return
        self.accept()

    def credentials(self) -> tuple[str, str]:
        return self.id_edit.text().strip(), self.pw_edit.text()


class RackPurchaseRequestPage(QWidget):
    """이미지 양식 기반 RACK 구매요청서 작성/엑셀 생성 위젯."""

    COL_LABEL = 0
    COL_ITEM = 1
    COL_QTY = 2
    COL_REF_ITEM = 3
    COL_REF_QTY = 4
    COL_REF_APPLY = 5
    COL_REQUEST = 6

    # (rack_0idx_autocomplete, rack_1idx_item_col, rack_1idx_qty_col)
    # row 2 = 인보이스 필요여부 (자동계산), row 3~ = 기존 데이터 행
    _ROW_RACK_MAP: Dict[int, Tuple] = {
        3:  (2,  3,  None),
        4:  (5,  6,  None),
        5:  (8,  9,  None),
        6:  (10, 11, None),
        7:  (22, 23, None),
        8:  (23, 24, None),
        9:  (24, 25, None),
        10: (25, 26, None),
        11: (26, 27, None),
        12: (None,28, None),
        13: (28, 29, None),
        14: (29, 30, None),
        15: (30, 31, None),
        16: (18, 19, None),
        17: (32, 33, 34),
        18: (47, 48, 49),
        19: (81, 82, 83),
        20: (52, 53, 54),
        21: (56, 57, 58),
        22: (60, 61, 62),
        23: (70, 71, 72),
        24: (None, None, 92),
        25: (94, 95, None),
        26: (96, 97, None),
        27: (None, 98, None),
        28: (99, 100, 101),
        29: (101, 102, 103),
        30: (103, 104, 105),
        31: (105, 106, 107),
    }

    def __init__(self) -> None:
        super().__init__()
        self.rack_template_path: Optional[str] = None
        self.rack_template_sheets: Dict[str, List[List[Any]]] = {}
        self.request_rows: List[Dict[str, Any]] = []
        self._last_generated_folder: Optional[str] = None
        self._matched_rack_row: Optional[List[Any]] = None
        self._gw_username: str = ""
        self._gw_password: str = ""
        self._build_ui()
        self._populate_table()

    def _build_ui(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(14, 14, 14, 14)
        outer.setSpacing(10)

        title_row = QHBoxLayout()
        title = bold_label("RACK구매요청서 작성", size=15)
        guide = QLabel("이미지 양식과 동일한 항목 틀에서 값을 입력한 뒤 Excel 요청서를 생성합니다.")
        guide.setStyleSheet("color:#555;")
        title_row.addWidget(title)
        title_row.addWidget(guide, 1)
        outer.addLayout(title_row)

        btn_row = QHBoxLayout()
        self.btn_load_template = QPushButton("통합양식 불러오기")
        self.btn_load_quote = QPushButton("견적의뢰DATA 불러오기")
        self.btn_generate = QPushButton("요청서 생성")
        self.btn_generate_overseas = QPushButton("국외요청서 생성")
        self.btn_generate_approval = QPushButton("결재상신용 생성")
        self.btn_approval_submit   = QPushButton("결재상신")
        self.btn_upload_history = QPushButton("발주이력 업로드")
        for btn, color in [
            (self.btn_load_template,    "#C8E6C9"),
            (self.btn_load_quote,       "#BBDEFB"),
            (self.btn_generate,         "#FFF176"),
            (self.btn_generate_overseas,"#B3E5FC"),
            (self.btn_generate_approval,"#FFE0B2"),
            (self.btn_approval_submit,  "#F8BBD0"),
            (self.btn_upload_history,   "#E1BEE7"),
        ]:
            btn.setMinimumHeight(42)
            f = btn.font(); f.setPointSize(11); f.setBold(True); btn.setFont(f)
            tint_button(btn, color)
            btn_row.addWidget(btn)
        btn_row.addStretch(1)
        outer.addLayout(btn_row)

        self.lbl_template_status = info_label("통합양식: 미로드")
        outer.addWidget(self.lbl_template_status)

        # ── 좌우 분할: 좌=의뢰DATA+테이블 / 우=생성된 요청서 목록(420px) ──
        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)

        # 좌측: 기존 의뢰파일DATA 패널 + 메인 테이블
        left_w = QWidget()
        left_v = QVBoxLayout(left_w)
        left_v.setContentsMargins(0, 0, 0, 0)
        left_v.setSpacing(6)
        left_v.addWidget(self._build_request_data_panel())

        self.table = QTableWidget(32, 7)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setVisible(False)
        self.table.setAlternatingRowColors(False)
        self.table.setWordWrap(False)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setShowGrid(True)
        self.table.setSpan(0, self.COL_ITEM, 1, 2)
        self.table.setSpan(0, self.COL_REF_ITEM, 1, 2)
        self.table.setColumnWidth(self.COL_LABEL, 110)
        self.table.setColumnWidth(self.COL_ITEM, 300)
        self.table.setColumnWidth(self.COL_QTY, 70)
        self.table.setColumnWidth(self.COL_REF_ITEM, 300)
        self.table.setColumnWidth(self.COL_REF_QTY, 70)
        self.table.setColumnWidth(self.COL_REF_APPLY, 60)
        self.table.setColumnWidth(self.COL_REQUEST, 150)
        self.table.horizontalHeader().setSectionResizeMode(self.COL_ITEM, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(self.COL_REF_ITEM, QHeaderView.Stretch)
        for r in range(self.table.rowCount()):
            self.table.setRowHeight(r, 24)
        self.table.setRowHeight(0, 30)
        left_v.addWidget(self.table, 1)
        splitter.addWidget(left_w)

        # 우측: 생성된 요청서 목록 (420px 고정)
        right_w = QWidget()
        right_w.setFixedWidth(420)
        right_v = QVBoxLayout(right_w)
        right_v.setContentsMargins(6, 0, 0, 0)
        right_v.setSpacing(4)

        # ── 그룹웨어 로그인 패널 ──────────────────────────────────────────────
        from PySide6.QtWidgets import QGroupBox, QFormLayout
        login_box = QGroupBox("그룹웨어 로그인")
        login_box.setStyleSheet("QGroupBox { font-weight: bold; }")
        login_box_v = QVBoxLayout(login_box)
        login_box_v.setSpacing(4)

        # 미로그인 상태 위젯
        self._login_form_w = QWidget()
        _lf = QFormLayout(self._login_form_w)
        _lf.setContentsMargins(0, 0, 0, 0)
        _lf.setSpacing(4)
        self._ed_gw_id = QLineEdit()
        self._ed_gw_id.setPlaceholderText("아이디")
        self._ed_gw_pw = QLineEdit()
        self._ed_gw_pw.setPlaceholderText("비밀번호")
        self._ed_gw_pw.setEchoMode(QLineEdit.Password)
        self._btn_gw_login = QPushButton("로그인")
        self._btn_gw_login.setFixedHeight(30)
        tint_button(self._btn_gw_login, "#C8E6C9")
        _lf.addRow("ID", self._ed_gw_id)
        _lf.addRow("PW", self._ed_gw_pw)
        _lf.addRow("", self._btn_gw_login)
        self._ed_gw_id.returnPressed.connect(self._btn_gw_login.click)
        self._ed_gw_pw.returnPressed.connect(self._btn_gw_login.click)
        login_box_v.addWidget(self._login_form_w)

        # 로그인 완료 상태 위젯 (숨김)
        self._login_status_w = QWidget()
        _ls = QHBoxLayout(self._login_status_w)
        _ls.setContentsMargins(0, 0, 0, 0)
        self._lbl_gw_user = QLabel("로그인됨: -")
        self._lbl_gw_user.setStyleSheet("color:#2e7d32; font-weight:bold;")
        self._btn_gw_logout = QPushButton("로그아웃")
        self._btn_gw_logout.setFixedHeight(26)
        tint_button(self._btn_gw_logout, "#FFCDD2")
        _ls.addWidget(self._lbl_gw_user, 1)
        _ls.addWidget(self._btn_gw_logout)
        login_box_v.addWidget(self._login_status_w)
        self._login_status_w.hide()

        right_v.addWidget(login_box)

        req_title_row = QHBoxLayout()
        req_title_lbl = bold_label("생성된 요청서", size=11)
        self.btn_refresh_req_list = QPushButton("새로고침")
        self.btn_refresh_req_list.setFixedHeight(24)
        self.btn_refresh_req_list.clicked.connect(self._refresh_req_list)
        req_title_row.addWidget(req_title_lbl)
        req_title_row.addStretch(1)
        req_title_row.addWidget(self.btn_refresh_req_list)
        right_v.addLayout(req_title_row)

        self.req_list = QListWidget()
        self.req_list.setToolTip("더블클릭하면 파일을 열 수 있습니다.")
        self.req_list.itemDoubleClicked.connect(self._open_req_file)
        right_v.addWidget(self.req_list, 1)
        splitter.addWidget(right_w)

        outer.addWidget(splitter, 1)

        self.btn_load_template.clicked.connect(self._load_rack_template)
        self.btn_generate.clicked.connect(self._generate_request)
        self.btn_generate_overseas.clicked.connect(self._generate_overseas_request)
        self.btn_load_quote.clicked.connect(self._load_quote_data)
        self.btn_generate_approval.clicked.connect(self._generate_approval_doc)
        self.btn_approval_submit.clicked.connect(self._do_approval_submit)
        self.btn_upload_history.clicked.connect(self._upload_order_history)
        self._btn_gw_login.clicked.connect(self._do_gw_login)
        self._btn_gw_logout.clicked.connect(self._do_gw_logout)

        # UI 렌더링 완료 후 파일 스캔 (블로킹 방지)
        QTimer.singleShot(0, self._refresh_req_list)

    def _build_request_data_panel(self) -> QFrame:
        """불러온 견적의뢰DATA를 표시하는 붉은 박스 영역."""
        frame = QFrame()
        frame.setFrameShape(QFrame.Box)
        frame.setLineWidth(2)
        frame.setStyleSheet("QFrame { border: 2px solid #D32F2F; }")
        frame.setFixedHeight(150)
        v = QVBoxLayout(frame)
        v.setContentsMargins(8, 6, 8, 8)
        v.setSpacing(6)
        title = bold_label("의뢰파일DATA", size=11)
        title.setStyleSheet("color:#B71C1C; border:none;")
        v.addWidget(title)
        self.request_data_table = QTableWidget(0, 9)
        self.request_data_table.setHorizontalHeaderLabels([
            "선택", "PR NO.", "항번", "투자정보", "수량(CH)", "공정", "5D", "FSC", "설비호기"
        ])
        self.request_data_table.verticalHeader().setVisible(False)
        self.request_data_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.request_data_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.request_data_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.request_data_table.setStyleSheet("QTableWidget { border: 1px solid #D32F2F; }")
        self.request_data_table.cellClicked.connect(self._on_request_row_clicked)
        self.request_data_table.cellDoubleClicked.connect(lambda row, _col: self._apply_request_row(row))
        h = self.request_data_table.horizontalHeader()
        for col in range(self.request_data_table.columnCount()):
            h.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        v.addWidget(self.request_data_table, 1)
        return frame

    def _populate_table(self) -> None:
        self.table.setUpdatesEnabled(False)
        try:
            self._populate_table_inner()
        finally:
            self.table.setUpdatesEnabled(True)

    def _populate_table_inner(self) -> None:
        self._set_item(0, self.COL_LABEL, "", "#DDE2E6", bold=True, editable=False)
        self._set_item(0, self.COL_ITEM, "RACK구매요청서", "#C7EAF4", bold=True, editable=False)
        self._set_item(0, self.COL_REF_ITEM, "Ref.", "#DDE2E6", bold=True, editable=False)
        self._set_item(0, self.COL_REF_APPLY, "", "#00B050", bold=True, editable=False)
        self.btn_apply_ref = QPushButton("적용")
        self.btn_apply_ref.setStyleSheet(
            "QPushButton {"
            "  background-color: #00E676; color: #000; border: 2px solid #00C853;"
            "  border-radius: 4px; font-weight: bold; padding: 2px 8px;"
            "}"
            "QPushButton:hover { background-color: #69F0AE; border-color: #00E676; }"
            "QPushButton:pressed { background-color: #00C853; }"
        )
        self.btn_apply_ref.clicked.connect(self._apply_ref_values)
        self.table.setCellWidget(0, self.COL_REF_APPLY, self.btn_apply_ref)
        self._set_item(0, self.COL_REQUEST, "비고", "#FFFFFF", bold=True, editable=False)

        for col, text, color in [
            (self.COL_ITEM, "항목", "#C7EAF4"), (self.COL_QTY, "수량", "#C7EAF4"),
            (self.COL_REF_ITEM, "항목", "#DDE2E6"), (self.COL_REF_QTY, "수량", "#DDE2E6"),
        ]:
            self._set_item(1, col, text, color, bold=True, editable=False)
        self._set_item(1, self.COL_LABEL, "", "#DDE2E6", editable=False)
        self._set_item(1, self.COL_REF_APPLY, "", "#DDE2E6", editable=False)
        self._set_item(1, self.COL_REQUEST, "", "#FFFFFF", editable=False)

        # Row 2: 인보이스 필요여부 (투자정보 기반 자동계산, 접수일자 위쪽)
        self._set_item(2, self.COL_LABEL, "인보이스 필요여부", "#EAC3E7", bold=True, editable=False)
        self._set_item(2, self.COL_ITEM, "", "#C7EAF4", editable=False)
        self._set_item(2, self.COL_QTY, "-", "#C7EAF4", editable=False)
        self._set_item(2, self.COL_REF_ITEM, "", "#DDE2E6", editable=False)
        self._set_item(2, self.COL_REF_QTY, "-", "#DDE2E6", editable=False)
        self._set_item(2, self.COL_REF_APPLY, "-", "#DDE2E6", editable=False)
        self._set_item(2, self.COL_REQUEST, "KIT투자 시 인보이스필요", "#FFFFFF", editable=False)

        # 기존 데이터 행: start=3 (row 2가 새 인보이스 행이므로)
        for i, label in enumerate(RACK_REQUEST_LEFT_LABELS, start=3):
            left_color = "#EAC3E7" if i <= 16 else "#F4DDCF"
            self._set_item(i, self.COL_LABEL, label, left_color, bold=True, editable=False)
            self._set_item(i, self.COL_ITEM, str(i - 2), "#C7EAF4")
            self._set_item(i, self.COL_QTY, str(i + 27), "#C7EAF4")
            self._set_item(i, self.COL_REF_ITEM, str(i + 56), "#DDE2E6")
            self._set_item(i, self.COL_REF_QTY, str(i + 85), "#DDE2E6")
            if i <= 16:
                self._set_item(i, self.COL_REF_APPLY, "-", "#DDE2E6", editable=False)
            else:
                self._set_item(i, self.COL_REF_APPLY, "", "#DDE2E6", editable=False)
                host = centered_checkbox(lambda _state, _row=i: self._on_ref_apply_changed(_row))
                cb = get_checkbox_from_cell(host)
                if cb:
                    cb.setChecked(True)
                self.table.setCellWidget(i, self.COL_REF_APPLY, host)
            self._set_item(i, self.COL_REQUEST, RACK_REQUEST_REF_NOTES[i - 3], "#FFFFFF", editable=False)

        # COL_ITEM 자동완성 콤보박스 (row 6=담당자 제외, row 27=간섭여부 별도 처리)
        for _row in range(3, 32):
            if _row == 6:
                continue  # 담당자: 뒤에서 manager combobox로 덮어씀
            if _row == 27:
                _cb27 = _NoScrollComboBox()
                _cb27.setEditable(True)
                _cb27.lineEdit().setAlignment(Qt.AlignCenter)
                _cb27.lineEdit().setReadOnly(True)
                _cb27.addItems(["", "간섭없음", "좌측간섭", "우측간섭", "좌/우간섭"])
                _cb27.setStyleSheet("QComboBox { background-color: #C7EAF4; }")
                self.table.setCellWidget(_row, self.COL_ITEM, _cb27)
                continue
            _cb = _NoScrollComboBox()
            _cb.setEditable(True)
            _cb.setInsertPolicy(QComboBox.NoInsert)
            _cb.lineEdit().setAlignment(Qt.AlignCenter)
            _cb.setStyleSheet("QComboBox { background-color: #C7EAF4; }")
            # FSC(16): 5D 자동 매칭 전용 — 수기 입력 불가
            if _row == 16:
                _cb.lineEdit().setReadOnly(True)
                _cb.setStyleSheet("QComboBox { background-color: #DDE8F0; }")
            else:
                _cmpl = _cb.completer()
                if _cmpl:
                    _cmpl.setFilterMode(Qt.MatchContains)
                    _cmpl.setCompletionMode(QCompleter.PopupCompletion)
            self.table.setCellWidget(_row, self.COL_ITEM, _cb)
            # 공정(7), 설비(10), 5D(15) 변경 시 실시간 Ref. 매칭
            if _row in (7, 10, 15):
                _cb.currentTextChanged.connect(self._on_key_field_changed)
            # 5D(15) 변경 시 FSC 자동 매칭
            if _row == 15:
                _cb.currentTextChanged.connect(self._auto_match_fsc)

        manager_row, manager_col = self._slot_pos(4)
        manager = _NoScrollComboBox()
        manager.setEditable(True)
        manager.lineEdit().setAlignment(Qt.AlignCenter)
        manager.addItems(RACK_REQUEST_MANAGERS)
        manager.setCurrentText(RACK_REQUEST_MANAGERS[0])
        manager.setStyleSheet("QComboBox { background-color: #C7EAF4; }")
        self.table.setCellWidget(manager_row, manager_col, manager)

        # RACK발주 매칭 REF 슬롯 초기화 (rows 17-31)
        for row in range(17, 32):
            item = self.table.item(row, self.COL_REF_ITEM)
            if item:
                item.setText("")
            item = self.table.item(row, self.COL_REF_QTY)
            if item:
                item.setText("")

        # COL_QTY 고정 "-" (rows 2-16, 인보이스 포함)
        for row in range(2, 17):
            self._set_item(row, self.COL_QTY, "-", "#C7EAF4", editable=False)

        # COL_REF_QTY 고정 "-" (rows 2-16)
        for row in range(2, 17):
            self._set_item(row, self.COL_REF_QTY, "-", "#DDE2E6", editable=False)

        # COL_REF_ITEM 초기값 "" (rows 3-16, 접수일자~RACK CH)
        for row in range(3, 17):
            item = self.table.item(row, self.COL_REF_ITEM)
            if item:
                item.setText("")

        # slot 80 (row 24, COL_REF_ITEM): 고정 "-"
        self._set_item(24, self.COL_REF_ITEM, "-", "#DDE2E6", editable=False)

        # slots 110-112 (rows 25-27, COL_REF_QTY): 고정 "-"
        for row in (25, 26, 27):
            self._set_item(row, self.COL_REF_QTY, "-", "#DDE2E6", editable=False)

        # slot 83 (row 27, COL_REF_ITEM): 간섭여부 드롭다운 (가운데 정렬)
        combo_83 = _NoScrollComboBox()
        combo_83.setEditable(True)
        combo_83.lineEdit().setAlignment(Qt.AlignCenter)
        combo_83.lineEdit().setReadOnly(True)
        combo_83.addItems(["", "간섭없음", "좌측간섭", "우측간섭", "좌/우간섭"])
        combo_83.setStyleSheet("QComboBox { background-color: #DDE2E6; }")
        self.table.setCellWidget(27, self.COL_REF_ITEM, combo_83)

        self._set_default_values()
        # _populate_table_inner 끝 — setUpdatesEnabled(True)는 _populate_table에서 처리

    def _set_default_values(self) -> None:
        """견적의뢰DATA와 무관한 기본값 설정 (통합양식 로드 후에도 재적용)."""
        inv_item = self.table.item(2, self.COL_ITEM)
        if inv_item:
            inv_item.setText("투자미확인")
        self._set_slot_value(1, datetime.now().strftime("%Y-%m-%d"))
        self._set_slot_value(2, "입력필요")
        self._set_slot_value(3, "구두발주")
        # 담당자/공정/세부공정/라인/설비/설비MODEL/설비호기/PUMP MODEL/수량(CH)/5D: 공란
        for slot in (4, 5, 6, 7, 8, 9, 10, 11, 12, 13):
            self._set_slot_value(slot, "")
        # FSC 공란 (5D 입력 시 자동 채워짐)
        fsc_widget = self.table.cellWidget(16, self.COL_ITEM)
        if isinstance(fsc_widget, QComboBox):
            fsc_widget.setCurrentText("")

    def _set_item(self, row: int, col: int, text: str, color: str, *, bold: bool = False,
                  editable: bool = True, size: Optional[int] = None) -> None:
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        item.setBackground(QBrush(QColor(color)))
        flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable
        if editable:
            flags |= Qt.ItemIsEditable
        item.setFlags(flags)
        font = item.font()
        font.setBold(bold)
        if size:
            font.setPointSize(size)
        item.setFont(font)
        self.table.setItem(row, col, item)

    def reset_page(self) -> None:
        """좌측 하단 초기화 버튼에서 호출되는 현재 페이지 초기화 진입점."""
        self.rack_template_path = None
        self.rack_template_sheets = {}
        self.request_rows = []
        self.lbl_template_status.setText("통합양식: 미로드")
        self.request_data_table.setRowCount(0)
        self._populate_table()

    def _is_ref_apply_checked(self, row: int) -> bool:
        cb = get_checkbox_from_cell(self.table.cellWidget(row, self.COL_REF_APPLY))
        return bool(cb and cb.isChecked())

    def _on_ref_apply_changed(self, row: int) -> None:
        """COL_REF_APPLY 체크박스 토글 시 해당 행의 Ref 값을 항목/수량에 즉시 반영."""
        if not self._is_ref_apply_checked(row):
            return
        ref_item = self._text(row, self.COL_REF_ITEM)
        ref_qty  = self._text(row, self.COL_REF_QTY)
        w = self.table.cellWidget(row, self.COL_ITEM)
        if isinstance(w, QComboBox):
            w.setCurrentText(ref_item)
        else:
            it = self.table.item(row, self.COL_ITEM)
            if it:
                it.setText(ref_item)
        it_qty = self.table.item(row, self.COL_QTY)
        if it_qty:
            it_qty.setText(ref_qty)

    def _clear_lower_rows(self) -> None:
        """rows 17-31(RACK CH~3단 모니터링 Cable) 항목/수량 값 공란으로 초기화."""
        for row in range(17, 32):
            w = self.table.cellWidget(row, self.COL_ITEM)
            if isinstance(w, QComboBox):
                w.setCurrentText("")
            else:
                it = self.table.item(row, self.COL_ITEM)
                if it:
                    it.setText("")
            it_qty = self.table.item(row, self.COL_QTY)
            if it_qty:
                it_qty.setText("")

    def _apply_all_checked_refs(self) -> None:
        """rows 17-31 중 체크된 행의 Ref 항목/수량을 COL_ITEM/COL_QTY에 즉시 반영."""
        for row in range(17, 32):
            self._on_ref_apply_changed(row)

    def _slot_pos(self, slot: int) -> Tuple[int, int]:
        """이미지 양식의 1~58 입력칸 번호를 테이블 row/col로 변환.
        row 2 = 인보이스 필요여부(자동계산), row 3~ = slot 1~ 대응."""
        if 1 <= slot <= 29:
            return slot + 2, self.COL_ITEM
        if 30 <= slot <= 58:
            return slot - 27, self.COL_QTY
        raise ValueError(f"지원하지 않는 RACK 요청서 칸 번호: {slot}")

    def _set_slot_value(self, slot: int, value: Any) -> None:
        row, col = self._slot_pos(slot)
        widget = self.table.cellWidget(row, col)
        if isinstance(widget, QComboBox):
            widget.setCurrentText(s(value))
            return
        item = self.table.item(row, col)
        if item:
            item.setText(s(value))

    def _reset_slot_values(self, slots: List[int]) -> None:
        for slot in slots:
            self._set_slot_value(slot, "")

    def _date_yyyy_mm_dd(self, value: Any) -> str:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        text = s(value)
        if not text:
            return ""
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
            try:
                return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
        if "T" in text:
            text = text.split("T", 1)[0]
        if " " in text:
            text = text.split(" ", 1)[0]
        return text.replace("/", "-")

    def _split_line_process(self, value: Any) -> Tuple[str, str]:
        text = s(value)
        if "_" not in text:
            return text, ""
        line, process = text.split("_", 1)
        return line, process

    def _pump_model_from_g(self, value: Any) -> str:
        parts = [part.strip() for part in s(value).split(",") if part.strip()]
        for idx, part in enumerate(parts):
            if part.upper() == "LOT" and idx + 1 < len(parts):
                return parts[idx + 1]
        return parts[-1] if parts else ""

    def _rack_count_from_v(self, value: Any) -> str:
        match = re.search(r"(\d+(?:\.\d+)?)\s*RACK", s(value), re.IGNORECASE)
        return match.group(1) if match else ""

    def _populate_item_combos(self) -> None:
        """통합양식 로드 후 COL_ITEM 콤보박스에 RACK발주 열 고유값을 채운다."""
        rack_rows = self.rack_template_sheets.get("RACK발주", [])
        for row, info in self._ROW_RACK_MAP.items():
            col_0idx = info[0]
            if col_0idx is None:
                continue
            widget = self.table.cellWidget(row, self.COL_ITEM)
            if not isinstance(widget, QComboBox):
                continue
            seen: set = {widget.itemText(i) for i in range(widget.count())}
            widget.blockSignals(True)
            for r in rack_rows:
                if col_0idx < len(r):
                    val = s(r[col_0idx])
                    if val and val not in seen:
                        widget.addItem(val)
                        seen.add(val)
            widget.blockSignals(False)

    def _lookup_fsc_from_template(self, value: Any) -> str:
        key = s(value)
        if not key:
            return ""
        for row in self.rack_template_sheets.get("FSC All List", []):
            if len(row) >= 8 and key in s(row[7]):
                return s(row[1]) if len(row) >= 2 else ""
        return ""

    def _apply_ref_values(self) -> None:
        """Ref. 적용 체크 행의 Ref 항목/수량을 좌측 수기입력 영역으로 복사."""
        applied = 0
        for row in range(2, self.table.rowCount()):
            if not self._is_ref_apply_checked(row):
                continue
            ref_item = self._text(row, self.COL_REF_ITEM)
            ref_qty  = self._text(row, self.COL_REF_QTY)
            w = self.table.cellWidget(row, self.COL_ITEM)
            if isinstance(w, QComboBox):
                w.setCurrentText(ref_item)
            else:
                it = self.table.item(row, self.COL_ITEM)
                if it:
                    it.setText(ref_item)
            it_qty = self.table.item(row, self.COL_QTY)
            if it_qty:
                it_qty.setText(ref_qty)
            applied += 1
        QMessageBox.information(self, "Ref. 적용", f"Ref. 항목/수량 {applied}건을 수기입력 칸에 반영했습니다.")

    def _load_rack_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "통합양식 선택", "", "Excel Files (*.xlsx)")
        if not path:
            return

        def _do_load(p):
            from openpyxl import load_workbook
            required = ["RACK발주", "RACK발주양식", "FSC All List"]
            wb = load_workbook(p, data_only=True)
            missing = [name for name in required if name not in wb.sheetnames]
            if missing:
                wb.close()
                raise ValueError("필수 시트 없음:\n" + "\n".join(missing))
            loaded: Dict[str, List[List[Any]]] = {}
            counts: List[str] = []
            for name in required:
                ws = wb[name]
                rows: List[List[Any]] = []
                for row in ws.iter_rows(values_only=True):
                    values = list(row)
                    if any(s(cell) for cell in values):
                        rows.append(values)
                loaded[name] = rows
                counts.append(f"{name} {len(rows)}행")
            wb.close()
            return loaded, counts

        def _on_result(res):
            loaded, counts = res
            self.rack_template_path = path
            self.rack_template_sheets = loaded
            self._populate_item_combos()
            self._set_default_values()
            self.lbl_template_status.setText(f"통합양식: {os.path.basename(path)} ({', '.join(counts)})")
            QMessageBox.information(self, "완료", "통합양식 LOAD 완료\n" + "\n".join(counts))

        def _on_error(e):
            if "필수 시트 없음" in e:
                QMessageBox.warning(self, "통합양식 오류", e.split("\n\n")[0])
            else:
                QMessageBox.critical(self, "불러오기 오류", f"통합양식을 읽을 수 없습니다.\n{e}")

        _BgWorker.run_with_progress(self, "통합양식 읽는 중...", _do_load, path,
                                    on_result=_on_result, on_error=_on_error)

    def _load_quote_data(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "견적의뢰DATA 선택", "", "Excel Files (*.xlsx)")
        if not path:
            return
        import excel_io

        def _on_result(res):
            sheet_name, rows = res
            if not rows:
                QMessageBox.warning(self, "견적의뢰DATA", "읽을 데이터가 없습니다.")
                return
            self.request_rows = rows
            self._fill_request_data_table(rows)
            self._apply_request_row(0, show_message=False)
            QMessageBox.information(self, "완료",
                f"견적의뢰DATA {len(rows)}건을 붉은 박스 영역에 표시했습니다.\n시트: {sheet_name}")

        _BgWorker.run_with_progress(self, "견적의뢰DATA 읽는 중...",
                                    excel_io.parse_request_xlsx, path,
                                    on_result=_on_result)

    def _fill_request_data_table(self, rows: List[Dict[str, Any]]) -> None:
        self.request_data_table.setRowCount(0)
        for i, rd in enumerate(rows):
            self.request_data_table.insertRow(i)
            self.request_data_table.setCellWidget(
                i, 0, centered_checkbox(lambda _state, _r=i: self._on_request_target_changed(_r))
            )
            values = [
                s(rd.get("D")),
                s(rd.get("E")),
                parse_invest_info(rd.get("G")),
                fmt_qty(to_float(rd.get("H"))) if rd.get("H") is not None else "",
                s(rd.get("K")),
                s(rd.get("V")),
                s(rd.get("X")),
                s(rd.get("Z")),
            ]
            for col, value in enumerate(values, start=1):
                item = QTableWidgetItem(value)
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                item.setTextAlignment(Qt.AlignCenter)
                self.request_data_table.setItem(i, col, item)
            self.request_data_table.resizeRowToContents(i)
        self.request_data_table.resizeColumnsToContents()
        if rows:
            self._set_request_checked(0, True)
            self.request_data_table.selectRow(0)

    def _set_request_checked(self, row: int, checked: bool) -> None:
        cb = get_checkbox_from_cell(self.request_data_table.cellWidget(row, 0))
        if cb:
            cb.blockSignals(True)
            cb.setChecked(checked)
            cb.blockSignals(False)

    def _is_request_checked(self, row: int) -> bool:
        cb = get_checkbox_from_cell(self.request_data_table.cellWidget(row, 0))
        return bool(cb and cb.isChecked())

    def _checked_request_rows(self) -> List[int]:
        return [r for r in range(self.request_data_table.rowCount()) if self._is_request_checked(r)]

    def _highlight_request_row(self, row: int, color: str) -> None:
        """request_data_table 특정 행의 배경색 설정 (color="" 이면 초기화)."""
        for col in range(1, self.request_data_table.columnCount()):
            item = self.request_data_table.item(row, col)
            if item:
                if color:
                    item.setBackground(QBrush(QColor(color)))
                else:
                    item.setBackground(QBrush())

    def _mark_request_generated(self, row: int) -> None:
        """생성 완료 행을 붉은 파스텔 음영으로 표시."""
        self._highlight_request_row(row, "#FFCDD2")

    def _on_request_row_clicked(self, row: int, _col: int) -> None:
        """클릭한 행만 파스텔 파랑; 이전 클릭 행 색 해제 (생성완료·체크 행 제외)."""
        n = self.request_data_table.rowCount()
        for r in range(n):
            item1 = self.request_data_table.item(r, 1)
            if item1 is None:
                continue
            bg = item1.background().color().name()
            if r == row:
                if bg != "#ffcdd2":
                    self._highlight_request_row(r, "#E3F2FD")
            else:
                if bg == "#e3f2fd" and not self._is_request_checked(r):
                    self._highlight_request_row(r, "")

    def _on_request_target_changed(self, row: int) -> None:
        if self._is_request_checked(row):
            self._highlight_request_row(row, "#E3F2FD")
            self._apply_request_row(row, show_message=False)
        else:
            self._highlight_request_row(row, "")
            self._reset_slot_values(list(range(1, 15)))
            self._fill_rack_ref_slots(None)
            self._clear_lower_rows()

    def _apply_request_row(self, row: int, show_message: bool = True) -> None:
        if row < 0 or row >= len(self.request_rows):
            return
        rd = self.request_rows[row]
        line, process = self._split_line_process(rd.get("K"))
        defaults = {
            1: datetime.now().strftime("%Y-%m-%d"),
            2: self._date_yyyy_mm_dd(rd.get("AA")),
            3: rd.get("D"),
            5: process,
            6: rd.get("N"),
            7: line,
            8: rd.get("X"),
            9: rd.get("Y"),
            10: rd.get("Z"),
            11: self._pump_model_from_g(rd.get("G")),
            12: self._rack_count_from_v(rd.get("V")),
            13: rd.get("V"),
            14: self._lookup_fsc_from_template(rd.get("V")),
        }
        for slot, value in defaults.items():
            self._set_slot_value(slot, value)
        self._set_request_checked(row, True)
        self.request_data_table.selectRow(row)

        # RACK발주 시트에서 공정/설비/5D로 매칭 행 조회 후 REF 슬롯 채우기
        # 공정: K열을 "_"로 split한 process 부분 사용 (라인_공정 → 공정만)
        rack_process = process
        rack_vendor  = s(rd.get("X"))
        rack_5d      = s(rd.get("V"))
        rack_row = self._lookup_rack_row(rack_process, rack_vendor, rack_5d)
        self._fill_rack_ref_slots(rack_row)
        self._apply_all_checked_refs()

        invest_info = parse_invest_info(rd.get("G"))
        invoice_text = "인보이스 필요" if "KIT" in invest_info.upper() else "해당 없음"
        inv_item = self.table.item(2, self.COL_ITEM)
        if inv_item:
            inv_item.setText(invoice_text)

        if show_message:
            if rack_row is not None:
                msg = f"{row + 1}번째 의뢰파일DATA를 반영했습니다.\n[RACK발주 매칭 성공] 공정={rack_process} / 설비={rack_vendor} / 5D={rack_5d}"
            else:
                msg = (f"{row + 1}번째 의뢰파일DATA를 반영했습니다.\n"
                       f"[RACK발주 매칭 없음] 공정={rack_process} / 설비={rack_vendor} / 5D={rack_5d}\n"
                       f"통합양식 RACK발주 시트에서 일치하는 행이 없습니다.")
            QMessageBox.information(self, "의뢰파일DATA 적용", msg)

    def _lookup_rack_row(self, process: str, vendor: str, code_5d: str) -> Optional[List[Any]]:
        """RACK발주 시트에서 공정/설비/5D가 일치하는 가장 마지막 행을 반환 (역순 탐색).

        비교 시 앞뒤 공백 제거 + 대소문자 무시.
        W열(col 23) = 공정, Z열(col 26) = 설비, AE열(col 31) = 5D.
        """
        rows = self.rack_template_sheets.get("RACK발주", [])
        proc = s(process).upper().replace(" ", "")
        vend = s(vendor).upper().replace(" ", "")
        d5   = s(code_5d).upper().replace(" ", "")
        if not (proc and vend and d5):
            logger.debug("RACK발주 조회 스킵: 공정=%r 설비=%r 5D=%r", process, vendor, code_5d)
            return None
        logger.debug("RACK발주 조회: 공정=%r 설비=%r 5D=%r (총 %d행)", proc, vend, d5, len(rows))
        for row in reversed(rows):
            if len(row) <= 30:
                continue
            w  = s(row[22]).upper().replace(" ", "")   # W열
            z  = s(row[25]).upper().replace(" ", "")   # Z열
            ae = s(row[30]).upper().replace(" ", "")   # AE열
            if w == proc and z == vend and ae == d5:
                logger.debug("RACK발주 매칭 성공: W=%r Z=%r AE=%r", row[22], row[25], row[30])
                return row
        logger.debug("RACK발주 매칭 없음: 공정=%r 설비=%r 5D=%r", proc, vend, d5)
        return None

    def _fill_rack_ref_slots(self, rack_row: Optional[List[Any]]) -> None:
        """찾은 RACK발주 행의 값으로 REF 슬롯(59~87, 102~116)을 채운다."""
        def _get(col_0idx: int) -> str:
            if rack_row is None or col_0idx >= len(rack_row):
                return ""
            val = rack_row[col_0idx]
            if isinstance(val, (datetime, date)):
                return val.strftime("%Y-%m-%d")
            return s(val)

        # COL_REF_ITEM (rows 3-16, slots 59-72): RACK발주 기본 정보 열
        basic_ref_cols = {
            3:  2,   # C열  (접수 일자)
            4:  5,   # F열  (납품 요청)
            5:  8,   # I열  (PR NO.)
            6:  10,  # K열  (담당자)
            7:  22,  # W열  (공 정)
            8:  23,  # X열  (세부 공정)
            9:  24,  # Y열  (라인)
            10: 25,  # Z열  (설 비)
            11: 26,  # AA열 (설비MODEL)
            12: 27,  # AB열 (설비호기)
            13: 28,  # AC열 (PUMP MODEL)
            14: 29,  # AD열 (수량(CH))
            15: 30,  # AE열 (5D)
            16: 18,  # S열  (FSC)
        }
        for row, col_0idx in basic_ref_cols.items():
            item = self.table.item(row, self.COL_REF_ITEM)
            if item:
                item.setText(_get(col_0idx))

        # COL_REF_ITEM (rows 17-31, slots 73-87)
        ref_item_cols = {
            17: 32,   # AG(33)
            18: 47,   # AV(48)
            19: 81,   # CD(82)
            20: 52,   # BA(53)
            21: 56,   # BE(57)
            22: 60,   # BI(61)
            23: 70,   # BS(71)
            # 24: slot 80 → 고정 "-"
            25: 94,   # CQ(95)
            26: 96,   # CS(97)
            # 27: 간섭여부 → QComboBox (CT열)
            28: 99,   # CV(100)
            29: 101,  # CX(102)
            30: 103,  # CZ(104)
            31: 105,  # DB(106)
        }
        for row, col_0idx in ref_item_cols.items():
            item = self.table.item(row, self.COL_REF_ITEM)
            if item:
                item.setText(_get(col_0idx))

        # slot 80 (row 24): 항상 "-"
        item = self.table.item(24, self.COL_REF_ITEM)
        if item:
            item.setText("-")

        # row 27 (간섭여부) COL_REF_ITEM: QComboBox → CT열(0-indexed 97)
        combo_27 = self.table.cellWidget(27, self.COL_REF_ITEM)
        if isinstance(combo_27, QComboBox):
            combo_27.setCurrentText(_get(97))

        # COL_REF_QTY (rows 17-31, slots 102-116)
        ref_qty_cols = {
            17: 33,   # AH(34)
            18: 48,   # AW(49)
            19: 82,   # CE(83)
            20: 53,   # BB(54)
            21: 57,   # BF(58)
            22: 61,   # BJ(62)
            23: 71,   # BT(72)
            24: 91,   # CN(92)
            # 25, 26, 27: slots 110-112 → 고정 "-"
            28: 100,  # CW(101)
            29: 102,  # CY(103)
            30: 104,  # DA(105)
            31: 106,  # DC(107)
        }
        for row, col_0idx in ref_qty_cols.items():
            item = self.table.item(row, self.COL_REF_QTY)
            if item:
                item.setText(_get(col_0idx))

        # slots 110-112 (rows 25-27): 항상 "-"
        for row in (25, 26, 27):
            item = self.table.item(row, self.COL_REF_QTY)
            if item:
                item.setText("-")

    def _text(self, row: int, col: int) -> str:
        widget = self.table.cellWidget(row, col)
        if isinstance(widget, QComboBox):
            return widget.currentText().strip()
        cb = get_checkbox_from_cell(widget)
        if cb:
            return "☑" if cb.isChecked() else ""
        item = self.table.item(row, col)
        return item.text().strip() if item else ""

    def _on_key_field_changed(self, _text: str = "") -> None:
        """공정/설비/5D 수기 수정 시 실시간 Ref. 매칭."""
        if not self.rack_template_sheets.get("RACK발주"):
            return
        process = self._text(7,  self.COL_ITEM)
        vendor  = self._text(10, self.COL_ITEM)
        code_5d = self._text(15, self.COL_ITEM)
        rack_row = self._lookup_rack_row(process, vendor, code_5d)
        self._fill_rack_ref_slots(rack_row)
        self._apply_all_checked_refs()

    def _auto_match_fsc(self, _text: str = "") -> None:
        """5D 입력 시 FSC All List에서 FSC를 실시간 자동 매칭."""
        code_5d = self._text(15, self.COL_ITEM)
        fsc = self._lookup_fsc_from_template(code_5d)
        if not fsc:
            return
        w = self.table.cellWidget(16, self.COL_ITEM)
        if isinstance(w, QComboBox):
            w.blockSignals(True)
            w.setCurrentText(fsc)
            w.blockSignals(False)
        else:
            it = self.table.item(16, self.COL_ITEM)
            if it:
                it.setText(fsc)

    def _upload_order_history(self) -> None:
        """발주이력 업로드: 선택한 요청서 xlsx의 RACK발주양식 2행을 통합양식 RACK발주 시트에 추가."""
        if not self.rack_template_path:
            QMessageBox.warning(self, "오류", "통합양식을 먼저 불러오세요.")
            return
        if not self._is_file_writable(self.rack_template_path):
            QMessageBox.warning(self, "파일 잠금",
                "통합양식 파일이 다른 프로그램(Excel)에서 열려 있어 수정할 수 없습니다.\n"
                "Excel에서 통합양식 파일을 닫은 후 다시 실행해 주세요.")
            return

        paths, _ = QFileDialog.getOpenFileNames(
            self, "발주이력 요청서 선택", "", "Excel Files (*.xlsx)")
        if not paths:
            return

        tmpl_path = self.rack_template_path

        def _do_upload(tmpl, src_paths):
            from openpyxl import load_workbook as _lw
            tmpl_wb = _lw(tmpl)
            if "RACK발주" not in tmpl_wb.sheetnames:
                tmpl_wb.close()
                raise ValueError("통합양식에 'RACK발주' 시트가 없습니다.")
            ws = tmpl_wb["RACK발주"]

            first_empty = ws.max_row + 1
            for r in range(1, ws.max_row + 2):
                if ws.cell(r, 3).value is None or str(ws.cell(r, 3).value).strip() == "":
                    first_empty = r
                    break

            existing_rows: List[Dict] = []
            for r in range(1, first_empty):
                if ws.cell(r, 3).value is not None and str(ws.cell(r, 3).value).strip():
                    rd = {c: ws.cell(r, c).value
                          for c in range(1, (ws.max_column or 0) + 1)
                          if ws.cell(r, c).value is not None}
                    existing_rows.append(rd)

            added = 0
            duplicates: List[str] = []
            errors: List[str] = []
            for path in src_paths:
                try:
                    src_wb = _lw(path, data_only=True)
                    if "RACK발주양식" not in src_wb.sheetnames:
                        errors.append(f"{os.path.basename(path)}: 'RACK발주양식' 시트 없음")
                        src_wb.close()
                        continue
                    src_ws = src_wb["RACK발주양식"]
                    new_row: Dict = {cell.column: cell.value
                                     for cell in src_ws[2] if cell.value is not None}
                    src_wb.close()

                    if any(ex == new_row for ex in existing_rows):
                        duplicates.append(os.path.basename(path))
                        continue

                    for col, val in new_row.items():
                        ws.cell(first_empty, col, val)
                    existing_rows.append(new_row)
                    first_empty += 1
                    added += 1
                except Exception as e:
                    errors.append(f"{os.path.basename(path)}: {e}")

            tmpl_wb.save(tmpl)
            tmpl_wb.close()
            return added, duplicates, errors

        def _on_result(res):
            added, duplicates, errors = res
            msg = f"{added}건 발주이력을 RACK발주 시트에 추가했습니다."
            if duplicates:
                msg += "\n\n중복 항목 발견 (등록 제외됨):\n" + "\n".join(duplicates)
            if errors:
                msg += "\n\n오류:\n" + "\n".join(errors)
            QMessageBox.information(self, "발주이력 업로드 완료", msg)

        _BgWorker.run_with_progress(self, "발주이력 업로드 중...", _do_upload, tmpl_path, paths,
                                    on_result=_on_result)

    def _refresh_req_list(self) -> None:
        """RACK구매요청서 폴더의 xlsx 파일 목록을 우측 패널에 갱신.
        결재상신용 파일(연주황 음영)과 일반 요청서를 구분하여 표시."""
        root = os.path.join(exe_dir(), "RACK구매요청서")

        def _scan():
            files: List[Tuple[float, str]] = []
            if not os.path.isdir(root):
                return files
            for dirpath, _, filenames in os.walk(root):
                for fn in filenames:
                    if fn.endswith(".xlsx") and not fn.startswith("~$"):
                        full = os.path.join(dirpath, fn)
                        files.append((os.path.getmtime(full), full))
            files.sort(reverse=True)
            return files

        def _on_done(files):
            self.req_list.clear()
            for _, full in files:
                rel = os.path.relpath(full, root)
                item = QListWidgetItem(rel)
                item.setData(Qt.UserRole, full)
                item.setToolTip(full)
                if not os.path.basename(full).endswith("_RACK구매요청서.xlsx"):
                    item.setBackground(QBrush(QColor(255, 224, 178)))
                self.req_list.addItem(item)

        worker = _BgWorker(_scan, parent=self)
        worker.result.connect(_on_done)
        self._req_list_worker = worker  # GC 방지
        worker.start()

    def _open_req_file(self, item: QListWidgetItem) -> None:
        """더블클릭한 요청서 파일을 기본 프로그램으로 열기."""
        path = item.data(Qt.UserRole)
        if path and os.path.exists(path):
            os.startfile(path)
        else:
            QMessageBox.warning(self, "파일 없음", f"파일을 찾을 수 없습니다:\n{path}")

    @staticmethod
    def _is_file_writable(path: str) -> bool:
        """파일이 쓰기 가능한지 확인 (다른 프로세스에 잠겨 있으면 False)."""
        try:
            with open(path, 'r+b'):
                return True
        except (IOError, PermissionError, OSError):
            return False

    def _do_generate_one(self) -> str:
        """현재 self.table 상태로 파일 1개 생성 후 경로 반환."""
        ymd      = datetime.now().strftime("%y%m%d")
        process  = safe_filename(self._text(7,  self.COL_ITEM)) or "RACK"
        line     = safe_filename(self._text(9,  self.COL_ITEM)) or ""
        equip_no = safe_filename(self._text(12, self.COL_ITEM)) or "설비호기"
        folder_mid = f"{ymd}_{line}_{process}" if line else f"{ymd}_{process}"
        folder = ensure_dir(os.path.join(
            exe_dir(), "RACK구매요청서", f"{folder_mid}_RACK구매요청서"))
        out_path = unique_path(os.path.join(folder, f"{equip_no}_RACK구매요청서.xlsx"))
        shutil.copy2(self.rack_template_path, out_path)
        self._direct_patch_xlsx(out_path)
        self._last_generated_folder = folder
        return out_path

    def _generate_request(self) -> None:
        if not self.rack_template_path:
            QMessageBox.warning(self, "오류", "통합양식을 먼저 불러오세요.")
            return

        checked_rows = self._checked_request_rows()

        if not checked_rows or len(checked_rows) == 1:
            # 체크 없음 또는 1개 → 현재 테이블 상태 그대로 생성 (수기 수정 반영)
            try:
                out_path = self._do_generate_one()
                QMessageBox.information(self, "완료",
                    f"RACK 구매요청서 생성 완료\n{os.path.basename(out_path)}")
                if checked_rows:
                    self._mark_request_generated(checked_rows[0])
            except Exception as e:
                QMessageBox.critical(self, "생성 오류", f"요청서 생성 중 오류:\n{e}")
                logger.error("요청서 생성 오류", exc_info=True)
        else:
            # 체크된 행 여러 개 → 각 행 적용 후 순서대로 생성
            # rows 17-31(RACK CH~3단모니터링Cable) 현재 값을 저장해 각 파일에 그대로 반영
            saved_lower: Dict[Tuple[int, int], str] = {}
            for r in range(17, 32):
                w = self.table.cellWidget(r, self.COL_ITEM)
                saved_lower[(r, self.COL_ITEM)] = (
                    w.currentText() if isinstance(w, QComboBox)
                    else (self.table.item(r, self.COL_ITEM).text()
                          if self.table.item(r, self.COL_ITEM) else ""))
                it_q = self.table.item(r, self.COL_QTY)
                saved_lower[(r, self.COL_QTY)] = it_q.text() if it_q else ""

            success, errors = [], []
            for row in checked_rows:
                try:
                    self._apply_request_row(row, show_message=False)
                    # rows 17-31 복원 (의뢰DATA 적용이 덮어쓴 값 되돌리기)
                    for r in range(17, 32):
                        item_val = saved_lower.get((r, self.COL_ITEM), "")
                        w = self.table.cellWidget(r, self.COL_ITEM)
                        if isinstance(w, QComboBox):
                            w.setCurrentText(item_val)
                        else:
                            it = self.table.item(r, self.COL_ITEM)
                            if it:
                                it.setText(item_val)
                        qty_val = saved_lower.get((r, self.COL_QTY), "")
                        it_q = self.table.item(r, self.COL_QTY)
                        if it_q:
                            it_q.setText(qty_val)
                    out_path = self._do_generate_one()
                    success.append(os.path.basename(out_path))
                    self._mark_request_generated(row)
                except Exception as e:
                    errors.append(f"행 {row + 1}: {e}")
                    logger.error("멀티 요청서 생성 오류 (행 %d)", row, exc_info=True)
            msg = f"총 {len(success)}개 생성 완료:\n" + "\n".join(success)
            if errors:
                msg += f"\n\n실패 {len(errors)}건:\n" + "\n".join(errors)
            QMessageBox.information(self, "생성 완료", msg)
        self._refresh_req_list()

    def _generate_overseas_request(self) -> None:
        """국외요청서 생성: 의뢰파일DATA 체크 여부와 무관하게 현재 입력값으로 요청서 1개 생성."""
        if not self.rack_template_path:
            QMessageBox.warning(self, "오류", "통합양식을 먼저 불러오세요.")
            return
        try:
            out_path = self._do_generate_one()
            QMessageBox.information(self, "완료",
                f"국외 RACK 구매요청서 생성 완료\n{os.path.basename(out_path)}")
        except Exception as e:
            QMessageBox.critical(self, "생성 오류", f"요청서 생성 중 오류:\n{e}")
            logger.error("국외요청서 생성 오류", exc_info=True)
        self._refresh_req_list()

    # ── xlsx 직접 조작 helper (ET 미사용, regex/raw bytes) ────────────────

    @staticmethod
    def _xlsx_find_sheet_file(wb_bytes: bytes, rels_bytes: bytes, sheet_name: str) -> Optional[str]:
        """workbook.xml / .rels에서 시트 파일의 zip 내부 경로 반환."""
        wb_str   = wb_bytes.decode('utf-8', errors='replace')
        rels_str = rels_bytes.decode('utf-8', errors='replace')
        rid = None
        for m in re.finditer(r'<sheet\b([^>]*/?>)', wb_str):
            attrs = m.group(1)
            nm_m  = re.search(r'\bname="([^"]*)"', attrs)
            rid_m = re.search(r'\br:id="([^"]*)"', attrs)
            if nm_m and rid_m and nm_m.group(1) == sheet_name:
                rid = rid_m.group(1)
                break
        if not rid:
            return None
        for m in re.finditer(r'<Relationship\b([^>]*/?>)', rels_str):
            attrs = m.group(1)
            id_m  = re.search(r'\bId="([^"]*)"', attrs)
            tgt_m = re.search(r'\bTarget="([^"]*)"', attrs)
            if id_m and id_m.group(1) == rid and tgt_m:
                tgt = tgt_m.group(1).lstrip('/')
                return tgt if tgt.startswith('xl/') else f'xl/{tgt}'
        return None

    @staticmethod
    def _xlsx_patch_wb_visibility(wb_bytes: bytes, visible_sheet: str) -> bytes:
        """workbook.xml에서 visible_sheet 외 모든 시트 veryHidden 처리.
        activeTab도 visible_sheet의 인덱스로 갱신하여 Synap 호환성 확보.
        """
        xml = wb_bytes.decode('utf-8', errors='replace')

        # visible_sheet의 0-based 인덱스 계산
        sheet_tags = re.findall(r'<sheet\b[^>]*/>', xml)
        active_idx = 0
        for i, tag in enumerate(sheet_tags):
            nm_m = re.search(r'\bname="([^"]*)"', tag)
            if nm_m and nm_m.group(1) == visible_sheet:
                active_idx = i
                break

        def repl(m: re.Match) -> str:
            tag = m.group(0)
            nm_m = re.search(r'\bname="([^"]*)"', tag)
            nm = nm_m.group(1) if nm_m else ''
            tag = re.sub(r'\s+state=(?:"[^"]*"|\'[^\']*\')', '', tag)
            if nm and nm != visible_sheet:
                tag = re.sub(r'(\s*/?>)$', r' state="veryHidden"\1', tag)
            return tag

        xml = re.sub(r'<sheet\b[^>]*/>', repl, xml)

        # activeTab을 visible_sheet 인덱스로 설정 (Synap이 hidden 시트 렌더 시도 방지)
        if re.search(r'<workbookView\b', xml):
            def _set_tab(m: re.Match) -> str:
                tag = m.group(0)
                tag = re.sub(r'\bactiveTab="\d+"', f'activeTab="{active_idx}"', tag)
                if 'activeTab=' not in tag:
                    tag = re.sub(r'(\s*/?>)$', f' activeTab="{active_idx}"\\1', tag)
                return tag
            xml = re.sub(r'<workbookView\b[^>]*/>', _set_tab, xml)

        if 'fullCalcOnLoad' not in xml:
            if '<calcPr' in xml:
                xml = re.sub(r'(<calcPr\b[^/]*?)(/?>)',
                             lambda m: m.group(1) + ' fullCalcOnLoad="1"' + m.group(2), xml)
            else:
                xml = xml.replace('</workbook>', '<calcPr fullCalcOnLoad="1"/></workbook>', 1)

        return xml.encode('utf-8')

    @staticmethod
    def _xlsx_patch_sheet_row(sheet_bytes: bytes, row_num: int,
                               col_vals: Dict[int, Any], get_col_letter,
                               cell_styles: Optional[Dict[str, str]] = None) -> bytes:
        """sheet XML 특정 행을 raw 문자열로 삽입/교체 (ET 미사용)."""
        xml = sheet_bytes.decode('utf-8', errors='replace')
        if not col_vals:
            return sheet_bytes

        styles = cell_styles or {}

        def _col_1idx(letters: str) -> int:
            n = 0
            for ch in letters:
                n = n * 26 + (ord(ch) - 64)
            return n

        all_cols = set(col_vals.keys()) | {_col_1idx(l) for l in styles}
        cells_xml = ''
        for col_idx in sorted(all_cols):
            col_l = get_col_letter(col_idx)
            ref   = f'{col_l}{row_num}'
            s_attr = f' s="{styles[col_l]}"' if col_l in styles else ''
            val   = col_vals.get(col_idx)
            if val is None:
                cells_xml += f'<c r="{ref}"{s_attr}/>'
            elif isinstance(val, (int, float)):
                num_s = str(int(val)) if float(val) == int(val) else str(val)
                cells_xml += f'<c r="{ref}"{s_attr}><v>{num_s}</v></c>'
            elif isinstance(val, datetime):
                serial = (val.date() - date(1899, 12, 30)).days
                cells_xml += f'<c r="{ref}"{s_attr}><v>{serial}</v></c>'
            elif isinstance(val, date):
                serial = (val - date(1899, 12, 30)).days
                cells_xml += f'<c r="{ref}"{s_attr}><v>{serial}</v></c>'
            else:
                sv = (str(val).replace('&', '&amp;')
                      .replace('<', '&lt;').replace('>', '&gt;'))
                if sv:
                    cells_xml += f'<c r="{ref}"{s_attr} t="inlineStr"><is><t>{sv}</t></is></c>'
                else:
                    cells_xml += f'<c r="{ref}"{s_attr}/>'

        row_xml = f'<row r="{row_num}">' + cells_xml + '</row>'
        rn = str(row_num)

        # 기존 같은 번호의 행 제거
        xml = re.sub(rf'<row\b[^>]*\br="{rn}"[^>]*/>', '', xml)
        xml = re.sub(rf'<row\b[^>]*\br="{rn}"[^>]*>.*?</row>', '', xml, flags=re.DOTALL)

        # 삽입 위치: row_num보다 큰 첫 번째 행 앞, 없으면 </sheetData> 앞
        insert_pos: Optional[int] = None
        for mm in re.finditer(r'<row\b[^>]*\br="(\d+)"', xml):
            if int(mm.group(1)) > row_num:
                insert_pos = mm.start()
                break

        if insert_pos is not None:
            xml = xml[:insert_pos] + row_xml + xml[insert_pos:]
        elif '</sheetData>' in xml:
            xml = xml.replace('</sheetData>', row_xml + '</sheetData>', 1)
        else:
            _rxml = row_xml
            xml = re.sub(r'<sheetData\s*/>', lambda _: f'<sheetData>{_rxml}</sheetData>', xml, count=1)

        return xml.encode('utf-8')

    @staticmethod
    def _xlsx_remove_external_links(files: dict, names: list) -> tuple:
        """zip에서 externalLinks 파일 및 모든 참조를 제거."""
        names = [n for n in names if not n.startswith('xl/externalLinks/')]
        for key in [k for k in files if k.startswith('xl/externalLinks/')]:
            del files[key]
        rels_key = 'xl/_rels/workbook.xml.rels'
        if rels_key in files:
            rs = files[rels_key].decode('utf-8', errors='replace')
            rs = re.sub(r'<Relationship\b[^>]*Type="[^"]*externalLink[^"]*"[^>]*/>', '', rs)
            files[rels_key] = rs.encode('utf-8')
        wb_key = 'xl/workbook.xml'
        if wb_key in files:
            ws = files[wb_key].decode('utf-8', errors='replace')
            ws = re.sub(r'<externalReferences\b[^>]*>.*?</externalReferences>', '', ws, flags=re.DOTALL)
            ws = re.sub(r'<externalReferences\s*/>', '', ws)
            files[wb_key] = ws.encode('utf-8')
        # [Content_Types].xml에서 externalLink Override 항목 제거
        ct_key = '[Content_Types].xml'
        if ct_key in files:
            ct = files[ct_key].decode('utf-8', errors='replace')
            ct = re.sub(r'<Override\b[^>]*PartName="[^"]*externalLink[^"]*"[^>]*/>', '', ct)
            files[ct_key] = ct.encode('utf-8')
        return files, names

    @staticmethod
    def _xlsx_strip_external_ref_formulas(files: dict) -> dict:
        """워크시트에서 [N] 외부링크 참조 수식의 <f> 태그와 고아 공유수식 인스턴스를 제거."""
        for name in list(files.keys()):
            if not re.match(r'xl/worksheets/sheet\d+\.xml$', name):
                continue
            xml = files[name].decode('utf-8', errors='replace')

            # Step 1: 외부링크 공유수식 마스터의 si 인덱스 수집
            dead_si: set = set()
            for m in re.finditer(r'<f\b([^>]*)>(.*?)</f>', xml, flags=re.DOTALL):
                if re.search(r'\[\d+\]', m.group(2)):
                    si_m = re.search(r'\bsi="(\d+)"', m.group(1))
                    if si_m:
                        dead_si.add(si_m.group(1))

            # Step 2: 외부링크 포함 <f>...</f> 제거
            def _drop(m: re.Match) -> str:
                return '' if re.search(r'\[\d+\]', m.group(0)) else m.group(0)
            new_xml = re.sub(r'<f(?:\s[^>]*)?>.*?</f>', _drop, xml, flags=re.DOTALL)

            # Step 3: 고아 공유수식 인스턴스 (자체종결 <f si="N"/>) 제거
            for si in dead_si:
                new_xml = re.sub(rf'<f\b[^>]*\bsi="{re.escape(si)}"[^>]*/>', '', new_xml)

            if new_xml != xml:
                files[name] = new_xml.encode('utf-8')
        return files

    @staticmethod
    def _xlsx_strip_all_formulas(files: dict) -> dict:
        """모든 워크시트의 수식을 제거하고 셀 타입을 정규화 (복구 메시지 완전 차단).

        모든 <c> 셀을 단일 패스로 처리하여 엣지케이스 누락 방지:
        - 수식 포함 셀: 수식 제거 후 t 속성을 값 타입에 맞게 재설정
        - t="str" 고아 셀(수식 없음): inlineStr 변환 또는 빈 셀
        - 자체종결 t="str" 셀: t 속성 제거
        """
        for name in list(files.keys()):
            if not re.match(r'xl/worksheets/sheet\d+\.xml$', name):
                continue
            xml = files[name].decode('utf-8', errors='replace')

            # ── 비자체종결 셀 처리 ───────────────────────────────────────────
            def _fix_cell(m: re.Match) -> str:
                full  = m.group(0)
                gt    = full.index('>')
                attrs = full[2:gt]        # <c 와 > 사이의 속성 문자열
                body  = full[gt + 1:-4]  # > 와 </c> 사이의 본문

                has_f    = '<f' in body
                has_tstr = 't="str"' in attrs

                if not has_f and not has_tstr:
                    return full  # 변경 불필요

                # 수식 태그 제거
                if has_f:
                    body = re.sub(r'<f(?:\s[^>]*)?>.*?</f>', '', body, flags=re.DOTALL)
                    body = re.sub(r'<f\b[^>]*/>', '', body)

                # t="str" 제거 (수식 없으면 OOXML 규격 위반)
                clean = re.sub(r'\s*\bt="str"', '', attrs)  # 뒤 \b 생략 (""는 word char 아님)

                vm  = re.search(r'<v>(.*?)</v>', body, re.DOTALL)
                val = vm.group(1) if vm else None

                if val is not None:
                    if has_tstr:
                        # 문자열 수식 결과 → inlineStr로 완전 재구성 (잔존 요소 없이 깔끔하게)
                        if val:
                            return f'<c{clean} t="inlineStr"><is><t>{val}</t></is></c>'
                        return f'<c{clean.rstrip()}/>'
                    else:
                        # 숫자/불리언 등 → <v>만 남기고 재구성
                        return f'<c{clean}><v>{val}</v></c>'
                else:
                    # 캐시 값 없음 → 빈 셀
                    return f'<c{clean.rstrip()}/>'

            xml = re.sub(r'<c\b[^>]*(?<!/)>.*?</c>', _fix_cell, xml, flags=re.DOTALL)

            # ── 자체종결 셀 안전망 ──────────────────────────────────────────
            # <c ... t="str" ... /> 에서 t="str" 제거 (빈 숫자형 셀)
            xml = re.sub(r'(<c\b[^>]*?)\s+t="str"([^>]*/>)', r'\1\2', xml)

            files[name] = xml.encode('utf-8')
        return files

    @staticmethod
    def _xlsx_delete_leading_cols(sheet_bytes: bytes, delete_count: int) -> bytes:
        """sheet XML에서 앞 delete_count개 열 삭제 후 나머지 열을 왼쪽으로 이동."""
        from openpyxl.utils import column_index_from_string as _ci, get_column_letter as _cl
        xml = sheet_bytes.decode('utf-8', errors='replace')
        del_letters = [_cl(i) for i in range(1, delete_count + 1)]
        del_pat = '|'.join(re.escape(l) for l in del_letters)

        # ── 셀 데이터 제거 ──────────────────────────────────────────────────
        xml = re.sub(rf'<c\b[^>]*\br="(?:{del_pat})\d+"[^>]*/>', '', xml)
        xml = re.sub(rf'<c\b[^>]*\br="(?:{del_pat})\d+"[^>]*(?<!/)>.*?</c>', '', xml, flags=re.DOTALL)

        def _shift_col(col_idx: int) -> int:
            return max(1, col_idx - delete_count)

        def _shift_cell(ref: str):
            """(new_ref, was_deleted) 반환."""
            rm = re.match(r'([A-Z]+)(\d+)', ref)
            if not rm:
                return ref, False
            ci = _ci(rm.group(1))
            if ci <= delete_count:
                return None, True
            return f'{_cl(ci - delete_count)}{rm.group(2)}', False

        # ── r="XN" 셀 참조 시프트 (cell 태그) ──────────────────────────────
        def shift_ref(m: re.Match) -> str:
            rm = re.match(r'([A-Z]+)(\d+)', m.group(1))
            if not rm:
                return m.group(0)
            ci = _ci(rm.group(1))
            if ci <= delete_count:
                return m.group(0)
            return f'r="{_cl(ci - delete_count)}{rm.group(2)}"'
        xml = re.sub(r'\br="([A-Z]+\d+)"', shift_ref, xml)

        # ── mergeCells 처리 ─────────────────────────────────────────────────
        def shift_merge(m: re.Match) -> str:
            ref = m.group(1)
            if ':' not in ref:
                return ''  # 단일 셀 병합 → 제거
            from_r, to_r = ref.split(':', 1)
            fm = re.match(r'([A-Z]+)(\d+)', from_r)
            tm = re.match(r'([A-Z]+)(\d+)', to_r)
            if not fm or not tm:
                return m.group(0)
            fc, frow = _ci(fm.group(1)), fm.group(2)
            tc, trow = _ci(tm.group(1)), tm.group(2)
            if tc <= delete_count:
                return ''  # 완전히 삭제 범위 → 제거
            new_fc = 1 if fc <= delete_count else fc - delete_count
            new_tc = tc - delete_count
            if new_fc == new_tc and frow == trow:
                return ''  # 단일 셀이 됨 → 병합 불필요
            return f'<mergeCell ref="{_cl(new_fc)}{frow}:{_cl(new_tc)}{trow}"/>'
        xml = re.sub(r'<mergeCell\b[^>]*\bref="([^"]+)"[^>]*/>', shift_merge, xml)
        xml = re.sub(r'<mergeCells[^>]*>\s*</mergeCells>', '', xml)

        # ── sqref 속성 처리 (dataValidation / conditionalFormatting) ────────
        def shift_sqref(sqref_val: str) -> str:
            new_parts = []
            for part in sqref_val.split():
                if ':' in part:
                    fr, tr = part.split(':', 1)
                    fm = re.match(r'([A-Z]+)(\d+)', fr)
                    tm = re.match(r'([A-Z]+)(\d+)', tr)
                    if fm and tm:
                        fc, frow = _ci(fm.group(1)), fm.group(2)
                        tc, trow = _ci(tm.group(1)), tm.group(2)
                        if tc <= delete_count:
                            continue
                        new_fc = 1 if fc <= delete_count else fc - delete_count
                        new_parts.append(f'{_cl(new_fc)}{frow}:{_cl(tc - delete_count)}{trow}')
                    else:
                        new_parts.append(part)
                else:
                    mm = re.match(r'([A-Z]+)(\d+)', part)
                    if mm:
                        ci = _ci(mm.group(1))
                        if ci <= delete_count:
                            continue
                        new_parts.append(f'{_cl(ci - delete_count)}{mm.group(2)}')
                    else:
                        new_parts.append(part)
            return ' '.join(new_parts)

        def repl_sqref(m: re.Match) -> str:
            new_val = shift_sqref(m.group(1))
            return f'sqref="{new_val}"' if new_val else 'sqref=""'
        xml = re.sub(r'\bsqref="([^"]+)"', repl_sqref, xml)
        # sqref가 빈 조건부서식/유효성 블록 제거
        xml = re.sub(r'<conditionalFormatting\b[^>]*sqref=""[^>]*/>', '', xml)
        xml = re.sub(r'<conditionalFormatting\b[^>]*sqref=""[^>]*>.*?</conditionalFormatting>',
                     '', xml, flags=re.DOTALL)
        xml = re.sub(r'<dataValidation\b[^>]*sqref=""[^>]*/>', '', xml)
        xml = re.sub(r'<dataValidation\b[^>]*sqref=""[^>]*>.*?</dataValidation>',
                     '', xml, flags=re.DOTALL)

        # ── <col> 요소 min/max 열 인덱스 시프트 ──────────────────────────────
        def shift_col_elem(m: re.Match) -> str:
            attrs = m.group(0)
            min_m = re.search(r'\bmin="(\d+)"', attrs)
            max_m = re.search(r'\bmax="(\d+)"', attrs)
            if not min_m or not max_m:
                return attrs
            col_min = int(min_m.group(1))
            col_max = int(max_m.group(1))
            if col_max <= delete_count:
                return ''  # 삭제 범위 내 → 제거
            new_min = max(1, col_min - delete_count)
            new_max = col_max - delete_count
            attrs = re.sub(r'\bmin="\d+"', f'min="{new_min}"', attrs)
            attrs = re.sub(r'\bmax="\d+"', f'max="{new_max}"', attrs)
            return attrs
        xml = re.sub(r'<col\b[^>]*/>', shift_col_elem, xml)
        # 비어진 <cols> 블록 제거
        xml = re.sub(r'<cols>\s*</cols>', '', xml)

        # ── sheetViews 완전 초기화: 열 삭제 후 뷰 상태(selection/pane)를 ────
        # ── 깨끗한 최소값으로 교체 — sqref="" / 잘못된 pane 오류 방지 ────────
        xml = re.sub(
            r'<sheetViews\b[^>]*>.*?</sheetViews>',
            '<sheetViews><sheetView workbookViewId="0">'
            '<selection activeCell="A1" sqref="A1"/></sheetView></sheetViews>',
            xml, flags=re.DOTALL)

        return xml.encode('utf-8')

    @staticmethod
    def _xlsx_shift_drawing_cols(files: dict, sheet_zip_path: str, delete_count: int) -> dict:
        """sheet와 연결된 drawing XML의 열 앵커를 왼쪽으로 보정 (_xlsx_delete_leading_cols 후 호출)."""
        # 시트 rels 경로: xl/worksheets/sheet1.xml → xl/worksheets/_rels/sheet1.xml.rels
        base, fname = sheet_zip_path.rsplit('/', 1) if '/' in sheet_zip_path else ('', sheet_zip_path)
        rels_path = f"{base}/_rels/{fname}.rels" if base else f"_rels/{fname}.rels"
        if rels_path not in files:
            return files

        rels_xml = files[rels_path].decode('utf-8', errors='replace')
        for m in re.finditer(r'<Relationship\b([^>]*/?>)', rels_xml):
            attrs  = m.group(1)
            type_m = re.search(r'\bType="([^"]*)"', attrs)
            tgt_m  = re.search(r'\bTarget="([^"]*)"', attrs)
            if not (type_m and tgt_m and '/drawing' in type_m.group(1)):
                continue

            tgt = tgt_m.group(1)
            # "../drawings/drawing1.xml" → "xl/drawings/drawing1.xml"
            if tgt.startswith('../'):
                drawing_path = base.rsplit('/', 1)[0] + '/' + tgt[3:] if '/' in base else tgt[3:]
            elif tgt.startswith('/'):
                drawing_path = tgt.lstrip('/')
            else:
                drawing_path = (base + '/' + tgt) if base else tgt

            if drawing_path not in files:
                continue

            draw_xml = files[drawing_path].decode('utf-8', errors='replace')

            def _shift(cm: re.Match) -> str:
                return f'<xdr:col>{max(0, int(cm.group(1)) - delete_count)}</xdr:col>'

            files[drawing_path] = re.sub(
                r'<xdr:col>(\d+)</xdr:col>', _shift, draw_xml).encode('utf-8')

        return files

    # ── xlsx 생성 / 통합양식 업데이트 ──────────────────────────────────────

    @staticmethod
    def _xlsx_load_and_clean(path: str, *, strip_all_formulas: bool = False) -> Tuple[dict, list]:
        """xlsx ZIP을 읽어 calcChain·externalLinks를 제거한 (files, names) 반환.

        strip_all_formulas=True: 모든 수식 제거 (출력 파일용 — 복구 메시지 원천 차단)
        strip_all_formulas=False: 외부링크+고아 공유수식만 제거 (템플릿 업데이트용)
        """
        import zipfile as _zf
        with _zf.ZipFile(path, 'r') as zf:
            names = zf.namelist()
            files = {n: zf.read(n) for n in names}
        files.pop('xl/calcChain.xml', None)
        names = [n for n in names if n != 'xl/calcChain.xml']
        files, names = RackPurchaseRequestPage._xlsx_remove_external_links(files, names)
        # 이전 버전 코드가 생성한 손상된 <c> 태그 수리:
        # 자체종결 셀(/>)이 인접 셀과 합쳐져 속성 중간에 / 가 삽입된 패턴 수정
        # 예: <c r="B2" s="364"/ t="s"> → <c r="B2" s="364" t="s">
        for _xml_key in list(files.keys()):
            if not _xml_key.endswith('.xml'):
                continue
            _raw = files[_xml_key].decode('utf-8', errors='replace')
            _fixed = re.sub(r'(<c\b[^>]*?)"/ ([a-zA-Z])', r'\1" \2', _raw, flags=re.DOTALL)
            if _fixed != _raw:
                files[_xml_key] = _fixed.encode('utf-8')
        if strip_all_formulas:
            files = RackPurchaseRequestPage._xlsx_strip_all_formulas(files)
        else:
            files = RackPurchaseRequestPage._xlsx_strip_external_ref_formulas(files)
        return files, names

    @staticmethod
    def _xlsx_expand_shared_strings(files: dict) -> dict:
        """모든 t="s" 셀을 인라인 문자열(inlineStr)로 펼침.

        손상된 sharedStrings.xml(이전 버그로 인덱스 오류)을 가진 파일을 로드할 때
        기존 t="s" 참조를 실제 문자열로 교체하여 인덱스 오류를 원천 제거.
        이후 _xlsx_inline_to_shared 가 깨끗한 상태에서 재구성함.
        """
        SS_KEY = 'xl/sharedStrings.xml'
        if SS_KEY not in files:
            return files

        ss_xml = files[SS_KEY].decode('utf-8', errors='replace')

        # 인덱스 → 텍스트 맵 구성 (단순 <t> 내용만 추출; rich-text는 첫 <t>로 대체)
        idx_to_str: Dict[int, str] = {}
        for idx, m in enumerate(re.finditer(r'<si\b[^>]*>(.*?)</si>', ss_xml, re.DOTALL)):
            tm = re.search(r'<t[^>]*>(.*?)</t>', m.group(1), re.DOTALL)
            if tm:
                raw = tm.group(1)
                plain = (raw.replace('&amp;', '&').replace('&lt;', '<')
                         .replace('&gt;', '>').replace('&quot;', '"').replace('&apos;', "'"))
                idx_to_str[idx] = plain

        def _expand(m: re.Match) -> str:
            full = m.group(0)
            if 't="s"' not in full:
                return full
            gt = full.index('>')
            attrs = full[2:gt]
            body  = full[gt + 1:-4]
            vm = re.search(r'<v>(\d+)</v>', body)
            if not vm:
                return full
            si_idx = int(vm.group(1))
            text = idx_to_str.get(si_idx, '')
            new_attrs = re.sub(r'\s*\bt="s"', '', attrs) + ' t="inlineStr"'
            if not text:
                return f'<c{re.sub(r" t=\"inlineStr\"", "", new_attrs).rstrip()}/>'
            esc = (text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))
            return f'<c{new_attrs}><is><t>{esc}</t></is></c>'

        for name in list(files.keys()):
            if not re.match(r'xl/worksheets/sheet\d+\.xml$', name):
                continue
            xml = files[name].decode('utf-8', errors='replace')
            new_xml = re.sub(r'<c\b[^>]*(?<!/)>.*?</c>', _expand, xml, flags=re.DOTALL)
            if new_xml != xml:
                files[name] = new_xml.encode('utf-8')

        return files

    @staticmethod
    def _xlsx_inline_to_shared(files: dict) -> dict:
        """시트의 t="inlineStr" 셀을 sharedStrings 방식으로 변환.
        Synap Document Viewer 등 inlineStr을 지원하지 않는 뷰어와의 호환성.
        """
        import zipfile as _zf

        SS_KEY = 'xl/sharedStrings.xml'
        SS_NS  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

        # 기존 sharedStrings 읽기
        if SS_KEY in files:
            ss_xml = files[SS_KEY].decode('utf-8', errors='replace')
        else:
            ss_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                f'<sst xmlns="{SS_NS}" count="0" uniqueCount="0"></sst>'
            )

        # 기존 항목 수: uniqueCount 속성이 아닌 실제 <si> 개수로 계산
        # (이전 버전 코드가 잘못된 uniqueCount를 저장했을 경우에도 안전하게 동작)
        existing_count = len(re.findall(r'<si\b', ss_xml))

        # 기존 문자열 → 인덱스 맵 (단순 <t>...</t> 패턴만 매핑)
        # enumerate로 실제 <si> 순서 인덱스 사용 (rich text 등 <t> 없는 항목도 인덱스 정확히 유지)
        str_to_idx: Dict[str, int] = {}
        for real_idx, m in enumerate(re.finditer(r'<si\b[^>]*>(.*?)</si>', ss_xml, flags=re.DOTALL)):
            tm = re.search(r'<t[^>]*>(.*?)</t>', m.group(1), re.DOTALL)
            if tm:
                raw = tm.group(1)
                plain = (raw.replace('&amp;', '&').replace('&lt;', '<')
                         .replace('&gt;', '>').replace('&quot;', '"').replace('&apos;', "'"))
                str_to_idx[plain] = real_idx

        # 새로 추가할 항목만 별도 리스트에 쌓음 (기존 항목은 절대 삭제하지 않음)
        new_strings: list = []
        new_str_to_idx: Dict[str, int] = {}

        def _get_idx(plain: str) -> int:
            if plain in str_to_idx:
                return str_to_idx[plain]
            if plain not in new_str_to_idx:
                idx = existing_count + len(new_strings)
                new_str_to_idx[plain] = idx
                esc = (plain.replace('&', '&amp;').replace('<', '&lt;')
                       .replace('>', '&gt;'))
                new_strings.append(f'<si><t>{esc}</t></si>')
            return new_str_to_idx[plain]

        changed_any = False
        for name in list(files.keys()):
            if not re.match(r'xl/worksheets/sheet\d+\.xml$', name):
                continue
            xml = files[name].decode('utf-8', errors='replace')
            changed = [False]

            def _conv(m: re.Match, _ch=changed) -> str:
                full = m.group(0)
                if 't="inlineStr"' not in full:
                    return full
                gt = full.index('>')
                attrs = full[2:gt]
                body  = full[gt + 1:-4]
                tm = re.search(r'<is>.*?<t[^>]*>(.*?)</t>.*?</is>', body, re.DOTALL)
                if not tm:
                    return full
                raw = tm.group(1)
                plain = (raw.replace('&amp;', '&').replace('&lt;', '<')
                         .replace('&gt;', '>').replace('&quot;', '"').replace('&apos;', "'"))
                idx = _get_idx(plain)
                new_attrs = re.sub(r'\s*t="inlineStr"', '', attrs) + ' t="s"'
                _ch[0] = True
                return f'<c{new_attrs}><v>{idx}</v></c>'

            new_xml = re.sub(r'<c\b[^>]*(?<!/)>.*?</c>', _conv, xml, flags=re.DOTALL)
            if changed[0]:
                files[name] = new_xml.encode('utf-8')
                changed_any = True

        if changed_any:
            total = existing_count + len(new_strings)
            ss_xml = re.sub(r'\bcount="\d+"', f'count="{total}"', ss_xml)
            ss_xml = re.sub(r'\buniqueCount="\d+"', f'uniqueCount="{total}"', ss_xml)
            # 기존 <si> 항목은 절대 삭제하지 않고, 새 항목만 뒤에 추가
            if new_strings:
                ss_xml = ss_xml.replace('</sst>', ''.join(new_strings) + '</sst>')
            files[SS_KEY] = ss_xml.encode('utf-8')

            # Content_Types에 sharedStrings 없으면 추가
            CT_KEY = '[Content_Types].xml'
            if CT_KEY in files:
                ct = files[CT_KEY].decode('utf-8', errors='replace')
                if 'sharedStrings' not in ct:
                    override = (
                        '<Override PartName="/xl/sharedStrings.xml"'
                        ' ContentType="application/vnd.openxmlformats-officedocument'
                        '.spreadsheetml.sharedStrings+xml"/>'
                    )
                    ct = ct.replace('</Types>', override + '</Types>')
                    files[CT_KEY] = ct.encode('utf-8')

            # xl/_rels/workbook.xml.rels에 sharedStrings 관계 없으면 추가
            WB_RELS = 'xl/_rels/workbook.xml.rels'
            if WB_RELS in files:
                wr = files[WB_RELS].decode('utf-8', errors='replace')
                if 'sharedStrings' not in wr:
                    rel = (
                        '<Relationship Id="rIdSS" '
                        'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                        'relationships/sharedStrings" '
                        'Target="sharedStrings.xml"/>'
                    )
                    wr = wr.replace('</Relationships>', rel + '</Relationships>')
                    files[WB_RELS] = wr.encode('utf-8')

        return files

    @staticmethod
    def _xlsx_save(path: str, files: dict, names: list) -> None:
        """(files, names)를 xlsx ZIP으로 저장.
        OOXML 스펙에 따라 [Content_Types].xml 은 ZIP_STORED(비압축) 필수.
        Synap 등 엄격한 뷰어는 이를 압축하면 미리보기 오류를 낸다.
        """
        import zipfile as _zf
        # [Content_Types].xml 을 항상 첫 번째로, 비압축으로 기록
        ordered = ['[Content_Types].xml'] + [n for n in names if n != '[Content_Types].xml']
        with _zf.ZipFile(path, 'w') as zf:
            for name in ordered:
                if name not in files:
                    continue
                ct = _zf.ZIP_STORED if name == '[Content_Types].xml' else _zf.ZIP_DEFLATED
                zf.writestr(name, files[name], compress_type=ct)

    @staticmethod
    def _xlsx_collect_row_styles(sheet_xml: str, row_num: int) -> Dict[str, str]:
        """sheet XML에서 지정 행의 열별 s 속성(스타일 인덱스) 수집."""
        styles: Dict[str, str] = {}
        for row_m in re.finditer(
                rf'<row\b[^>]*\br="{row_num}"[^>]*>(.*?)</row>', sheet_xml, flags=re.DOTALL):
            for cell_m in re.finditer(r'<c\b([^>]*)>', row_m.group(1)):
                attrs = cell_m.group(1)
                ref_m = re.search(r'\br="([A-Z]+)\d+"', attrs)
                s_m   = re.search(r'\bs="(\d+)"', attrs)
                if ref_m and s_m:
                    styles[ref_m.group(1)] = s_m.group(1)
            break
        return styles

    def _direct_patch_xlsx(self, path: str) -> None:
        """drawings 손상·calcChain 오류 없이 xlsx를 직접 수정 (regex 방식)."""
        from openpyxl.utils import get_column_letter as _gcl

        files, names = self._xlsx_load_and_clean(path, strip_all_formulas=True)

        wb_bytes   = files.get('xl/workbook.xml', b'')
        rels_bytes = files.get('xl/_rels/workbook.xml.rels', b'')
        files['xl/workbook.xml'] = self._xlsx_patch_wb_visibility(wb_bytes, 'RACK발주양식')
        rack_file = self._xlsx_find_sheet_file(wb_bytes, rels_bytes, 'RACK발주양식')

        row2_vals: Dict[int, Any] = {}
        for row, (_, item_col, qty_col) in self._ROW_RACK_MAP.items():
            for col, tbl_col in ((item_col, self.COL_ITEM), (qty_col, self.COL_QTY)):
                if col is None:
                    continue
                txt = self._text(row, tbl_col)
                if txt == '':
                    continue
                try:
                    v: Any = int(txt) if '.' not in txt else float(txt)
                except (ValueError, TypeError):
                    v = txt
                row2_vals[col] = v

        rack_ch_item = self._text(17, self.COL_ITEM)
        rack_ch_qty  = self._text(17, self.COL_QTY)
        if rack_ch_item:
            row2_vals[41] = rack_ch_item  # AO열(41): RACK CH 항목
        if rack_ch_qty:
            try:
                rack_ch_qty_v: Any = int(rack_ch_qty) if '.' not in rack_ch_qty else float(rack_ch_qty)
            except (ValueError, TypeError):
                rack_ch_qty_v = rack_ch_qty
            row2_vals[42] = rack_ch_qty_v  # AP열(42): RACK CH 수량

        row2_vals[88] = "전기공통"  # CJ열(88) 항상 기입

        invoice_text = self._text(2, self.COL_ITEM)
        if invoice_text:
            row2_vals[2] = invoice_text  # B열(2): 인보이스 필요여부
        if invoice_text == "인보이스 필요":
            row2_vals[95] = "KIT투자 시 인보이스필요"  # CQ열(95): 비고

        # 데이터 입력된 열 중 비어 있는 셀은 "-"로 채움
        all_data_cols: set = set()
        for _, item_col, qty_col in self._ROW_RACK_MAP.values():
            if item_col is not None:
                all_data_cols.add(item_col)
            if qty_col is not None:
                all_data_cols.add(qty_col)
        for col in all_data_cols:
            if col not in row2_vals:
                row2_vals[col] = "-"

        if rack_file and rack_file in files:
            try:
                sheet_xml = files[rack_file].decode('utf-8', errors='replace')
                row2_styles = self._xlsx_collect_row_styles(sheet_xml, 2)
                files[rack_file] = self._xlsx_patch_sheet_row(
                    files[rack_file], 2, row2_vals, _gcl, row2_styles)
            except Exception as e:
                logger.warning("RACK발주양식 시트 패치 실패: %s", e)

        # 기존 t="s" 셀을 모두 inlineStr로 펼침 (손상된 sharedStrings 인덱스 제거)
        files = self._xlsx_expand_shared_strings(files)
        # inlineStr → sharedStrings 변환 (Synap 미리보기 호환)
        files = self._xlsx_inline_to_shared(files)
        self._xlsx_save(path, files, names)


    def _generate_approval_doc(self) -> None:
        """결재상신용: 여러 요청서 RACK발주양식 2행을 합쳐 A·B열 삭제 후 저장 (zipfile 방식)."""
        from openpyxl.utils import get_column_letter as _gcl

        start_dir = self._last_generated_folder or exe_dir()
        paths, _ = QFileDialog.getOpenFileNames(
            self, "RACK 구매요청서 선택 (2개 이상)", start_dir, "Excel Files (*.xlsx)"
        )
        if not paths:
            return
        if len(paths) < 2:
            QMessageBox.warning(self, "선택 오류", "2개 이상의 파일을 선택하세요.")
            return

        folder = os.path.dirname(paths[0])
        folder_name = os.path.basename(folder)
        process = folder_name
        if "_RACK구매요청서" in folder_name:
            base = folder_name.split("_RACK구매요청서")[0]
            parts = base.split("_")
            process = "_".join(parts[1:]) if len(parts) > 1 else base

        today    = datetime.now().strftime("%Y-%m-%d")
        out_path = unique_path(os.path.join(
            folder, f"{process} Rack 구매요청서 ({today}).xlsx"))

        try:
            all_row_vals: List[Dict[int, Any]] = []
            for path in paths:
                try:
                    from openpyxl import load_workbook
                    src_wb = load_workbook(path, data_only=True)
                    ws_r   = src_wb["RACK발주양식"]
                    rv = {c: ws_r.cell(2, c).value
                          for c in range(1, ws_r.max_column + 1)
                          if ws_r.cell(2, c).value is not None}
                    src_wb.close()
                    all_row_vals.append(rv)
                except Exception as e:
                    logger.warning("결재상신용 파일 읽기 실패 (%s): %s", path, e)
                    all_row_vals.append({})

            shutil.copy2(paths[0], out_path)
            files, names = self._xlsx_load_and_clean(out_path, strip_all_formulas=True)

            wb_bytes   = files.get('xl/workbook.xml', b'')
            rels_bytes = files.get('xl/_rels/workbook.xml.rels', b'')
            files['xl/workbook.xml'] = self._xlsx_patch_wb_visibility(wb_bytes, 'RACK발주양식')
            rack_file  = self._xlsx_find_sheet_file(wb_bytes, rels_bytes, 'RACK발주양식')

            if not rack_file or rack_file not in files:
                QMessageBox.critical(self, "오류", "RACK발주양식 시트를 찾을 수 없습니다.")
                return

            row2_styles = self._xlsx_collect_row_styles(
                files[rack_file].decode('utf-8', errors='replace'), 2)

            for i, rv in enumerate(all_row_vals[1:], start=3):
                try:
                    files[rack_file] = self._xlsx_patch_sheet_row(
                        files[rack_file], i, rv, _gcl, row2_styles)
                except Exception as e:
                    logger.warning("결재상신용 행 추가 실패 (row=%d): %s", i, e)

            files[rack_file] = self._xlsx_delete_leading_cols(files[rack_file], 2)
            files = self._xlsx_shift_drawing_cols(files, rack_file, 2)

            sheet_str = files[rack_file].decode('utf-8', errors='replace')
            sheet_str = re.sub(r'(<col\b[^>]*?)\s+hidden="1"', r'\1', sheet_str)
            files[rack_file] = sheet_str.encode('utf-8')

            files = self._xlsx_expand_shared_strings(files)
            files = self._xlsx_inline_to_shared(files)
            self._xlsx_save(out_path, files, names)

            self._refresh_req_list()
            QMessageBox.information(self, "완료",
                f"결재상신용 생성 완료\n{os.path.basename(out_path)}")
        except Exception as e:
            QMessageBox.critical(self, "생성 오류", f"결재상신용 생성 중 오류:\n{e}")
            logger.error("결재상신용 생성 오류", exc_info=True)

    # ── 그룹웨어 로그인 관리 ──────────────────────────────────────────────────
    def _do_gw_login(self) -> None:
        uid = self._ed_gw_id.text().strip()
        pwd = self._ed_gw_pw.text().strip()
        if not uid or not pwd:
            QMessageBox.warning(self, "입력 오류", "아이디와 비밀번호를 입력해 주세요.")
            return

        import approval_auto as _aa

        self._btn_gw_login.setEnabled(False)
        self._btn_gw_login.setText("확인 중…")

        def _check(_):
            return _aa.check_login(uid, pwd)

        def _on_done(result):
            success, msg = result
            self._btn_gw_login.setEnabled(True)
            self._btn_gw_login.setText("로그인")
            if success:
                self._gw_username = uid
                self._gw_password = pwd
                self._lbl_gw_user.setText(f"로그인됨: {uid}")
                self._login_form_w.hide()
                self._login_status_w.show()
            else:
                QMessageBox.warning(self, "로그인 실패", f"로그인 실패\n\n{msg}")

        _BgWorker.run_with_progress(
            self, "로그인 확인 중…",
            _check, None,
            on_result=_on_done,
        )

    def _do_gw_logout(self) -> None:
        self._gw_username = ""
        self._gw_password = ""
        self._ed_gw_id.clear()
        self._ed_gw_pw.clear()
        self._login_status_w.hide()
        self._login_form_w.show()

    # ── 결재상신 자동화 ────────────────────────────────────────────────────────
    def _do_approval_submit(self) -> None:
        """결재상신용 xlsx를 선택 → 전자결재 시스템에 자동으로 구매요청서(NPN) 결재상신."""
        # 선행 조건 검사를 파일 선택 전에 수행 (순서 중요: 안내 없이 조용히 종료되는 것 방지)

        # selenium/webdriver-manager 설치 확인
        try:
            import selenium  # noqa: F401
        except ImportError:
            QMessageBox.critical(
                self, "패키지 없음",
                "결재상신 기능에는 selenium 패키지가 필요합니다.\n\n"
                "pip install selenium webdriver-manager\n\n"
                "설치 후 프로그램을 재시작하세요."
            )
            return

        import approval_auto as _aa

        if not self._gw_username or not self._gw_password:
            QMessageBox.warning(self, "로그인 필요", "먼저 우측 패널에서 그룹웨어 로그인을 해주세요.")
            return

        start_dir = self._last_generated_folder or exe_dir()
        paths, _ = QFileDialog.getOpenFileNames(
            self, "결재상신할 RACK 구매요청서 선택", start_dir, "Excel Files (*.xlsx)"
        )
        if not paths:
            return
        username, password = self._gw_username, self._gw_password

        # 각 파일에 대해 순서대로 결재상신 (파일이 여러 개면 확인)
        if len(paths) > 1:
            reply = QMessageBox.question(
                self, "다중 파일",
                f"{len(paths)}개 파일에 대해 순서대로 결재상신을 진행합니다.\n"
                "계속하시겠습니까?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return

        def _run_all(file_paths):
            results = []
            for fp in file_paths:
                try:
                    data = _aa.read_rack_row2(fp)
                    fname = os.path.splitext(os.path.basename(fp))[0]
                    _aa.run_approval(
                        xlsx_path    = fp,
                        title        = fname,
                        pr_no        = data["pr_no"],
                        sales_person = data["sales_person"],
                        line         = data["line"],
                        process      = data["process"],
                        equipment    = data["equipment"],
                        equip_model  = data["equip_model"],
                        remark       = data["remark"],
                        po_no        = "-",
                        username     = username,
                        password     = password,
                    )
                    results.append(f"✅ {fname}")
                except Exception as e:
                    results.append(f"❌ {os.path.basename(fp)}: {e}")
            return results

        def _on_done(results):
            _aa.start_window_monitor()
            msg = "\n".join(results) if results else "완료"
            QMessageBox.information(self, "결재상신 결과", msg)

        _BgWorker.run_with_progress(
            self, "결재상신 진행 중… (Chrome이 자동으로 실행됩니다)",
            _run_all, paths,
            on_result=_on_done,
        )

