"""
approval_auto.py
================
Selenium 기반 전자결재 자동화 모듈.
구매요청서(NPN) 결재상신 워크플로를 자동으로 진행한다.

연결 방식
---------
앱 전용 Chrome 프로필을 별도로 실행한다(기존 Chrome 창에 영향 없음).
ID/PW를 받아 자동 로그인 후 결재상신을 진행한다.

요구 패키지:
    pip install selenium webdriver-manager
"""
from __future__ import annotations

import os
import sys
import time
import logging
from typing import Callable, Dict, Optional

logger = logging.getLogger("QuoteApp.approval")

APPROVAL_URL = "https://groupware.lotvacuum.com/app/approval"
FORM_NAME    = "구매요청서(NPN)"

_active_driver = None  # 열려 있는 드라이버 추적 (두 번째 실행 전 정리용)


# ──────────────────────────────────────────────────────────────────────────────
# 내부 유틸
# ──────────────────────────────────────────────────────────────────────────────
def _profile_dir() -> str:
    """앱 전용 Chrome 프로필 경로 (AppData\\Roaming — 쓰기 권한 보장)."""
    appdata = os.environ.get("APPDATA") or os.path.expanduser("~")
    return os.path.join(appdata, "quote-auto", "approval_profile")


def _find_chrome_binary() -> Optional[str]:
    candidates = [
        os.path.join(os.environ.get("PROGRAMFILES", ""),
                     "Google", "Chrome", "Application", "chrome.exe"),
        os.path.join(os.environ.get("PROGRAMFILES(X86)", ""),
                     "Google", "Chrome", "Application", "chrome.exe"),
        os.path.join(os.environ.get("LOCALAPPDATA", ""),
                     "Google", "Chrome", "Application", "chrome.exe"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return None


# ──────────────────────────────────────────────────────────────────────────────
# WebDriver 생성
# ──────────────────────────────────────────────────────────────────────────────
def _patch_chrome_shortcut() -> bool:
    """
    사용자 바탕화면에 Chrome 디버깅 바로가기를 만든다.
    - 사용자 바탕화면의 기존 Chrome 바로가기 수정 시도
    - 실패(권한 없음)하면 사용자 바탕화면에 새 바로가기 생성
    반환값: 성공 True / 실패 False
    """
    import glob
    try:
        import win32com.client
    except ImportError:
        return False

    shell = win32com.client.Dispatch("WScript.Shell")
    user_desktop = shell.SpecialFolders("Desktop")  # 사용자 개인 바탕화면만 사용
    debug_arg = "--remote-debugging-port=9222"

    # ① 사용자 바탕화면의 Chrome 바로가기 수정 시도
    for lnk in glob.glob(os.path.join(user_desktop, "*.lnk")):
        try:
            sc = shell.CreateShortcut(lnk)
            if "chrome" not in (sc.TargetPath or "").lower():
                continue
            if debug_arg in (sc.Arguments or ""):
                logger.info("Chrome 바로가기 이미 설정됨: %s", lnk)
                return True
            sc.Arguments = ((sc.Arguments or "") + " " + debug_arg).strip()
            sc.Save()
            logger.info("Chrome 바로가기 수정 완료: %s", lnk)
            return True
        except Exception as e:
            logger.warning("바로가기 수정 실패 (%s): %s", lnk, e)

    # ② 수정 실패 또는 바로가기 없음 → 사용자 바탕화면에 새 바로가기 생성
    chrome_bin = _find_chrome_binary()
    if not chrome_bin:
        return False
    try:
        new_lnk = os.path.join(user_desktop, "Google Chrome (결재).lnk")
        sc = shell.CreateShortcut(new_lnk)
        sc.TargetPath = chrome_bin
        sc.Arguments = debug_arg
        sc.WorkingDirectory = os.path.dirname(chrome_bin)
        sc.Description = "Chrome (결재상신용)"
        sc.Save()
        logger.info("Chrome 바로가기 새로 생성: %s", new_lnk)
        return True
    except Exception as e:
        logger.warning("바로가기 생성 실패: %s", e)
        return False


def _automation_profile_dir() -> str:
    """자동화 전용 Chrome 프로필 경로."""
    appdata = os.environ.get("APPDATA") or os.path.expanduser("~")
    return os.path.join(appdata, "quote-auto", "chrome_profile")


def _find_cloudoc_extension_id() -> Optional[str]:
    """사용자 기본 Chrome 프로필에서 ClouDoc 확장프로그램 ID를 찾아 반환한다."""
    import json

    ext_base = os.path.join(
        os.environ.get("LOCALAPPDATA", ""),
        "Google", "Chrome", "User Data", "Default", "Extensions"
    )
    if not os.path.isdir(ext_base):
        return None

    for ext_id in os.listdir(ext_base):
        ext_id_path = os.path.join(ext_base, ext_id)
        if not os.path.isdir(ext_id_path):
            continue
        for version in os.listdir(ext_id_path):
            manifest_path = os.path.join(ext_id_path, version, "manifest.json")
            if not os.path.isfile(manifest_path):
                continue
            try:
                with open(manifest_path, encoding="utf-8") as f:
                    manifest = json.load(f)
                name = manifest.get("name", "").lower()
                if "cloudoc" in name:
                    logger.info("ClouDoc 확장프로그램 ID 발견: %s", ext_id)
                    return ext_id
            except Exception:
                continue
    return None


def _is_cloudoc_installed_in_profile(profile_dir: str) -> bool:
    """자동화 프로필에 ClouDoc이 설치되어 있는지 확인한다."""
    import json

    ext_base = os.path.join(profile_dir, "Default", "Extensions")
    if not os.path.isdir(ext_base):
        return False

    for ext_id in os.listdir(ext_base):
        for version in os.listdir(os.path.join(ext_base, ext_id)):
            manifest_path = os.path.join(ext_base, ext_id, version, "manifest.json")
            if not os.path.isfile(manifest_path):
                continue
            try:
                with open(manifest_path, encoding="utf-8") as f:
                    manifest = json.load(f)
                if "cloudoc" in manifest.get("name", "").lower():
                    return True
            except Exception:
                continue
    return False


def _create_driver(offscreen: bool = False):
    """
    자동화 전용 Chrome 프로필로 WebDriver를 생성한다.
    ClouDoc 확장프로그램이 프로필에 없으면 웹스토어에서 자동 설치 후 재시작한다.
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.by import By
    except ImportError:
        raise RuntimeError(
            "selenium 패키지가 없습니다.\n"
            "pip install selenium webdriver-manager"
        )

    try:
        from webdriver_manager.chrome import ChromeDriverManager
        service = Service(ChromeDriverManager().install())
    except Exception:
        service = Service()

    chrome_bin = _find_chrome_binary()
    if not chrome_bin:
        raise RuntimeError("Chrome 실행 파일을 찾을 수 없습니다.")

    # 이전 드라이버가 열려 있으면 종료 (프로필 잠금 해제)
    global _active_driver
    if _active_driver is not None:
        try:
            _active_driver.quit()
        except Exception:
            pass
        _active_driver = None
        time.sleep(2)  # Chrome이 프로필 잠금을 완전히 해제할 때까지 대기

    profile_dir = _automation_profile_dir()
    os.makedirs(profile_dir, exist_ok=True)

    def _build_opts(visible: bool = True) -> Options:
        opts = Options()
        opts.binary_location = chrome_bin
        opts.add_argument(f"--user-data-dir={profile_dir}")
        opts.add_argument("--profile-directory=Default")
        opts.add_argument("--no-first-run")
        opts.add_argument("--no-default-browser-check")
        opts.add_argument("--disable-notifications")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        if not visible:
            opts.add_argument("--start-minimized")
        return opts

    # ClouDoc 설치 여부 확인 → 없으면 웹스토어에서 설치
    if not _is_cloudoc_installed_in_profile(profile_dir):
        logger.info("ClouDoc 미설치 → 웹스토어 설치 시작")
        ext_id = _find_cloudoc_extension_id()

        driver = webdriver.Chrome(service=service, options=_build_opts(visible=True))
        driver.set_window_size(1000, 700)

        if ext_id:
            store_url = f"https://chromewebstore.google.com/detail/{ext_id}"
        else:
            store_url = "https://chromewebstore.google.com/search/ClouDoc"

        driver.get(store_url)

        # 설치하기 버튼 클릭
        INSTALL_BTN_XPATH = (
            '//*[@id="yDmH0d"]/c-wiz/div/div/main/div'
            '/section[1]/section/div/div[4]/div/div/button/span[4]'
        )
        try:
            install_btn = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, INSTALL_BTN_XPATH))
            )
            driver.execute_script("arguments[0].click();", install_btn)
            logger.info("설치하기 버튼 클릭.")
            # 확장 프로그램 추가 확인 팝업 → Enter 키로 처리 (네이티브 다이얼로그)
            time.sleep(1.5)
            try:
                import pyautogui
                pyautogui.press("tab")    # 취소 → 확장 프로그램 추가 버튼으로 이동
                time.sleep(0.3)
                pyautogui.press("enter")  # 확장 프로그램 추가 확인
                logger.info("확장 프로그램 추가 확인 팝업 Tab+Enter 처리.")
            except Exception as e:
                logger.warning("확인 팝업 Enter 실패: %s", e)
        except Exception as e:
            logger.warning("설치하기 버튼 클릭 실패: %s", e)

        # 설치 완료 대기 (최대 3분)
        try:
            WebDriverWait(driver, 180).until(
                lambda d: _is_cloudoc_installed_in_profile(profile_dir)
            )
            logger.info("ClouDoc 설치 완료.")
        except Exception:
            logger.warning("ClouDoc 설치 대기 시간 초과.")
        finally:
            driver.quit()
            time.sleep(1)

    # 정상 실행
    driver = webdriver.Chrome(service=service, options=_build_opts(visible=not offscreen))
    if not offscreen:
        driver.set_window_size(1400, 950)
    _active_driver = driver
    return driver


# ──────────────────────────────────────────────────────────────────────────────
# 로그인
# ──────────────────────────────────────────────────────────────────────────────
def _is_login_page(driver) -> bool:
    try:
        url = driver.current_url.lower()
        if any(kw in url for kw in ("login", "signin", "auth", "sso")):
            return True
        src = driver.page_source.lower()
        return 'type="password"' in src or 'id="password"' in src
    except Exception:
        return False


def _do_login(
    driver,
    wait,
    username: str,
    password: str,
    progress_cb: Optional[Callable[[str], None]] = None,
) -> None:
    """로그인 페이지에서 계정/비밀번호를 자동 입력하고 로그인한다."""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC

    def _log(msg: str) -> None:
        logger.info(msg)
        if progress_cb:
            progress_cb(msg)

    _log("로그인 정보 입력 중…")

    # 계정 입력
    account_field = wait.until(EC.presence_of_element_located((By.ID, "username")))
    account_field.click()
    account_field.clear()
    account_field.send_keys(username)

    # 비밀번호 입력
    pw_field = driver.find_element(By.ID, "password")
    pw_field.click()
    pw_field.clear()
    pw_field.send_keys(password)

    # 로그인 버튼 클릭
    login_btn = wait.until(EC.element_to_be_clickable((By.ID, "login_submit")))
    driver.execute_script("arguments[0].click();", login_btn)
    _log("로그인 버튼 클릭…")

    # 로그인 완료 대기 (최대 30초)
    for _ in range(30):
        time.sleep(1)
        if not _is_login_page(driver):
            _log("로그인 완료!")
            time.sleep(1)
            return

    raise RuntimeError(
        "로그인 실패. 계정/비밀번호를 확인하세요.\n"
        "로그인 페이지에서 벗어나지 못했습니다."
    )


# ──────────────────────────────────────────────────────────────────────────────
# 메인 자동화 함수
# ──────────────────────────────────────────────────────────────────────────────
def run_approval(
    xlsx_path: str,
    title: str,
    pr_no: str,
    sales_person: str,
    line: str,
    process: str,
    equipment: str,
    equip_model: str,
    remark: str,
    po_no: str = "",
    username: str = "",
    password: str = "",
    progress_cb: Optional[Callable[[str], None]] = None,
) -> None:
    """
    전자결재 구매요청서(NPN) 결재상신 자동화.

    Parameters
    ----------
    username / password : 그룹웨어 로그인 계정 (없으면 수동 로그인 대기)
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    def _log(msg: str) -> None:
        logger.info(msg)
        if progress_cb:
            progress_cb(msg)

    def _fill_by_id(field_id: str, value: str) -> None:
        if not value:
            return
        try:
            el = driver.find_element(By.ID, field_id)
            driver.execute_script("arguments[0].scrollIntoView(true);", el)
            el.click()
            el.clear()
            el.send_keys(value)
        except Exception as e:
            logger.warning("필드 입력 실패 (id=%s): %s", field_id, e)

    driver = _create_driver(offscreen=True)
    wait   = WebDriverWait(driver, 30)

    try:
        # ── STEP 1: 결재 홈 이동 ─────────────────────────────────────────────
        _log("① 전자결재 페이지 이동 중…")
        driver.get(APPROVAL_URL)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(2)

        # ── 로그인 처리 ───────────────────────────────────────────────────────
        if _is_login_page(driver):
            if username and password:
                _do_login(driver, wait, username, password, progress_cb)
                if APPROVAL_URL not in driver.current_url:
                    driver.get(APPROVAL_URL)
                    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                    time.sleep(2)
            else:
                _log("⚠️ 로그인이 필요합니다. Chrome에서 로그인해 주세요…")
                for _ in range(300):
                    time.sleep(1)
                    if not _is_login_page(driver):
                        _log("로그인 완료!")
                        break
                else:
                    raise TimeoutError("로그인 대기 시간 초과 (5분)")
                if APPROVAL_URL not in driver.current_url:
                    driver.get(APPROVAL_URL)
                    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                    time.sleep(2)

        # ── STEP 2: 새 결재 진행 버튼 클릭 ──────────────────────────────────
        _log("② 새 결재 진행 버튼 클릭…")
        btn_new = wait.until(EC.element_to_be_clickable((By.ID, "writeBtn")))
        driver.execute_script("arguments[0].click();", btn_new)
        time.sleep(1)

        # ── STEP 3: 구매요청서(NPN) 직접 선택 ──────────────────────────────
        _log(f"③ '{FORM_NAME}' 선택…")
        # gpopupLayer 모달이 완전히 로드될 때까지 대기
        wait.until(EC.visibility_of_element_located((By.ID, "gpopupLayer")))
        time.sleep(1)

        # Full XPath로 직접 클릭 (id 방식 폴백 포함)
        FORM_FULL_XPATH = (
            "/html/body/div[3]/div/div[2]/div[1]/div/div[1]"
            "/ul/li[1]/ul/li[7]/ul/li[10]/a"
        )
        try:
            item_el = wait.until(EC.element_to_be_clickable((By.XPATH, FORM_FULL_XPATH)))
            driver.execute_script("arguments[0].scrollIntoView(true);", item_el)
            driver.execute_script("arguments[0].click();", item_el)
        except Exception:
            # 폴백: id 속성으로 재시도
            item_el = wait.until(EC.element_to_be_clickable((By.ID, "FORM_8637")))
            driver.execute_script("arguments[0].scrollIntoView(true);", item_el)
            driver.execute_script("arguments[0].click();", item_el)
        time.sleep(0.5)

        # ── STEP 4: 확인 버튼 클릭 ──────────────────────────────────────────
        _log("④ 확인 버튼 클릭…")
        confirm_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='gpopupLayer']/footer/a[1]")
        ))
        driver.execute_script("arguments[0].click();", confirm_btn)

        # ── STEP 5: 팝업 창 전환 ────────────────────────────────────────────
        _log("⑤ 팝업 창 대기 중…")
        main_window = driver.current_window_handle
        wait.until(lambda d: len(d.window_handles) > 1)
        popup_handle = next(
            h for h in driver.window_handles if h != main_window
        )
        driver.switch_to.window(popup_handle)
        driver.set_window_position(100, 50)
        driver.set_window_size(1400, 900)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(2)

        # ── STEP 6: 제목 입력 ────────────────────────────────────────────────
        _log("⑥ 제목 입력…")
        title_input = wait.until(EC.presence_of_element_located((By.ID, "subject")))
        title_input.click()
        title_input.clear()
        title_input.send_keys(title)

        # ── STEP 7: 구매요청 양식 데이터 입력 ───────────────────────────────
        _log("⑦ 구매요청 양식 데이터 입력…")
        _fill_by_id("addContentTable_1_3",  pr_no)
        _fill_by_id("addContentTable_1_4",  po_no)
        _fill_by_id("addContentTable_1_5",  sales_person)
        _fill_by_id("addContentTable_1_6",  line)
        _fill_by_id("addContentTable_1_7",  process)
        _fill_by_id("addContentTable_1_8",  equipment)
        _fill_by_id("addContentTable_1_9",  equip_model)
        _fill_by_id("addContentTable_1_10", remark)

        # ── STEP 8: 파일 첨부 ────────────────────────────────────────────────
        _log("⑧ 파일 첨부 중…")
        try:
            drop_zone = driver.find_element(By.ID, "dropZone")
            file_input = drop_zone.find_element(By.XPATH, ".//input[@type='file']")
            driver.execute_script(
                "arguments[0].style.display='block';"
                "arguments[0].style.visibility='visible';"
                "arguments[0].style.opacity='1';",
                file_input,
            )
            file_input.send_keys(xlsx_path)
            time.sleep(2)
            _log("   파일 첨부 완료.")
        except Exception as e:
            logger.warning("파일 첨부 중 오류: %s", e)

        # ── STEP 9: 작성 완료 안내 ──────────────────────────────────────────
        _log("⑨ 구매요청서 작성 완료.")
        import tkinter as tk
        from tkinter import messagebox
        _root = tk.Tk()
        _root.withdraw()
        _root.attributes("-topmost", True)
        _root.lift()
        messagebox.showinfo(
            "구매요청서 작성완료",
            "구매요청서 작성완료!\n검토 후 결재요청 버튼을 누르세요.",
            parent=_root,
        )
        _root.destroy()

    except Exception as e:
        import traceback
        # 오류 발생 시에만 브라우저 종료
        try:
            driver.quit()
        except Exception:
            pass
        raise RuntimeError(
            f"결재상신 자동화 중 오류:\n{e}\n\n{traceback.format_exc()}"
        ) from e
    # finally 에서 driver.quit() 제거 → 브라우저를 열어둬서 사용자가 직접 결재요청


def _create_login_driver():
    """
    로그인 검증 전용 Chrome 드라이버.
    별도 임시 프로필을 사용하므로 자동화 프로필 잠금과 무관하게 동작한다.
    ClouDoc 설치 과정도 건너뛴다.
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
    except ImportError:
        raise RuntimeError("selenium 패키지가 없습니다.\npip install selenium webdriver-manager")

    try:
        from webdriver_manager.chrome import ChromeDriverManager
        service = Service(ChromeDriverManager().install())
    except Exception:
        service = Service()

    chrome_bin = _find_chrome_binary()
    if not chrome_bin:
        raise RuntimeError("Chrome 실행 파일을 찾을 수 없습니다.")

    # 로그인 전용 프로필 (자동화 프로필과 분리)
    appdata = os.environ.get("APPDATA") or os.path.expanduser("~")
    login_profile_dir = os.path.join(appdata, "quote-auto", "login_profile")
    os.makedirs(login_profile_dir, exist_ok=True)

    opts = Options()
    opts.binary_location = chrome_bin
    opts.add_argument(f"--user-data-dir={login_profile_dir}")
    opts.add_argument("--profile-directory=Default")
    opts.add_argument("--no-first-run")
    opts.add_argument("--no-default-browser-check")
    opts.add_argument("--disable-notifications")
    opts.add_argument("--disable-extensions")  # ClouDoc 불필요
    opts.add_argument("--headless=new")         # 완전 숨김 (창 없음)
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1280,800")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)

    return webdriver.Chrome(service=service, options=opts)


def check_login(
    username: str,
    password: str,
) -> tuple:
    """
    로그인 검증. 반환: (success: bool, message: str)
    별도 login_profile을 사용하므로 자동화 프로필 잠금과 충돌하지 않는다.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    driver = None
    try:
        driver = _create_login_driver()
        driver.get(APPROVAL_URL)
        wait = WebDriverWait(driver, 15)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(1)

        if _is_login_page(driver):
            _do_login(driver, wait, username, password)
            time.sleep(2)

            # 로그인 오류 메시지 확인
            try:
                err_el = driver.find_element(
                    By.XPATH,
                    '//*[@id="loginForm"]/section/div[3]/span[2]'
                )
                err_text = err_el.text.strip()
                if err_text:
                    return False, err_text
            except Exception:
                pass

            if _is_login_page(driver):
                return False, "로그인 실패"

        return True, ""
    except Exception as e:
        return False, str(e)
    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception:
                pass
        time.sleep(1)


# ──────────────────────────────────────────────────────────────────────────────
# Excel 데이터 읽기 (openpyxl)
# ──────────────────────────────────────────────────────────────────────────────
def read_rack_row2(xlsx_path: str) -> Dict[str, str]:
    """
    RACK발주양식 시트 2행에서 결재상신에 필요한 컬럼 값을 반환.
    반환 키: pr_no, sales_person, line, process, equipment, equip_model, remark
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["RACK발주양식"] if "RACK발주양식" in wb.sheetnames else wb.active
    row2 = ws[2]

    def _cell(col_idx: int) -> str:
        v = row2[col_idx - 1].value if col_idx <= len(row2) else None
        return str(v).strip() if v is not None else ""

    # Z열(26번째) 전체에서 비어있지 않은 셀 수 - 1 (헤더 제외)
    z_count = sum(
        1 for r in ws.iter_rows(min_col=26, max_col=26, values_only=True)
        if r[0] is not None and str(r[0]).strip() != ""
    ) - 1

    result = {
        "pr_no":        _cell(7),    # G열
        "sales_person": _cell(9),    # I열
        "line":         _cell(23),   # W열
        "process":      _cell(21),   # U열
        "equipment":    _cell(24),   # X열
        "equip_model":  _cell(25),   # Y열
        "remark":       f"세부List 유첨 (총 {max(z_count, 0)}건)",
    }
    wb.close()
    return result
