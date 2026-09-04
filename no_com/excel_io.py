"""
excel_io.py  [no-COM edition]
==============================
Excel 입출력 전담 모듈.

- openpyxl  : 읽기 + 쓰기 전담
              (파싱, 견적서/갑지 생성 모두 openpyxl)
- Excel COM : PDF 변환 전용 (excel_to_merged_pdf 만 사용)

핵심 전략
---------
템플릿의 이미지·서식은 shutil.copy2 로 원본 ZIP 구조를 통째로
복사한 뒤 openpyxl 로 셀 값만 수정한다.
수식 재계산은 wb.calculation.fullCalcOnLoad = True 를 설정해
파일을 열 때 Excel 이 자동 처리하도록 위임한다.
"""

from __future__ import annotations

import logging
import math
import os
import re
import shutil
import stat
import threading
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core import (
    CN, DOM, US, SheetName, QuoteState,
    ensure_dir, safe_filename, s, to_float, unique_path,
    exe_dir, parse_invest_info,
)

logger = logging.getLogger("QuoteApp")


# ──────────────────────────────────────────────
# 생성 중 출력 파일 추적 — 실패 시 미완성 파일 정리
# ──────────────────────────────────────────────
# 템플릿을 출력 경로로 복사한 뒤 편집 도중 예외가 나면, 가공되지 않은
# 템플릿 사본이 '대외비_….xlsx' 이름으로 출력 폴더에 남는다.
# 사용자가 그걸 갑지 원본으로 다시 선택하면 문제가 연쇄된다.
_inflight = threading.local()


def _copy_template(src: str, dst: str) -> str:
    """템플릿을 출력 경로로 복사하고 실패 시 정리 대상으로 등록한다."""
    shutil.copy2(src, dst)
    # copy2 는 읽기전용 속성까지 복사한다. 공유 드라이브의 마스터 템플릿이
    # 읽기전용이면 사본도 읽기전용이 되어 wb.save() 가 PermissionError 로 죽는다.
    try:
        os.chmod(dst, os.stat(dst).st_mode | stat.S_IWRITE)
    except OSError:
        pass
    _inflight.path = dst
    return dst


def _discard_inflight() -> None:
    """생성에 실패한 미완성 출력 파일을 삭제한다."""
    p = getattr(_inflight, "path", None)
    _inflight.path = None
    if p and os.path.exists(p):
        try:
            os.remove(p)
            logger.warning("생성 실패 — 미완성 파일 삭제: %s", p)
        except OSError as e:
            logger.warning("미완성 파일 삭제 실패 (%s): %s", p, e)


def _clear_inflight() -> None:
    """생성 성공 — 추적 해제."""
    _inflight.path = None


def inspect_template_losses(path: str) -> List[str]:
    """
    openpyxl 왕복(load→save) 시 소실되는 파트를 미리 찾아 사람이 읽을 목록으로 반환.

    openpyxl 은 아래를 보존하지 못하고 조용히 버린다. 생성된 견적서에서
    그림·머리글 로고·인쇄 설정이 사라져도 아무 오류가 나지 않으므로,
    통합양식을 읽는 시점에 미리 알려준다.

      - EMF/WMF 벡터 이미지 → 도형이 통째로 사라진다. 복원 불가.
                              엑셀에서 PNG 그림으로 교체해야 한다.

    ※ 머리글/바닥글 이미지(VML)도 openpyxl 이 버리지만
      restore_header_footer_images() 로 저장 후 되살리므로 경고하지 않는다.
    ※ 인쇄 설정은 용지·배율·방향·여백·printOptions 가 모두 보존된다.
      printerSettings*.bin(프린터 드라이버 고유값)만 빠지며 실사용에 영향 없다.

    반환: 경고 문구 리스트 (비어 있으면 소실 없음)
    """
    import zipfile
    from xml.etree import ElementTree as ET

    _RELNS = "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
    out: List[str] = []
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()

            # ① EMF/WMF 이미지 — 어느 시트에서 쓰이는지까지 찾아 알려준다
            vec = [n for n in names
                   if n.startswith("xl/media/") and n.lower().endswith((".emf", ".wmf"))]
            if vec:
                def _rels(p):
                    try:
                        return {r.get("Id"): (r.get("Target") or "")
                                for r in ET.fromstring(z.read(p)).iter(_RELNS)}
                    except Exception:
                        return {}
                wbrel = _rels("xl/_rels/workbook.xml.rels")
                sheets = re.findall(
                    r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"',
                    z.read("xl/workbook.xml").decode("utf8", "replace"))
                users = {}
                for sname, rid in sheets:
                    fn = (wbrel.get(rid) or "").split("/")[-1]
                    for tgt in _rels(f"xl/worksheets/_rels/{fn}.rels").values():
                        if "drawings/" not in tgt or not tgt.endswith(".xml"):
                            continue
                        dn = tgt.split("/")[-1]
                        for v in _rels(f"xl/drawings/_rels/{dn}.rels").values():
                            base = v.split("/")[-1]
                            if any(base == m.split("/")[-1] for m in vec):
                                users.setdefault(base, []).append(sname)
                for m in vec:
                    base = m.split("/")[-1]
                    where = ", ".join(users.get(base, [])) or "(사용 시트 불명)"
                    out.append(f"벡터 이미지 {base} → {where} 에서 사라집니다")


    except Exception as e:
        logger.warning("템플릿 소실 검사 실패 (%s): %s", path, e)
    return out


def check_pump_price_consistency(path: str) -> List[str]:
    """
    앱이 쓰는 '품목' 단가와 템플릿 '사양서'가 VLOOKUP 하는 'Pump 단가표' 단가를 대조.

    [왜 필요한가]
    PUMP 단가가 두 경로로 갈라져 있다.
      · 앱   : 품목 시트의 분류(A열)=Q-Code 로 찾은 행들을 STEP5 에 PUMP 로 추가
      · 템플릿: 사양서!F15 = VLOOKUP(A15,'Pump 단가표'!F:M,6,0)
    둘이 어긋나면 인쇄되는 사양서 금액과 갑지 금액이 조용히 달라진다.
    (ver.1.8 실측: 102개 중 3개 불일치, 최대 2,840만원/CH 차이)

    반환: 불일치 설명 리스트 (비어 있으면 정상)
    """
    out: List[str] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if "품목" not in wb.sheetnames or "Pump 단가표" not in wb.sheetnames:
                return out

            by_class: Dict[str, List[float]] = {}
            for a, _b, c in wb["품목"].iter_rows(max_col=3, values_only=True):
                key = s(a)
                if key:
                    by_class.setdefault(key, []).append(to_float(c))

            # Pump 단가표: C열=자재코드, F열=모델명, K열=단가
            # 템플릿 수식 체인을 그대로 흉내낸다:
            #   A15 = VLOOKUP(Q-Code, C:F, 4, 0)   → 모델명 (첫 일치 행)
            #   F15 = VLOOKUP(A15,    F:M, 6, 0)   → 단가   (첫 일치 행)
            # VLOOKUP 은 첫 일치만 쓰므로 중복 행이 있어도 첫 행 기준으로 본다.
            code_to_model: Dict[str, str] = {}
            model_to_price: Dict[str, float] = {}
            for row in wb["Pump 단가표"].iter_rows(max_col=11, values_only=True):
                code, model = s(row[2]), s(row[5])
                if code and code.upper().startswith("Q") and code not in code_to_model:
                    code_to_model[code] = model
                if model and model not in model_to_price:
                    model_to_price[model] = to_float(row[10])

            for code, model in code_to_model.items():
                if code not in by_class:
                    continue
                tbl = model_to_price.get(model, 0.0)
                app = sum(by_class[code])
                if abs(app - tbl) < 1:
                    continue
                why = (f"품목 시트에 {len(by_class[code])}행이 있어 합산됨"
                       if len(by_class[code]) > 1 else "단가가 서로 다름")
                out.append(
                    f"{code}({model}): 사양서 {tbl:,.0f} vs 앱 {app:,.0f} "
                    f"(차이 {app - tbl:+,.0f}) — {why}"
                )
        finally:
            wb.close()
    except Exception as e:
        logger.warning("PUMP 단가 대조 실패 (%s): %s", path, e)
    return out


def _clear_clipboard() -> None:
    """
    Windows 클립보드를 강제로 비운다.

    CopyPicture()/ExportAsFixedFormat() 후 COM 세션이 종료되면
    클립보드에 죽은 프로세스 소유의 CF_ENHMETAFILE(EMF) 핸들이 남는다.
    이 stale 핸들이 있는 상태에서 별도 Excel 창이 Ctrl+C를 시도하면
    Windows가 이전 클립보드를 초기화하다가 액세스 위반 → Excel 전체 종료.
    COM 세션 종료 전, 또는 CopyPicture 직후에 반드시 호출한다.
    """
    try:
        import ctypes
        u32 = ctypes.windll.user32
        if u32.OpenClipboard(0):
            u32.EmptyClipboard()
            u32.CloseClipboard()
    except Exception:
        pass


# ── COM 지연 로딩 (PDF 변환 전용) ───────────────────────────────
# 모듈 임포트 시 DLL LoadLibrary를 호출하지 않도록 지연.
# GIL 점유로 인한 시작 시 UI 동결 방지 (AV/Defender 환경에서 10~120초 블로킹 가능).
pythoncom: Any = None
win32: Any = None
COM_AVAILABLE: Optional[bool] = None  # None=미확인, True=사용가능, False=불가


def _ensure_com() -> bool:
    """COM DLL을 최초 필요 시점에만 로드한다."""
    global COM_AVAILABLE, pythoncom, win32
    if COM_AVAILABLE is not None:
        return COM_AVAILABLE
    try:
        import pythoncom as _pc
        import win32com.client as _w32
        pythoncom = _pc
        win32 = _w32
        COM_AVAILABLE = True
    except Exception:
        pythoncom = None
        win32 = None
        COM_AVAILABLE = False
    return COM_AVAILABLE


# ═══════════════════════════════════════════════════════════════
# 읽기 (openpyxl) — 기존과 동일
# ═══════════════════════════════════════════════════════════════

def parse_items_sheet(ws: Worksheet) -> Tuple[
    List[Dict[str, Any]],
    Dict[str, List[Dict[str, Any]]],
    Dict[str, float],
    Dict[str, int],
]:
    """품목 시트 파싱. 반환: (rows, by_class, price_by_spec, spec_order_index)"""
    # iter_rows(values_only=True) 사용: ws.cell(r,c) 임의 접근은 read_only
    # 워크시트에서 매 호출마다 스트림을 처음부터 다시 훑어 O(n²)이 된다
    # (3천 행 기준 100배+ 저하 실측). read_only=True 로 여는 호출부와 짝을 이룬다.
    def _is_header(a: str, b: str, c: str) -> bool:
        h = (a + b + c).replace(" ", "").lower()
        return "분류" in h and ("단가" in h or "가격" in h)

    row_iter = ws.iter_rows(max_col=3, values_only=True)
    try:
        r1a, r1b, r1c = next(row_iter)
    except StopIteration:
        return [], {}, {}, {}
    header = _is_header(s(r1a), s(r1b), s(r1c))
    if not header:
        row_iter = ws.iter_rows(max_col=3, values_only=True)   # 1행부터 다시

    rows: List[Dict[str, Any]] = []
    by_class: Dict[str, List[Dict[str, Any]]] = {}
    price_by_spec: Dict[str, float] = {}
    order_index: Dict[str, int] = {}
    order = 0

    for a_raw, b_raw, c in row_iter:
        a = s(a_raw)
        b = s(b_raw)
        if not (a or b or c):
            continue
        price = to_float(c)
        row = {"A": a, "B": b, "C": price}
        rows.append(row)
        if b and b not in order_index:
            order_index[b] = order
            order += 1
        if a:
            by_class.setdefault(a, []).append(row)
        if b:
            price_by_spec[b] = price

    return rows, by_class, price_by_spec, order_index


def parse_code_map_sheet(ws: Worksheet) -> Dict[Tuple[str, str, str], List[Tuple[str, float]]]:
    """코드매핑 시트 파싱. 반환: {(공정, 설비사, 5D): [(ACC, 수량), ...]}"""
    def _is_header(a, b, c, d, e) -> bool:
        h = (a + b + c + d + e).replace(" ", "").lower()
        return "공정" in h and "설비" in h and ("acc" in h or "수량" in h or "5d" in h)

    row_iter = ws.iter_rows(max_col=5, values_only=True)
    try:
        r1 = next(row_iter)
    except StopIteration:
        return {}
    if not _is_header(*(s(v) for v in r1)):
        row_iter = ws.iter_rows(max_col=5, values_only=True)   # 1행부터 다시

    cmap: Dict[Tuple[str, str, str], List[Tuple[str, float]]] = {}
    for proc_raw, vendor_raw, code5d_raw, acc_raw, qty_raw in row_iter:
        proc   = s(proc_raw)
        vendor = s(vendor_raw)
        code5d = s(code5d_raw)
        acc    = s(acc_raw)
        qty    = to_float(qty_raw)
        if proc and vendor and code5d and acc:
            cmap.setdefault((proc, vendor, code5d), []).append((acc, qty))
    return cmap


def append_code_map_rows(
    template_path: str,
    process: str,
    vendor: str,
    code5d: str,
    items: List[Tuple[str, float]],
) -> Tuple[List[Tuple[str, float]], List[Tuple[str, float]]]:
    """
    통합양식(마스터 템플릿) 코드매핑 시트에 (공정,설비사,5D,품목,수량) 행을 추가한다.
    items 중 (공정,설비사,5D,품목,수량)이 완전히 동일한 행이 이미 있으면 그 항목은
    건너뛴다. 새로 추가할 행이 하나도 없으면(전부 중복) 파일을 저장하지 않는다.

    template_path 는 여러 견적서가 공유하는 마스터 파일이라 직접 덮어쓴다.
    openpyxl 저장은 머리글/바닥글 그림(VML)을 읽지도 쓰지도 못해 저장 시
    사라지므로, 저장 전에 만든 백업으로 restore_header_footer_images 를 호출해
    되살린다 — 이 백업은 복원용인 동시에 만약을 대비한 안전장치이기도 하다.

    반환: (added, skipped) — 실제로 추가된 (품목,수량) 목록 / 중복이라 건너뛴 목록.
    """
    wb = load_workbook(template_path)
    ws = wb[SheetName.CODE_MAP]
    existing = parse_code_map_sheet(ws)
    key = (process, vendor, code5d)
    seen = set(existing.get(key, []))

    added: List[Tuple[str, float]] = []
    skipped: List[Tuple[str, float]] = []
    for spec, qty in items:
        pair = (spec, qty)
        if pair in seen:
            skipped.append(pair)
            continue
        added.append(pair)
        seen.add(pair)   # 이번 호출 안에서의 자체 중복도 방지

    if not added:
        wb.close()
        return added, skipped

    backup_path = unique_path(
        f"{os.path.splitext(template_path)[0]}"
        f".backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        f"{os.path.splitext(template_path)[1]}"
    )
    shutil.copy2(template_path, backup_path)

    row = ws.max_row + 1
    for spec, qty in added:
        qty_val = int(qty) if float(qty).is_integer() else qty
        ws.cell(row, 1, process)
        ws.cell(row, 2, vendor)
        ws.cell(row, 3, code5d)
        ws.cell(row, 4, spec)
        ws.cell(row, 5, qty_val)
        row += 1

    wb.save(template_path)
    wb.close()
    restore_header_footer_images(backup_path, template_path)
    return added, skipped


def parse_request_xlsx(path: str) -> Tuple[str, List[Dict[str, Any]]]:
    """
    의뢰파일 파싱.
    반환: (active_sheet_name, row_dicts)
    row_dict keys: row_idx, D, E, F, G, H, J, K, N, V, X, Y, Z, AA
    """
    # read_only=True + iter_rows(): ws.cell(r,c) 임의 접근을 read_only 모드에서
    # 쓰면 매 호출이 스트림을 처음부터 다시 훑어 O(n²)이 된다(실측: 3천 행에서
    # 100배+ 저하). 의뢰파일은 수백~수천 행이 흔하므로 반드시 iter_rows.
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    sheet_name = ws.title
    NCOL = 27  # AA

    def _is_header(r1: tuple) -> bool:
        checks = [
            s(r1[25]).replace(" ", "").lower(),   # Z (26)
            s(r1[10]).replace(" ", "").lower(),   # K (11)
            s(r1[23]).replace(" ", "").lower(),   # X (24)
            s(r1[21]).replace(" ", "").lower(),   # V (22)
        ]
        return any("설비" in c or "공정" in c or "5d" in c or "코드" in c for c in checks)

    row_iter = ws.iter_rows(max_col=NCOL, values_only=True)
    try:
        first = next(row_iter)
    except StopIteration:
        wb.close()
        return sheet_name, []
    start = 2 if _is_header(first) else 1
    if start == 1:
        row_iter = ws.iter_rows(max_col=NCOL, values_only=True)   # 1행부터 다시

    rows: List[Dict[str, Any]] = []
    for idx, vals in enumerate(row_iter, start=start):
        z = vals[25]  # Z (26)
        if s(z) == "":
            continue
        rows.append({
            "row_idx": idx,
            "D": vals[3],   "E": vals[4],   "F": vals[5],
            "G": vals[6],   "H": vals[7],   "J": vals[9],
            "K": vals[10],  "N": vals[13],  "R": vals[17],
            "V": vals[21],  "X": vals[23],  "Y": vals[24],
            "Z": z,         "AA": vals[26],
        })
    wb.close()
    return sheet_name, rows


# ═══════════════════════════════════════════════════════════════
# openpyxl 공통 헬퍼
# ═══════════════════════════════════════════════════════════════

def _hide_rows(ws: Worksheet, start: int, end: int, filled_count: int) -> None:
    """start~end 행 중 filled_count 이후 행을 숨김."""
    for i in range(end - start + 1):
        ws.row_dimensions[start + i].hidden = (i >= filled_count)


# 생성물에서 삭제하는 내부 참조 시트.
#
# veryHidden 은 '숨김'일 뿐 파일에는 그대로 남는다. 엑셀에서 숨기기 해제하거나
# 압축을 풀면 그대로 읽히므로, 대외비로 나가는 견적서에 전사 단가표가 실려 나간다.
# 아래 세 시트는 앱이 STEP1 에서 읽어 쓰는 입력 자료일 뿐 생성물에는 불필요하고,
# 통합양식 ver.1.8 기준으로 어떤 수식·정의된이름·데이터유효성·조건부서식도
# 이들을 참조하지 않는 것을 확인했다.
#
# ※ 'Pump 단가표' / '악세서리 단가표' 는 사양서·입고검수확인서 수식이 VLOOKUP 으로
#    참조하므로 절대 삭제하면 안 된다(삭제 시 #REF!).
_DROP_SHEETS = ("품목", "코드매핑", "용량 및 등급")


_VML_RELTYPE = ("http://schemas.openxmlformats.org/officeDocument/2006/"
                "relationships/vmlDrawing")


def restore_header_footer_images(src_template: str, out_path: str) -> int:
    """
    openpyxl 이 버린 머리글/바닥글 이미지(VML)를 원본 템플릿에서 되살린다.

    openpyxl 은 legacyDrawingHF(머리글/바닥글 그림)를 읽지도 쓰지도 못한다.
    머리글 텍스트의 '&G'(그림 삽입) 지시자는 남는데 그림 파트가 사라져
    인쇄물에서 로고가 빠진다. 저장이 끝난 파일에 원본의 VML 파트와
    참조만 되살린다(셀 내용은 건드리지 않는다).

    반환: 복원한 시트 수
    """
    import zipfile
    from xml.etree import ElementTree as ET
    _RELNS = "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"

    def _rels(z, p):
        try:
            return [(r.get("Id"), r.get("Type"), r.get("Target"))
                    for r in ET.fromstring(z.read(p)).iter(_RELNS)]
        except Exception:
            return []

    def _sheetfiles(z):
        wbrel = {i: t for i, _t, t in _rels(z, "xl/_rels/workbook.xml.rels")}
        return {n: (wbrel.get(r) or "").split("/")[-1] for n, r in
                re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"',
                           z.read("xl/workbook.xml").decode("utf8", "replace"))}

    try:
        with zipfile.ZipFile(src_template) as zs:
            src_names = set(zs.namelist())
            want = {}
            for name, fn in _sheetfiles(zs).items():
                for _id, ty, tgt in _rels(zs, f"xl/worksheets/_rels/{fn}.rels"):
                    if ty != _VML_RELTYPE:
                        continue
                    vml = tgt.split("/")[-1]
                    parts = {f"xl/drawings/{vml}": zs.read(f"xl/drawings/{vml}")}
                    rp = f"xl/drawings/_rels/{vml}.rels"
                    if rp in src_names:
                        parts[rp] = zs.read(rp)
                        for _i, _t, mt in _rels(zs, rp):
                            m = "xl/" + mt.replace("../", "")
                            if m in src_names:
                                parts[m] = zs.read(m)
                    want[name] = (vml, parts)
        if not want:
            return 0

        with zipfile.ZipFile(out_path) as zo:
            omap = _sheetfiles(zo)
            data = {n: zo.read(n) for n in zo.namelist()}

        added = 0
        for name, (vml, parts) in want.items():
            fn = omap.get(name)
            sheet = f"xl/worksheets/{fn}" if fn else None
            if not sheet or sheet not in data:
                continue           # 삭제된 시트는 건너뛴다
            data.update(parts)

            rp = f"xl/worksheets/_rels/{fn}.rels"
            cur = data.get(
                rp,
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                b'<Relationships xmlns="http://schemas.openxmlformats.org/'
                b'package/2006/relationships"></Relationships>').decode("utf8")
            used = set(re.findall(r'Id="(rId\d+)"', cur))
            rid = next(f"rId{i}" for i in range(1, 10000) if f"rId{i}" not in used)
            data[rp] = cur.replace(
                "</Relationships>",
                f'<Relationship Id="{rid}" Type="{_VML_RELTYPE}" '
                f'Target="../drawings/{vml}"/></Relationships>').encode("utf8")

            x = data[sheet].decode("utf8")
            if "<legacyDrawingHF" not in x:
                # openpyxl 은 r: 네임스페이스를 선언하지 않는 시트가 있어
                # 그냥 r:id 를 쓰면 'unbound prefix' 로 파일이 깨진다 → 인라인 선언.
                tag = (f'<legacyDrawingHF xmlns:r="http://schemas.openxmlformats.org'
                       f'/officeDocument/2006/relationships" r:id="{rid}"/>')
                # 스키마 순서: drawing → legacyDrawing → legacyDrawingHF
                m = re.search(r'<legacyDrawing\b[^>]*/>|<drawing\b[^>]*/>', x)
                x = (x[:m.end()] + tag + x[m.end():]) if m else \
                    x.replace("</worksheet>", tag + "</worksheet>")
                data[sheet] = x.encode("utf8")
            added += 1

        ct = data["[Content_Types].xml"].decode("utf8")
        for ext, mime in (
            ("vml", "application/vnd.openxmlformats-officedocument.vmlDrawing"),
            ("jpeg", "image/jpeg"), ("jpg", "image/jpeg"), ("png", "image/png"),
            ("gif", "image/gif"), ("emf", "image/x-emf"),
        ):
            if f'Extension="{ext}"' not in ct:
                ct = ct.replace(
                    "</Types>",
                    f'<Default Extension="{ext}" ContentType="{mime}"/></Types>')
        data["[Content_Types].xml"] = ct.encode("utf8")

        tmp = out_path + ".tmp"
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zw:
            for n, d in data.items():
                zw.writestr(n, d)
        os.replace(tmp, out_path)
        return added
    except Exception as e:
        # 복원은 부가 기능이다. 실패해도 생성물 자체는 유효하므로 로그만 남긴다.
        logger.warning("머리글/바닥글 이미지 복원 실패 (%s): %s", out_path, e)
        try:
            if os.path.exists(out_path + ".tmp"):
                os.remove(out_path + ".tmp")
        except OSError:
            pass
        return 0


def _drop_internal_sheets(wb) -> None:
    """대외비 유출·용량 축소를 위해 내부 참조 시트를 삭제한다."""
    for name in _DROP_SHEETS:
        if name in wb.sheetnames:
            try:
                del wb[name]
            except Exception as e:
                logger.warning("시트 삭제 실패 (%s): %s", name, e)


def _show_only_sheets(wb, keep: List[str]) -> None:
    """keep 목록에 없는 시트를 veryHidden으로 설정."""
    keep_set = set(keep)
    for name in wb.sheetnames:
        wb[name].sheet_state = "visible" if name in keep_set else "veryHidden"


def _reset_sheet_selections(wb) -> None:
    """
    모든 시트의 SheetView selection을 A1 단일 셀로 초기화한다.

    [문제]
    openpyxl은 템플릿에 저장된 selection.sqref 값을 그대로 보존한다.
    템플릿이 다중 비연속 범위가 선택된 채로 저장되어 있으면(예: "A1 C3:D5"),
    생성된 xlsx도 동일한 상태를 갖게 된다.
    Excel이 이 파일을 열면 다중 선택 상태가 복원되어:
      - 복사·붙여넣기: "이 작업은 다중선택 범위에서 작동하지 않습니다"
      - 시트 삽입·삭제: 차단
      - 시트 이동/복사: Excel 비정상 종료 (PrintArea 정의명과 충돌)

    [수정]
    저장 직전에 모든 시트의 selection 을 단일 A1 으로 강제 초기화한다.
    """
    from openpyxl.worksheet.views import Selection
    for ws in wb.worksheets:
        sv = ws.sheet_view
        if not sv.selection:
            sv.selection.append(Selection(activeCell="A1", sqref="A1"))
        else:
            first = sv.selection[0]
            first.activeCell = "A1"
            first.sqref = "A1"
            sv.selection[:] = [first]


def _write_req_row_openpyxl(wb_copy, state: QuoteState, rd: Dict[str, Any],
                             req_ws_cache: Optional[Any] = None) -> None:
    """
    의뢰파일의 특정 행을 견적의뢰복사본 시트 2행에 복사 (openpyxl).
    req_ws_cache: 외부에서 미리 열어 전달된 Worksheet 객체(멀티 생성 시 재사용).
                  None이면 여기서 직접 open/close 한다.
    """
    req_path = state.request_path
    if not req_path or not req_path.lower().endswith(".xlsx"):
        return
    try:
        row_idx = int(rd.get("row_idx", 0))
    except (TypeError, ValueError):
        return
    if row_idx <= 0:
        return

    def _copy_row(req_ws) -> None:
        copy_ws = wb_copy[SheetName.REQ_COPY]
        for c in range(1, 15):
            copy_ws.cell(2, c).value = None
        copy_ws.cell(2, 18).value = None
        for c in range(23, 28):
            copy_ws.cell(2, c).value = None
        for c in range(1, 15):
            copy_ws.cell(2, c).value = req_ws.cell(row_idx, c).value
        copy_ws.cell(2, 18).value = req_ws.cell(row_idx, 18).value
        for c in range(23, 28):
            copy_ws.cell(2, c).value = req_ws.cell(row_idx, c).value

    if req_ws_cache is not None:
        _copy_row(req_ws_cache)
        return

    req_wb = load_workbook(req_path, data_only=True)
    try:
        sheet_name = state.request_sheet_name
        req_ws = (
            req_wb[sheet_name]
            if sheet_name and sheet_name in req_wb.sheetnames
            else req_wb.active
        )
        _copy_row(req_ws)
    finally:
        req_wb.close()


# ═══════════════════════════════════════════════════════════════
# 견적서 생성 (openpyxl 기반, COM 없음)
# ═══════════════════════════════════════════════════════════════

def _pin_images(wb) -> None:
    """
    이미지 바이트를 메모리에 고정해 같은 워크북을 여러 번 저장할 수 있게 한다.

    openpyxl 은 첫 save 때 이미지 원본 파일 핸들을 닫아버리므로, 두 번째
    save 에서 'ValueError: I/O operation on closed file' 로 죽는다.
    일괄 생성에서 워크북을 재사용하려면 반드시 선행해야 한다.
    """
    import io
    for ws in wb.worksheets:
        for img in getattr(ws, "_images", []):
            try:
                data = img._data()
            except Exception:
                continue
            img.ref = io.BytesIO(data)
            img._data = lambda _d=data: _d


def open_template_for_reuse(template_path: str):
    """일괄 생성용 — 템플릿을 1회만 읽어 재사용 가능한 워크북으로 돌려준다."""
    wb = load_workbook(template_path)
    _pin_images(wb)
    return wb


def _fill_domestic(state: QuoteState,
                   rd: Dict[str, Any],
                   items: List[Dict[str, Any]],
                   req_ws_cache: Optional[Any] = None,
                   wb_cache: Optional[Any] = None) -> str:
    """
    국내 견적서: 템플릿을 복사한 뒤 openpyxl 로 데이터 기입.
    반환: 저장된 xlsx 경로.

    wb_cache: open_template_for_reuse() 로 만든 워크북. 주면 건마다 템플릿을
              다시 읽지 않고 재사용한다(일괄 생성에서 60% 이상 단축).
              건별 상태는 아래 ②③④⑤ 에서 전부 덮어쓰므로 누수되지 않지만,
              도중에 예외가 나면 워크북이 더러워지므로 호출자가 다시 열어야 한다.
    """
    ymd        = datetime.now().strftime("%y%m%d")
    pr         = safe_filename(s(rd.get("D")))
    itemno     = safe_filename(s(rd.get("E")))
    lineproc   = safe_filename(s(rd.get("K")))
    investor_fn = safe_filename(s(rd.get("J")))
    tool       = safe_filename(s(rd.get("Z")))
    folder = ensure_dir(
        os.path.join(exe_dir(), "견적서", f"{ymd}_LOT베큠_{lineproc}_{investor_fn}")
    )
    path = unique_path(
        os.path.join(folder, f"대외비_{pr}-{itemno}_{ymd}_LOT베큠_{lineproc}_{investor_fn}_{tool}.xlsx")
    )

    # ① 템플릿 준비 — 재사용 워크북이 있으면 복사·로드를 건너뛴다
    if wb_cache is not None:
        wb = wb_cache
        _inflight.path = path      # 저장 중 실패 시 정리 대상으로 등록
    else:
        # 템플릿 전체를 복사 — 이미지·서식·수식 구조 모두 보존
        _copy_template(state.template_path, path)
        wb = load_workbook(path)

    ws_spec = wb[SheetName.SPEC]
    ws_sign = wb[SheetName.SIGN_SPEC]
    ws_incoming = wb[SheetName.INCOMING]

    # ② 의뢰파일 행 → 견적의뢰복사본 시트 복사
    _write_req_row_openpyxl(wb, state, rd, req_ws_cache)

    # ③ 투자자명
    investor = s(state.investor_name) or "채승철"
    ws_spec["B4"] = f"{investor} 님 / 설비구매그룹"

    # ③-1 옵션에서 지정한 설비수량/실반입라인/Exh Size — 값이 있을 때만 기입.
    #     0/빈 문자열(=옵션 미설정)이면 템플릿 원본 셀을 건드리지 않는다.
    if state.equip_qty:
        ws_spec["A43"] = int(state.equip_qty)
    if state.actual_line:
        ws_sign["B6"] = state.actual_line
    if state.exh_size:
        ws_incoming["D27"] = int(state.exh_size)

    rack_items = [r for r in items if r["cat"] != "PUMP" and r["role"] != "CREDIT"]
    credits    = [r for r in items if r["role"] == "CREDIT"]

    # ④ 기존 데이터 클리어
    for rr in range(DOM.SPEC_START, DOM.SPEC_END + 1):
        for col in (DOM.COL_SPEC, DOM.COL_QTY, DOM.COL_PRICE, DOM.COL_AMT):
            ws_spec[f"{col}{rr}"] = None

    # ⑤ 품목 기입
    #
    # [템플릿(ver.1.8) 실제 수식 구조]
    #     사양서!E15 = 견적의뢰복사본!H2                     (수량)
    #     사양서!F15 = VLOOKUP(A15,'Pump 단가표'!F:M,6,0)    (PUMP CH당 단가)
    #     사양서!G15 = E15*F15                              ← PUMP 총액
    #     사양서!G16 = SUM(G17:G41)                         ← Rack system 합계
    #     사양서!F43 = ROUNDDOWN((G16+G15)/견적의뢰복사본!H2,-3)
    #
    # [Pump Credit 처리]
    # Pump Credit 을 G17:G41 에 실금액으로 넣으면 G16(=Rack system 합계)에서
    # 차감되어, PUMP 값이어야 할 금액이 Rack 쪽에서 빠지는 것처럼 보인다.
    # 그래서 PUMP 는 credit 반영된 CH당 단가를 F15 에 직접 기입해 G15 에서
    # 차감되도록 하고, 아래 라인은 표시 전용(G=0)으로 남긴다.
    #   · Rack Credit → 기존대로 G열에 실금액 (G16 에서 차감)
    #   · Pump Credit → F15 하드코딩 + 표시행 G=0 (이중 차감 방지)
    pump_credits = [r for r in credits if "Pump" in r.get("spec", "")]
    rack_credits = [r for r in credits if "Pump" not in r.get("spec", "")]

    pump_items      = [r for r in items if r["cat"] == "PUMP" and r["role"] != "CREDIT"]
    pump_spec       = pump_items[0]["spec"] if pump_items else None
    pump_amt        = sum(r["amt"] for r in pump_items)
    rack_amt        = sum(r["amt"] for r in rack_items)
    pump_credit_amt = sum(c["amt"] for c in pump_credits)   # 통상 음수
    rack_credit_amt = sum(c["amt"] for c in rack_credits)   # 통상 음수

    h_qty = to_float(rd.get("H"))
    if h_qty <= 0:
        h_qty = 1.0

    # PUMP CH당 단가 — credit 반영 후 정수(원)로 확정한다.
    # 소수점이 남으면 견적서에 13,900,000.333 같은 값이 찍히고,
    # G15=E15*F15 결과도 원 단위로 떨어지지 않는다.
    p_unit_net = round((pump_amt + pump_credit_amt) / h_qty)
    # 표시용 CH당 Credit 단가 (음수)
    _pump_credit_unit = round(pump_credit_amt / h_qty) if pump_credits else 0

    out_list = rack_items[:25]
    for cr in rack_credits:
        out_list.append({"spec": cr["spec"], "qty": 0.0, "price": 0.0, "amt": cr["amt"]})
    for cr in pump_credits:
        # 표시 전용: CH당 단가는 F열에, 금액(G열)은 0 — G16 에 영향 없음
        out_list.append({
            "spec": cr["spec"], "qty": 0.0,
            "price": _pump_credit_unit, "amt": 0, "display_only": True,
        })
    out_list = out_list[:25]

    # 템플릿 용량 초과 시 조용히 잘리면 사양서 합계와 갑지 금액이 어긋난다.
    if len(rack_items) + len(credits) > 25:
        raise ValueError(
            f"사양서 품목 칸(25행)을 초과했습니다: "
            f"품목 {len(rack_items)}건 + Credit {len(credits)}건. "
            f"품목을 줄이거나 템플릿 행을 늘려주세요."
        )

    def _won(v) -> int:
        """금액은 원 단위 정수로 — 소수점이 찍히지 않게 한다."""
        return int(round(to_float(v)))

    for i, r in enumerate(out_list):
        rr = DOM.SPEC_START + i
        ws_spec[f"{DOM.COL_SPEC}{rr}"] = r["spec"]
        if r.get("display_only"):
            # Pump Credit 표시행: CH당 단가만 보여주고 금액은 0
            # (F15 에서 이미 차감했으므로 G16 에 또 넣으면 이중 차감)
            ws_spec[f"{DOM.COL_QTY}{rr}"]   = None
            ws_spec[f"{DOM.COL_PRICE}{rr}"] = _won(r["price"])
            ws_spec[f"{DOM.COL_AMT}{rr}"]   = 0
        elif r.get("qty") == 0.0 and r.get("price") == 0.0 and "amt" in r:
            ws_spec[f"{DOM.COL_AMT}{rr}"] = _won(r["amt"])
        else:
            ws_spec[f"{DOM.COL_QTY}{rr}"]   = r["qty"]
            ws_spec[f"{DOM.COL_PRICE}{rr}"]  = _won(r["price"])
            ws_spec[f"{DOM.COL_AMT}{rr}"]    = _won(r["amt"])

    # ⑤-b PUMP CH당 단가를 F15 에 직접 기입 — Pump Credit 을 G15 에서 차감시킨다.
    # 템플릿 F15 는 'Pump 단가표' VLOOKUP 수식이지만, credit 이 있으면 그 값을
    # 그대로 두면 credit 이 반영되지 않는다. 값으로 덮어써 G15=E15*F15 가
    # net 금액이 되게 한다. credit 이 없으면 수식을 그대로 둔다.
    if pump_credits:
        ws_spec[f"{DOM.COL_PRICE}15"] = p_unit_net

    # ⑤-a 견적의뢰복사본 수식 셀 → 계산값 직접 기입 ─────────────────────────
    # openpyxl 저장 시 수식 캐시(cached value)가 소멸된다.
    # generate_cover 에서 data_only=True 로 읽으면 수식 셀(O/P/Q/S/T/U)이 None 반환 →
    # 갑지DATA 해당 열 공백. 계산값을 값(value)으로 직접 써서 이 경로를 막는다.
    #
    # 재현 수식:
    #   S2 = 사양서!F43,  T2 = S2*H2,  U2 = ROUNDUP((P2*H2+Q2)/H2,-3)
    #
    # 아래 값들은 Excel 이 재계산할 값과 반드시 일치해야 한다.
    # 어긋나면 인쇄되는 사양서와 갑지 금액이 서로 다른 견적서가 나간다.
    copy_ws = wb[SheetName.REQ_COPY]

    # 템플릿 수식 그대로 재현한다. Excel 이 재계산할 값과 어긋나면
    # 인쇄되는 사양서와 갑지 금액이 서로 다른 견적서가 나간다.
    #   G15 = E15*F15  = H2 × F15
    #        · Pump Credit 있음 → F15 는 위에서 기입한 p_unit_net (credit 반영)
    #        · 없음            → F15 는 Pump 단가표 VLOOKUP (= pump 단가)
    #   G16 = SUM(G17:G41) = rack_items + Rack Credit  (Pump Credit 은 0 으로 표시만)
    #   F43 = ROUNDDOWN((G16+G15)/H2, -3)
    #
    # ※ p_unit_net 은 이미 정수로 반올림했으므로 g15 도 그 값에서 직접 만든다.
    #   (A/h)*h 식으로 되돌리면 부동소수점 1 ulp 오차로 floor 가 1,000원
    #   낮게 떨어질 수 있다.
    if pump_credits:
        p_unit = p_unit_net                 # F15 에 실제로 쓴 값
        g15    = p_unit_net * h_qty         # Excel 의 E15*F15 와 정확히 동일
    else:
        p_unit = round(pump_amt / h_qty) if h_qty else 0
        g15    = pump_amt

    g16   = rack_amt + rack_credit_amt      # Pump Credit 은 G열 0 이라 미포함
    q_val = g16                             # Q2 = 사양서!G16

    # S = 사양서!F43 = ROUNDDOWN((G15+G16)/H2, -3)
    s_raw = (g15 + g16) / h_qty
    s_val = math.floor(s_raw / 1000) * 1000   # ROUNDDOWN(..., -3)
    # T = S2 * H2
    t_val = s_val * h_qty
    # U = ROUNDUP((P2*H2+Q2)/H2, -3) — 템플릿 수식 그대로 (Q2 경유로 credit 반영)
    u_raw = (p_unit * h_qty + q_val) / h_qty
    u_val = math.ceil(u_raw / 1000) * 1000    # ROUNDUP(..., -3)

    # 금액은 모두 원 단위 정수로 기입한다 — 소수점이 남으면 갑지·견적서에
    # 13,900,000.333 같은 값이 찍힌다.
    copy_ws.cell(2, 15).value = pump_spec         # O: PUMP 메인모듈 (=사양서!A15)
    copy_ws.cell(2, 16).value = int(p_unit)       # P: PUMP CH당 단가 (=사양서!F15)
    copy_ws.cell(2, 17).value = int(q_val)        # Q: Rack 합계 (=사양서!G16)
    copy_ws.cell(2, 19).value = int(s_val)        # S: 견적단가 (=사양서!F43)
    copy_ws.cell(2, 20).value = int(t_val)        # T: 견적금액 (=S2*H2)
    copy_ws.cell(2, 21).value = int(u_val)        # U: 견적단가(Check)

    # ⑥ 수식 재계산을 파일 열 때 Excel 에 위임
    wb.calculation.fullCalcOnLoad = True

    # ⑦ 빈 행 숨기기 (SPEC / SIGN_SPEC 모두 같은 개수)
    filled = len(out_list)
    _hide_rows(ws_spec, DOM.SPEC_START, DOM.SPEC_END, filled)
    _hide_rows(ws_sign, DOM.SIGN_START, DOM.SIGN_END, filled)

    # ⑧ 불필요한 시트 숨기기
    _drop_internal_sheets(wb)
    _show_only_sheets(wb, [
        SheetName.SPEC, SheetName.SIGN_SPEC,
        SheetName.INCOMING, SheetName.REQ_COPY,
    ])

    _reset_sheet_selections(wb)
    wb.save(path)
    if wb_cache is None:
        wb.close()          # 재사용 워크북은 호출자가 닫는다
    restore_header_footer_images(state.template_path, path)
    logger.info("국내 견적서 생성: %s", path)
    return path


def _fill_china(state: QuoteState, items: List[Dict[str, Any]]) -> str:
    """중국 견적서: 템플릿 복사 후 openpyxl 로 데이터 기입."""
    info   = state.cn_info
    ymd    = datetime.now().strftime("%y%m%d")
    line   = safe_filename(info.get("line", ""))
    tool   = safe_filename(info.get("tool", ""))
    folder = ensure_dir(
        os.path.join(exe_dir(), "견적서", f"{ymd}_중국_SCS_{line}")
    )
    path = unique_path(
        os.path.join(folder, f"대외비_{ymd}_중국_SCS_{line}_{tool}.xlsx")
    )

    _copy_template(state.template_path, path)
    wb = load_workbook(path)
    ws = wb[SheetName.QUOTE_CN]

    ws[CN.MAKER_CELL]   = info.get("maker", "")
    ws[CN.PROCESS_CELL] = info.get("process", "")
    ws[CN.TOOL_CELL]    = info.get("tool", "")

    pump    = [r for r in items if r["role"] != "CREDIT" and r["cat"] == "PUMP"]
    others  = [r for r in items if r["role"] != "CREDIT" and r["cat"] != "PUMP"]
    credits = [r for r in items if r["role"] == "CREDIT"]
    pump_credits = [r for r in credits if "Pump" in r.get("spec", "")]
    rack_credits = [r for r in credits if "Pump" not in r.get("spec", "")]

    # 펌프 구간 클리어 & 기입 (Pump Credit 포함)
    for rr in range(CN.PUMP_START, CN.PUMP_END + 1):
        ws[f"{CN.COL_SPEC}{rr}"]  = None
        ws[f"{CN.COL_PRICE}{rr}"] = None
        ws[f"{CN.COL_QTY}{rr}"]   = None

    pump_out = list(pump[:3])
    for cr in pump_credits:
        pump_out.append({"spec": cr["spec"], "qty": 0.0, "price": cr["amt"]})
    for i, r in enumerate(pump_out[:3]):
        rr = CN.PUMP_START + i
        ws[f"{CN.COL_SPEC}{rr}"]  = r["spec"]
        ws[f"{CN.COL_PRICE}{rr}"] = r["price"]
        ws[f"{CN.COL_QTY}{rr}"]   = None if float(r.get("qty", 0)) == 0 else r["qty"]

    # 랙 구간 클리어 & 기입 (Rack Credit만 포함)
    for rr in range(CN.RACK_START, CN.RACK_END + 1):
        ws[f"{CN.COL_SPEC}{rr}"]  = None
        ws[f"{CN.COL_PRICE}{rr}"] = None
        ws[f"{CN.COL_QTY}{rr}"]   = None

    out_list = others[:20]
    for cr in rack_credits:
        out_list.append({"spec": cr["spec"], "qty": 0.0, "price": cr["amt"]})
    out_list = out_list[:20]

    for i, r in enumerate(out_list):
        rr = CN.RACK_START + i
        ws[f"{CN.COL_SPEC}{rr}"]  = r["spec"]
        ws[f"{CN.COL_PRICE}{rr}"] = r["price"]
        ws[f"{CN.COL_QTY}{rr}"]   = None if float(r.get("qty", 0)) == 0 else r["qty"]

    wb.calculation.fullCalcOnLoad = True
    _hide_rows(ws, CN.RACK_START, CN.RACK_END, len(out_list))
    _drop_internal_sheets(wb)
    _show_only_sheets(wb, [SheetName.QUOTE_CN])
    _reset_sheet_selections(wb)
    wb.save(path)
    wb.close()
    restore_header_footer_images(state.template_path, path)
    logger.info("중국 견적서 생성: %s", path)
    return path


def _fill_us(state: QuoteState, items: List[Dict[str, Any]]) -> str:
    """미국 견적서: 템플릿 복사 후 openpyxl 로 데이터 기입."""
    info   = state.us_info
    ymd    = datetime.now().strftime("%y%m%d")
    site   = safe_filename(info.get("site", ""))
    tool   = safe_filename(info.get("tool", ""))
    folder = ensure_dir(
        os.path.join(exe_dir(), "견적서", f"{ymd}_미국_{site}")
    )
    path = unique_path(
        os.path.join(folder, f"대외비_{ymd}_{site}_{tool}_Quotation.xlsx")
    )

    _copy_template(state.template_path, path)
    wb = load_workbook(path)
    ws = wb[SheetName.QUOTE_US]

    ws[US.MAKER_CELL]    = info.get("maker", "")
    ws[US.PROCESS_CELL]  = info.get("process", "")
    ws[US.TOOL_CELL]     = info.get("tool", "")
    ws[US.EXCHANGE_CELL] = float(info.get("exchange", 0))
    y, m = info.get("base_ym", (datetime.now().year, datetime.now().month))
    ws[US.DATE_CELL] = f"{int(y)}-{int(m):02d}-1"

    pump    = [r for r in items if r["role"] != "CREDIT" and r["cat"] == "PUMP"]
    others  = [r for r in items if r["role"] != "CREDIT" and r["cat"] != "PUMP"]
    credits = [r for r in items if r["role"] == "CREDIT"]
    pump_credits = [r for r in credits if "Pump" in r.get("spec", "")]
    rack_credits = [r for r in credits if "Pump" not in r.get("spec", "")]

    # 펌프 단일 행 (Pump Credit은 RACK 구간 맨 앞 줄로 표기)
    ws[US.PUMP_SPEC]  = None
    ws[US.PUMP_PRICE] = None
    ws[US.PUMP_QTY]   = None
    if pump:
        ws[US.PUMP_SPEC]  = pump[0]["spec"]
        ws[US.PUMP_PRICE] = pump[0]["price"]
        ws[US.PUMP_QTY]   = pump[0]["qty"]

    # 랙 구간 클리어 & 기입 (Pump Credit → 맨 앞, Rack Credit → 맨 뒤)
    for rr in range(US.RACK_START, US.RACK_END + 1):
        ws[f"{US.COL_SPEC}{rr}"]  = None
        ws[f"{US.COL_PRICE}{rr}"] = None
        ws[f"{US.COL_QTY}{rr}"]   = None

    out_list = []
    for cr in pump_credits:
        out_list.append({"spec": cr["spec"], "qty": 0.0, "price": cr["amt"]})
    out_list += others[:20 - len(pump_credits)]
    for cr in rack_credits:
        out_list.append({"spec": cr["spec"], "qty": 0.0, "price": cr["amt"]})
    out_list = out_list[:20]

    for i, r in enumerate(out_list):
        rr = US.RACK_START + i
        ws[f"{US.COL_SPEC}{rr}"]  = r["spec"]
        ws[f"{US.COL_PRICE}{rr}"] = r["price"]
        ws[f"{US.COL_QTY}{rr}"]   = None if float(r.get("qty", 0)) == 0 else r["qty"]

    wb.calculation.fullCalcOnLoad = True
    _hide_rows(ws, US.RACK_START, US.RACK_END, len(out_list))
    _drop_internal_sheets(wb)
    _show_only_sheets(wb, [SheetName.QUOTE_US])
    _reset_sheet_selections(wb)
    wb.save(path)
    wb.close()
    restore_header_footer_images(state.template_path, path)
    logger.info("미국 견적서 생성: %s", path)
    return path


# ═══════════════════════════════════════════════════════════════
# 공개 API: 견적서 생성
# ═══════════════════════════════════════════════════════════════

def generate_quote(
    state   : QuoteState,
    qtype   : str,
    items   : List[Dict[str, Any]],
    rd      : Dict[str, Any],
) -> str:
    """견적서 1건 생성 → 저장 → 경로 반환."""
    try:
        if qtype == "중국":
            out = _fill_china(state, items)
        elif qtype == "미국":
            out = _fill_us(state, items)
        else:
            out = _fill_domestic(state, rd, items)
    except BaseException:
        _discard_inflight()   # 미완성 템플릿 사본 제거
        raise
    _clear_inflight()
    return out


def generate_quote_multi(
    state        : QuoteState,
    items        : List[Dict[str, Any]],
    request_rows : List[Dict[str, Any]],
    progress_cb  : Optional[Callable[[int, int, str], None]] = None,
) -> List[Tuple[Dict[str, Any], str, str]]:
    """
    국내 견적서 여러 건을 순서대로 생성.
    Excel 을 전혀 실행하지 않으므로 COM 버전 대비 크게 빠르다.
    반환: [(rd, saved_path, error), ...]
      성공: error="" / 실패: saved_path="" 이고 error 에 사유
    """
    results: List[Tuple[Dict[str, Any], str, str]] = []
    total = len(request_rows)

    # 의뢰파일을 한 번만 열어 전 행에 재사용 — 반복 open/close 방지
    req_ws_cache = None
    req_wb_cache = None
    req_path = state.request_path
    if req_path and req_path.lower().endswith(".xlsx"):
        try:
            req_wb_cache = load_workbook(req_path, data_only=True)
            sn = state.request_sheet_name
            req_ws_cache = (
                req_wb_cache[sn]
                if sn and sn in req_wb_cache.sheetnames
                else req_wb_cache.active
            )
        except Exception as e:
            logger.warning("의뢰파일 사전 로드 실패 (행별 열기로 대체): %s", e)
            req_wb_cache = None
            req_ws_cache = None

    # 템플릿도 한 번만 열어 재사용 — 건마다 전체 파싱/직렬화를 반복하던 것이
    # 일괄 생성 시간의 대부분이었다(실측 1.40초/건 → 0.47초/건).
    wb_cache = None
    if state.template_path:
        try:
            wb_cache = open_template_for_reuse(state.template_path)
        except Exception as e:
            logger.warning("템플릿 사전 로드 실패 (건별 로드로 대체): %s", e)
            wb_cache = None

    try:
        for i, rd in enumerate(request_rows):
            try:
                path = _fill_domestic(state, rd, items, req_ws_cache, wb_cache)
                _clear_inflight()
                results.append((rd, path, ""))
                if progress_cb:
                    progress_cb(i + 1, total, os.path.basename(path))
            except Exception as e:
                _discard_inflight()   # 미완성 출력 파일 제거
                logger.error("멀티 생성 실패: %s", e, exc_info=True)
                results.append((rd, "", str(e)))
                if progress_cb:
                    progress_cb(i + 1, total, f"[실패] {e}")
                # 실패 지점에 따라 워크북이 중간 상태로 남을 수 있다.
                # 다음 건이 그 상태를 물려받지 않도록 새로 연다.
                if wb_cache is not None:
                    try:
                        wb_cache.close()
                    except Exception:
                        pass
                    try:
                        wb_cache = open_template_for_reuse(state.template_path)
                    except Exception as e2:
                        logger.warning("템플릿 재로드 실패 (건별 로드로 대체): %s", e2)
                        wb_cache = None
    finally:
        for _wb in (req_wb_cache, wb_cache):
            if _wb is not None:
                try:
                    _wb.close()
                except Exception:
                    pass

    return results


def _generate_cover_impl(
    template_path : str,
    folder        : str,
    source_files  : List[str],
    investor_name : str = "채승철",
    progress_cb   : Optional[Callable[[int, int, str], None]] = None,
    warranty_years: int = 2,
) -> Tuple[str, List[Tuple[str, str]]]:
    """
    갑지 생성.
    source_files 각각의 견적의뢰복사본 2행 데이터를 openpyxl 로 읽어
    갑지DATA 시트에 쌓은 뒤 저장.

    반환: (저장 경로, 읽기 실패 목록[(파일명, 사유)])
      실패 목록이 비어 있지 않으면 그만큼 갑지에서 누락된 것이므로
      호출자가 반드시 사용자에게 알려야 한다.
    """
    folder_name = os.path.basename(folder.rstrip("\\/"))
    out_path = unique_path(os.path.join(folder, f"대외비_{folder_name}_갑지.xlsx"))

    input_abs = {os.path.abspath(p).lower() for p in source_files}
    while os.path.abspath(out_path).lower() in input_abs:
        out_path = unique_path(out_path)

    def _natural_key(s: str):
        return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', s)]

    out_base = os.path.basename(out_path).lower()
    clean = sorted(
        [p for p in source_files if os.path.basename(p).lower() != out_base],
        key=lambda p: _natural_key(os.path.basename(p)),
    )
    if not clean:
        raise RuntimeError("처리할 엑셀 파일이 없습니다.")

    # ── COM 으로 수식 계산값 미리 수집 ────────────────────────────────────
    # openpyxl 저장 시 수식 캐시(cached value)가 소멸된다.
    # data_only=True 로 열면 O/P/Q/S/T/U 열(15,16,17,19,20,21)이 None 반환.
    # COM 세션을 열어 Excel 이 fullCalcOnLoad 재계산한 값을 직접 읽는다.
    # {abs_path.lower(): {col: value}}
    _FORMULA_COLS = (15, 16, 17, 19, 20, 21)
    _com_vals: Dict[str, Dict[int, Any]] = {}

    if _ensure_com():
        _xl = None
        _wb_c = _ws_c = None
        try:
            pythoncom.CoInitialize()
            _xl = win32.DispatchEx("Excel.Application")
            _xl.Visible = False
            _xl.DisplayAlerts = False
            for _p in clean:
                _pkey = os.path.abspath(_p).lower()
                _row: Dict[int, Any] = {}
                try:
                    _wb_c = _xl.Workbooks.Open(
                        os.path.abspath(_p), ReadOnly=True, UpdateLinks=0, AddToMru=False
                    )
                    try:
                        _ws_c = _wb_c.Worksheets(SheetName.REQ_COPY)
                        for _c in _FORMULA_COLS:
                            try:
                                _row[_c] = _ws_c.Cells(2, _c).Value
                            except Exception:
                                pass
                    finally:
                        _ws_c = None          # 시트 프록시 먼저 해제
                        try:
                            _wb_c.Close(False)
                        except Exception:
                            pass
                        _wb_c = None
                except Exception as _e:
                    logger.warning("COM 수식값 읽기 실패 (%s): %s", _p, _e)
                _com_vals[_pkey] = _row
        except Exception as _e:
            logger.warning("COM 초기화 실패 — 수식 셀은 openpyxl 폴백: %s", _e)
        finally:
            # 해제 순서가 중요하다.
            # 살아있는 COM 프록시가 남은 채 Quit() 하면 EXCEL.EXE 가 죽지 않고,
            # CoUninitialize() 를 프록시 해제보다 먼저 부르면 그 뒤의 해제가
            # 이미 정리된 아파트먼트에서 일어나 보이지 않는 좀비 프로세스가 남는다.
            #   프록시 해제 → gc → Quit() → 참조 제거 → gc → CoUninitialize()
            _clear_clipboard()
            _ws_c = None
            _wb_c = None
            import gc
            gc.collect()
            if _xl is not None:
                try:
                    _xl.Quit()
                except Exception:
                    pass
                _xl = None
                gc.collect()
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    # 템플릿 복사 후 openpyxl 로 편집
    _copy_template(template_path, out_path)
    wb       = load_workbook(out_path)
    ws_data  = wb[SheetName.COVER_DATA]
    ws_cover = wb[SheetName.COVER]

    # 갑지DATA 초기화 — iter_rows 공개 API 사용 (내부 _cells 직접 접근 금지)
    max_data_row = min(ws_data.max_row or 1, 2000)
    if max_data_row >= 2:
        for row_cells in ws_data.iter_rows(min_row=2, max_row=max_data_row,
                                            min_col=1, max_col=27):
            for cell in row_cells:
                if cell.value is not None:
                    cell.value = None

    # 각 source 파일의 견적의뢰복사본 2행 읽기
    write_row = 2
    total_src = len(clean)
    failed: List[Tuple[str, str]] = []   # (파일명, 사유)
    for idx, path in enumerate(clean):
        if progress_cb:
            progress_cb(idx + 1, total_src, os.path.basename(path))
        src_wb = None
        try:
            src_wb = load_workbook(path, data_only=True)
            if SheetName.REQ_COPY not in src_wb.sheetnames:
                raise KeyError(f"'{SheetName.REQ_COPY}' 시트 없음")
            src_ws = src_wb[SheetName.REQ_COPY]
            for c in range(1, 28):  # A:AA
                ws_data.cell(write_row, c).value = src_ws.cell(2, c).value
            # COM 계산값으로 수식 셀(O/P/Q/S/T/U) 덮어쓰기
            for _c, _v in _com_vals.get(os.path.abspath(path).lower(), {}).items():
                if _v is not None:
                    ws_data.cell(write_row, _c).value = _v
            write_row += 1
        except Exception as e:
            # 읽기에 실패한 원본은 갑지에서 누락된다. 조용히 넘기면
            # 갑지가 한 건 모자란 채 정상으로 보이므로 호출자에게 돌려준다.
            logger.warning("갑지 원본 읽기 실패 (%s): %s", path, e)
            failed.append((os.path.basename(path), str(e)))
        finally:
            if src_wb is not None:
                try:
                    src_wb.close()
                except Exception:
                    pass

    if not failed and write_row == 2:
        raise ValueError("갑지에 기록할 원본이 없습니다.")

    # 투자자명
    ws_cover["B7"] = f"삼성전자 ㈜ / {investor_name} 님"

    # 보증기간 (기본값: 2 Year after delivery)
    ws_cover["I12"] = f"{warranty_years} Year after delivery"

    # 갑지DATA A2:A31 을 직접 확인하여 빈 행을 COVER(20~49) + COVER_DATA(2~31) 양쪽에 숨기기
    for i in range(30):
        a_val = ws_data.cell(2 + i, 1).value
        blank = a_val is None or str(a_val).strip() == ""
        ws_cover.row_dimensions[20 + i].hidden = blank
        ws_data.row_dimensions[2 + i].hidden   = blank

    wb.calculation.fullCalcOnLoad = True
    _drop_internal_sheets(wb)
    _show_only_sheets(wb, [SheetName.COVER_DATA, SheetName.COVER])

    # 열릴 때 COVER 시트만 활성 탭으로 — 여러 시트가 tabSelected=True 상태로
    # 저장되면 Excel이 '그룹' 모드로 파일을 열어버린다.
    for _sn in wb.sheetnames:
        wb[_sn].sheet_view.tabSelected = (_sn == SheetName.COVER)
    wb.active = wb[SheetName.COVER]

    _reset_sheet_selections(wb)
    wb.save(out_path)
    wb.close()
    restore_header_footer_images(template_path, out_path)
    logger.info("갑지 생성: %s (원본 %d건, 누락 %d건)",
                out_path, write_row - 2, len(failed))
    return out_path, failed


def _export_cover_data_sheet_impl(src_path: str) -> str:
    """
    갑지 파일에서 갑지DATA(SheetName.COVER_DATA) 시트만 추출하여 새 xlsx 로 저장.

    파일명 규칙:
      원본 파일명 끝의 '_갑지' 를 제거한다.
      예) 대외비_260512_LOT베큠_P3_CVD_권혁성_갑지.xlsx
        → 대외비_260512_LOT베큠_P3_CVD_권혁성.xlsx

    '_갑지' 가 없는 파일명이면 '_제출용' 을 접미사로 붙여 구분한다.
    동일 경로에 같은 이름이 이미 있으면 unique_path() 로 번호를 붙인다.

    반환: 저장된 파일 경로
    """
    import re

    stem, ext = os.path.splitext(os.path.basename(src_path))
    new_stem = re.sub(r"_갑지(?:_\d+)?$", "", stem)
    if new_stem == stem:
        raise ValueError(f"갑지 파일이 아닙니다: '{os.path.basename(src_path)}'")
    out_path = unique_path(os.path.join(os.path.dirname(src_path), new_stem + ext))

    # 원본을 통째로 복사 → 서식·스타일 완전 보존
    _copy_template(src_path, out_path)

    wb = load_workbook(out_path)
    keep = SheetName.COVER_DATA
    if keep not in wb.sheetnames:
        wb.close()
        os.remove(out_path)
        raise ValueError(f"'{keep}' 시트를 찾을 수 없습니다: {os.path.basename(src_path)}")

    for name in list(wb.sheetnames):
        if name != keep:
            del wb[name]

    ws = wb[keep]
    ws.sheet_state = "visible"
    wb.active = ws
    _reset_sheet_selections(wb)
    wb.save(out_path)
    wb.close()
    logger.info("제출용 엑셀 생성: %s", out_path)
    return out_path


# ═══════════════════════════════════════════════════════════════
# COM 헬퍼 — PDF 변환 전용 (변경 없음)
# ═══════════════════════════════════════════════════════════════

def generate_cover(*args, **kwargs) -> Tuple[str, List[Tuple[str, str]]]:
    """generate_cover 래퍼 — 실패 시 미완성 갑지 파일을 지운다."""
    try:
        out = _generate_cover_impl(*args, **kwargs)
    except BaseException:
        _discard_inflight()
        raise
    _clear_inflight()
    return out


def export_cover_data_sheet(*args, **kwargs) -> str:
    """export_cover_data_sheet 래퍼 — 실패 시 미완성 파일을 지운다."""
    try:
        out = _export_cover_data_sheet_impl(*args, **kwargs)
    except BaseException:
        _discard_inflight()
        raise
    _clear_inflight()
    return out


class ExcelCOM:
    """
    Excel COM 세션 컨텍스트 매니저.
    _ExcelLoaderThread(전자서명 시트 캡처)가 공유 Application 객체로 사용한다.
    """

    def __init__(self) -> None:
        self._excel = None

    def __enter__(self) -> "ExcelCOM":
        if not _ensure_com():
            raise RuntimeError("pywin32가 설치되어 있지 않습니다.")
        pythoncom.CoInitialize()
        try:
            self._excel = win32.DispatchEx("Excel.Application")
            self._excel.Visible = False
            self._excel.DisplayAlerts = False
            # 자동화 스위치: 이 세션에서 여는 워크북(이 앱이 만든 파일)은
            # 전부 fullCalcOnLoad=True 가 저장되어 있어, 스위치 없이 열면
            # Workbooks.Open 마다 전체 재계산이 돈다 — 파일 수만큼 반복되는
            # 비용이라 가장 이득이 큰 최적화다. 세션 종료 시 프로세스 자체를
            # Quit() 하므로 복원할 필요는 없다.
            try:
                self._excel.Calculation = -4135     # xlCalculationManual
                self._excel.EnableEvents = False
                self._excel.ScreenUpdating = False
            except Exception:
                # 일부 Excel 버전/보안 정책에서 속성 설정이 막힐 수 있다.
                # 최적화일 뿐 필수 기능이 아니므로 실패해도 세션은 계속한다.
                logger.warning("Excel 자동화 스위치 설정 실패 — 계속 진행", exc_info=True)
        except BaseException:
            # DispatchEx 실패(Excel 미설치·손상) 시 __exit__ 가 호출되지 않으므로
            # 여기서 아파트먼트를 되돌리지 않으면 호출할 때마다 하나씩 누수된다.
            self._excel = None
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
            raise
        return self

    def __exit__(self, *_) -> None:
        # 순서 중요: 프록시 해제 → Quit() → 참조 제거 → CoUninitialize()
        # 반대로 하면 보이지 않는 EXCEL.EXE 가 파일 잠금을 쥔 채 남는다.
        import gc
        gc.collect()
        if self._excel is not None:
            try:
                self._excel.Quit()
            except Exception:
                pass
            self._excel = None
            gc.collect()
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass

    @property
    def app(self):
        return self._excel

    def open(self, path: str, read_only: bool = False):
        return self._excel.Workbooks.Open(
            path, ReadOnly=read_only, UpdateLinks=0, AddToMru=False
        )


# ═══════════════════════════════════════════════════════════════
# 공개 API: 전자서명용 임시 PDF 생성 (COM 유지)
# ═══════════════════════════════════════════════════════════════

def excel_to_merged_pdf(xlsx_path: str, tmp_dir: str, file_index: int,
                        xl_app=None) -> str:
    """
    xlsx 의 ESIGN_TARGET 시트(원래 Visible인 것)만 PDF 로 내보낸다.
    반환: PDF 경로 (대상 시트 없으면 "")

    워크북 단위 ExportAsFixedFormat 1회 호출:
      - 비대상 시트를 일시적으로 VeryHidden 처리 → workbook 전체 export
      - Close(False) 로 디스크 파일은 변경하지 않음
      - RenameFile 발생 횟수: 파일당 1회 (기존 시트 수만큼 → 1회로 감소)

    xl_app: 공유 Excel.Application COM 객체. None이면 자체 ExcelCOM 사용.
    """
    import fitz

    out_pdf = os.path.join(tmp_dir, f"tmp_{file_index:03d}.pdf")

    def _process(app) -> str:
        wb = app.Workbooks.Open(xlsx_path, ReadOnly=False, UpdateLinks=0, AddToMru=False)
        try:
            target_set = set(SheetName.ESIGN_TARGET)
            to_hide = []

            # 원래 Visible 이면서 ESIGN_TARGET에 속하는 시트만 유지
            has_visible_target = False
            for ws in wb.Worksheets:
                is_target  = ws.Name in target_set
                was_visible = int(ws.Visible) == -1   # xlSheetVisible = -1
                if is_target and was_visible:
                    has_visible_target = True
                else:
                    to_hide.append(ws)

            if not has_visible_target:
                return ""

            # 비대상 시트를 VeryHidden (xlVeryHidden = 2)
            for ws in to_hide:
                try:
                    ws.Visible = 2
                except Exception:
                    pass

            # 워크북 전체를 PDF 1회 export (ExportAsFixedFormat 1회 = RenameFile 1회)
            wb.ExportAsFixedFormat(0, out_pdf)
            return out_pdf if os.path.exists(out_pdf) else ""
        finally:
            _clear_clipboard()   # ExportAsFixedFormat 이 클립보드를 사용하는 경우 대비
            try:
                wb.Close(False)   # 디스크 파일 변경 없음
            except Exception:
                pass

    if xl_app is not None:
        return _process(xl_app)

    with ExcelCOM() as xl:
        return _process(xl.app)


def _print_area_range(ws):
    """PrintArea(R1C1 또는 A1 형식)를 Range 객체로 반환. 없으면 UsedRange."""
    import re
    pa = ws.PageSetup.PrintArea
    if not pa:
        return ws.UsedRange
    if "!" in pa:
        pa = pa.split("!")[-1]
    # R1C1 형식 감지: "R숫자C숫자:R숫자C숫자" 또는 단일 "R숫자C숫자"
    m = re.match(r'R(\d+)C(\d+)(?::R(\d+)C(\d+))?$', pa.strip())
    if m:
        r1, c1 = int(m.group(1)), int(m.group(2))
        r2 = int(m.group(3)) if m.group(3) else r1
        c2 = int(m.group(4)) if m.group(4) else c1
        return ws.Range(ws.Cells(r1, c1), ws.Cells(r2, c2))
    # A1 형식
    try:
        return ws.Range(pa)
    except Exception:
        return ws.UsedRange


def excel_capture_sheets_to_pngs(xlsx_path: str, tmp_dir: str, file_index: int,
                                  xl_app=None, progress_cb=None,
                                  should_cancel=None) -> List[str]:
    """
    xlsx ESIGN_TARGET 가시 시트를 클립보드로 캡처 → PNG 저장.
    ExportAsFixedFormat/PrintOut 미사용 → RenameFile 없음.

    xl_app       : 공유 Excel.Application COM 객체. None이면 자체 ExcelCOM 사용.
    progress_cb  : (완료 시트수, 전체 시트수, 시트명) 콜백. 시트 캡처가 끝날 때마다 호출.
    should_cancel: 인자 없이 bool 반환하는 콜백. True 면 남은 시트를 건너뛰고 지금까지
                   캡처한 것만 반환한다 (파일 단위보다 촘촘한 취소 체크).
    반환: 저장된 PNG 경로 리스트 (순서 = ESIGN_TARGET 순서)
    """
    try:
        from PIL import ImageGrab
    except ImportError:
        logger.error("Pillow(PIL) 없음 — pip install pillow")
        return []

    png_paths: List[str] = []

    def _process(app) -> List[str]:
        wb = app.Workbooks.Open(xlsx_path, ReadOnly=True, UpdateLinks=0, AddToMru=False)
        try:
            # ESIGN_TARGET 시트 우선, 없으면 보이는 시트 전체 캡처.
            # 이름 조회 + Visible 확인을 한 번만 하고 ws 프록시를 그대로 들고 있는다
            # (기존엔 이 목록을 만들 때 한 번, 캡처 루프에서 또 한 번 조회했다).
            target_sheets = []   # [(name, ws), ...]
            for name in SheetName.ESIGN_TARGET:
                try:
                    ws = wb.Worksheets(name)
                    if int(ws.Visible) == -1:
                        target_sheets.append((name, ws))
                except Exception:
                    continue
            if not target_sheets:
                for ws in wb.Worksheets:
                    try:
                        if int(ws.Visible) == -1:
                            target_sheets.append((ws.Name, ws))
                    except Exception:
                        continue

            total = len(target_sheets)
            for idx, (name, ws) in enumerate(target_sheets):
                if should_cancel is not None:
                    try:
                        if should_cancel():
                            break
                    except Exception:
                        pass
                safe_name = "".join(c if c not in r'\/:*?"<>|' else "_" for c in name)
                png_path = os.path.join(
                    tmp_dir, f"cap_{file_index:03d}_{idx:02d}_{safe_name}.png")
                try:
                    ws.Activate()
                    # 페이지 나누기 미리보기 → 기본 보기로 전환 후 캡처.
                    # [주의] View 는 창 속성이지만, 시트마다 마지막 저장 시점의
                    # 보기 모드를 따로 기억하고 있어 Activate() 로 다른 시트로
                    # 넘어가면 그 시트의 저장된 모드(페이지 나누기 미리보기)가
                    # 되살아날 수 있다. 워크북당 1회로 줄였다가 "1 페이지"
                    # 워터마크가 다시 나타나는 회귀가 있었다 — 반드시 시트마다
                    # 매번 설정해야 한다.
                    try:
                        app.ActiveWindow.View = 1  # xlNormalView
                    except Exception:
                        pass
                    rng = _print_area_range(ws)
                    rng.CopyPicture(Appearance=1, Format=2)  # xlScreen, xlBitmap
                    img = ImageGrab.grabclipboard()
                    # ── 클립보드 즉시 해제 ──────────────────────────────────────
                    # CopyPicture 후 Windows 클립보드에 CF_ENHMETAFILE(EMF) 핸들이
                    # 남아 있는 상태에서 COM 세션(xl.Quit)이 종료되면 stale 핸들이
                    # 발생한다. 이후 사용자가 별도 Excel에서 Ctrl+C를 시도하면
                    # Windows가 이 핸들을 해제하려다 액세스 위반 → Excel 전체 종료.
                    # 이미지를 읽은 직후 클립보드를 비워 이 경로를 차단한다.
                    _clear_clipboard()
                    if img is not None:
                        # 이전 세션 잠금 파일이 남아 있으면 삭제 시도 후 저장
                        _dst = png_path
                        if os.path.exists(_dst):
                            try:
                                os.remove(_dst)
                            except OSError:
                                _dst = unique_path(_dst)
                        # compress_level=1: 잠깐 쓰고 지울 임시 파일이라 압축률보다
                        # 저장/전송 속도가 낫다(기본값 6 대비 체감 저하 없음).
                        img.save(_dst, "PNG", compress_level=1)
                        png_paths.append(_dst)
                    else:
                        logger.warning("클립보드 캡처 실패 (%s / %s)", xlsx_path, name)
                except Exception as e:
                    logger.warning("시트 캡처 실패 (%s / %s): %s", xlsx_path, name, e)
                if progress_cb is not None:
                    try:
                        progress_cb(idx + 1, total, name)
                    except Exception:
                        pass
        finally:
            _clear_clipboard()   # 워크북 닫기 전 최종 클립보드 해제
            try:
                wb.Close(False)
            except Exception:
                pass
        return png_paths

    if xl_app is not None:
        return _process(xl_app)

    with ExcelCOM() as xl:
        return _process(xl.app)
