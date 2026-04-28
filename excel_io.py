"""
excel_io.py
===========
Excel 입출력 전담 모듈.
- openpyxl  : 읽기 전용 (템플릿/의뢰파일 파싱)
- Excel COM : 쓰기 전용 (견적서 생성, 갑지 생성, PDF 변환)

UI에 일절 의존하지 않는다.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core import (
    CN, DOM, US, SheetName, QuoteState,
    ensure_dir, safe_filename, s, to_float, unique_path,
    exe_dir, parse_invest_info,
)

logger = logging.getLogger("QuoteApp")

# ── COM 임포트 (Windows 전용) ──────────────────
try:
    import pythoncom
    import win32com.client as win32
    COM_AVAILABLE = True
except ImportError:
    pythoncom = None        # type: ignore
    win32 = None            # type: ignore
    COM_AVAILABLE = False


# ═══════════════════════════════════════════════
# 읽기 (openpyxl)
# ═══════════════════════════════════════════════

def parse_items_sheet(ws: Worksheet) -> Tuple[
    List[Dict[str, Any]],
    Dict[str, List[Dict[str, Any]]],
    Dict[str, float],
    Dict[str, int],
]:
    """
    품목 시트 파싱.
    반환: (rows, by_class, price_by_spec, spec_order_index)
    """
    def _is_header(a: str, b: str, c: str) -> bool:
        h = (a + b + c).replace(" ", "").lower()
        return "분류" in h and ("단가" in h or "가격" in h)

    start = 2 if _is_header(
        s(ws.cell(1, 1).value), s(ws.cell(1, 2).value), s(ws.cell(1, 3).value)
    ) else 1

    rows: List[Dict[str, Any]] = []
    by_class: Dict[str, List[Dict[str, Any]]] = {}
    price_by_spec: Dict[str, float] = {}
    order_index: Dict[str, int] = {}
    order = 0

    for r in range(start, ws.max_row + 1):
        a = s(ws.cell(r, 1).value)
        b = s(ws.cell(r, 2).value)
        c = ws.cell(r, 3).value
        if not (a or b or c):
            continue
        price = to_float(c)
        row = {"A": a, "B": b, "C": price}
        rows.append(row)
        if b and b not in order_index:
            order_index[b] = order
            order += 1
        if a:
            by_class.setdefault(a, []).append(row)
        if b:
            price_by_spec[b] = price

    return rows, by_class, price_by_spec, order_index


def parse_code_map_sheet(ws: Worksheet) -> Dict[Tuple[str, str, str], List[Tuple[str, float]]]:
    """
    코드매핑 시트 파싱.
    반환: {(공정, 설비사, 5D): [(ACC, 수량), ...]}
    """
    def _is_header(a, b, c, d, e) -> bool:
        h = (a + b + c + d + e).replace(" ", "").lower()
        return "공정" in h and "설비" in h and ("acc" in h or "수량" in h or "5d" in h)

    start = 2 if _is_header(
        s(ws.cell(1, 1).value), s(ws.cell(1, 2).value), s(ws.cell(1, 3).value),
        s(ws.cell(1, 4).value), s(ws.cell(1, 5).value),
    ) else 1

    cmap: Dict[Tuple[str, str, str], List[Tuple[str, float]]] = {}
    for r in range(start, ws.max_row + 1):
        proc   = s(ws.cell(r, 1).value)
        vendor = s(ws.cell(r, 2).value)
        code5d = s(ws.cell(r, 3).value)
        acc    = s(ws.cell(r, 4).value)
        qty    = to_float(ws.cell(r, 5).value)
        if proc and vendor and code5d and acc:
            cmap.setdefault((proc, vendor, code5d), []).append((acc, qty))
    return cmap


def parse_request_xlsx(path: str) -> Tuple[str, List[Dict[str, Any]]]:
    """
    의뢰파일 파싱.
    반환: (active_sheet_name, row_dicts)
    row_dict keys: row_idx, D, E, F, H, J, K, V, X, Z
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    sheet_name = ws.title

    def _is_header() -> bool:
        checks = [
            s(ws.cell(1, 26).value).replace(" ", "").lower(),
            s(ws.cell(1, 11).value).replace(" ", "").lower(),
            s(ws.cell(1, 24).value).replace(" ", "").lower(),
            s(ws.cell(1, 22).value).replace(" ", "").lower(),
        ]
        return any("설비" in c or "공정" in c or "5d" in c or "코드" in c for c in checks)

    start = 2 if _is_header() else 1
    rows: List[Dict[str, Any]] = []
    for r in range(start, ws.max_row + 1):
        z = ws.cell(r, 26).value
        if s(z) == "":
            continue
        rows.append({
            "row_idx": r,
            "D": ws.cell(r,  4).value,
            "E": ws.cell(r,  5).value,
            "F": ws.cell(r,  6).value,
            "G": ws.cell(r,  7).value,
            "H": ws.cell(r,  8).value,
            "J": ws.cell(r, 10).value,
            "K": ws.cell(r, 11).value,
            "V": ws.cell(r, 22).value,
            "X": ws.cell(r, 24).value,
            "Z": z,
        })
    return sheet_name, rows


# ═══════════════════════════════════════════════
# COM 헬퍼
# ═══════════════════════════════════════════════

class ExcelCOM:
    """
    Excel COM 세션 컨텍스트 매니저.

    with ExcelCOM() as xl:
        wb = xl.open(path)
        ...
        wb.SaveAs(out)
    """

    def __init__(self) -> None:
        self._excel = None

    def __enter__(self) -> "ExcelCOM":
        if not COM_AVAILABLE:
            raise RuntimeError("pywin32가 설치되어 있지 않습니다.")
        pythoncom.CoInitialize()
        self._excel = win32.DispatchEx("Excel.Application")
        self._excel.Visible = False
        self._excel.DisplayAlerts = False
        return self

    def __exit__(self, *_) -> None:
        try:
            self._excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass

    # ── 속성 ────────────────────────────────────
    @property
    def app(self):
        return self._excel

    # ── 워크북 ──────────────────────────────────
    def open(self, path: str, read_only: bool = False):
        return self._excel.Workbooks.Open(
            path, ReadOnly=read_only, UpdateLinks=0, AddToMru=False
        )

    def recalc(self) -> None:
        try:
            self._excel.CalculateFullRebuild()
        except Exception:
            try:
                self._excel.Calculate()
            except Exception:
                logger.warning("Excel 재계산 실패")

    # ── 시트 가시성 ─────────────────────────────
    @staticmethod
    def show_only(wb, keep: List[str]) -> None:
        """keep에 없는 시트는 모두 VeryHidden으로."""
        keep_set = set(keep)
        for ws in wb.Worksheets:
            ws.Visible = -1 if str(ws.Name) in keep_set else 2  # -1=Visible, 2=VeryHidden
        try:
            wb.Worksheets(keep[0]).Activate()
        except Exception:
            pass

    # ── 행 숨기기 ────────────────────────────────
    @staticmethod
    def hide_blank_rows(ws, col: str, r_start: int, r_end: int) -> None:
        for r in range(r_start, r_end + 1):
            try:
                v = ws.Range(f"{col}{r}").Value
                blank = (
                    v is None
                    or (isinstance(v, (int, float)) and float(v) == 0.0)
                    or str(v).strip().replace(",", "") in ("", "0")
                )
                ws.Rows(r).Hidden = blank
            except Exception:
                logger.debug("행 숨기기 실패: %s%d", col, r)

    # ── 저장 경로 충돌 방지 ──────────────────────
    @staticmethod
    def close_if_open(excel, path: str) -> None:
        """같은 경로의 워크북이 열려 있으면 닫는다."""
        target = os.path.abspath(path).lower()
        try:
            for i in range(excel.Workbooks.Count, 0, -1):
                wb = excel.Workbooks(i)
                try:
                    fn = str(wb.FullName or "").strip()
                except Exception:
                    fn = ""
                if fn and os.path.abspath(fn).lower() == target:
                    wb.Close(False)
        except Exception:
            pass


# ═══════════════════════════════════════════════
# 견적서 생성 로직
# ═══════════════════════════════════════════════

def _write_req_row(xl: ExcelCOM, template_wb, state: QuoteState, rd: Dict[str, Any]) -> None:
    """의뢰파일의 특정 행을 견적의뢰복사본 시트 2행에 복사."""
    req_path = state.request_path
    if not req_path or not req_path.lower().endswith(".xlsx"):
        return
    try:
        row_idx = int(rd.get("row_idx", 0))
    except (TypeError, ValueError):
        return
    if row_idx <= 0:
        return

    req_wb = None
    try:
        req_wb = xl.open(req_path, read_only=True)
        sheet_name = state.request_sheet_name
        try:
            req_ws = req_wb.Worksheets(sheet_name) if sheet_name else req_wb.ActiveSheet
        except Exception:
            req_ws = req_wb.ActiveSheet

        copy_ws = template_wb.Worksheets(SheetName.REQ_COPY)
        copy_ws.Range("A2:N2").Value  = None
        copy_ws.Range("R2").Value     = None
        copy_ws.Range("W2:AA2").Value = None

        vals_an  = req_ws.Range(f"A{row_idx}:N{row_idx}").Value
        val_r    = req_ws.Range(f"R{row_idx}").Value
        vals_waa = req_ws.Range(f"W{row_idx}:AA{row_idx}").Value

        if vals_an  is not None: copy_ws.Range("A2:N2").Value  = vals_an
        if val_r    is not None: copy_ws.Range("R2").Value     = val_r
        if vals_waa is not None: copy_ws.Range("W2:AA2").Value = vals_waa
    finally:
        if req_wb is not None:
            try:
                req_wb.Close(False)
            except Exception:
                pass


def _fill_domestic(xl: ExcelCOM, wb, state: QuoteState,
                   rd: Dict[str, Any], items: List[Dict[str, Any]]) -> str:
    """국내 견적서 시트에 데이터 기입. 반환: 저장 경로(미저장)."""
    ws_spec = wb.Worksheets(SheetName.SPEC)
    ws_sign = wb.Worksheets(SheetName.SIGN_SPEC)

    _write_req_row(xl, wb, state, rd)

    investor = s(state.investor_name) or "채승철"
    ws_spec.Range("B4").Value = f"{investor} 님 / 설비구매그룹"

    rack_items = [r for r in items if r["cat"] != "PUMP" and r["role"] != "CREDIT"]
    credits    = [r for r in items if r["role"] == "CREDIT"]

    # 기존 데이터 클리어
    for rr in range(DOM.SPEC_START, DOM.SPEC_END + 1):
        for col in (DOM.COL_SPEC, DOM.COL_QTY, DOM.COL_PRICE, DOM.COL_AMT):
            ws_spec.Range(f"{col}{rr}").Value = None

    out_list = rack_items[:25]
    for cr in credits:
        out_list.append({"spec": cr["spec"], "qty": 0.0, "price": 0.0, "amt": cr["amt"]})
    out_list = out_list[:25]

    for i, r in enumerate(out_list):
        rr = DOM.SPEC_START + i
        ws_spec.Range(f"{DOM.COL_SPEC}{rr}").Value = r["spec"]
        if r.get("qty") == 0.0 and r.get("price") == 0.0 and "amt" in r:
            ws_spec.Range(f"{DOM.COL_AMT}{rr}").Value = r["amt"]
        else:
            ws_spec.Range(f"{DOM.COL_QTY}{rr}").Value   = r["qty"]
            ws_spec.Range(f"{DOM.COL_PRICE}{rr}").Value  = r["price"]
            ws_spec.Range(f"{DOM.COL_AMT}{rr}").Value    = r["amt"]

    xl.recalc()
    ExcelCOM.hide_blank_rows(ws_spec, DOM.COL_SPEC, DOM.SPEC_START, DOM.SPEC_END)
    ExcelCOM.hide_blank_rows(ws_sign, DOM.COL_SPEC, DOM.SIGN_START, DOM.SIGN_END)
    ExcelCOM.show_only(wb, [SheetName.SPEC, SheetName.SIGN_SPEC,
                            SheetName.INCOMING, SheetName.REQ_COPY])

    ymd      = datetime.now().strftime("%y%m%d")
    pr       = safe_filename(s(rd.get("D")))
    itemno   = safe_filename(s(rd.get("E")))
    lineproc = safe_filename(s(rd.get("K")))
    investor = safe_filename(s(rd.get("J")))
    tool     = safe_filename(s(rd.get("Z")))
    folder   = ensure_dir(os.path.join(exe_dir(), f"{ymd}_LOT베큠_{lineproc}_{investor}"))
    return os.path.join(folder, f"{pr}-{itemno}_{ymd}_LOT베큠_{lineproc}_{investor}_{tool}.xlsx")


def _fill_china(xl: ExcelCOM, wb, state: QuoteState, items: List[Dict[str, Any]]) -> str:
    """중국 견적서 시트에 데이터 기입."""
    info = state.cn_info
    ws = wb.Worksheets(SheetName.QUOTE_CN)

    ws.Range(CN.MAKER_CELL).Value   = info.get("maker", "")
    ws.Range(CN.PROCESS_CELL).Value = info.get("process", "")
    ws.Range(CN.TOOL_CELL).Value    = info.get("tool", "")

    pump    = [r for r in items if r["role"] != "CREDIT" and r["cat"] == "PUMP"]
    others  = [r for r in items if r["role"] != "CREDIT" and r["cat"] != "PUMP"]
    credits = [r for r in items if r["role"] == "CREDIT"]

    for rr in range(CN.PUMP_START, CN.PUMP_END + 1):
        ws.Range(f"{CN.COL_SPEC}{rr}").Value = None
        ws.Range(f"{CN.COL_PRICE}{rr}").Value = None
        ws.Range(f"{CN.COL_QTY}{rr}").Value  = None

    for i, r in enumerate(pump[:3]):
        rr = CN.PUMP_START + i
        ws.Range(f"{CN.COL_SPEC}{rr}").Value  = r["spec"]
        ws.Range(f"{CN.COL_PRICE}{rr}").Value = r["price"]
        ws.Range(f"{CN.COL_QTY}{rr}").Value   = r["qty"]

    for rr in range(CN.RACK_START, CN.RACK_END + 1):
        ws.Range(f"{CN.COL_SPEC}{rr}").Value = None
        ws.Range(f"{CN.COL_PRICE}{rr}").Value = None
        ws.Range(f"{CN.COL_QTY}{rr}").Value  = None

    out_list = others[:20]
    for cr in credits:
        out_list.append({"spec": cr["spec"], "qty": 0.0, "price": cr["amt"]})
    out_list = out_list[:20]

    for i, r in enumerate(out_list):
        rr = CN.RACK_START + i
        ws.Range(f"{CN.COL_SPEC}{rr}").Value  = r["spec"]
        ws.Range(f"{CN.COL_PRICE}{rr}").Value = r["price"]
        ws.Range(f"{CN.COL_QTY}{rr}").Value   = None if float(r.get("qty", 0)) == 0 else r["qty"]

    xl.recalc()
    ExcelCOM.hide_blank_rows(ws, CN.COL_SPEC, CN.RACK_START, CN.RACK_END)
    ExcelCOM.show_only(wb, [SheetName.QUOTE_CN])

    ymd    = datetime.now().strftime("%y%m%d")
    line   = safe_filename(info.get("line", ""))
    tool   = safe_filename(info.get("tool", ""))
    folder = ensure_dir(os.path.join(exe_dir(), f"{ymd}_중국_SCS_{line}"))
    return os.path.join(folder, f"{ymd}_중국_SCS_{line}_{tool}.xlsx")


def _fill_us(xl: ExcelCOM, wb, state: QuoteState, items: List[Dict[str, Any]]) -> str:
    """미국 견적서 시트에 데이터 기입."""
    info = state.us_info
    ws = wb.Worksheets(SheetName.QUOTE_US)

    ws.Range(US.MAKER_CELL).Value    = info.get("maker", "")
    ws.Range(US.PROCESS_CELL).Value  = info.get("process", "")
    ws.Range(US.TOOL_CELL).Value     = info.get("tool", "")
    ws.Range(US.EXCHANGE_CELL).Value = float(info.get("exchange", 0))
    y, m = info.get("base_ym", (datetime.now().year, datetime.now().month))
    ws.Range(US.DATE_CELL).Value = f"{int(y)}-{int(m):02d}-1"

    pump    = [r for r in items if r["role"] != "CREDIT" and r["cat"] == "PUMP"]
    others  = [r for r in items if r["role"] != "CREDIT" and r["cat"] != "PUMP"]
    credits = [r for r in items if r["role"] == "CREDIT"]

    ws.Range(US.PUMP_SPEC).Value  = None
    ws.Range(US.PUMP_PRICE).Value = None
    ws.Range(US.PUMP_QTY).Value   = None
    if pump:
        ws.Range(US.PUMP_SPEC).Value  = pump[0]["spec"]
        ws.Range(US.PUMP_PRICE).Value = pump[0]["price"]
        ws.Range(US.PUMP_QTY).Value   = pump[0]["qty"]

    for rr in range(US.RACK_START, US.RACK_END + 1):
        ws.Range(f"{US.COL_SPEC}{rr}").Value  = None
        ws.Range(f"{US.COL_PRICE}{rr}").Value = None
        ws.Range(f"{US.COL_QTY}{rr}").Value   = None

    out_list = others[:20]
    for cr in credits:
        out_list.append({"spec": cr["spec"], "qty": 0.0, "price": cr["amt"]})
    out_list = out_list[:20]

    for i, r in enumerate(out_list):
        rr = US.RACK_START + i
        ws.Range(f"{US.COL_SPEC}{rr}").Value  = r["spec"]
        ws.Range(f"{US.COL_PRICE}{rr}").Value = r["price"]
        ws.Range(f"{US.COL_QTY}{rr}").Value   = None if float(r.get("qty", 0)) == 0 else r["qty"]

    xl.recalc()
    ExcelCOM.hide_blank_rows(ws, US.COL_SPEC, US.RACK_START, US.RACK_END)
    ExcelCOM.show_only(wb, [SheetName.QUOTE_US])

    ymd    = datetime.now().strftime("%y%m%d")
    site   = safe_filename(info.get("site", ""))
    tool   = safe_filename(info.get("tool", ""))
    folder = ensure_dir(os.path.join(exe_dir(), f"{ymd}_미국_{site}"))
    return os.path.join(folder, f"{ymd}_{site}_{tool}_Quotation.xlsx")


# ═══════════════════════════════════════════════
# 공개 API: 견적서 생성
# ═══════════════════════════════════════════════

def generate_quote(
    state       : QuoteState,
    qtype       : str,
    items       : List[Dict[str, Any]],
    rd          : Dict[str, Any],
) -> str:
    """
    견적서 1건 생성 → 저장 → 경로 반환.
    템플릿을 열고 데이터를 기입한 뒤 새 경로로 SaveAs.
    """
    with ExcelCOM() as xl:
        wb = xl.open(state.template_path)
        xl.recalc()
        try:
            if qtype == "중국":
                path = _fill_china(xl, wb, state, items)
            elif qtype == "미국":
                path = _fill_us(xl, wb, state, items)
            else:
                path = _fill_domestic(xl, wb, state, rd, items)

            path = unique_path(path)
            ExcelCOM.close_if_open(xl.app, path)
            wb.SaveAs(path)
            return path
        finally:
            try:
                wb.Close(False)
            except Exception:
                pass


def generate_quote_multi(
    state           : QuoteState,
    items           : List[Dict[str, Any]],
    request_rows    : List[Dict[str, Any]],
    progress_cb     : Optional[Callable[[int, int, str], None]] = None,
) -> List[Tuple[Dict[str, Any], str]]:
    """
    국내 견적서 여러 건을 COM 인스턴스 1개로 순서대로 생성.
    반환: [(rd, saved_path), ...]  (실패하면 path="")
    progress_cb(done, total, basename)
    """
    results: List[Tuple[Dict[str, Any], str]] = []
    total = len(request_rows)
    with ExcelCOM() as xl:
        for i, rd in enumerate(request_rows):
            wb = None
            try:
                wb = xl.open(state.template_path)
                xl.recalc()
                path = _fill_domestic(xl, wb, state, rd, items)
                path = unique_path(path)
                ExcelCOM.close_if_open(xl.app, path)
                wb.SaveAs(path)
                results.append((rd, path))
                logger.info("멀티 생성: %s", path)
                if progress_cb:
                    progress_cb(i + 1, total, os.path.basename(path))
            except Exception as e:
                logger.error("멀티 생성 실패: %s", e, exc_info=True)
                results.append((rd, ""))
                if progress_cb:
                    progress_cb(i + 1, total, f"[실패] {e}")
            finally:
                if wb is not None:
                    try:
                        wb.Close(False)
                    except Exception:
                        pass
    return results


def generate_cover(
    template_path   : str,
    folder          : str,
    source_files    : List[str],
    investor_name   : str = "채승철",
    progress_cb     : Optional[Callable[[int, int, str], None]] = None,
) -> str:
    """
    갑지 생성.
    source_files 각각의 견적의뢰복사본 2행 데이터를 갑지DATA에 쌓아
    '폴더명_갑지.xlsx' 로 저장.
    """
    folder_name = os.path.basename(folder.rstrip("\\/"))
    out_path = unique_path(os.path.join(folder, f"{folder_name}_갑지.xlsx"))

    # 입력 파일 목록에서 출력 파일과 겹치는 항목 제거
    input_abs = {os.path.abspath(p).lower() for p in source_files}
    while os.path.abspath(out_path).lower() in input_abs:
        out_path = unique_path(out_path)

    out_base = os.path.basename(out_path).lower()
    clean = sorted(
        [p for p in source_files if os.path.basename(p).lower() != out_base],
        key=lambda p: os.path.basename(p).lower(),
    )
    if not clean:
        raise RuntimeError("처리할 엑셀 파일이 없습니다.")

    with ExcelCOM() as xl:
        wb = xl.open(template_path)
        try:
            ws_data  = wb.Worksheets(SheetName.COVER_DATA)
            ws_cover = wb.Worksheets(SheetName.COVER)

            try:
                ws_data.Range("A2:AA2000").Value = None
            except Exception:
                pass

            write_row = 2
            total_src = len(clean)
            for idx, path in enumerate(clean):
                src = None
                try:
                    if progress_cb:
                        progress_cb(idx + 1, total_src, os.path.basename(path))
                    src = xl.open(path, read_only=True)
                    vals = src.Worksheets(SheetName.REQ_COPY).Range("A2:AA2").Value
                    ws_data.Range(f"A{write_row}:AA{write_row}").Value = vals
                    write_row += 1
                except Exception as e:
                    logger.warning("갑지 원본 읽기 실패 (%s): %s", path, e)
                finally:
                    if src is not None:
                        try:
                            src.Close(False)
                        except Exception:
                            pass

            xl.recalc()

            ws_cover.Range("B7").Value = f"삼성전자 ㈜ / {investor_name} 님"

            for r in range(20, 50):
                try:
                    v = ws_cover.Range(f"B{r}").Value
                    blank = (
                        v is None
                        or (isinstance(v, (int, float)) and float(v) == 0)
                        or str(v).strip().replace(",", "") in ("", "0")
                    )
                    ws_cover.Rows(r).Hidden = blank
                except Exception:
                    pass

            ExcelCOM.show_only(wb, [SheetName.COVER_DATA, SheetName.COVER])
            ExcelCOM.close_if_open(xl.app, out_path)
            wb.SaveAs(out_path)
            return out_path
        finally:
            try:
                wb.Close(False)
            except Exception:
                pass


# ═══════════════════════════════════════════════
# 공개 API: 전자서명용 임시 PDF 생성
# ═══════════════════════════════════════════════

def excel_to_merged_pdf(xlsx_path: str, tmp_dir: str, file_index: int) -> str:
    """
    xlsx의 특정 4개 시트(Visible인 것만)를 개별 PDF로 내보낸 후 merge.
    반환: 병합된 PDF 경로 (시트가 없으면 "")
    """
    import fitz  # PyMuPDF

    out_pdf = os.path.join(tmp_dir, f"tmp_{file_index:03d}.pdf")

    with ExcelCOM() as xl:
        wb = xl.open(xlsx_path, read_only=True)
        try:
            merged = fitz.open()
            for name in SheetName.ESIGN_TARGET:
                try:
                    ws = wb.Worksheets(name)
                except Exception:
                    continue
                if int(ws.Visible) != -1:   # -1 = xlSheetVisible
                    continue
                safe_name = "".join(c if c not in r'\/:*?"<>|' else "_" for c in name)
                sheet_pdf = os.path.join(tmp_dir, f"tmp_{file_index:03d}_{safe_name}.pdf")
                try:
                    ws.ExportAsFixedFormat(0, sheet_pdf)
                except Exception as e:
                    logger.warning("PDF Export 실패 (%s): %s", name, e)
                    continue
                if os.path.exists(sheet_pdf):
                    src = fitz.open(sheet_pdf)
                    merged.insert_pdf(src)
                    src.close()

            if len(merged) == 0:
                merged.close()
                return ""
            merged.save(out_pdf)
            merged.close()
            return out_pdf
        finally:
            try:
                wb.Close(False)
            except Exception:
                pass
