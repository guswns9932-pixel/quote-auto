"""
pages.py
========
세 개의 메인 페이지와 옵션 다이얼로그.
비즈니스 로직은 core / excel_io 로 위임한다.
"""

from __future__ import annotations

import logging
import os
import traceback
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from PySide6.QtCore import Qt, QPointF
from PySide6.QtGui import QBrush, QColor, QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView, QCheckBox, QComboBox, QDialog,
    QDialogButtonBox, QDoubleSpinBox, QFileDialog,
    QFormLayout, QFrame, QGridLayout, QHBoxLayout,
    QHeaderView, QInputDialog, QLabel, QLineEdit,
    QListWidget, QListWidgetItem, QMessageBox,
    QPushButton, QSizePolicy, QSpinBox, QTabWidget,
    QTableWidgetItem, QTextEdit,
    QVBoxLayout, QWidget,
)

import excel_io
from core import (
    LogDB, LogEntry, QuoteState, SheetName,
    ensure_dir, exe_dir, fmt_krw, fmt_qty,
    normalize_token, s, safe_filename, to_float, unique_path,
)
from widgets import (
    DraggableItemsTable, DroppableQuoteTable, PdfView,
    SignatureItem, bold_label, centered_checkbox,
    get_checkbox_from_cell, info_label, labeled_frame,
    PasswordDialog,
)

logger = logging.getLogger("QuoteApp")

PROCESS_CHOICES = ["", "CVD", "DIFF", "IMP", "ETCH", "METAL", "CLEAN", "GCS"]
VENDOR_CHOICES  = [
    "", "AMAT", "ASM", "AXCELIS", "EUGENETECH", "WONIK_IPS",
    "KOKUSAI", "LAM", "SEMES", "TEL", "TES", "ULVAC", "GCS",
]


# ══════════════════════════════════════════════
# 옵션 다이얼로그 (Credit / 중국 / 미국)
# ══════════════════════════════════════════════

class OptionsDialog(QDialog):
    def __init__(self, parent, state: QuoteState, focus: str = "credit") -> None:
        super().__init__(parent)
        self.setWindowTitle("옵션 입력")
        self.state = state
        self.setModal(True)
        self.resize(520, 340)

        root = QVBoxLayout(self)
        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        # Credit 탭
        tab_cr = QWidget()
        f1 = QFormLayout(tab_cr)
        self._pump = QLineEdit(str(int(state.pump_credit)) if state.pump_credit else "")
        self._rack = QLineEdit(str(int(state.rack_credit)) if state.rack_credit else "")
        for ed in (self._pump, self._rack):
            ed.setPlaceholderText("미입력 시 0(무시)")
        f1.addRow("Pump Credit(원)", self._pump)
        f1.addRow("Rack Credit(원)", self._rack)
        self.tabs.addTab(tab_cr, "Credit")

        # 중국 탭
        tab_cn = QWidget()
        f2 = QFormLayout(tab_cn)
        cn = state.cn_info or {}
        self._cn_maker   = QLineEdit(cn.get("maker", ""))
        self._cn_process = QLineEdit(cn.get("process", ""))
        self._cn_tool    = QLineEdit(cn.get("tool", ""))
        self._cn_line    = QLineEdit(cn.get("line", ""))
        f2.addRow("Maker",    self._cn_maker)
        f2.addRow("Process",  self._cn_process)
        f2.addRow("설비호기", self._cn_tool)
        f2.addRow("라인",     self._cn_line)
        self.tabs.addTab(tab_cn, "중국")

        # 미국 탭
        tab_us = QWidget()
        f3 = QFormLayout(tab_us)
        us = state.us_info or {}
        self._us_maker   = QLineEdit(us.get("maker", ""))
        self._us_process = QLineEdit(us.get("process", ""))
        self._us_tool    = QLineEdit(us.get("tool", ""))
        self._exchange   = QDoubleSpinBox()
        self._exchange.setRange(0, 1e9); self._exchange.setDecimals(2)
        self._exchange.setValue(float(us.get("exchange", 0) or 0))
        now = datetime.now()
        y, m = us.get("base_ym", (now.year, now.month))
        self._year  = QSpinBox(); self._year.setRange(2000, 2100); self._year.setValue(int(y))
        self._month = QSpinBox(); self._month.setRange(1, 12);      self._month.setValue(int(m))
        self._site  = QComboBox(); self._site.addItems(["Taylor", "Austin"])
        site = us.get("site", "Taylor")
        self._site.setCurrentText(site if site in ("Taylor", "Austin") else "Taylor")
        ym_row = QWidget()
        ym_h = QHBoxLayout(ym_row); ym_h.setContentsMargins(0,0,0,0)
        ym_h.addWidget(self._year);  ym_h.addWidget(QLabel("년"))
        ym_h.addWidget(self._month); ym_h.addWidget(QLabel("월"))
        ym_h.addStretch(1)
        f3.addRow("Maker",         self._us_maker)
        f3.addRow("Process",       self._us_process)
        f3.addRow("설비호기",      self._us_tool)
        f3.addRow("환율",          self._exchange)
        f3.addRow("기준일(년/월)", ym_row)
        f3.addRow("Site",          self._site)
        self.tabs.addTab(tab_us, "미국")

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        root.addWidget(bb)
        self.tabs.setCurrentIndex({"credit": 0, "cn": 1, "us": 2}.get(focus, 0))

    @staticmethod
    def _money(text: str) -> float:
        t = (text or "").strip().replace(",", "")
        try:
            return float(t) if t else 0.0
        except ValueError:
            return 0.0

    def apply(self) -> None:
        self.state.pump_credit = self._money(self._pump.text())
        self.state.rack_credit = self._money(self._rack.text())
        self.state.cn_info = {
            "maker": self._cn_maker.text().strip(), "process": self._cn_process.text().strip(),
            "tool":  self._cn_tool.text().strip(),  "line":    self._cn_line.text().strip(),
        }
        self.state.us_info = {
            "maker": self._us_maker.text().strip(), "process": self._us_process.text().strip(),
            "tool":  self._us_tool.text().strip(),  "exchange": float(self._exchange.value()),
            "base_ym": (int(self._year.value()), int(self._month.value())),
            "site": self._site.currentText(),
        }


# ══════════════════════════════════════════════
# STEP5 테이블 관리 믹스인
# ══════════════════════════════════════════════

class Step5Manager:
    """
    STEP5 테이블의 행 CRUD + TOTAL/CREDIT 행 관리.
    QuoteBuilderPage 가 다중 상속하여 사용한다.

    하위 클래스가 반드시 제공해야 하는 속성:
        self.step5_table      : DroppableQuoteTable
        self.chk_step5_all    : QCheckBox
        self.state            : QuoteState
        self._step4_highlight : set
    하위 클래스가 구현해야 하는 메서드:
        self._render_items_table(scroll_top: bool)
    """

    # ── 행 종류 판별 ────────────────────────────
    def _row_role(self, row: int) -> str:
        it = self.step5_table.item(row, 1)
        return s(it.data(Qt.UserRole)) if it else ""

    def _is_total(self, row: int)  -> bool: return self._row_role(row) == "TOTAL"
    def _is_credit(self, row: int) -> bool: return self._row_role(row) == "CREDIT"
    def _is_item(self, row: int)   -> bool: return self._row_role(row) == "ITEM"

    # ── 셀 setter ───────────────────────────────
    def _set_qty(self, row: int, qty: float, blank: bool = False) -> None:
        it = QTableWidgetItem("" if blank else fmt_qty(qty))
        it.setData(Qt.UserRole, qty); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(row, 3, it)

    def _set_price(self, row: int, price: float, blank: bool = False) -> None:
        it = QTableWidgetItem("" if blank else fmt_krw(price))
        it.setData(Qt.UserRole, price); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(row, 4, it)

    def _set_amt(self, row: int, amt: float) -> None:
        it = QTableWidgetItem(fmt_krw(amt))
        it.setData(Qt.UserRole, amt); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(row, 5, it)

    def _get_float(self, row: int, col: int) -> float:
        it = self.step5_table.item(row, col)
        return float(it.data(Qt.UserRole) or 0.0) if it else 0.0

    # ── TOTAL 행 보장 ────────────────────────────
    def _ensure_total(self) -> None:
        n = self.step5_table.rowCount()
        if n == 0 or not self._is_total(n - 1):
            self._append_total()

    def _append_total(self) -> None:
        r = self.step5_table.rowCount()
        self.step5_table.insertRow(r)
        it_cat = QTableWidgetItem("TOTAL")
        it_cat.setData(Qt.UserRole, "TOTAL"); it_cat.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(r, 1, it_cat)
        it_spec = QTableWidgetItem("총 합계")
        it_spec.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(r, 2, it_spec)
        self._set_qty(r, 0, blank=True); self._set_price(r, 0, blank=True); self._set_amt(r, 0)

    # ── 합계 재계산 ──────────────────────────────
    def _recalc_totals(self) -> None:
        self._ensure_total()
        total_qty = total_amt = 0.0
        last = self.step5_table.rowCount() - 1
        for r in range(last):
            if self._is_total(r): continue
            total_amt += self._get_float(r, 5)
            if not self._is_credit(r):
                total_qty += self._get_float(r, 3)
        self._set_qty(last, total_qty, blank=True)
        self._set_amt(last, total_amt)

    def _recalc_row(self, row: int) -> None:
        self._set_amt(row, self._get_float(row, 3) * self._get_float(row, 4))

    # ── 행 추가 ─────────────────────────────────
    def _add_row(self, category: str, spec: str, qty: float, unit_price: float) -> None:
        """같은 spec 이면 수량 누적, 없으면 TOTAL 행 위에 새 행 삽입."""
        self._ensure_total()
        last = self.step5_table.rowCount() - 1
        for r in range(last):
            if self._is_total(r) or self._is_credit(r): continue
            it = self.step5_table.item(r, 2)
            if it and it.text().strip() == spec.strip():
                self._set_qty(r, self._get_float(r, 3) + qty)
                self._recalc_row(r); self._recalc_totals(); return

        ins = last
        self.step5_table.insertRow(ins)
        cb = QCheckBox()
        cb.stateChanged.connect(self._on_step5_check_changed)
        self.step5_table.setCellWidget(ins, 0, cb)
        it_cat = QTableWidgetItem(category)
        it_cat.setData(Qt.UserRole, "ITEM"); it_cat.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(ins, 1, it_cat)
        it_spec = QTableWidgetItem(spec)
        it_spec.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(ins, 2, it_spec)
        self._set_qty(ins, qty); self._set_price(ins, unit_price)
        self._set_amt(ins, qty * unit_price); self._recalc_totals()

    # ── Credit 행 ────────────────────────────────
    def _clear_credit_rows(self) -> None:
        for r in range(self.step5_table.rowCount() - 1, -1, -1):
            if self._is_credit(r): self.step5_table.removeRow(r)
        self._ensure_total()

    def _add_credit_row(self, label: str, amount: float) -> None:
        self._ensure_total()
        ins = self.step5_table.rowCount() - 1
        self.step5_table.insertRow(ins)
        red = QBrush(QColor(200, 0, 0))
        it_cat = QTableWidgetItem("CREDIT")
        it_cat.setData(Qt.UserRole, "CREDIT"); it_cat.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        self.step5_table.setItem(ins, 1, it_cat)
        it_spec = QTableWidgetItem(label)
        it_spec.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable); it_spec.setForeground(red)
        self.step5_table.setItem(ins, 2, it_spec)
        self.step5_table.setItem(ins, 3, QTableWidgetItem(""))
        self.step5_table.setItem(ins, 4, QTableWidgetItem(""))
        it_amt = QTableWidgetItem(fmt_krw(amount))
        it_amt.setData(Qt.UserRole, float(amount)); it_amt.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        it_amt.setForeground(red); self.step5_table.setItem(ins, 5, it_amt)

    def _refresh_credits(self) -> None:
        self._clear_credit_rows()
        if self.state.pump_credit: self._add_credit_row("Pump Credit", -abs(self.state.pump_credit))
        if self.state.rack_credit: self._add_credit_row("Rack Credit", -abs(self.state.rack_credit))

    # ── STEP5 스냅샷 ─────────────────────────────
    def _snapshot_items(self) -> List[Dict[str, Any]]:
        rows = []
        for r in range(self.step5_table.rowCount()):
            if self._is_total(r): continue
            cat_it  = self.step5_table.item(r, 1)
            spec_it = self.step5_table.item(r, 2)
            rows.append({
                "role":  s(cat_it.data(Qt.UserRole)) if cat_it else "",
                "cat":   s(cat_it.text())             if cat_it else "",
                "spec":  s(spec_it.text())            if spec_it else "",
                "qty":   self._get_float(r, 3),
                "price": self._get_float(r, 4),
                "amt":   self._get_float(r, 5),
            })
        return rows

    # ── 전체선택 동기화 ──────────────────────────
    def _sync_step5_all_chk(self) -> None:
        selectable = checked = 0
        last = self.step5_table.rowCount() - 1
        for r in range(last + 1):
            if self._is_total(r) or self._is_credit(r): continue
            cb = self.step5_table.cellWidget(r, 0)
            if isinstance(cb, QCheckBox):
                selectable += 1
                if cb.isChecked(): checked += 1
        self.chk_step5_all.blockSignals(True)
        if   selectable == 0 or checked == 0: self.chk_step5_all.setCheckState(Qt.Unchecked)
        elif checked == selectable:            self.chk_step5_all.setCheckState(Qt.Checked)
        else:                                  self.chk_step5_all.setCheckState(Qt.PartiallyChecked)
        self.chk_step5_all.blockSignals(False)

    def _on_step5_check_changed(self, _state: int) -> None:
        """개별 체크박스 변경 → 전체선택 상태 동기화."""
        self._sync_step5_all_chk()

    def _on_step5_select_all(self, checked: bool) -> None:
        """전체선택 체크박스 핸들러."""
        last = self.step5_table.rowCount() - 1
        for r in range(last + 1):
            if self._is_total(r) or self._is_credit(r): continue
            cb = self.step5_table.cellWidget(r, 0)
            if isinstance(cb, QCheckBox):
                cb.blockSignals(True); cb.setChecked(checked); cb.blockSignals(False)
        self._sync_step5_all_chk()

    # ── STEP4 하이라이트 동기화 ──────────────────
    def _sync_step4_highlight(self, scroll_top: bool = True) -> None:
        transferred: set = set()
        for r in range(self.step5_table.rowCount()):
            if self._is_total(r) or self._is_credit(r): continue
            it = self.step5_table.item(r, 2)
            if it and it.text().strip():
                transferred.add(it.text().strip())
        self._step4_highlight = transferred
        self._render_items_table(scroll_top=scroll_top)

    def _render_items_table(self, scroll_top: bool = True) -> None:
        """하위 클래스(QuoteBuilderPage)에서 구현."""
        raise NotImplementedError("_render_items_table must be implemented by subclass")


# ══════════════════════════════════════════════
# 견적서 작성 페이지
# ══════════════════════════════════════════════

class QuoteBuilderPage(Step5Manager, QWidget):

    def __init__(self, db: LogDB) -> None:
        QWidget.__init__(self)
        self.db    = db
        self.state = QuoteState()

        self._step4_highlight : set                     = set()
        self._filtered_items  : List[Tuple[str, float]] = []
        self._spec_order      : Dict[str, int]          = {}

        self._build_ui()
        self._ensure_total()
        self._log("준비: STEP1 → STEP2 → STEP4/드래그 → STEP5 → 생성")

        sc = QShortcut(QKeySequence.Delete, self)
        sc.activated.connect(self._delete_checked_rows)

    # ══════════════════════════════════════════
    # UI 구성
    # ══════════════════════════════════════════

    def _build_ui(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(14, 14, 14, 14)
        outer.setSpacing(12)

        grid = QGridLayout()
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(12)

        # Row 0: 버튼 4개
        self.btn_step1 = self._action_btn("STEP 1\n통합양식 LOAD", self._load_template)
        self.btn_step2 = self._action_btn("STEP 2\n의뢰파일 LOAD", self._load_request)
        self.btn_step2.setEnabled(False)
        self._step3_frame = self._build_step3()
        self._step3_frame.setEnabled(False)
        guide = labeled_frame("견적서작성 사용법", min_h=110)
        guide.layout().addWidget(QLabel(
            "① STEP1: 통합양식 LOAD\n② STEP2: 의뢰파일 LOAD\n"
            "③ 의뢰행 더블클릭 → 조건 자동입력\n"
            "④ STEP4 더블클릭/드래그 → STEP5\n⑤ 견적서 생성 버튼"))
        grid.addWidget(self.btn_step1,    0, 0)
        grid.addWidget(self.btn_step2,    0, 1)
        grid.addWidget(self._step3_frame, 0, 2)
        grid.addWidget(guide,             0, 3)

        # Row 1: 의뢰 / STEP4 / 우측
        grid.addWidget(self._build_req_panel(),   1, 0)
        grid.addWidget(self._build_step4_panel(), 1, 1)
        grid.addWidget(self._build_right_panel(), 1, 2, 1, 2)

        # Row 2: 로그 / 옵션+견적정보
        grid.addWidget(self._build_worklog(),            2, 0, 1, 2)
        grid.addWidget(self._build_bottom_right_panel(), 2, 2, 1, 2)

        grid.setColumnStretch(0, 3); grid.setColumnStretch(1, 5)
        grid.setColumnStretch(2, 5); grid.setColumnStretch(3, 2)
        grid.setRowStretch(1, 10)
        outer.addLayout(grid, 1)

    @staticmethod
    def _action_btn(label: str, slot) -> QPushButton:
        btn = QPushButton(label); btn.setMinimumHeight(110)
        f = btn.font(); f.setPointSize(12); f.setBold(True); btn.setFont(f)
        btn.clicked.connect(slot); return btn

    # ── STEP3 ─────────────────────────────────
    def _build_step3(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2); frame.setMinimumHeight(110)
        v = QVBoxLayout(frame); v.setContentsMargins(12,12,12,12); v.setSpacing(6)
        v.addWidget(bold_label("STEP 3\n조건 입력"))
        row1 = QHBoxLayout()
        self.cb_process = QComboBox(); self.cb_process.addItems(PROCESS_CHOICES)
        self.cb_vendor  = QComboBox(); self.cb_vendor.addItems(VENDOR_CHOICES)
        row1.addWidget(QLabel("공정")); row1.addWidget(self.cb_process, 1)
        row1.addWidget(QLabel("설비사")); row1.addWidget(self.cb_vendor, 1)
        row2 = QHBoxLayout()
        self.ed_code = QLineEdit(); self.ed_code.setPlaceholderText("5D")
        self.btn_apply_map = QPushButton("코드매핑 적용 → STEP5(RACK)"); self.btn_apply_map.setEnabled(False)
        row2.addWidget(QLabel("5D")); row2.addWidget(self.ed_code, 1); row2.addWidget(self.btn_apply_map)
        v.addLayout(row1); v.addLayout(row2)
        self.cb_process.currentTextChanged.connect(self._on_step3_changed)
        self.cb_vendor.currentTextChanged.connect(self._on_step3_changed)
        self.ed_code.textChanged.connect(self._on_step3_changed)
        self.btn_apply_map.clicked.connect(self._apply_code_map)
        return frame

    # ── 의뢰파일 패널 ─────────────────────────
    def _build_req_panel(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2)
        frame.setMinimumHeight(420); frame.setFixedWidth(230)
        v = QVBoxLayout(frame); v.setContentsMargins(12,12,12,12); v.setSpacing(8); v.setAlignment(Qt.AlignTop)
        v.addWidget(bold_label("의뢰파일DATA", size=12))
        top = QHBoxLayout()
        self.chk_req_all = QCheckBox("전체선택"); self.chk_req_all.clicked.connect(self._on_req_select_all)
        top.addWidget(self.chk_req_all); top.addStretch(1); v.addLayout(top)
        self.req_table = self._plain_table(3, ["선택", "설비호기(Z)", "견적서작성여부"])
        self.req_table.setColumnWidth(0, 42)
        self.req_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.req_table.cellDoubleClicked.connect(self._on_req_double_click)
        self._style_req_table()
        v.addWidget(self.req_table, 0, Qt.AlignTop); v.addStretch(1)
        return frame

    # ── STEP4 패널 ────────────────────────────
    def _build_step4_panel(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2); frame.setMinimumHeight(420)
        v = QVBoxLayout(frame); v.setContentsMargins(12,12,12,12); v.setSpacing(8)
        v.addWidget(bold_label("STEP 4\n품목(규격/단가) + 필터  (더블클릭/드래그 → STEP5)"))
        self.ed_filter = QLineEdit(); self.ed_filter.setPlaceholderText("규격 필터(실시간) - 포함 검색")
        self.ed_filter.textChanged.connect(self._apply_filter)
        self.items_table = DraggableItemsTable(0, 2)
        self.items_table.setHorizontalHeaderLabels(["규격", "단가(원)"])
        self.items_table.verticalHeader().setVisible(False)
        self.items_table.setAlternatingRowColors(True)
        self.items_table.setEditTriggers(DraggableItemsTable.NoEditTriggers)
        self.items_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.items_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.items_table.setDragEnabled(True); self.items_table.setWordWrap(False)
        self.items_table.setStyleSheet("QTableWidget::item:selected{color:black;}")
        h = self.items_table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.Stretch); h.setSectionResizeMode(1, QHeaderView.Fixed)
        self.items_table.setColumnWidth(1, 80)
        self.items_table.cellDoubleClicked.connect(self._on_item_double_click)
        v.addWidget(self.ed_filter); v.addWidget(self.items_table, 1)
        return frame

    # ── 우측 패널 (STEP5 + 버튼 + STEP7) ──────
    def _build_right_panel(self) -> QWidget:
        w = QWidget(); g = QGridLayout(w)
        g.setContentsMargins(0,0,0,0); g.setHorizontalSpacing(12)

        step5 = labeled_frame("STEP5 견적작성 대상", min_h=320)
        step5.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        sl = step5.layout()
        top5 = QHBoxLayout()
        self.chk_step5_all = QCheckBox("전체선택")
        self.chk_step5_all.clicked.connect(self._on_step5_select_all)
        top5.addWidget(self.chk_step5_all); top5.addStretch(1); sl.addLayout(top5)

        self.step5_table = DroppableQuoteTable()
        self.step5_table.setColumnCount(6)
        self.step5_table.setHorizontalHeaderLabels(["선택","분류","품목","수량","단가(원)","금액(원)"])
        self.step5_table.verticalHeader().setVisible(False)
        self.step5_table.setAlternatingRowColors(True)
        self.step5_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.step5_table.setEditTriggers(DroppableQuoteTable.NoEditTriggers)
        self.step5_table.cellDoubleClicked.connect(self._on_step5_double_click)
        self.step5_table.drop_callback = self._on_drop_to_step5
        h5 = self.step5_table.horizontalHeader()
        for i, mode in enumerate([QHeaderView.ResizeToContents]*2 + [QHeaderView.Stretch] + [QHeaderView.ResizeToContents]*3):
            h5.setSectionResizeMode(i, mode)
        self.step5_table.setColumnWidth(0, 46)
        sl.addWidget(self.step5_table, 1)

        side = QWidget(); sv = QVBoxLayout(side); sv.setContentsMargins(0,0,0,0); sv.setSpacing(12)
        sv.addWidget(self._build_generate_panel()); sv.addWidget(self._build_step7(), 1)
        g.addWidget(step5, 0, 0); g.addWidget(side, 0, 1)
        g.setColumnStretch(0, 4); g.setColumnStretch(1, 2)
        return w

    def _build_generate_panel(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2); frame.setMinimumHeight(110)
        h = QHBoxLayout(frame); h.setContentsMargins(6,6,6,6); h.setSpacing(6)
        for label, slot in [("견적서 생성", self._generate_quote), ("갑지 생성", self._generate_cover)]:
            btn = QPushButton(label); btn.setMinimumHeight(90)
            f = btn.font(); f.setPointSize(12); f.setBold(True); btn.setFont(f)
            btn.clicked.connect(slot); h.addWidget(btn, 1)
        return frame

    def _build_step7(self) -> QFrame:
        frame = labeled_frame("STEP 7\n완성된 견적서 List", min_h=260)
        self.list_done = QListWidget()
        self.list_done.itemDoubleClicked.connect(self._open_done)
        frame.layout().addWidget(self.list_done, 1); return frame

    def _build_worklog(self) -> QFrame:
        frame = labeled_frame("작업로그", min_h=140)
        frame.setFixedHeight(140); frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.log_view = QTextEdit(); self.log_view.setReadOnly(True)
        frame.layout().addWidget(self.log_view, 1); return frame

    def _build_bottom_right_panel(self) -> QWidget:
        w = QWidget(); w.setFixedHeight(140)
        h = QHBoxLayout(w); h.setContentsMargins(0,0,0,0); h.setSpacing(12)
        h.addWidget(self._build_options_panel(), 1)
        h.addWidget(self._build_quote_info_panel())
        return w

    def _build_options_panel(self) -> QFrame:
        frame = labeled_frame("옵션", min_h=55)
        frame.setMaximumHeight(140); frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        row = QHBoxLayout()
        self.btn_credit = QPushButton("Credit"); self.btn_credit.setMinimumHeight(32)
        self.btn_credit.clicked.connect(self._open_credit_dialog)
        self.chk_qt_kr = QCheckBox("국내"); self.chk_qt_cn = QCheckBox("중국"); self.chk_qt_us = QCheckBox("미국")
        self.chk_qt_kr.setChecked(True)
        for chk, qt in [(self.chk_qt_kr,"국내"),(self.chk_qt_cn,"중국"),(self.chk_qt_us,"미국")]:
            chk.clicked.connect(lambda checked, t=qt: self._on_qt_checked(t, checked))
        row.addWidget(self.btn_credit); row.addSpacing(10); row.addWidget(QLabel("견적서 타입"))
        row.addWidget(self.chk_qt_kr); row.addWidget(self.chk_qt_cn); row.addWidget(self.chk_qt_us)
        row.addStretch(1); frame.layout().addLayout(row); return frame

    def _build_quote_info_panel(self) -> QFrame:
        frame = QFrame(); frame.setFrameShape(QFrame.Box); frame.setLineWidth(2)
        frame.setFixedHeight(140); frame.setFixedWidth(260)
        v = QVBoxLayout(frame); v.setContentsMargins(12,6,12,6); v.setSpacing(4)
        v.addWidget(bold_label("견적서정보", size=9))
        self.lbl_pr = info_label("PR: -"); self.lbl_item = info_label("항번: -")
        self.lbl_line = info_label("라인공정: -"); self.lbl_investor = info_label("투자자: -")
        self.lbl_more = info_label("")
        for lbl in (self.lbl_pr, self.lbl_item, self.lbl_line, self.lbl_investor): v.addWidget(lbl)
        v.addStretch(1); v.addWidget(self.lbl_more); return frame

    @staticmethod
    def _plain_table(cols: int, headers: List[str]):
        from PySide6.QtWidgets import QTableWidget
        t = QTableWidget(0, cols); t.setHorizontalHeaderLabels(headers)
        t.verticalHeader().setVisible(False); t.setEditTriggers(QTableWidget.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.setSelectionMode(QAbstractItemView.SingleSelection); t.setWordWrap(False); return t

    def _style_req_table(self) -> None:
        self.req_table.setAlternatingRowColors(False)
        self.req_table.setStyleSheet(
            "QTableWidget{background:transparent;border:none;}"
            "QTableWidget::item{background:transparent;color:black;}"
            "QTableWidget::item:selected{color:black;background:rgba(180,200,230,140);}")

    # ══════════════════════════════════════════
    # STEP1
    # ══════════════════════════════════════════

    def _load_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "통합양식 선택", "", "Excel Files (*.xlsx)")
        if not path: return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=False)
        except Exception as e:
            QMessageBox.critical(self, "오류", f"파일을 열 수 없습니다.\n{e}"); return
        missing = [n for n in SheetName.REQUIRED if n not in wb.sheetnames]
        if missing:
            QMessageBox.warning(self, "통합양식 오류", "필수 시트 없음:\n" + "\n".join(missing)); return
        try:
            rows, by_class, price_by_spec, order = excel_io.parse_items_sheet(wb[SheetName.ITEMS])
            cmap = excel_io.parse_code_map_sheet(wb[SheetName.CODE_MAP])
        except Exception as e:
            logger.error("통합양식 파싱 오류", exc_info=True)
            QMessageBox.critical(self, "파싱 오류", str(e)); return
        self.state.template_path = path; self.state.items_rows = rows
        self.state.items_by_class = by_class; self.state.price_by_spec = price_by_spec
        self.state.code_map = cmap; self._spec_order = order
        self._apply_filter(); self.btn_step2.setEnabled(True); self._step3_frame.setEnabled(True)
        self._log(f"STEP1 완료: {path}  (품목 {len(rows)}건, 코드매핑 키 {len(cmap)}개)")
        QMessageBox.information(self, "완료", "통합양식 LOAD 완료")

    # ══════════════════════════════════════════
    # STEP2
    # ══════════════════════════════════════════

    def _load_request(self) -> None:
        if not self.state.template_path:
            QMessageBox.warning(self, "순서 오류", "STEP1을 먼저 완료하세요."); return
        path, _ = QFileDialog.getOpenFileName(self, "의뢰파일 선택", "", "Excel Files (*.xlsx)")
        if not path: return
        try:
            sheet_name, rows = excel_io.parse_request_xlsx(path)
        except Exception as e:
            logger.error("의뢰파일 파싱 오류", exc_info=True)
            QMessageBox.critical(self, "오류", f"파일을 읽을 수 없습니다.\n{e}"); return
        if not rows:
            QMessageBox.warning(self, "의뢰파일", "읽을 데이터가 없습니다."); return
        self.state.request_path = path; self.state.request_sheet_name = sheet_name
        self.state.request_rows = rows
        self._fill_req_table(rows); self._step3_frame.setEnabled(True)
        self._log(f"STEP2 완료: {path}  ({len(rows)}건, 시트={sheet_name})")
        QMessageBox.information(self, "완료", "의뢰파일 LOAD 완료")

    def _fill_req_table(self, rows: List[Dict[str, Any]]) -> None:
        self.req_table.setRowCount(0)
        self.chk_req_all.blockSignals(True); self.chk_req_all.setCheckState(Qt.Unchecked); self.chk_req_all.blockSignals(False)
        for i, rd in enumerate(rows):
            self.req_table.insertRow(i)
            host = centered_checkbox(lambda _s, _r=i: self._on_req_check())
            self.req_table.setCellWidget(i, 0, host)
            for col, val in [(1, s(rd.get("Z"))), (2, "미작성")]:
                it = QTableWidgetItem(val); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                self.req_table.setItem(i, col, it)
        self.req_table.resizeColumnsToContents(); self.req_table.setColumnWidth(0, 42)
        self._style_req_table(); self._update_quote_info()

    def _is_req_checked(self, row: int) -> bool:
        cb = get_checkbox_from_cell(self.req_table.cellWidget(row, 0))
        return bool(cb and cb.isChecked())

    def _checked_req_rows(self) -> List[int]:
        return [r for r in range(self.req_table.rowCount()) if self._is_req_checked(r)]

    def _on_req_check(self) -> None:
        n = sum(1 for r in range(self.req_table.rowCount()) if self._is_req_checked(r))
        total = self.req_table.rowCount()
        self.chk_req_all.blockSignals(True)
        if   n == 0:     self.chk_req_all.setCheckState(Qt.Unchecked)
        elif n == total: self.chk_req_all.setCheckState(Qt.Checked)
        else:            self.chk_req_all.setCheckState(Qt.PartiallyChecked)
        self.chk_req_all.blockSignals(False); self._update_quote_info()

    def _on_req_select_all(self, checked: bool) -> None:
        for r in range(self.req_table.rowCount()):
            cb = get_checkbox_from_cell(self.req_table.cellWidget(r, 0))
            if cb: cb.blockSignals(True); cb.setChecked(checked); cb.blockSignals(False)
        self._on_req_check()

    def _update_quote_info(self) -> None:
        checked = self._checked_req_rows()
        if not checked or not self.state.request_rows:
            self.lbl_pr.setText("PR: -"); self.lbl_item.setText("항번: -")
            self.lbl_line.setText("라인공정: -"); self.lbl_investor.setText("투자자: -")
            self.lbl_more.setText(""); return
        first = self.state.request_rows[checked[0]]
        self.lbl_pr.setText(f"PR: {s(first.get('D')) or '-'}")
        self.lbl_item.setText(f"항번: {s(first.get('E')) or '-'}")
        self.lbl_line.setText(f"라인공정: {s(first.get('K')) or '-'}")
        self.lbl_investor.setText(f"투자자: {s(first.get('J')) or '-'}")
        self.lbl_more.setText(f"외 {len(checked)-1}건" if len(checked) > 1 else "")

    def _on_req_double_click(self, row: int, _col: int) -> None:
        if row >= len(self.state.request_rows): return
        rd = self.state.request_rows[row]

        def _match(text: str, choices: List[str]) -> Optional[str]:
            t = normalize_token(text)
            if not t: return None
            for c in sorted((x for x in choices if x), key=len, reverse=True):
                if normalize_token(c) in t: return c
            return None

        p = _match(s(rd.get("K")), PROCESS_CHOICES)
        if p:
            idx = self.cb_process.findText(p)
            if idx >= 0: self.cb_process.setCurrentIndex(idx)
        v = _match(s(rd.get("X")), VENDOR_CHOICES)
        if v:
            idx = self.cb_vendor.findText(v)
            if idx >= 0: self.cb_vendor.setCurrentIndex(idx)
        vcode = s(rd.get("V"))
        if vcode: self.ed_code.setText(vcode)

        f_class = s(rd.get("F")); qty = to_float(rd.get("H"))
        if f_class:
            matches = self.state.items_by_class.get(f_class, [])
            for row_data in matches:
                spec = s(row_data.get("B")); price = float(row_data.get("C", 0))
                if spec: self._add_row("PUMP", spec, qty, price)
            if matches:
                self._log(f"PUMP 자동: 분류={f_class}, {len(matches)}개, 수량={fmt_qty(qty)}")
                self._sync_step4_highlight()
        self._log(f"더블클릭 → 공정={self.cb_process.currentText()} / 설비사={self.cb_vendor.currentText()} / 5D={self.ed_code.text()}")

    # ══════════════════════════════════════════
    # STEP3
    # ══════════════════════════════════════════

    def _on_step3_changed(self) -> None:
        self.state.process = s(self.cb_process.currentText()) or None
        self.state.vendor  = s(self.cb_vendor.currentText())  or None
        self.state.code_5d = s(self.ed_code.text())           or None
        ready = bool(self.state.process and self.state.vendor and self.state.code_5d)
        self.btn_apply_map.setEnabled(ready and bool(self.state.code_map))

    def _apply_code_map(self) -> None:
        key = (self.state.process, self.state.vendor, self.state.code_5d)
        entries = self.state.code_map.get(key)
        if not entries:
            near = sum(1 for k in self.state.code_map if k[:2] == key[:2])
            QMessageBox.warning(self, "코드매핑 없음", f"키: {key}\n동일 공정/설비사 키 수: {near}"); return
        for acc, qty in entries:
            price = float(self.state.price_by_spec.get(acc, 0))
            self._add_row("RACK", acc, qty, price)
        self._log(f"코드매핑 적용 {key} → {len(entries)}개"); self._sync_step4_highlight()

    # ══════════════════════════════════════════
    # STEP4
    # ══════════════════════════════════════════

    def _apply_filter(self) -> None:
        key = s(self.ed_filter.text()).lower()
        self._filtered_items = [
            (s(r.get("B")), float(r.get("C", 0)))
            for r in self.state.items_rows
            if s(r.get("B")) and key in s(r.get("B")).lower()
        ]
        self._render_items_table(scroll_top=False)

    def _render_items_table(self, scroll_top: bool = True) -> None:
        """Step5Manager 추상 메서드 구현: STEP4 테이블 재렌더링."""
        highlighted = self._step4_highlight

        def order(spec: str) -> int:
            return self._spec_order.get(spec, 10**9)

        sorted_rows = sorted(
            self._filtered_items,
            key=lambda x: (0 if x[0] in highlighted else 1, order(x[0])))

        self.items_table.blockSignals(True)
        self.items_table.setRowCount(len(sorted_rows))
        for i, (spec, price) in enumerate(sorted_rows):
            for col, val in [(0, spec), (1, fmt_krw(price))]:
                it = QTableWidgetItem(val); it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                if spec in highlighted: it.setBackground(QBrush(QColor(220, 220, 220)))
                self.items_table.setItem(i, col, it)
        self.items_table.blockSignals(False)
        self.items_table.setColumnWidth(1, 80)
        if scroll_top: self.items_table.scrollToTop()

    def _on_item_double_click(self, row: int, _col: int) -> None:
        spec_it  = self.items_table.item(row, 0); price_it = self.items_table.item(row, 1)
        if not spec_it: return
        self._ask_qty_add(spec_it.text().strip(), to_float(price_it.text() if price_it else "0"))

    def _on_drop_to_step5(self, spec: str, price: float) -> None:
        self._ask_qty_add(spec, price)

    def _ask_qty_add(self, spec: str, unit_price: float) -> None:
        qty, ok = QInputDialog.getDouble(self, "수량 입력", f"수량 ({spec})", 1.0, 0, 1_000_000, 3)
        if ok: self._add_row("RACK", spec, qty, unit_price); self._sync_step4_highlight()

    # ══════════════════════════════════════════
    # STEP5 이벤트
    # ══════════════════════════════════════════

    def _on_step5_double_click(self, row: int, col: int) -> None:
        if self._is_total(row) or self._is_credit(row): return
        if col == 1:
            it = self.step5_table.item(row, 1)
            if it:
                it.setText("RACK" if it.text().strip() == "PUMP" else "PUMP")
                self._recalc_totals(); self._sync_step4_highlight(scroll_top=False)
        elif col == 3:
            cur = self._get_float(row, 3)
            val, ok = QInputDialog.getDouble(self, "수량 수정", "수량:", cur, 0, 1_000_000, 3)
            if ok: self._set_qty(row, val); self._recalc_row(row); self._recalc_totals(); self._sync_step4_highlight()

    def _delete_checked_rows(self) -> None:
        to_del = [
            r for r in range(self.step5_table.rowCount() - 1)
            if not self._is_total(r) and not self._is_credit(r)
            and isinstance(self.step5_table.cellWidget(r, 0), QCheckBox)
            and self.step5_table.cellWidget(r, 0).isChecked()
        ]
        if not to_del:
            sel = self.step5_table.selectionModel().selectedRows()
            if sel:
                r = sel[0].row()
                if not self._is_total(r) and not self._is_credit(r): to_del = [r]
        if not to_del: return
        reply = QMessageBox.question(self, "삭제 확인", f"{len(to_del)}개 품목을 삭제할까요?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes: return
        for r in sorted(to_del, reverse=True): self.step5_table.removeRow(r)
        self._ensure_total(); self._refresh_credits(); self._recalc_totals()
        self.chk_step5_all.blockSignals(True); self.chk_step5_all.setCheckState(Qt.Unchecked); self.chk_step5_all.blockSignals(False)
        self._sync_step4_highlight()

    # ══════════════════════════════════════════
    # 옵션
    # ══════════════════════════════════════════

    def _on_qt_checked(self, selected: str, checked: bool) -> None:
        chk_map = {"국내": self.chk_qt_kr, "중국": self.chk_qt_cn, "미국": self.chk_qt_us}
        if not checked:
            chk = chk_map[selected]; chk.blockSignals(True); chk.setChecked(True); chk.blockSignals(False); return
        for name, chk in chk_map.items():
            chk.blockSignals(True); chk.setChecked(name == selected); chk.blockSignals(False)
        self.state.quote_type = selected; self._log(f"견적서 타입 = {selected}")
        if selected in ("중국","미국"): self._open_options(focus={"중국":"cn","미국":"us"}[selected])

    def _open_options(self, focus: str = "credit") -> None:
        dlg = OptionsDialog(self, self.state, focus=focus)
        if dlg.exec() == QDialog.Accepted:
            dlg.apply(); self._refresh_credits(); self._recalc_totals()
            self._sync_step4_highlight(scroll_top=False); self._log("[옵션] 입력값 반영")

    def _open_credit_dialog(self) -> None:
        dlg = QDialog(self); dlg.setWindowTitle("Credit 입력"); dlg.setFixedWidth(420)
        v = QVBoxLayout(dlg)
        form = QFormLayout()
        ed_pump = QLineEdit("" if not self.state.pump_credit else fmt_krw(self.state.pump_credit))
        ed_rack = QLineEdit("" if not self.state.rack_credit else fmt_krw(self.state.rack_credit))
        for ed in (ed_pump, ed_rack): ed.setPlaceholderText("미입력 시 0(무시)")
        form.addRow("Pump Credit(원)", ed_pump); form.addRow("Rack Credit(원)", ed_rack)
        v.addLayout(form)
        lbl_before = QLabel(); lbl_after = QLabel(); lbl_after.setStyleSheet("font-weight:bold;")
        v.addWidget(lbl_before); v.addWidget(lbl_after)

        def _money(t: str) -> float:
            try: return float(t.strip().replace(",","")) if t.strip() else 0.0
            except ValueError: return 0.0

        def _refresh() -> None:
            base   = sum(self._get_float(r,5) for r in range(self.step5_table.rowCount()) if not self._is_total(r))
            credit = _money(ed_pump.text()) + _money(ed_rack.text())
            lbl_before.setText(f"현재 총금액: {fmt_krw(base)} 원")
            lbl_after.setText(f"Credit 반영: {fmt_krw(base-credit)} 원  (감액: {fmt_krw(credit)} 원)")

        ed_pump.textChanged.connect(_refresh); ed_rack.textChanged.connect(_refresh)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel); v.addWidget(bb)

        def _ok() -> None:
            self.state.pump_credit = _money(ed_pump.text()); self.state.rack_credit = _money(ed_rack.text())
            self._refresh_credits(); self._recalc_totals(); dlg.accept()

        bb.accepted.connect(_ok); bb.rejected.connect(dlg.reject); _refresh(); dlg.exec()

    # ══════════════════════════════════════════
    # STEP6 – 생성
    # ══════════════════════════════════════════

    def _generate_quote(self) -> None:
        if not self.state.template_path:
            QMessageBox.warning(self, "오류", "STEP1 통합양식을 먼저 LOAD 하세요."); return
        if not excel_io.COM_AVAILABLE:
            QMessageBox.critical(self, "오류", "Excel COM(win32)이 없습니다."); return
        self._refresh_credits(); self._recalc_totals()
        items = self._snapshot_items(); qtype = self.state.quote_type or "국내"
        if qtype == "중국" and not self.state.cn_info.get("tool"): self._open_options("cn")
        if qtype == "미국"  and not self.state.us_info.get("tool"):  self._open_options("us")
        try:
            checked = self._checked_req_rows()
            if qtype == "국내" and self.state.request_rows and len(checked) >= 2:
                rds = [self.state.request_rows[i] for i in checked if i < len(self.state.request_rows)]
                results = excel_io.generate_quote_multi(self.state, items, rds)
                for rd, path in results:
                    if path:
                        self._add_done(path); self._log(f"저장: {path}")
                        self.state.last_output_dir = os.path.dirname(path)
                        self._save_log(qtype, rd, path, items, "국내 멀티")
                        try:
                            idx = self.state.request_rows.index(rd)
                            it  = self.req_table.item(idx, 2)
                            if it: it.setText("작성완료")
                        except (ValueError, AttributeError): pass
                QMessageBox.information(self, "완료", f"국내 견적서 {len(results)}건 생성 완료"); return
            rd   = self._pick_rd()
            path = excel_io.generate_quote(self.state, qtype, items, rd)
            self._add_done(path); self._log(f"저장: {path}")
            self.state.last_output_dir = os.path.dirname(path)
            self._save_log(qtype, rd, path, items, "단건")
            QMessageBox.information(self, "완료", f"견적서 생성 완료\n{os.path.basename(path)}")
        except Exception as e:
            logger.error("견적서 생성 실패", exc_info=True)
            self._save_error_log(qtype, f"{e}\n\n{traceback.format_exc()}")
            QMessageBox.critical(self, "생성 오류", f"{e}\n\n{traceback.format_exc()}")

    def _generate_cover(self) -> None:
        if not self.state.template_path:
            QMessageBox.warning(self, "오류", "STEP1 통합양식을 먼저 LOAD 하세요."); return
        if not excel_io.COM_AVAILABLE:
            QMessageBox.critical(self, "오류", "Excel COM(win32)이 없습니다."); return
        start  = self.state.last_output_dir or exe_dir()
        paths, _ = QFileDialog.getOpenFileNames(self, "갑지 생성 대상 엑셀파일 (다중)", start, "Excel Files (*.xlsx)")
        if not paths: return
        folder = os.path.commonpath(paths)
        if os.path.isfile(folder): folder = os.path.dirname(folder)
        try:
            out = excel_io.generate_cover(self.state.template_path, folder, paths)
            self.db.insert_simple(LogEntry(
                created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                action="COVER", quote_type="", template_path=s(self.state.template_path),
                output_path=out, message=f"갑지 생성: {len(paths)}개"))
            QMessageBox.information(self, "완료", f"갑지 생성 완료\n{out}")
        except Exception as e:
            logger.error("갑지 생성 실패", exc_info=True)
            self._save_error_log("", f"{e}\n\n{traceback.format_exc()}")
            QMessageBox.critical(self, "오류", f"{e}\n\n{traceback.format_exc()}")

    def _save_log(self, qtype, rd, path, items, msg) -> None:
        try:
            self.db.insert(LogEntry(
                created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                action="QUOTE", quote_type=qtype,
                pr=s(rd.get("D")), itemno=s(rd.get("E")), lineproc=s(rd.get("K")),
                investor=s(rd.get("J")), tool=s(rd.get("Z")),
                template_path=s(self.state.template_path), request_path=s(self.state.request_path),
                output_path=path, message=msg), items)
        except Exception as e: logger.error("로그 저장 실패: %s", e)

    def _save_error_log(self, qtype, msg) -> None:
        try:
            self.db.insert_simple(LogEntry(
                created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                action="ERROR", quote_type=qtype,
                template_path=s(self.state.template_path), request_path=s(self.state.request_path),
                message=msg))
        except Exception as e: logger.error("에러 로그 저장 실패: %s", e)

    def _pick_rd(self) -> Dict[str, Any]:
        checked = self._checked_req_rows()
        if checked and checked[0] < len(self.state.request_rows): return self.state.request_rows[checked[0]]
        if self.state.request_rows: return self.state.request_rows[0]
        return {}

    def _add_done(self, path: str) -> None:
        it = QListWidgetItem(os.path.basename(path)); it.setData(Qt.UserRole, path); self.list_done.addItem(it)

    def _open_done(self, item: QListWidgetItem) -> None:
        path = item.data(Qt.UserRole)
        if path:
            try: os.startfile(path)
            except Exception as e: QMessageBox.critical(self, "오류", str(e))

    def _log(self, text: str) -> None:
        self.log_view.append(text); logger.debug(text)


# ══════════════════════════════════════════════
# 전자서명 페이지
# ══════════════════════════════════════════════

class ESignPage(QWidget):

    SIGN_W = 346
    SIGN_H = 86

    def __init__(self) -> None:
        super().__init__()
        self._code        : str  = ""
        self._signs       : List = []
        self._files       : List[str] = []
        self._base_folder : str  = ""
        self._pdfs        : List[str] = []
        self._cur_file    : int  = 0
        self._cur_page    : int  = 0
        self._pdf_doc            = None
        self._sign_items  : Dict[Tuple[int,int], List[SignatureItem]] = {}
        self._render_sz   : Dict[Tuple[int,int], Tuple[int,int]]     = {}
        self._bg_item            = None
        self._shown_key   : Optional[Tuple[int,int]] = None
        self._build_ui()

    def _build_ui(self) -> None:
        outer = QVBoxLayout(self); outer.setContentsMargins(14,14,14,14); outer.setSpacing(10)
        top = QHBoxLayout()
        self.btn_code   = QPushButton("승인코드 LOAD")
        self.btn_excel  = QPushButton("엑셀 LOAD")
        self.btn_save   = QPushButton("PDF 저장")
        self.lbl_status = QLabel("준비")
        for b in (self.btn_code, self.btn_excel, self.btn_save): top.addWidget(b); top.addSpacing(8)
        top.addStretch(1); top.addWidget(self.lbl_status); outer.addLayout(top)
        mid = QHBoxLayout()
        self.file_list = QListWidget(); self.file_list.setFixedWidth(360); mid.addWidget(self.file_list)
        from PySide6.QtWidgets import QGraphicsScene
        self.scene = QGraphicsScene(self)
        self.view  = PdfView(self); self.view.setScene(self.scene); self.view.setAlignment(Qt.AlignCenter)
        self.view.setFocusPolicy(Qt.StrongFocus); self.view.setFocus(); mid.addWidget(self.view, 1)
        outer.addLayout(mid, 1)
        self.btn_code.clicked.connect(self._load_code)
        self.btn_excel.clicked.connect(self._load_excels)
        self.btn_save.clicked.connect(self._save_pdf)
        self.file_list.currentRowChanged.connect(self._on_select_file)
        self.view.on_prev = self._prev_page; self.view.on_next = self._next_page
        self.view.on_double_click = self._add_sign

    def _load_code(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "승인코드 TXT", "", "Text Files (*.txt)")
        if not path: return
        try: self._code = open(path, encoding="utf-8").read().strip()
        except Exception as e: QMessageBox.critical(self, "오류", str(e)); return
        folder = os.path.dirname(path); exts = (".png",".jpg",".jpeg",".bmp",".webp")
        imgs = sorted([os.path.join(folder,f) for f in os.listdir(folder) if f.lower().endswith(exts)],
                       key=lambda p: os.path.basename(p).lower())
        def _pick(kws):
            for kw in kws:
                for p in imgs:
                    if kw in os.path.basename(p).lower(): return p
            return None
        from PySide6.QtGui import QPixmap
        p1 = _pick(["서명1","sign1","signature1","stamp1"]); p2 = _pick(["서명2","sign2","signature2","stamp2"])
        if not p1 or not p2:
            p1 = p1 or (imgs[0] if imgs else None); p2 = p2 or (imgs[1] if len(imgs)>1 else None)
        if not p1 or not p2:
            QMessageBox.warning(self, "안내", "서명 이미지 2개를 찾지 못했습니다.")
            self._signs = []; self.lbl_status.setText("승인코드 OK, 서명 없음"); return
        pm1 = QPixmap(p1).scaled(self.SIGN_W,self.SIGN_H,Qt.IgnoreAspectRatio,Qt.SmoothTransformation)
        pm2 = QPixmap(p2).scaled(self.SIGN_W,self.SIGN_H,Qt.IgnoreAspectRatio,Qt.SmoothTransformation)
        if pm1.isNull() or pm2.isNull(): QMessageBox.critical(self,"오류","서명 이미지 로드 실패"); return
        self._signs = [pm1, pm2]
        self.lbl_status.setText(f"승인코드 OK / {os.path.basename(p1)}, {os.path.basename(p2)}")
        QMessageBox.information(self, "완료", "승인코드 LOAD 완료\n전자서명 ON")

    def _load_excels(self) -> None:
        if not excel_io.COM_AVAILABLE: QMessageBox.critical(self,"오류","Excel COM이 없습니다."); return
        paths, _ = QFileDialog.getOpenFileNames(self, "엑셀 선택(다중)", "", "Excel Files (*.xlsx)")
        if not paths: return
        paths = sorted(paths, key=lambda p:(0 if "갑지" in os.path.basename(p).lower() else 1, os.path.basename(p).lower()))
        base = os.path.commonpath(paths)
        if os.path.isfile(base): base = os.path.dirname(base)
        self._base_folder = base; self._files = paths; self._pdfs = []; self._sign_items.clear()
        self.file_list.blockSignals(True); self.file_list.clear()
        for p in paths:
            it = QListWidgetItem(os.path.basename(p)); it.setData(Qt.UserRole, p); self.file_list.addItem(it)
        self.file_list.blockSignals(False)
        tmp = ensure_dir(os.path.join(base, "_esign_tmp_pdf"))
        for i, xlsx in enumerate(paths):
            try: self._pdfs.append(excel_io.excel_to_merged_pdf(xlsx, tmp, i+1))
            except Exception as e: logger.error("임시 PDF 실패 (%s): %s", xlsx, e, exc_info=True); self._pdfs.append("")
        self.lbl_status.setText(f"{len(paths)}개 로드 완료")
        if self.file_list.count() > 0: self.file_list.setCurrentRow(0)

    def _on_select_file(self, row: int) -> None:
        if row < 0 or row >= len(self._files): return
        self._cur_file = row; self._cur_page = 0; self._open_pdf(); self._render()

    def _open_pdf(self) -> None:
        if self._pdf_doc:
            try: self._pdf_doc.close()
            except Exception: pass
            self._pdf_doc = None
        if self._cur_file >= len(self._pdfs): return
        pdf_path = self._pdfs[self._cur_file]
        if not pdf_path or not os.path.exists(pdf_path):
            self.lbl_status.setText("표시할 시트 없음(스킵)"); self.scene.clear(); return
        try:
            import fitz; self._pdf_doc = fitz.open(pdf_path)
        except Exception as e: QMessageBox.critical(self, "오류", f"PDF 열기 실패\n{e}")

    def _render(self) -> None:
        if not self._pdf_doc: return
        self._cur_page = max(0, min(self._cur_page, len(self._pdf_doc)-1))
        import fitz
        from PySide6.QtGui import QImage, QPixmap
        page = self._pdf_doc.load_page(self._cur_page)
        vp_w = max(1, self.view.viewport().width())
        zoom = max(2.0, min(4.0, vp_w*2 / max(1.0, float(page.rect.width))))
        pix  = page.get_pixmap(matrix=fitz.Matrix(zoom,zoom), alpha=False)
        img  = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format_RGB888)
        pm   = QPixmap.fromImage(img.copy())
        self._render_sz[(self._cur_file, self._cur_page)] = (pm.width(), pm.height())
        if self._shown_key is not None:
            for it in list(self._sign_items.get(self._shown_key, [])):
                try:
                    if it.scene() is self.scene: self.scene.removeItem(it)
                except RuntimeError: pass
        if self._bg_item is not None:
            try:
                if self._bg_item.scene() is self.scene: self.scene.removeItem(self._bg_item)
            except Exception: pass
        from PySide6.QtWidgets import QGraphicsPixmapItem
        bg = QGraphicsPixmapItem(pm); bg.setZValue(0); bg.setAcceptedMouseButtons(Qt.NoButton)
        self.scene.addItem(bg); self._bg_item = bg; self._shown_key = (self._cur_file, self._cur_page)
        for it in list(self._sign_items.get(self._shown_key, [])):
            try: self.scene.addItem(it); it.setZValue(10)
            except RuntimeError: pass
        self.scene.setSceneRect(bg.boundingRect()); self.view.resetTransform()
        scale = vp_w / max(1, pm.width()); self.view.scale(scale, scale); self.view.setFocus()
        self.view.verticalScrollBar().setValue(self.view.verticalScrollBar().minimum())
        self.lbl_status.setText(f"파일 {self._cur_file+1}/{len(self._files)} / 페이지 {self._cur_page+1}/{len(self._pdf_doc)}")

    def _next_page(self) -> None:
        if not self._pdf_doc: return
        if self._cur_page+1 < len(self._pdf_doc): self._cur_page += 1; self._render()
        elif self._cur_file+1 < len(self._files): self.file_list.setCurrentRow(self._cur_file+1)

    def _prev_page(self) -> None:
        if not self._pdf_doc: return
        if self._cur_page-1 >= 0: self._cur_page -= 1; self._render()
        elif self._cur_file-1 >= 0:
            self.file_list.setCurrentRow(self._cur_file-1)
            if self._pdf_doc: self._cur_page = max(0, len(self._pdf_doc)-1); self._render()

    def _add_sign(self, scene_pos: QPointF) -> None:
        if not self._signs:
            QMessageBox.information(self, "안내", "승인코드 LOAD 후 서명 이미지가 필요합니다."); return
        if not self._pdf_doc: return
        dlg = PasswordDialog(self, self._code)
        if dlg.exec() != QDialog.Accepted or not dlg.verified: return
        from PySide6.QtWidgets import QApplication as _App
        idx = min(1 if (_App.keyboardModifiers() & Qt.ShiftModifier) else 0, len(self._signs)-1)
        item = SignatureItem(self._signs[idx], self._cur_page); item.setZValue(10)
        item.setPos(QPointF(scene_pos.x()-self.SIGN_W/2, scene_pos.y()-self.SIGN_H/2))
        self.scene.addItem(item)
        self._sign_items.setdefault((self._cur_file, self._cur_page), []).append(item)
        self.scene.update()

    def _save_pdf(self) -> None:
        if not self._pdf_doc or not self._files:
            QMessageBox.information(self, "안내", "먼저 엑셀을 LOAD 하세요."); return
        folder_name = os.path.basename(self._base_folder.rstrip("\\/"))
        out = unique_path(os.path.join(self._base_folder, f"{folder_name}.pdf"))
        try:
            self._build_pdf(out)
            QMessageBox.information(self, "완료", f"저장 완료:\n{out}"); self.lbl_status.setText("PDF 저장 완료")
        except Exception as e:
            logger.error("PDF 저장 실패", exc_info=True)
            QMessageBox.critical(self, "오류", f"{e}\n\n{traceback.format_exc()}")

    def _build_pdf(self, out: str) -> None:
        import fitz
        from PySide6.QtCore import QBuffer, QByteArray, QIODevice
        final = fitz.open()
        for fi, pdf_path in enumerate(self._pdfs):
            if not pdf_path or not os.path.exists(pdf_path): continue
            src = fitz.open(pdf_path)
            try:
                for pno in range(len(src)):
                    final.insert_pdf(src, from_page=pno, to_page=pno); dst = final[-1]
                    signs = self._sign_items.get((fi,pno), [])
                    if not signs: continue
                    iw, ih = self._render_sz.get((fi,pno), (0,0))
                    if iw <= 0 or ih <= 0:
                        pix = src.load_page(pno).get_pixmap(matrix=fitz.Matrix(2.0,2.0), alpha=False)
                        iw, ih = pix.width, pix.height
                    pr = dst.rect
                    for it in list(signs):
                        try: x, y = float(it.pos().x()), float(it.pos().y())
                        except RuntimeError: continue
                        rect = fitz.Rect((x/iw)*pr.width, (y/ih)*pr.height,
                                          ((x+self.SIGN_W)/iw)*pr.width, ((y+self.SIGN_H)/ih)*pr.height)
                        ba = QByteArray(); buf = QBuffer(ba); buf.open(QIODevice.WriteOnly)
                        it.pixmap().save(buf, "PNG"); dst.insert_image(rect, stream=ba.data())
            finally: src.close()
        final.save(out); final.close()


# ══════════════════════════════════════════════
# 견적 LOG 페이지
# ══════════════════════════════════════════════

class LogPage(QWidget):

    _COL_HEADERS = ["", "ID", "생성일", "견적타입", "PR", "항번", "설비호기", "파일"]

    def __init__(self, db: LogDB) -> None:
        super().__init__(); self.db = db; self._build_ui(); self.reload()

    def _build_ui(self) -> None:
        outer = QVBoxLayout(self); outer.setContentsMargins(14,14,14,14); outer.setSpacing(10)
        bar = QHBoxLayout()
        self.btn_all    = QPushButton("전체선택")
        self.btn_none   = QPushButton("전체해제")
        self.btn_reload = QPushButton("새로고침")
        bar.addWidget(self.btn_all); bar.addWidget(self.btn_none); bar.addWidget(self.btn_reload); bar.addSpacing(12)
        self.ed_pr    = QLineEdit(); self.ed_pr.setPlaceholderText("PR 검색");  self.ed_pr.setFixedWidth(140)
        self.ed_tool  = QLineEdit(); self.ed_tool.setPlaceholderText("설비호기"); self.ed_tool.setFixedWidth(160)
        self.ed_item  = QLineEdit(); self.ed_item.setPlaceholderText("품목(규격)")
        self.cb_qtype = QComboBox(); self.cb_qtype.addItems(["","국내","중국","미국"]); self.cb_qtype.setFixedWidth(90)
        self.cb_cat   = QComboBox(); self.cb_cat.addItems([""]); self.cb_cat.setFixedWidth(110)
        self.btn_export = QPushButton("합산 Export")
        for lbl, w in [("PR",self.ed_pr),("설비호기",self.ed_tool),("품목",self.ed_item),("견적타입",self.cb_qtype),("분류(합산)",self.cb_cat)]:
            bar.addWidget(QLabel(lbl)); bar.addWidget(w); bar.addSpacing(6)
        bar.addWidget(self.btn_export); bar.addStretch(1); outer.addLayout(bar)
        self.tbl_logs = self._make_table(8, self._COL_HEADERS)
        h = self.tbl_logs.horizontalHeader()
        for i in range(7): h.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        h.setSectionResizeMode(7, QHeaderView.Stretch)
        outer.addWidget(self.tbl_logs, 3)
        outer.addWidget(bold_label("체크된 견적서 품목 합산 (금액 내림차순)", size=10))
        self.tbl_items = self._make_table(5, ["분류","품목(규격)","총수량","단가(추정)","총금액"])
        hi = self.tbl_items.horizontalHeader(); hi.setStretchLastSection(False); hi.setSectionResizeMode(QHeaderView.Fixed)
        for col, w in enumerate([90,820,90,120,505]): self.tbl_items.setColumnWidth(col, w)
        outer.addWidget(self.tbl_items, 2)
        self.btn_all.clicked.connect(lambda: self._set_all(True))
        self.btn_none.clicked.connect(lambda: self._set_all(False))
        self.btn_reload.clicked.connect(self.reload); self.btn_export.clicked.connect(self._export)
        for w in (self.ed_pr, self.ed_tool, self.ed_item): w.textChanged.connect(self.reload)
        self.cb_qtype.currentIndexChanged.connect(self.reload)
        self.cb_cat.currentIndexChanged.connect(self._refresh_items)
        self.tbl_logs.cellDoubleClicked.connect(self._open_file)

    @staticmethod
    def _make_table(cols: int, headers: List[str]):
        from PySide6.QtWidgets import QTableWidget
        t = QTableWidget(0, cols); t.setHorizontalHeaderLabels(headers)
        t.verticalHeader().setVisible(False); t.setEditTriggers(QTableWidget.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.setSelectionMode(QAbstractItemView.SingleSelection); t.setWordWrap(False); return t

    def reload(self) -> None:
        logs = sorted(
            self.db.fetch_logs(pr_kw=self.ed_pr.text().strip(), tool_kw=self.ed_tool.text().strip(),
                               item_kw=self.ed_item.text().strip(), quote_type=self.cb_qtype.currentText().strip()),
            key=lambda x: (s(x.get("pr")), s(x.get("itemno")), s(x.get("created_at"))))
        self.tbl_logs.blockSignals(True); self.tbl_logs.setRowCount(len(logs))
        for r, lg in enumerate(logs):
            log_id = int(lg["id"]); cb = QCheckBox(); cb.stateChanged.connect(self._refresh_items)
            self.tbl_logs.setCellWidget(r, 0, cb)
            it_id = QTableWidgetItem(str(log_id)); it_id.setData(Qt.UserRole, log_id)
            self.tbl_logs.setItem(r, 1, it_id)
            for col, key in [(2,"created_at"),(3,"quote_type"),(4,"pr"),(5,"itemno"),(6,"tool"),(7,"output_path")]:
                self.tbl_logs.setItem(r, col, QTableWidgetItem(s(lg.get(key))))
        self.tbl_logs.blockSignals(False); self._refresh_cat_combo(); self._refresh_items()

    def _refresh_cat_combo(self) -> None:
        all_ids = [self._log_id(r) for r in range(self.tbl_logs.rowCount())]
        all_ids = [i for i in all_ids if i is not None]
        cats = set(self.db.fetch_categories(all_ids)) if all_ids else set()
        cur = self.cb_cat.currentText()
        self.cb_cat.blockSignals(True); self.cb_cat.clear(); self.cb_cat.addItem("")
        for c in sorted(cats): self.cb_cat.addItem(c)
        if cur in cats: self.cb_cat.setCurrentText(cur)
        self.cb_cat.blockSignals(False)

    def _log_id(self, row: int) -> Optional[int]:
        it = self.tbl_logs.item(row, 1)
        if not it: return None
        try: return int(it.data(Qt.UserRole) or it.text())
        except Exception: return None

    def _checked_ids(self) -> List[int]:
        ids = []
        for r in range(self.tbl_logs.rowCount()):
            cb = self.tbl_logs.cellWidget(r, 0)
            if isinstance(cb, QCheckBox) and cb.isChecked():
                lid = self._log_id(r)
                if lid is not None: ids.append(lid)
        return ids

    def _set_all(self, checked: bool) -> None:
        self.tbl_logs.blockSignals(True)
        for r in range(self.tbl_logs.rowCount()):
            cb = self.tbl_logs.cellWidget(r, 0)
            if isinstance(cb, QCheckBox): cb.blockSignals(True); cb.setChecked(checked); cb.blockSignals(False)
        self.tbl_logs.blockSignals(False); self._refresh_items()

    def _refresh_items(self) -> None:
        ids = self._checked_ids()
        if not ids: self.tbl_items.setRowCount(0); return
        items = self.db.fetch_items(ids)
        cat_f = self.cb_cat.currentText().strip()
        if cat_f: items = [it for it in items if s(it.get("cat")) == cat_f]
        agg: Dict[Tuple[str,str], Dict] = {}
        for it in items:
            cat = s(it.get("cat")); spec = s(it.get("spec"))
            if not spec: continue
            qty = to_float(it.get("qty")); price = to_float(it.get("price"))
            amt = to_float(it.get("amt"), qty*price); key = (cat, spec)
            a = agg.setdefault(key, {"qty":0.0,"amt":0.0,"pn":0.0,"pd":0.0})
            a["qty"] += qty; a["amt"] += amt
            if qty > 0: a["pn"] += price*qty; a["pd"] += qty
        rows = [(cat, spec, a["qty"], a["pn"]/a["pd"] if a["pd"] else 0.0, a["amt"]) for (cat,spec),a in agg.items()]
        rows.sort(key=lambda x: x[4], reverse=True)
        self.tbl_items.setRowCount(len(rows))
        for r, (cat,spec,qty,pe,amt) in enumerate(rows):
            for col, val in [(0,cat),(1,spec),(2,fmt_qty(qty)),(3,fmt_krw(pe) if pe else ""),(4,fmt_krw(amt))]:
                self.tbl_items.setItem(r, col, QTableWidgetItem(val))

    def _export(self) -> None:
        if self.tbl_items.rowCount() == 0: QMessageBox.information(self,"안내","체크된 로그가 없습니다."); return
        default = f"품목합산_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "저장", default, "Excel Files (*.xlsx)")
        if not path: return
        if not path.lower().endswith(".xlsx"): path += ".xlsx"
        try:
            from openpyxl import Workbook; from openpyxl.utils import get_column_letter
            wb = Workbook(); ws = wb.active; ws.title = "품목합산"
            ws.append(["분류","품목(규격)","총수량","단가(추정)","총금액"])
            for r in range(self.tbl_items.rowCount()):
                ws.append([(self.tbl_items.item(r,c).text() if self.tbl_items.item(r,c) else "") for c in range(5)])
            for col in range(1,6):
                max_len = max((len(str(ws.cell(row=rr,column=col).value or "")) for rr in range(1,ws.max_row+1)), default=10)
                ws.column_dimensions[get_column_letter(col)].width = min(max_len+2, 80)
            wb.save(path); QMessageBox.information(self, "완료", f"저장 완료:\n{path}")
        except Exception as e: logger.error("Export 실패", exc_info=True); QMessageBox.critical(self,"오류",str(e))

    def _open_file(self, row: int, _col: int) -> None:
        it = self.tbl_logs.item(row, 7)
        if not it: return
        path = it.text().strip()
        if not path: return
        try: os.startfile(path)
        except Exception:
            d = os.path.dirname(path)
            if d and os.path.isdir(d):
                try: os.startfile(d)
                except Exception as e: logger.warning("파일 열기 실패: %s", e)
